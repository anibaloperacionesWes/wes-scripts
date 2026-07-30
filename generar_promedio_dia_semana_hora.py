"""
Lee la hoja DETALLE (node_id, fecha, hora, m³/hora) y genera Excel con:
- promedio m³/h por hora y por día de la semana
- promedio de total diario por día de semana
- conteo de muestras por celda

Salida por defecto en la carpeta ``calculo de regulaciones/``.

Uso:
  python generar_promedio_dia_semana_hora.py -i "ruta\\archivo.xlsx"
"""
from __future__ import annotations

import argparse
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parent
OUT_DIR_DEFAULT = ROOT / "calculo de regulaciones"

WD_NAMES = (
    "Lunes",
    "Martes",
    "Miercoles",
    "Jueves",
    "Viernes",
    "Sabado",
    "Domingo",
)


def _parse_date(v) -> date:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()


def _parse_hour(v) -> int:
    if isinstance(v, datetime):
        return v.hour
    s = str(v)
    if ":" in s:
        return int(s.split(":")[0])
    return int(float(s))


def _parse_m3(v) -> float:
    if v is None:
        return 0.0
    s = str(v).strip().replace(",", ".")
    return float(s) if s else 0.0


def generar(path_in: Path, path_out: Path) -> Path:
    wb = load_workbook(path_in, data_only=True)
    if "DETALLE" not in wb.sheetnames:
        wb.close()
        raise KeyError("El Excel debe tener hoja DETALLE")
    ws = wb["DETALLE"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    sum_by = defaultdict(float)
    count_by = defaultdict(int)
    sum_daily = defaultdict(float)
    count_days = defaultdict(set)

    for r in rows:
        if not r or r[0] is None:
            continue
        d = _parse_date(r[2])
        h = _parse_hour(r[3])
        m3 = _parse_m3(r[4])
        wd = d.weekday()
        sum_by[(wd, h)] += m3
        count_by[(wd, h)] += 1
        sum_daily[(wd, d)] += m3
        count_days[wd].add(d)

    out_wb = Workbook()
    ws1 = out_wb.active
    ws1.title = "Promedio_dia_hora"
    headers = ["Hora"] + list(WD_NAMES)
    for c, h in enumerate(headers, 1):
        ws1.cell(1, c, h)
    for h in range(24):
        ws1.cell(h + 2, 1, f"{h:02d}:00")
        for wd in range(7):
            n = count_by.get((wd, h), 0)
            avg = (sum_by[(wd, h)] / n) if n else 0.0
            ws1.cell(h + 2, wd + 2, round(avg, 4))

    ws2 = out_wb.create_sheet("Promedio_total_diario")
    ws2.append(["Dia_semana", "Promedio_m3_dia", "Dias_muestra"])
    for wd in range(7):
        dias = sorted(count_days.get(wd, set()))
        if dias:
            vals = [sum_daily[(wd, d)] for d in dias]
            avg = sum(vals) / len(vals)
            n = len(vals)
        else:
            avg = 0.0
            n = 0
        ws2.append([WD_NAMES[wd], round(avg, 4), n])

    ws3 = out_wb.create_sheet("Muestras_por_hora")
    ws3.append(["Hora"] + list(WD_NAMES))
    for h in range(24):
        ws3.append([f"{h:02d}:00"] + [count_by.get((wd, h), 0) for wd in range(7)])

    for wsx in (ws1, ws2, ws3):
        for col in wsx.columns:
            maxlen = 0
            letter = col[0].column_letter
            for cell in col:
                v = "" if cell.value is None else str(cell.value)
                maxlen = max(maxlen, len(v))
            wsx.column_dimensions[letter].width = min(max(12, maxlen + 2), 28)

    path_out.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(path_out)
    return path_out


def main() -> int:
    ap = argparse.ArgumentParser(description="Promedios por día de semana y hora desde DETALLE")
    ap.add_argument(
        "-i",
        "--excel",
        type=Path,
        default=Path(r"c:\Users\aniba\Downloads\e832e010-4508-4e0b-b17d-c3e05da2e0f2.xlsx"),
        help="Excel con hoja DETALLE",
    )
    ap.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Ruta .xlsx de salida (default: calculo de regulaciones/promedio_...)",
    )
    args = ap.parse_args()
    src = args.excel.resolve()
    if not src.is_file():
        print(f"[ERROR] No existe: {src}")
        return 1
    out = args.output
    if out is None:
        stem = src.stem[:40] if len(src.stem) > 40 else src.stem
        out = OUT_DIR_DEFAULT / f"promedio_por_dia_semana_y_hora_{stem}.xlsx"
    else:
        out = out.resolve()

    p = generar(src, out)
    print(p.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

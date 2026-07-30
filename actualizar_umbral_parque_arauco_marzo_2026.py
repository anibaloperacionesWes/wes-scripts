"""
Actualiza el Excel `reports/HORARIOS CONTROL NOCTURNO.xlsx` con el promedio de consumo
entre 00:01 y 05:59 para todos los puntos de Parque Arauco (companyId=000025).

Metodología:
- Para cada nodo se consultan datos horarios por día usando:
  /wes/api/acl-node/v1/nodes/{node_id}/dates.measures.csv?start=DDMMYYYY&end=DDMMYYYY
- Para la ventana 00:01 a 05:59 se usa la misma conversión que usa el reporte de control:
  "00:01 a 05:59" -> horas [0,1,2,3,4,5]
- Se calcula el promedio mensual como:
  promedio = (suma diaria de esas horas / 6) promedio de todos los días disponibles

Salida:
- Se guarda como archivo nuevo (no sobrescribe el Excel original).
"""

from __future__ import annotations

import os
import shutil
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import requests
from openpyxl import load_workbook


ENTITY_BASE_URL = "http://104.248.53.141:7001/wes/api/acl-entities/v1"
NODE_BASE_URL = "http://104.248.53.141:7003/wes/api/acl-node/v1"


DEFAULT_EXCEL = Path("reports") / "HORARIOS CONTROL NOCTURNO.xlsx"
COMPANY_ID_PARQUE_ARAUCO = "000025"

# Ventana solicitada: 00:01 a 05:59 => horas [0..5]
HOURS_EVAL = [0, 1, 2, 3, 4, 5]
HORARIO_TEXT = "00:01 a 05:59"

MONTH_START = datetime(2026, 3, 1)
MONTH_END = datetime(2026, 3, 31)


@dataclass(frozen=True)
class NodeInfo:
    node_id: str
    node_name: str


def excel_abierto() -> bool:
    """
    Detecta si Excel está abierto (Windows). Se usa para decidir si sobrescribimos o no.
    """
    if os.environ.get("WES_FORCE_OVERWRITE", "0").strip() == "1":
        return False
    if sys.platform != "win32":
        return False
    try:
        import subprocess

        out = subprocess.check_output(
            ["tasklist", "/FI", "IMAGENAME eq excel.exe"],
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="ignore",
        )
        return "excel.exe" in out.lower()
    except Exception:
        return False


def obtener_nodos_parque_arauco() -> List[NodeInfo]:
    url = f"{ENTITY_BASE_URL}/companies/{COMPANY_ID_PARQUE_ARAUCO}"
    resp = requests.get(url, timeout=20)
    resp.raise_for_status()
    data = resp.json()
    nodes = data.get("nodes", [])
    out: List[NodeInfo] = []
    for n in nodes:
        node_id = str(n.get("nodeId") or "").strip()
        node_name = str(n.get("name") or "").strip()
        if node_id and node_name:
            out.append(NodeInfo(node_id=node_id, node_name=node_name))
    if not out:
        raise RuntimeError("No se encontraron nodos para Parque Arauco (000025)")
    return out


def obtener_horarios_dia(node_id: str, dia: datetime, session: requests.Session) -> Dict[int, float]:
    """
    Devuelve dict {hour:int -> value:float} para un día.
    """
    date_str = dia.strftime("%d%m%Y")
    url = f"{NODE_BASE_URL}/nodes/{node_id}/dates.measures.csv"
    resp = session.get(url, params={"start": date_str, "end": date_str}, timeout=20)
    resp.raise_for_status()
    lines = resp.text.strip().splitlines()
    hourly: Dict[int, float] = {}
    for line in lines[1:]:  # skip header TIME,VALUE
        if not line.strip():
            continue
        parts = line.split(",")
        if len(parts) < 2:
            continue
        time_str = parts[0].strip()
        value_str = parts[1].strip()
        if "T" not in time_str:
            continue
        try:
            hour = int(time_str.split("T")[1].split(":")[0])
            value = float(value_str)
        except Exception:
            continue
        hourly[hour] = value
    return hourly


def promedio_mes_parque_arauco_por_nodo(
    node: NodeInfo,
    hour_list: Iterable[int],
    month_start: datetime,
    month_end: datetime,
    session: requests.Session,
    max_reintentos_dia: int = 2,
) -> float:
    days_ok = 0
    sum_day_avg = 0.0
    hour_list = list(hour_list)
    if not hour_list:
        raise ValueError("hour_list no puede estar vacío")
    hour_count = len(hour_list)

    cur = month_start
    while cur <= month_end:
        ok = False
        for attempt in range(max_reintentos_dia + 1):
            try:
                hourly = obtener_horarios_dia(node.node_id, cur, session)
                day_sum = 0.0
                for h in hour_list:
                    day_sum += float(hourly.get(h, 0.0))

                day_avg = day_sum / float(hour_count)
                sum_day_avg += day_avg
                days_ok += 1
                ok = True
                break
            except Exception:
                if attempt >= max_reintentos_dia:
                    ok = False
                else:
                    time.sleep(1.5 * (attempt + 1))

        # Si falla el día, simplemente no se cuenta
        cur += timedelta(days=1)

    if days_ok == 0:
        raise RuntimeError(f"Sin datos para el nodo {node.node_id}")
    return sum_day_avg / float(days_ok)


def _find_header_columns(ws) -> Dict[str, int]:
    """
    Busca columnas por encabezado (fila 2) para:
    CLIENTE, NOMBRE DEL COLEGIO O LICEO, ID, HORARIO DE CORTE, UMBRAL DE ALERTA.
    """
    header_row = 2
    wanted = {
        "CLIENTE": None,
        "NOMBRE": None,
        "ID": None,
        "HORARIO": None,
        "UMBRAL": None,
    }
    for col in range(1, ws.max_column + 1):
        v = ws.cell(header_row, col).value
        if v is None:
            continue
        sv = str(v).strip().upper()
        if sv == "CLIENTE":
            wanted["CLIENTE"] = col
        elif "ID" == sv:
            wanted["ID"] = col
        elif "HORARIO DE CORTE" in sv or "HORARIO" in sv:
            wanted["HORARIO"] = col
        elif "UMBRAL" in sv or "ALERTA" in sv:
            wanted["UMBRAL"] = col
        elif "NOMBRE" in sv:
            wanted["NOMBRE"] = col

    if any(wanted[k] is None for k in wanted):
        raise ValueError(f"Encabezados no encontrados en {ws.title}: {wanted}")
    return wanted  # type: ignore[return-value]


def _format_umbral(valor: float, decimales: int = 2) -> str:
    s = f"{valor:.{decimales}f}".replace(".", ",")
    return f">{s}"


def actualizar_excel(
    excel_path: Path,
    nodes: List[NodeInfo],
    avg_by_node_id: Dict[str, float],
    overwrite_schedule: bool = True,
) -> Path:
    if not excel_path.exists():
        raise FileNotFoundError(str(excel_path))

    excel_tmp = None
    try:
        # Copia temporal si el archivo está siendo usado por Excel/OneDrive
        if excel_abierto():
            suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
            out_path = excel_path.with_name(excel_path.stem + f"_PA_MARZO2026_{suffix}" + excel_path.suffix)
            shutil.copy2(excel_path, out_path)
            excel_tmp = out_path
        else:
            # Igual usamos copia temporal para que el original no quede corrupto si algo falla
            suffix = datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_tmp = excel_path.with_name(excel_path.stem + f"_PA_MARZO2026_{suffix}" + excel_path.suffix)
            shutil.copy2(excel_path, excel_tmp)

        wb = load_workbook(excel_tmp, data_only=True)
        ws = wb[wb.sheetnames[0]]

        cols = _find_header_columns(ws)
        id_col = cols["ID"]
        cliente_col = cols["CLIENTE"]
        nombre_col = cols["NOMBRE"]
        horario_col = cols["HORARIO"]
        umbral_col = cols["UMBRAL"]

        # Índice existente por node_id
        index: Dict[str, int] = {}
        for r in range(3, ws.max_row + 1):
            v = ws.cell(r, id_col).value
            if v is None:
                continue
            node_id = str(v).strip()
            if node_id:
                index[node_id] = r

        # Actualizar o agregar filas
        next_row = ws.max_row + 1
        for node in nodes:
            if node.node_id not in avg_by_node_id:
                continue
            avg = avg_by_node_id[node.node_id]
            row = index.get(node.node_id)
            if row is None:
                row = next_row
                next_row += 1

                ws.cell(row, cliente_col).value = "PARQUE ARAUCO"
                ws.cell(row, nombre_col).value = node.node_name
                ws.cell(row, id_col).value = node.node_id

            if overwrite_schedule:
                ws.cell(row, horario_col).value = HORARIO_TEXT
            ws.cell(row, umbral_col).value = _format_umbral(avg)

        wb.save(excel_tmp)
        return Path(excel_tmp)
    finally:
        # No limpiamos el temporal porque excel_tmp ya es el output.
        pass


def main() -> None:
    excel_path = Path(os.environ.get("WES_EXCEL_PATH", str(DEFAULT_EXCEL)))
    overwrite_schedule = os.environ.get("WES_OVERWRITE_SCHEDULE", "1").strip() != "0"
    threads = int(os.environ.get("WES_THREADS", "3"))

    print("=" * 70)
    print("ACTUALIZAR UMBRAL PARQUE ARAUCO - MARZO 2026")
    print("=" * 70)
    print(f"Excel: {excel_path}")
    print(f"Parque Arauco companyId: {COMPANY_ID_PARQUE_ARAUCO}")
    print(f"Ventana horas (eval): {HOURS_EVAL} => {HORARIO_TEXT}")
    print(f"Periodo: {MONTH_START:%Y-%m-%d} a {MONTH_END:%Y-%m-%d}")
    print(f"Hilos: {threads}")
    print()

    nodes = obtener_nodos_parque_arauco()
    print(f"[OK] Nodos Parque Arauco: {len(nodes)}")

    # Cálculo
    avg_by_node: Dict[str, float] = {}
    session = requests.Session()

    # Nota: para evitar saturar la API, calculamos en serie por nodo por defecto.
    # Si quieres paralelizar, podemos ajustar a ThreadPoolExecutor con límite.
    # Por ahora, dejamos secuencial (más estable).
    for i, node in enumerate(nodes, start=1):
        print(f"[{i}/{len(nodes)}] Calculando nodo {node.node_id} - {node.node_name} ...")
        avg = promedio_mes_parque_arauco_por_nodo(
            node=node,
            hour_list=HOURS_EVAL,
            month_start=MONTH_START,
            month_end=MONTH_END,
            session=session,
        )
        avg_by_node[node.node_id] = avg
        print(f"   Promedio: {avg:.4f} => umbral {_format_umbral(avg)}")

    print()
    out_path = actualizar_excel(
        excel_path=excel_path,
        nodes=nodes,
        avg_by_node_id=avg_by_node,
        overwrite_schedule=overwrite_schedule,
    )
    print(f"[OK] Excel actualizado: {out_path}")


if __name__ == "__main__":
    main()


"""
Dado un listado de correos, consulta WES (/users?email=) y determina a qué mall(es)
de Parque Arauco (000025) tiene acceso el usuario, según allowedNodes.

Salida: Excel (.xlsx) con columnas:
  - Email_input
  - Email_api (si se resolvió)
  - Nombre
  - Malls_PA (lista o "Todos")
  - Nodos_PA_count
  - Observacion
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Set, Tuple

import requests

from generar_reporte_word import get_mall_name_for_parque_arauco

ENTITY_BASE = "http://104.248.53.141:7001/wes/api/acl-entities/v1"
COMPANY_PA = "000025"
PREFIX_PA = "000025-"


def _cargar_nodos_pa() -> Dict[str, str]:
    r = requests.get(f"{ENTITY_BASE}/companies/{COMPANY_PA}", timeout=60)
    r.raise_for_status()
    out: Dict[str, str] = {}
    for n in r.json().get("nodes") or []:
        nid = str(n.get("nodeId", "") or "").strip()
        if not nid.startswith(PREFIX_PA):
            continue
        out[nid] = str(n.get("name", "") or "").strip()
    return out


def _malls_total_pa(nodes_pa: Dict[str, str]) -> Set[str]:
    malls: Set[str] = set()
    for nid, nombre in nodes_pa.items():
        mall = get_mall_name_for_parque_arauco(nid, nombre).strip()
        if mall:
            malls.add(mall)
    return malls


def _fetch_user(email: str) -> Tuple[Optional[dict], str]:
    """
    Retorna (user_json|None, observacion).
    Si el correo tiene coma en dominio (typo común), intenta reemplazar ',' -> '.' una vez.
    """
    email = email.strip()
    if not email:
        return None, "Email vacío"

    def _get(e: str) -> requests.Response:
        return requests.get(f"{ENTITY_BASE}/users", params={"email": e}, timeout=30)

    try:
        r = _get(email)
        if r.status_code == 200:
            u = r.json()
            if isinstance(u, dict):
                u.pop("password", None)
            return u, ""
        if r.status_code == 404:
            if "," in email:
                alt = email.replace(",", ".")
                r2 = _get(alt)
                if r2.status_code == 200:
                    u = r2.json()
                    if isinstance(u, dict):
                        u.pop("password", None)
                    return u, f"Email corregido para API: {alt}"
                if r2.status_code == 404:
                    return None, "No existe en API (404)"
                return None, f"HTTP {r2.status_code} al consultar email corregido"
            return None, "No existe en API (404)"
        return None, f"HTTP {r.status_code}"
    except requests.RequestException as e:
        return None, f"Error HTTP: {e}"


def _malls_usuario_pa(user: dict, nodes_pa: Dict[str, str]) -> Tuple[Set[str], int]:
    allowed = user.get("allowedNodes") or []
    if not isinstance(allowed, list):
        return set(), 0
    malls: Set[str] = set()
    c = 0
    for nid in allowed:
        if not isinstance(nid, str) or not nid.startswith(PREFIX_PA):
            continue
        c += 1
        nombre = nodes_pa.get(nid, "")
        mall = get_mall_name_for_parque_arauco(nid, nombre).strip()
        if mall:
            malls.add(mall)
    return malls, c


def _normalizar_email_cell(s: str) -> str:
    # Quitar espacios y normalizar underscores/espacios que vienen desde Excel/copia.
    s = (s or "").strip()
    s = s.replace(" ", "")
    return s


def _leer_emails_desde_archivo(path: Path) -> List[str]:
    emails: List[str] = []
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        emails.append(_normalizar_email_cell(line))
    return emails


def main() -> int:
    parser = argparse.ArgumentParser(description="Usuarios -> malls PA por allowedNodes")
    parser.add_argument(
        "--emails",
        type=Path,
        default=Path("reports")
        / "Parque_Arauco"
        / "Reportes_agregados_reunion_abril_2026"
        / "emails_pa_linkes_parauco.txt",
        help="Archivo de correos (1 por línea)",
    )
    parser.add_argument(
        "--salida-base",
        type=Path,
        default=Path("reports")
        / "Parque_Arauco"
        / "Reportes_agregados_reunion_abril_2026"
        / "usuarios_pa_malls",
        help="Ruta base sin extensión (se generan .xlsx y .csv)",
    )
    args = parser.parse_args()

    if not args.emails.exists():
        print(f"[ERROR] No existe archivo de emails: {args.emails}")
        return 1

    emails = _leer_emails_desde_archivo(args.emails)
    if not emails:
        print("[ERROR] Archivo de emails vacío.")
        return 1

    nodes_pa = _cargar_nodos_pa()
    malls_total = _malls_total_pa(nodes_pa)

    filas: List[dict] = []
    for email in emails:
        u, obs = _fetch_user(email)
        if not u:
            # El usuario no existe en API: se omite del Excel final.
            continue

        malls, n_pa = _malls_usuario_pa(u, nodes_pa)
        malls_txt = ", ".join(sorted(malls, key=lambda x: x.casefold()))
        if malls and malls_total and malls == malls_total:
            malls_txt = "Todos"

        filas.append(
            {
                "Email_input": email,
                "Email_api": str(u.get("username") or "").strip(),
                "Nombre": f"{u.get('name','')} {u.get('lastName','')}".strip(),
                "Malls_PA": malls_txt,
                "Nodos_PA_count": n_pa,
                "Observacion": obs,
            }
        )

    out_base: Path = args.salida_base
    out_base.parent.mkdir(parents=True, exist_ok=True)
    out_xlsx = out_base.with_suffix(".xlsx")

    cols = [
        "Email_input",
        "Email_api",
        "Nombre",
        "Malls_PA",
        "Nodos_PA_count",
        "Observacion",
    ]

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Usuarios_PA_Malls"
        for c, titulo in enumerate(cols, 1):
            ws.cell(row=1, column=c, value=titulo)
        for r, fila in enumerate(filas, 2):
            for c, key in enumerate(cols, 1):
                ws.cell(row=r, column=c, value=fila.get(key, ""))
        widths = (28, 28, 26, 32, 14, 60)
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        wb.save(out_xlsx)
    except ImportError:
        print("[ERROR] openpyxl no instalado; no se pudo generar el .xlsx.")
        return 1

    print(f"[OK] {len(filas)} usuario(s) existente(s) en API (filtrados)")
    if out_xlsx.exists():
        print(f"     XLSX: {out_xlsx.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


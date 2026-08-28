# -*- coding: utf-8 -*-
"""
Puntos CPA (monitoreo con control): sitios con horario programado en el Excel
de Drive «Horarios de contron hidrico clientes wes».

El resto de puntos activos WES se consideran solo monitoreo.
"""

from __future__ import annotations

import io
import re
import unicodedata
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Set, Tuple

DRIVE_HORARIOS_ID = "1eM03xh4Pmqx5YTTKWiOku8_NT0115ausiWOKNaLc8Lk"
CACHE_XLSX = Path(__file__).resolve().parent / "reports" / "Horarios_control_hidrico_clientes_wes.xlsx"

# Nombres del Excel de horarios (typos incluidos) → nodeId WES.
CPA_ALIAS = {
    "171 antonio hermidas fabres": "000008-01",
    "antonio hermidas fabres": "000008-01",
    "antonio hermida fabres": "000008-01",
    "carlos fernandes pena": "000008-03",
    "carlos fernandez pena": "000008-03",
    "colegio tobalaba": "000008-04",
    "tobalaba": "000008-04",
    "colegio santa maria": "000008-05",
    "santa maria": "000008-05",
    "luis arrieta canas": "000008-06",
    "luis arrieta cana": "000008-06",
    "colegio erasmo escala": "000008-07",
    "erasmo escala": "000008-07",
    "juan bautista pastene": "000008-09",
    "juan bautista pasten": "000008-09",
    "juan bautista pastenes": "000008-09",
    "matilde huichinavas": "000008-10",
    "matilde huici navas": "000008-10",
    "ce valle germoso": "000008-11",
    "ce valle hermoso": "000008-11",
    "valle hermoso": "000008-11",
    "union nacional arabe": "000008-12",
    "juan pablo segundo": "000008-14",
    "juan pablo ii": "000008-14",
    "juan pablo 2": "000008-14",
    "pae estanquer norte locales": "000025-01",
    "pae estanque norte locales": "000025-01",
    "estanque norte locales": "000025-01",
    "mae sala de bomba estanque sur": "000025-19",
    "sala de bomba estanque sur": "000025-19",
    "pizza hut": "000025-07",
    "san ignacio 500": "000025-18",
    "liceo alto cordillera la florida": "000028-01",
    "alto cordillera": "000028-01",
    "escuela alexander fleming": "000022-00",
    "alexander fleming": "000022-00",
    "derco matriz principal": "000012-06",
    "quilicura matriz principal": "000012-06",
    "las tarrias providencia": "000006-01",
    "lastarria": "000006-01",
    "liceo lastarria": "000006-01",
    "carmela carvajal provi": "000006-02",
    "carmela carvajal": "000006-02",
    "liceo 7 luisa saavedra": "000006-04",
    "luisa saavedra": "000006-04",
    "liceo juan pablo duarte": "000006-05",
    "juan pablo duarte": "000006-05",
    "lo valledor p1": "000002-01",
    "club hause uc": "000021-01",
    "club house uc": "000021-01",
    "club house cduc": "000021-01",
    "raymundo tupper": "000021-03",
    "raimundo tupper": "000021-03",
    "agunsa modulo d": "000020-02",
    "modulo d": "000020-02",
    "escuela lo velazques": "000017-04",
    "escuela lo velasquez": "000017-04",
    "lo velazquez": "000017-04",
    "lo velasquez": "000017-04",
    "gym renca": "000017-05",
    "gimnasio renca": "000017-05",
    "picina municipal renca": "000017-06",
    "piscina municipal renca": "000017-06",
    "piscina municipal": "000017-06",
    "iccp": "000017-07",
    "icco": "000017-08",
    "eugenio maria de hostos": "000024-01",
}

SKIP_TITULOS = {
    "corporacion penalolen colegios",
    "corporacion penalolen",
}

STOP = {
    "colegio",
    "liceo",
    "escuela",
    "esc",
    "de",
    "la",
    "el",
    "los",
    "las",
    "del",
    "san",
    "sala",
    "provi",
    "providencia",
    "renca",
}

SINONIMOS = {
    "segundo": "ii",
    "gym": "gimnasio",
    "picina": "piscina",
    "germoso": "hermoso",
    "pastene": "pasten",
    "pastenes": "pasten",
    "huichinavas": "huici",
    "velazques": "velasquez",
    "velazquez": "velasquez",
    "tarrias": "lastarria",
    "cana": "canas",
    "hause": "house",
}


def _slug(texto: str) -> str:
    t = unicodedata.normalize("NFKD", texto or "")
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    t = t.lower()
    t = re.sub(r"[^a-z0-9]+", " ", t)
    return t.strip()


def _tokens(slug: str) -> Set[str]:
    out: Set[str] = set()
    for t in slug.split():
        t = SINONIMOS.get(t, t)
        if len(t) >= 3 and t not in STOP:
            out.add(t)
    return out


def _es_celda_horario(val: object) -> bool:
    s = str(val or "").strip().lower()
    if not s:
        return True
    if s.startswith("#######") or s.startswith("ev."):
        return True
    if re.match(r"^\d+\s*=", s):
        return True
    return False


def _titulos_desde_xlsx(path: Path) -> List[str]:
    """Lee TODAS las columnas: los colegios van en A, D, G, J, M, P, S, V, Y, AA, AC."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out: List[str] = []
    vistos: Set[str] = set()
    for row in ws.iter_rows(values_only=True):
        for val in row:
            if _es_celda_horario(val):
                continue
            t = str(val).strip()
            slug = _slug(t)
            if not slug or slug in SKIP_TITULOS or slug in vistos:
                continue
            if "?" in t:
                continue
            vistos.add(slug)
            out.append(t)
    return out


def _descargar_drive(destino: Path) -> Path:
    from googleapiclient.http import MediaIoBaseDownload

    from wes_google_drive import obtener_servicio_drive

    svc = obtener_servicio_drive()
    request = svc.files().export_media(
        fileId=DRIVE_HORARIOS_ID,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _status, done = downloader.next_chunk()
    destino.parent.mkdir(parents=True, exist_ok=True)
    destino.write_bytes(buf.getvalue())
    return destino


def _resolver_node_id(
    titulo: str,
    nombres: Dict[str, str],
) -> Optional[str]:
    slug = _slug(titulo)
    if slug in CPA_ALIAS:
        return CPA_ALIAS[slug]
    for key, nid in CPA_ALIAS.items():
        if key in slug or slug in key:
            return nid
    tokens = _tokens(slug)
    if not tokens:
        return None
    best_nid = None
    best_score = 0
    for nid, name in nombres.items():
        ntok = _tokens(_slug(name))
        score = len(tokens & ntok)
        if score > best_score:
            best_score = score
            best_nid = nid
    if best_score >= 2:
        return best_nid
    if best_score == 1 and best_nid:
        ntok = _tokens(_slug(nombres.get(best_nid, "")))
        comunes = tokens & ntok
        if len(comunes) != 1:
            return None
        token = next(iter(comunes))
        unicos = [
            nid
            for nid, name in nombres.items()
            if token in _tokens(_slug(name))
        ]
        if len(unicos) == 1:
            return unicos[0]
    return None


def cargar_ids_cpa(
    nombres_nodos: Optional[Dict[str, str]] = None,
    *,
    excel: Optional[Path] = None,
) -> Tuple[Set[str], List[Tuple[str, str, str]]]:
    """
    Returns:
      set de nodeId CPA
      lista (titulo_excel, nodeId, nombre_wes) para auditoría
    """
    path = excel or CACHE_XLSX
    if excel is None:
        try:
            path = _descargar_drive(CACHE_XLSX)
            print(f"[INFO] Horarios CPA desde Drive → {path}")
        except Exception as exc:
            if path.is_file():
                print(f"[WARN] No se pudo bajar Drive ({exc}); uso caché {path}")
            else:
                raise

    nombres = nombres_nodos or {}
    titulos = _titulos_desde_xlsx(path)
    ids: Set[str] = set()
    detalle: List[Tuple[str, str, str]] = []
    for titulo in titulos:
        nid = _resolver_node_id(titulo, nombres)
        if not nid:
            detalle.append((titulo, "", "NO MAPEADO"))
            print(f"[WARN] Título CPA sin nodeId: {titulo}")
            continue
        ids.add(nid)
        detalle.append((titulo, nid, nombres.get(nid, "")))
    return ids, detalle


def clasificar_nodos(
    node_ids: Iterable[str],
    ids_cpa: Set[str],
) -> Tuple[List[str], List[str]]:
    cpa: List[str] = []
    solo: List[str] = []
    for nid in sorted(set(node_ids)):
        if nid in ids_cpa:
            cpa.append(nid)
        else:
            solo.append(nid)
    return cpa, solo

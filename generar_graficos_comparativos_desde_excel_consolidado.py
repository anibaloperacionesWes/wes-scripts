"""
Lee la hoja ``Consolidado`` del Excel generado con
``generar_excel_auditoria_consolidado_dos_periodos.py --toda-la-carpeta``
y produce PNG comparativos (promedio 24 h entre periodos, barras por total acumulado,
y siete graficos ``area + lineas`` como el 03, uno por dia de la semana homologo).

Convencion de periodos (14 columnas = 7+7): **23–29 mar = con WES**, **06–12 abr = sin WES**.

Por defecto elimina PNG previos en ``graficos_comparativos/`` y PNG sueltos en la carpeta
de auditoria (no borra nada dentro de ``csv_descarga_api/``). Use ``--no-limpiar-png-raiz-auditoria``
para conservar PNG en la raiz de esa carpeta.

Uso:
  python generar_graficos_comparativos_desde_excel_consolidado.py
  python generar_graficos_comparativos_desde_excel_consolidado.py -i "ruta\\consolidado.xlsx"
"""
from __future__ import annotations

import argparse
from datetime import date, datetime
from pathlib import Path
from typing import List, Sequence, Tuple

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

AUDIT_DIR = (
    Path(__file__).resolve().parent
    / "reports"
    / "reporte de auditoria"
    / "auditoria_puntos_renca_abril_2026"
    / "Auditoria ICCO abril"
)

DEFAULT_XLSX = AUDIT_DIR / "consolidado_revision_todos_los_csv_descarga_api.xlsx"
OUT_DIR = AUDIT_DIR / "graficos_comparativos"

# Etiquetas por defecto (14 dias: 7 marzo con WES + 7 abril sin WES)
LABEL_P1 = "Con WES 23–29 mar 2026"
LABEL_P2 = "Sin WES 06–12 abr 2026"

# Estética informe CIH / auditoría: Con WES = azul, Sin WES = rojo-marrón
COLOR_CON_WES = "#2a6fad"
COLOR_SIN_WES = "#a0503d"
COLOR_CON_WES_OSCURO = "#154a7a"
COLOR_SIN_WES_OSCURO = "#6b2f26"
COLOR_GRID = "#b8b8b8"


def _aplicar_marco_y_grid(ax) -> None:
    ax.set_facecolor("white")
    ax.grid(True, linestyle="--", color=COLOR_GRID, linewidth=0.85, alpha=0.95, zorder=0)
    ax.set_axisbelow(True)
    for spine in ax.spines.values():
        spine.set_visible(True)
        spine.set_edgecolor("black")
        spine.set_linewidth(0.9)


def _leyenda_informe(ax) -> None:
    leg = ax.legend(
        loc="upper right",
        fontsize=10,
        frameon=True,
        fancybox=False,
        framealpha=1.0,
        facecolor="white",
        edgecolor="black",
    )
    leg.get_frame().set_linewidth(0.8)


def _dibujar_comparativo_area_lineas(
    ax,
    horas: np.ndarray,
    y_con: np.ndarray,
    y_sin: np.ndarray,
    etiqueta_con: str,
    etiqueta_sin: str,
    titulo: str,
) -> None:
    """Área semitransparente + líneas (mismo criterio que informes CIH / ejemplo Lunes homólogo)."""
    _aplicar_marco_y_grid(ax)
    # Rellenos: primero Sin WES, luego Con WES para que el azul quede visible al cruzarse
    ax.fill_between(horas, 0, y_sin, alpha=0.32, color=COLOR_SIN_WES, zorder=1)
    ax.fill_between(horas, 0, y_con, alpha=0.32, color=COLOR_CON_WES, zorder=2)
    ax.plot(horas, y_con, "-", color=COLOR_CON_WES_OSCURO, linewidth=2.0, label=etiqueta_con, zorder=3)
    ax.plot(horas, y_sin, "-", color=COLOR_SIN_WES_OSCURO, linewidth=2.0, label=etiqueta_sin, zorder=4)
    ax.set_xlabel("h", fontsize=11)
    ax.set_ylabel("m³/h", fontsize=11)
    ax.set_title(titulo, fontsize=13, fontweight="bold", pad=10)
    ax.set_xlim(-0.5, 23.5)
    ax.set_xticks(list(range(0, 24, 2)))
    ax.set_ylim(bottom=0)
    _leyenda_informe(ax)

_NOMBRE_DIA_SEMANA_ES: Tuple[str, ...] = (
    "Lunes",
    "Martes",
    "Miércoles",
    "Jueves",
    "Viernes",
    "Sábado",
    "Domingo",
)

# Slug ASCII para nombre de archivo (Windows / consolas)
_NOMBRE_DIA_ARCHIVO: Tuple[str, ...] = (
    "Lunes",
    "Martes",
    "Miercoles",
    "Jueves",
    "Viernes",
    "Sabado",
    "Domingo",
)


def _parse_fecha_celda(val) -> date:
    if val is None:
        raise ValueError("celda fecha vacia")
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    return datetime.strptime(s, "%d-%m-%Y").date()


def leer_matriz_consolidado(path_xlsx: Path) -> Tuple[List[date], List[List[float]]]:
    """Fila 3 = fechas (desde col B); filas 4–27 = 24 horas x N columnas."""
    wb = load_workbook(path_xlsx, read_only=True, data_only=True)
    if "Consolidado" not in wb.sheetnames:
        raise KeyError(f"No hay hoja Consolidado en {path_xlsx}")
    ws = wb["Consolidado"]
    fechas: List[date] = []
    c = 2
    while True:
        v = ws.cell(row=3, column=c).value
        if v is None or str(v).strip() == "":
            break
        fechas.append(_parse_fecha_celda(v))
        c += 1
    if not fechas:
        raise ValueError("No se leyeron fechas en fila 3.")
    n = len(fechas)
    mats: List[List[float]] = []
    for j in range(n):
        col = []
        for h in range(24):
            row = 4 + h
            val = ws.cell(row=row, column=2 + j).value
            col.append(float(val) if val is not None else 0.0)
        mats.append(col)
    wb.close()
    return fechas, mats


def _partir_dos_periodos(
    fechas: Sequence[date], mats: Sequence[Sequence[float]]
) -> Tuple[Tuple[str, List[List[float]]], Tuple[str, List[List[float]]]]:
    """Si hay 14 columnas ordenadas: primeras 7 = con WES (mar), siguientes 7 = sin WES (abr)."""
    if len(fechas) != len(mats):
        raise ValueError("fechas y matrices no coinciden")
    n = len(fechas)
    if n == 14:
        mid = 7
        return (
            (LABEL_P1, list(mats[:mid])),
            (LABEL_P2, list(mats[mid:])),
        )
    if n % 2 == 0 and n >= 2:
        mid = n // 2
        return (
            (f"Periodo A ({fechas[0]} … {fechas[mid - 1]})", list(mats[:mid])),
            (f"Periodo B ({fechas[mid]} … {fechas[-1]})", list(mats[mid:])),
        )
    raise ValueError(
        f"Se esperaban 14 columnas (7+7) o un numero par de dias; hay {n}."
    )


def _promedio_horario(vectores: Sequence[Sequence[float]]) -> np.ndarray:
    """Media por hora 0..23 sobre los dias."""
    if not vectores:
        return np.zeros(24)
    a = np.array(vectores, dtype=float)
    return a.mean(axis=0)


def _total_acumulado_m3h(vectores: Sequence[Sequence[float]]) -> float:
    """Suma de todos los m3/h (aprox. m3 acumulados en la rejilla horaria)."""
    return float(np.sum(vectores))


def totales_rejilla_desde_excel_consolidado(path_xlsx: Path) -> Tuple[float, float, int]:
    """
    Lee la hoja ``Consolidado`` y devuelve (Σ periodo con WES, Σ periodo sin WES, dias por periodo).
    """
    fechas, mats = leer_matriz_consolidado(path_xlsx)
    (_, v1), (_, v2) = _partir_dos_periodos(fechas, mats)
    n = len(v1)
    return _total_acumulado_m3h(v1), _total_acumulado_m3h(v2), n


def generar_png_barras_rejilla_totales(out_dir: Path, t_con_wes: float, t_sin_wes: float) -> Path:
    """
    PNG ``02_barras_total_rejilla_por_periodo.png`` (titulo y estilo alineados al informe).
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    fig, ax = plt.subplots(figsize=(7, 5))
    x = ["Con WES", "Sin WES"]
    vals = [t_con_wes, t_sin_wes]
    colors = [COLOR_CON_WES, COLOR_SIN_WES]
    bars = ax.bar(x, vals, color=colors, width=0.55)
    ax.set_ylabel("Σ (m³/h)", fontsize=11)
    ax.set_title("Total acumulado con WES V/S sin WES", fontsize=12, fontweight="bold")
    ax.tick_params(axis="x", labelsize=11)
    for b, v in zip(bars, vals):
        ax.text(
            b.get_x() + b.get_width() / 2,
            b.get_height(),
            f"{v:.1f}",
            ha="center",
            va="bottom",
            fontsize=11,
            fontweight="bold",
        )
    plt.tight_layout()
    p2path = out_dir / "02_barras_total_rejilla_por_periodo.png"
    fig.savefig(p2path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    return p2path


def _limpiar_pngs_carpeta_graficos(out_dir: Path) -> int:
    out_dir.mkdir(parents=True, exist_ok=True)
    n = 0
    for p in out_dir.glob("*.png"):
        p.unlink()
        n += 1
    return n


def _limpiar_pngs_raiz_auditoria(base: Path) -> int:
    """Quita PNG sueltos en la carpeta de auditoria (no toca subcarpetas salvo graficos)."""
    n = 0
    for p in base.glob("*.png"):
        p.unlink()
        n += 1
    return n


def generar_pngs(
    fechas: List[date],
    mats: List[List[float]],
    out_dir: Path,
) -> List[Path]:
    (lab1, v1), (lab2, v2) = _partir_dos_periodos(fechas, mats)
    p1 = _promedio_horario(v1)
    p2 = _promedio_horario(v2)
    t1 = _total_acumulado_m3h(v1)
    t2 = _total_acumulado_m3h(v2)
    horas = np.arange(24)
    guardados: List[Path] = []

    # 1) Perfil 24 h promedio (área + líneas, mismo formato que días homólogos)
    fig, ax = plt.subplots(figsize=(11, 5.5))
    fig.patch.set_facecolor("white")
    _dibujar_comparativo_area_lineas(
        ax,
        horas,
        p1,
        p2,
        lab1,
        lab2,
        "Comparativo: perfil horario promedio (área + líneas)",
    )
    plt.tight_layout()
    p = out_dir / "01_promedio_24h_dos_periodos.png"
    fig.savefig(p, dpi=150, bbox_inches="tight")
    plt.close(fig)
    guardados.append(p)

    # 2) Barras totales acumulados (suma m3/h en la rejilla)
    p2path = generar_png_barras_rejilla_totales(out_dir, t1, t2)
    guardados.append(p2path)

    # 3) Mismo dato que 01 (variante sólo visual / archivo histórico)
    fig, ax = plt.subplots(figsize=(11, 5.5))
    fig.patch.set_facecolor("white")
    _dibujar_comparativo_area_lineas(
        ax,
        horas,
        p1,
        p2,
        lab1,
        lab2,
        "Comparativo: área + líneas (mismo dato que 01)",
    )
    plt.tight_layout()
    p3 = out_dir / "03_area_promedio_24h.png"
    fig.savefig(p3, dpi=150, bbox_inches="tight")
    plt.close(fig)
    guardados.append(p3)

    # 4–10) Un gráfico por día homólogo: col j vs col (mitad+j) — 7+7 o 4+4 (ICCO CPA).
    n_all = len(fechas)
    if n_all == len(mats) and n_all in (8, 14) and n_all % 2 == 0:
        mitad = n_all // 2
        for j in range(mitad):
            d_con = fechas[j]
            d_sin = fechas[mitad + j]
            nombre_dia = _NOMBRE_DIA_SEMANA_ES[d_con.weekday()]
            slug_dia = _NOMBRE_DIA_ARCHIVO[d_con.weekday()]
            p_con = np.array(mats[j], dtype=float)
            p_sin = np.array(mats[mitad + j], dtype=float)
            lab1_d = f"Con WES ({d_con:%d-%m-%Y})"
            lab2_d = f"Sin WES ({d_sin:%d-%m-%Y})"

            fig, ax = plt.subplots(figsize=(11, 5.5))
            fig.patch.set_facecolor("white")
            _dibujar_comparativo_area_lineas(
                ax,
                horas,
                p_con,
                p_sin,
                lab1_d,
                lab2_d,
                f"Comparativo: {nombre_dia} — área + líneas (día homólogo)",
            )
            plt.tight_layout()
            num = 4 + j
            path_d = out_dir / f"{num:02d}_area_{slug_dia}.png"
            fig.savefig(path_d, dpi=150, bbox_inches="tight")
            plt.close(fig)
            guardados.append(path_d)

    return guardados


def main() -> int:
    ap = argparse.ArgumentParser(
        description="PNG comparativos desde hoja Consolidado del Excel de auditoria."
    )
    ap.add_argument(
        "-i",
        "--excel",
        type=Path,
        default=DEFAULT_XLSX,
        help="Excel con hoja Consolidado",
    )
    ap.add_argument(
        "-o",
        "--out-dir",
        type=Path,
        default=OUT_DIR,
        help="Carpeta para PNG (se borran *.png previos aqui)",
    )
    ap.add_argument(
        "--no-limpiar-png-raiz-auditoria",
        action="store_true",
        help="No borrar *.png sueltos en la carpeta de auditoria (solo limpia graficos_comparativos/).",
    )
    args = ap.parse_args()
    path_xlsx = args.excel.resolve()
    if not path_xlsx.is_file():
        raise SystemExit(f"No existe: {path_xlsx}")

    out_dir = args.out_dir.resolve()
    n_old = _limpiar_pngs_carpeta_graficos(out_dir)
    n_root = 0
    if not args.no_limpiar_png_raiz_auditoria:
        n_root = _limpiar_pngs_raiz_auditoria(AUDIT_DIR.resolve())

    fechas, mats = leer_matriz_consolidado(path_xlsx)
    paths = generar_pngs(fechas, mats, out_dir)
    print(f"Excel: {path_xlsx}")
    print(f"PNG eliminados (carpeta graficos): {n_old}")
    if n_root:
        print(f"PNG eliminados (raiz auditoria): {n_root}")
    for p in paths:
        print(p.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

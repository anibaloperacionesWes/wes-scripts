"""
Exporta Excel Tobalaba (000008-04) — consumo nocturno mayo 2026.
Compara suma CSV UTC 0-7 (como en Excel del usuario) vs hora Chile (informe).
"""
from __future__ import annotations

import sys
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from generar_reporte_word import (
    _chile_hours_from_dates_measures_csv_text,
    _dt_to_chile,
    _reconcile_chile_hours_with_total_m3,
    _total_m3_from_json_for_chile_day,
    _utc_calendar_dates_for_chile_day,
    _value_by_time_sum_duplicate_rows,
    acl_node_base_url,
    calculate_nocturnal_metrics,
    parse_date,
)

NODE = "000008-04"
NOMBRE = "Tobalaba"
DESDE = date(2026, 5, 1)
HASTA = date(2026, 5, 31)
OUT = (
    ROOT
    / "reports"
    / "CORMUP"
    / "ABREGADO"
    / f"Tobalaba_nocturno_000008-04_{DESDE.strftime('%Y%m%d')}_{HASTA.strftime('%Y%m%d')}.xlsx"
)


def _style_header(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _fetch_csv_app_por_fecha(dia: date) -> str:
    """Como al bajar CSV en la app: start=end=fecha del día (ddmmyyyy)."""
    url = f"{acl_node_base_url()}/nodes/{NODE}/dates.measures.csv"
    ds = dia.strftime("%d%m%Y")
    r = requests.get(url, params=[("start", ds), ("end", ds)], timeout=30)
    r.raise_for_status()
    return r.text


def _fetch_csv_chile_completo(dia: date) -> str:
    """Fusión 1-2 días UTC para cubrir las 24 h del día civil Chile (usa el informe)."""
    url = f"{acl_node_base_url()}/nodes/{NODE}/dates.measures.csv"
    parts: list[str] = []
    for ud in _utc_calendar_dates_for_chile_day(dia):
        ds = ud.strftime("%d%m%Y")
        r = requests.get(url, params=[("start", ds), ("end", ds)], timeout=30)
        r.raise_for_status()
        lines = r.text.strip().split("\n")
        if not parts:
            parts.append(lines[0])
        parts.extend(lines[1:])
    return "\n".join(parts)


def _suma_utc_0_7(csv_text: str) -> float:
    """Mismo criterio que filtrar en Excel TIME entre T00:00Z y T07:00Z (sin pasar a Chile)."""
    total = 0.0
    by_time = _value_by_time_sum_duplicate_rows(csv_text)
    for time_str, val in by_time.items():
        try:
            dt = datetime.fromisoformat(time_str.replace("Z", "+00:00"))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            if 0 <= dt.hour <= 7:
                total += float(val)
        except (ValueError, TypeError):
            pass
    return total


def _detalle_utc_0_7(csv_text: str) -> list[tuple[str, int, int, float]]:
    rows: list[tuple[str, int, int, float]] = []
    by_time = _value_by_time_sum_duplicate_rows(csv_text)
    for time_str in sorted(by_time.keys()):
        try:
            dt = datetime.fromisoformat(time_str.replace("Z", "+00:00"))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            ch = _dt_to_chile(dt)
            if 0 <= dt.hour <= 7:
                rows.append((time_str, dt.hour, ch.hour, float(by_time[time_str])))
        except (ValueError, TypeError):
            pass
    return rows


def main() -> int:
    print(f"[INFO] Generando Excel — {NODE} {NOMBRE} — {DESDE} a {HASTA}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen_diario"
    ws.append(
        [
            "Fecha",
            "Día",
            "H0 Chile",
            "H1 Chile",
            "H2 Chile",
            "H3 Chile",
            "H4 Chile",
            "H5 Chile",
            "H6 Chile",
            "H7 Chile",
            "Suma 0-6 Chile (informe Word)",
            "Suma 0-7 Chile",
            "Suma UTC 0-7 (tu filtro CSV)",
            "totalM3 día",
            "Reconciliado",
        ]
    )
    _style_header(ws)

    total_06 = total_07 = total_utc07 = 0.0
    cur = DESDE
    while cur <= HASTA:
        csv_app = _fetch_csv_app_por_fecha(cur)
        csv_chile = _fetch_csv_chile_completo(cur)
        raw_h = _chile_hours_from_dates_measures_csv_text(csv_chile, cur)
        tj = _total_m3_from_json_for_chile_day(NODE, cur)
        recon_h, adj = _reconcile_chile_hours_with_total_m3(dict(raw_h), tj)
        s06 = sum(recon_h.get(h, 0.0) for h in range(7))
        s07 = sum(recon_h.get(h, 0.0) for h in range(8))
        s_utc = _suma_utc_0_7(csv_app)
        dias = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"][cur.weekday()]
        ws.append(
            [
                cur.isoformat(),
                dias,
                round(recon_h.get(0, 0), 2),
                round(recon_h.get(1, 0), 2),
                round(recon_h.get(2, 0), 2),
                round(recon_h.get(3, 0), 2),
                round(recon_h.get(4, 0), 2),
                round(recon_h.get(5, 0), 2),
                round(recon_h.get(6, 0), 2),
                round(recon_h.get(7, 0), 2),
                round(s06, 2),
                round(s07, 2),
                round(s_utc, 2),
                round(float(tj or 0), 2) if tj else None,
                "Sí" if adj else "No",
            ]
        )
        total_06 += s06
        total_07 += s07
        total_utc07 += s_utc
        cur += timedelta(days=1)

    ws.append([])
    ws.append(
        [
            "TOTAL mayo 2026",
            "",
            "", "", "", "", "", "", "", "",
            round(total_06, 2),
            round(total_07, 2),
            round(total_utc07, 2),
            "",
            "",
        ]
    )
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    metrics = calculate_nocturnal_metrics(NODE, parse_date("01/05/2026"), parse_date("31/05/2026", end_of_day=True))
    ws.append(["calculate_nocturnal_metrics", "", "", "", "", "", "", "", "", "", round(metrics["consumo_nocturno_total"], 2)])

    # Detalle horario UTC 0-7 por día
    ws2 = wb.create_sheet("Detalle_UTC_0a7")
    ws2.append(["Fecha Chile", "TIME UTC", "Hora UTC", "Hora Chile", "VALUE m3"])
    _style_header(ws2)
    cur = DESDE
    while cur <= HASTA:
        csv_day = _fetch_csv_app_por_fecha(cur)
        for time_str, h_utc, h_chile, val in _detalle_utc_0_7(csv_day):
            ws2.append([cur.isoformat(), time_str, h_utc, h_chile, round(val, 4)])
        cur += timedelta(days=1)

    ws3 = wb.create_sheet("Explicacion")
    ws3.append(["Hallazgo"])
    _style_header(ws3)
    notas = [
        f"Tu suma manual (141,8 m³) coincide EXACTO con filtrar el CSV por marcas UTC T00:00Z a T07:00Z sin convertir a hora Chile.",
        f"Esa suma UTC 0-7 en mayo 2026 = {round(total_utc07, 1)} m³.",
        f"El consumo de madrugada en Chile (03:00-06:59) aparece en el CSV como UTC 07:00-10:00, fuera de tu filtro.",
        f"El informe Word suma horas Chile 0-6 (00:00-06:59 local) = {round(total_06, 1)} m³.",
        "El CSV mensual (01/05-31/05 en una sola descarga) trae TOTALES DIARIOS, no horarios; para nocturno hay que bajar día a día.",
        "Ejemplo 05-may: UTC T07Z=4,5 m³ es Chile 03:00; no entra si filtras solo hasta T07Z en sentido estricto o si miras solo 0-2 Chile.",
    ]
    for n in notas:
        ws3.append([n])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"[OK] Excel: {OUT}")
    print(f"  UTC 0-7 (tu método): {round(total_utc07, 1)} m³")
    print(f"  Chile 0-6 (informe): {round(total_06, 1)} m³")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

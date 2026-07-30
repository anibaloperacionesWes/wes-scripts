"""
Genera un Excel: una hoja con horas 00:00-23:00 en filas y 7 columnas de dias (23 al 29).

Por defecto rellena consumo horario (m3/h) desde la API WES (misma logica que el informe:
``get_hourly_measures_for_day`` en generar_reporte_word). El CSV ``dates.measures.csv`` lo devuelve
el **backend** WES; la app puede usar ademas el JSON ``/nodes/measures/dates`` (``measures``).

Uso:
  python generar_excel_plantilla_semana_horaria.py
  python generar_excel_plantilla_semana_horaria.py --sin-datos
  python generar_excel_plantilla_semana_horaria.py --node-id 000017-08 -o reports/mi.xlsx

Mismo dato que la app (CSV descargado):
  1) Crea una carpeta, p. ej. ``C:\\MisCSVWes``
  2) Guarda un .csv por dia civil Chile con nombre ``2026-03-25.csv`` (texto TIME,VALUE igual que la app).
     Si un dia necesitas dos archivos UTC, concatena las filas en un solo archivo o pega el contenido
     de ambas descargas en ese ``AAAA-MM-DD.csv``.
  3) Ejecuta:
     python generar_excel_plantilla_semana_horaria.py --csv-dir C:\\MisCSVWes

Opcional: ``--api-base-url https://...`` si tu app usa otro servidor que el de este proyecto
(env ``WES_API_BASE_URL``).
"""
from __future__ import annotations

import argparse
import os
from calendar import monthrange
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# Mismo nodo por defecto que auditoria ICCO Renca
NODE_DEFAULT = "000017-08"


def _vector_m3h_dia_chile(node_id: str, dia: date) -> List[float]:
    """24 valores m3/h en hora civil Chile (misma logica que el informe)."""
    from generar_reporte_word import get_hourly_measures_for_day

    target = datetime.combine(dia, datetime.min.time())
    hourly_list = get_hourly_measures_for_day(node_id, target) or []
    por_hora: Dict[int, float] = defaultdict(float)
    for h, v in hourly_list:
        hi = int(h)
        if 0 <= hi < 24:
            por_hora[hi] += float(v)
    return [float(por_hora.get(h, 0.0)) for h in range(24)]


def _vectores_consumo_dias(node_id: str, dias: Sequence[date]) -> List[List[float]]:
    return [_vector_m3h_dia_chile(node_id, d) for d in dias]


def generar_excel(
    out_path: Path,
    *,
    year: int,
    month: int,
    dia_inicio: int,
    valores_m3h_por_dia: Optional[Sequence[Sequence[float]]] = None,
    titulo_nota: str = "Consumo horario (m3/h) — hora Chile [h, h+1)",
) -> None:
    _, ultimo = monthrange(year, month)
    if dia_inicio < 1 or dia_inicio + 6 > ultimo:
        raise ValueError(f"Rango invalido: el mes tiene hasta dia {ultimo}.")
    dias = [date(year, month, dia_inicio + i) for i in range(7)]
    if valores_m3h_por_dia is not None:
        if len(valores_m3h_por_dia) != 7:
            raise ValueError("Se esperan 7 listas de 24 valores m3/h.")
        for i, v in enumerate(valores_m3h_por_dia):
            if len(v) != 24:
                raise ValueError(f"Dia indice {i}: se esperan 24 valores horarios.")

    wb = Workbook()
    ws = wb.active
    ws.title = "Semana"

    ws["A1"] = "Hora"
    for col, d in enumerate(dias, start=2):
        c = get_column_letter(col)
        ws[f"{c}1"] = d.strftime("%d-%m-%Y")
        ws[f"{c}1"].font = Font(bold=True)
        ws.column_dimensions[c].width = 14

    ws["A1"].font = Font(bold=True)
    ws.column_dimensions["A"].width = 10

    for h in range(24):
        row = h + 2
        ws.cell(row=row, column=1, value=f"{h:02d}:00")
        for col in range(2, 9):
            di = col - 2
            if valores_m3h_por_dia is not None:
                val = float(valores_m3h_por_dia[di][h])
                cell = ws.cell(row=row, column=col, value=round(val, 4))
                cell.number_format = "0.0000"
            else:
                ws.cell(row=row, column=col, value="")

    ws["A26"] = titulo_nota
    ws["A26"].font = Font(italic=True, size=9)

    ws.freeze_panes = "B2"
    for row in ws.iter_rows(min_row=1, max_row=25, min_col=1, max_col=8):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> int:
    p = argparse.ArgumentParser(description="Plantilla Excel: 7 dias x 24 horas")
    p.add_argument("--year", type=int, default=2026)
    p.add_argument("--month", type=int, default=3)
    p.add_argument("--dia-inicio", type=int, default=23, help="Primer dia (inclusive), 7 dias seguidos")
    p.add_argument("--node-id", default=NODE_DEFAULT, help="Nodo WES para leer m3/h (default ICCO Renca)")
    p.add_argument(
        "--sin-datos",
        action="store_true",
        help="Solo plantilla vacia (sin llamar a la API).",
    )
    p.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Ruta del .xlsx (default: reports/plantilla_horaria_...) ",
    )
    p.add_argument(
        "--csv-dir",
        type=Path,
        default=None,
        help="Carpeta con archivos AAAA-MM-DD.csv (mismo contenido que exportas desde la app). "
        "Equivale a la variable de entorno WES_MEDIDAS_CSV_DIR.",
    )
    p.add_argument(
        "--api-base-url",
        default=None,
        metavar="URL",
        help="Base URL del API acl-node (mismo host que la app si difiere). Env WES_API_BASE_URL.",
    )
    args = p.parse_args()

    if args.api_base_url:
        os.environ["WES_API_BASE_URL"] = str(args.api_base_url).strip().rstrip("/")
    if args.csv_dir is not None:
        os.environ["WES_MEDIDAS_CSV_DIR"] = str(args.csv_dir.resolve())

    d0 = date(args.year, args.month, args.dia_inicio)
    d6 = date(args.year, args.month, args.dia_inicio + 6)
    root = Path(__file__).resolve().parent
    out = args.output or (
        root / "reports" / f"plantilla_horaria_{d0:%Y%m%d}_a_{d6:%Y%m%d}.xlsx"
    )

    valores = None
    nota_base = "Consumo horario (m3/h) - hora Chile [h, h+1)"
    nota = nota_base
    if not args.sin_datos:
        dias = [date(args.year, args.month, args.dia_inicio + i) for i in range(7)]
        valores = _vectores_consumo_dias(args.node_id, dias)
        if args.csv_dir is not None:
            nota = f"{nota_base} | nodo {args.node_id} | CSV local {args.csv_dir}"
        else:
            nota = f"{nota_base} | nodo {args.node_id} | API WES"

    generar_excel(
        out,
        year=args.year,
        month=args.month,
        dia_inicio=args.dia_inicio,
        valores_m3h_por_dia=valores,
        titulo_nota=nota,
    )
    print(out.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

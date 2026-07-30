"""Genera Excel de propuesta regulación diurna lun-vie 90% para 000006-01 (Inst. Lastarria)."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

OUT_DIR = Path(__file__).resolve().parent / "calculo de regulaciones"
NODE_ID = "000006-01"
NOMBRE = "Inst. Lastarria"
FACTOR = 0.90
PCT = 90


def main() -> Path:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d")
    out = OUT_DIR / f"propuesta_regulacion_diurna_{NODE_ID}_{stamp}.xlsx"

    wb = Workbook()
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws1 = wb.active
    ws1.title = "Programacion_WES"
    headers = [
        "Dia",
        "Hora_inicio",
        "Hora_fin",
        "node_id",
        "dispositivo",
        "Factor_sobre_referencia",
        "Pct_vs_referencia_actual",
    ]
    for col, h in enumerate(headers, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4472C4")

    dias = ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes"]
    row = 2
    for dia in dias:
        ws1.cell(row=row, column=1, value=dia)
        ws1.cell(row=row, column=2, value="08:00")
        ws1.cell(row=row, column=3, value="19:00")
        ws1.cell(row=row, column=4, value=NODE_ID)
        ws1.cell(row=row, column=5, value=NOMBRE)
        ws1.cell(row=row, column=6, value=FACTOR)
        ws1.cell(row=row, column=7, value=PCT)
        for c in range(1, 8):
            ws1.cell(row=row, column=c).border = border
        row += 1
    for c in range(1, 8):
        ws1.cell(row=1, column=c).border = border
    ws1.column_dimensions["A"].width = 12
    for col_letter in "BCDEFG":
        ws1.column_dimensions[col_letter].width = 20
    ws1.column_dimensions["E"].width = 24

    ws2 = wb.create_sheet("Detalle_hora_a_hora")
    h2 = ["Hora_bloque", "node_id", "dispositivo", "Factor", "Aplica"]
    for col, h in enumerate(h2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="E67E22")
    r = 2
    for h in range(8, 20):
        ws2.cell(r, 1, value=f"{h:02d}:00 - {h+1:02d}:00")
        ws2.cell(r, 2, value=NODE_ID)
        ws2.cell(r, 3, value=NOMBRE)
        ws2.cell(r, 4, value=FACTOR)
        ws2.cell(r, 5, value="Solo Lunes a Viernes")
        for c in range(1, 6):
            ws2.cell(r, c).border = border
        r += 1
    for c in range(1, 6):
        ws2.cell(1, c).border = border
    for col_letter in "ABCDE":
        ws2.column_dimensions[col_letter].width = 24

    ws3 = wb.create_sheet("Criterios")
    lines = [
        ("Titulo", "Propuesta regulacion diurna — Inst. Lastarria (000006-01)"),
        ("Generado", datetime.now().strftime("%d-%m-%Y %H:%M")),
        ("", ""),
        ("Objetivo", "Reducir respecto de referencia actual (100%) a 90% en ventana diurna laboral."),
        ("Dias", "Lunes a Viernes (no sabado ni domingo)."),
        ("Ventana", "08:00 a 19:00 (ajustable a 08:00-18:00)."),
        ("Factor WES", "0,90 sobre caudal o consigna de referencia."),
        ("Fuera de ventana", "100% de referencia o politica nocturna vigente."),
        ("", ""),
        ("Fuente", "Analisis consumo horario (ej. export DETALLE API / Excel auditoria)."),
    ]
    for i, (a, b) in enumerate(lines, 1):
        ws3.cell(i, 1, value=a)
        ws3.cell(i, 2, value=b)
    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 72

    wb.save(out)
    return out


if __name__ == "__main__":
    p = main()
    print(p.resolve())

"""
Genera un Excel estilo ``Detalle consumo Lo Valledor.xlsx``:
- **Consolidado**: matriz hora x día de semana + bloque % / m³·h⁻¹ / L·min⁻¹ (como el ejemplo).
- **Detalles x Dia**: por cada día de la semana, columnas por cada fecha real del periodo + Promedio.

Entradas:
- Detalle horario: Excel con hoja ``DETALLE`` (columnas: node_id, nombre, fecha, hora, m³/h), o bien
- ``--api --node-id … --desde YYYY-MM-DD --hasta YYYY-MM-DD`` para el mismo criterio que los informes WES.

Promedios hora × día de semana: si no se pasa ``--promedio``, se calculan desde el detalle (recomendado).

Incluye **tres hojas** ``Propuesta N°1`` … ``N°3`` con la misma matriz y tabla % / m³·h⁻¹ / L·min⁻¹,
pero **umbrales verde/amarillo/rojo** distintos (ver ``PROPUESTAS_REGULACION`` en el código).

Valores de caudal en **m³/h** (y detalle por día / promedio) con **2 decimales**; litros/min en la tabla % con formato **0,00**.

Salida: ``calculo de regulaciones/Detalle_consumo_<node>_<stamp>.xlsx``
"""
from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from datetime import date, datetime, time, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
CALC_DIR = ROOT / "calculo de regulaciones"

# Decimales mostrados en m³/h, detalle por día y tabla de referencia (Excel)
DECIMALES_M3 = 2
FMT_M3 = "0.00"
FMT_LMIN = "0.00"

# Ventana diurna para semáforo de regulación (mismo criterio operativo)
HORA_INI_REG = 6
HORA_FIN_REG = 23  # inclusive
# v / máximo(06-23h) ese día de semana → umbral de regulación sugerida
# ≥ RATIO_VERDE: cerca del pico del día → conservar al menos ~90 % del caudal de ref.
# entre RATIO_ROJO y RATIO_VERDE: zona intermedia (ej. ~1,4 m³/h vs pico) → ~60 %
# < RATIO_ROJO: bajo vs pico → mayor margen, ~30 %
RATIO_VERDE = 0.74
RATIO_ROJO = 0.40

# Tres hojas «Propuesta N°»: mismo dato; distinto par verde/rojo (ratio vs máx. 06–23 h ese día)
PROPUESTAS_REGULACION: tuple[tuple[int, str, float, float, str], ...] = (
    (1, "Criterio base", 0.74, 0.40, "Equilibrio entre conservar caudal y margen de reducción."),
    (2, "Más conservador", 0.68, 0.35, "Umbral verde más amplio: más horas en verde (~90 %)."),
    (3, "Más exigente", 0.80, 0.45, "Umbral verde más estrecho: más horas en amarillo/rojo (~60 % / ~30 %)."),
)

FILL_VERDE = PatternFill("solid", fgColor="C6EFCE")
FILL_AMARILLO = PatternFill("solid", fgColor="FFF2CC")
FILL_ROJO = PatternFill("solid", fgColor="FFC7CE")
# Sin relleno cuando no hay consumo (evita semáforo en periodos en cero)
FILL_SIN_SEMAFORO = PatternFill(fill_type="none")

# Cuadro referencia % / M³·h⁻¹ / L·min (encabezado + filas 100→90→60→30)
FILL_TABLA_HDR = PatternFill("solid", fgColor="4472C4")
FONT_TABLA_HDR = Font(bold=True, color="FFFFFF")
FILL_TABLA_ROW_100 = PatternFill("solid", fgColor="D9E1F2")
FILL_TABLA_ROW_90 = PatternFill("solid", fgColor="C6EFCE")
FILL_TABLA_ROW_60 = PatternFill("solid", fgColor="FFF2CC")
FILL_TABLA_ROW_30 = PatternFill("solid", fgColor="FFC7CE")
_TABLA_BORDE = Side(style="thin", color="FFADADAD")
BORDE_TABLA = Border(
    left=_TABLA_BORDE, right=_TABLA_BORDE, top=_TABLA_BORDE, bottom=_TABLA_BORDE
)

# Por debajo de esto se considera consumo nulo para colorear (m³/h)
UMBRAL_CERO_M3H = 1e-6


def _es_periodo_cero_m3h(v: float) -> bool:
    return abs(float(v or 0.0)) < UMBRAL_CERO_M3H


def _r_m3(v: float | None) -> float:
    return round(float(v or 0.0), DECIMALES_M3)


def _normalizar_matriz_promedios(prom_mat: list[list[object]]) -> list[list[object]]:
    """Unifica redondeo de m³/h en la matriz 24×7 (fila 0 = encabezados)."""
    out: list[list[object]] = [list(prom_mat[0])]
    for ri in range(1, len(prom_mat)):
        row = prom_mat[ri]
        new_row: list[object] = [row[0]]
        for c in range(1, 8):
            new_row.append(_r_m3(row[c] if c < len(row) else 0.0))
        out.append(new_row)
    return out


# Encabezado como en el ejemplo (Sábado con tilde opcional)
WD_HEADER = ("Lunes", "Martes", "Miercoles", "Jueves", "Viernes", "Sábado", "Domingo")

# Ajustes para análisis de regulación (solo el nodo indicado)
NODE_ID_LASTARRIA_REG = "000006-01"
# Lun–vie: horas 0..7 (00:00–07:00) → 0 m³/h; Sáb–Dom: todas las horas → 0
HORA_FIN_BLOQUE_NOCTURNO_LASTARRIA = 7  # inclusive

NODE_ID_CARMELA_REG = "000006-02"
# Lun–vie: horas 0..4 (00:00–04:00) → 0 m³/h
HORA_FIN_BLOQUE_NOCTURNO_CARMELA = 4  # inclusive


def _normalizar_id_nodo(nid: str) -> str:
    return str(nid).strip().replace("0000006-", "000006-")


def _regulacion_mascara_lastarria_aplica(nid: str) -> bool:
    return _normalizar_id_nodo(nid) == NODE_ID_LASTARRIA_REG


def _regulacion_mascara_carmela_aplica(nid: str) -> bool:
    return _normalizar_id_nodo(nid) == NODE_ID_CARMELA_REG


def _mascara_consumo_regulacion_lastarria(
    por_hora: dict[tuple[date, int], float],
) -> dict[tuple[date, int], float]:
    """
    Para 000006-01: lun–vie fuerza 0 en 00:00–07:00; sábado y domingo todo el día en 0.
    """
    out: dict[tuple[date, int], float] = {}
    for (d, h), v in por_hora.items():
        wd = d.weekday()
        if wd >= 5:
            out[(d, h)] = 0.0
        elif wd <= 4 and 0 <= h <= HORA_FIN_BLOQUE_NOCTURNO_LASTARRIA:
            out[(d, h)] = 0.0
        else:
            out[(d, h)] = v
    return out


def _mascara_consumo_regulacion_carmela(
    por_hora: dict[tuple[date, int], float],
) -> dict[tuple[date, int], float]:
    """Para 000006-02: lun–vie 00:00–04:00 → 0 m³/h."""
    out: dict[tuple[date, int], float] = {}
    for (d, h), v in por_hora.items():
        wd = d.weekday()
        if wd <= 4 and 0 <= h <= HORA_FIN_BLOQUE_NOCTURNO_CARMELA:
            out[(d, h)] = 0.0
        else:
            out[(d, h)] = v
    return out


def _parse_date(v) -> date:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()


def _parse_hour(v) -> int:
    if isinstance(v, datetime):
        return v.hour
    s = str(v)
    if ":" in s:
        return int(s.split(":")[0])
    return int(float(s))


def _parse_m3(v) -> float:
    if v is None:
        return 0.0
    s = str(v).strip().replace(",", ".")
    return float(s) if s else 0.0


def _cargar_detalle(
    path: Path,
    *,
    node_id_filter: str | None = None,
    fecha_min: date | None = None,
    fecha_max: date | None = None,
) -> tuple[str, str, dict[tuple[date, int], float], date | None, date | None]:
    wb = load_workbook(path, data_only=True)
    ws = wb["DETALLE"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()
    node_id = ""
    nombre = ""
    acc: dict[tuple[date, int], float] = defaultdict(float)
    dmin: date | None = None
    dmax: date | None = None
    nf = node_id_filter.strip() if node_id_filter else None
    for r in rows:
        if not r or r[0] is None:
            continue
        nid = str(r[0]).strip()
        if nf and nid != nf:
            continue
        node_id = nid
        nombre = str(r[1] or "").strip()
        d = _parse_date(r[2])
        if fecha_min is not None and d < fecha_min:
            continue
        if fecha_max is not None and d > fecha_max:
            continue
        h = _parse_hour(r[3])
        m3 = _parse_m3(r[4])
        acc[(d, h)] += m3
        dmin = d if dmin is None or d < dmin else dmin
        dmax = d if dmax is None or d > dmax else dmax
    return node_id, nombre, dict(acc), dmin, dmax


def _fetch_detalle_api(node_id: str, fecha_min: date, fecha_max: date) -> tuple[str, str, dict[tuple[date, int], float], date, date]:
    """Descarga m³/h por hora Chile desde el API (misma lógica que informes WES)."""
    sys.path.insert(0, str(ROOT))
    from generar_reporte_word import get_hourly_measures_for_day, get_node_name  # noqa: WPS433

    nombre = get_node_name(node_id) or ""
    acc: dict[tuple[date, int], float] = defaultdict(float)
    d = fecha_min
    while d <= fecha_max:
        dt = datetime.combine(d, datetime.min.time())
        rows = get_hourly_measures_for_day(node_id, dt) or []
        for h, m3 in rows:
            acc[(d, int(h))] += float(m3)
        d += timedelta(days=1)
    return node_id, nombre.strip(), dict(acc), fecha_min, fecha_max


def _compute_prom_mat(por_hora_fecha: dict[tuple[date, int], float]) -> list[list[object]]:
    """Misma agregación que ``generar_promedio_dia_semana_hora`` (promedio m³/h por wd × hora)."""
    sum_by: defaultdict[tuple[int, int], float] = defaultdict(float)
    count_by: defaultdict[tuple[int, int], int] = defaultdict(int)
    for (d, h), m3 in por_hora_fecha.items():
        wd = d.weekday()
        sum_by[(wd, h)] += m3
        count_by[(wd, h)] += 1
    mat: list[list[object]] = []
    mat.append(["Hora"] + list(WD_HEADER))
    for hh in range(24):
        row: list[object] = [f"{hh:02d}:00"]
        for wd in range(7):
            n = count_by.get((wd, hh), 0)
            avg = (sum_by[(wd, hh)] / n) if n else 0.0
            row.append(_r_m3(avg))
        mat.append(row)
    return mat


def _cargar_promedios(path: Path) -> list[list[object]]:
    wb = load_workbook(path, data_only=True)
    ws = wb["Promedio_dia_hora"]
    mat: list[list[object]] = []
    for r in range(1, 26):
        row = []
        for c in range(1, 9):
            row.append(ws.cell(r, c).value)
        mat.append(row)
    wb.close()
    return mat


def _umbral_regulacion(
    ratio: float,
    ratio_verde: float = RATIO_VERDE,
    ratio_rojo: float = RATIO_ROJO,
) -> tuple[str, PatternFill]:
    """ratio = m³/h ÷ max(06-23) ese día de semana."""
    if ratio >= ratio_verde:
        return "90%", FILL_VERDE
    if ratio >= ratio_rojo:
        return "60%", FILL_AMARILLO
    return "30%", FILL_ROJO


def _escribir_tabla_pct_m3_litros(
    ws,
    ref_100: float,
    *,
    fila_enc_pct: int,
    col_pct: int = 10,
    col_m3: int = 11,
    col_lmin: int = 12,
) -> None:
    """Bloque % / M³·h⁻¹ / L·min⁻¹ con fórmulas como en tu Excel (100→90→60; 30 % encadenado al 90 %)."""
    ws.cell(fila_enc_pct, col_pct, "%")
    ws.cell(fila_enc_pct, col_m3, "M3 x hora")
    ws.cell(fila_enc_pct, col_lmin, "Litros x minuto")
    # Filas 100, 90, 60, 30 bajo la fila de encabezado
    r100 = fila_enc_pct + 1
    r90 = fila_enc_pct + 2
    r60 = fila_enc_pct + 3
    r30 = fila_enc_pct + 4
    pct_col = get_column_letter(col_pct)
    m3_col = get_column_letter(col_m3)

    ref_r = _r_m3(ref_100) if ref_100 > 0 else 1.0
    ws.cell(r100, col_pct, 100)
    ws.cell(r100, col_m3, ref_r)
    c_m3_100 = ws.cell(r100, col_m3)
    c_m3_100.number_format = FMT_M3
    c_l_100 = ws.cell(r100, col_lmin, f"={m3_col}{r100}*1000/60")
    c_l_100.number_format = FMT_LMIN

    ws.cell(r90, col_pct, 90)
    c_m3_90 = ws.cell(r90, col_m3, f"={pct_col}{r90}*{m3_col}{r100}/{pct_col}{r100}")
    c_m3_90.number_format = FMT_M3
    c_l_90 = ws.cell(r90, col_lmin, f"={m3_col}{r90}*1000/60")
    c_l_90.number_format = FMT_LMIN

    ws.cell(r60, col_pct, 60)
    c_m3_60 = ws.cell(r60, col_m3, f"={pct_col}{r60}*{m3_col}{r100}/{pct_col}{r100}")
    c_m3_60.number_format = FMT_M3
    c_l_60 = ws.cell(r60, col_lmin, f"={m3_col}{r60}*1000/60")
    c_l_60.number_format = FMT_LMIN

    ws.cell(r30, col_pct, 30)
    c_m3_30 = ws.cell(r30, col_m3, f"={pct_col}{r30}*{m3_col}{r90}/{pct_col}{r90}")
    c_m3_30.number_format = FMT_M3
    c_l_30 = ws.cell(r30, col_lmin, f"={m3_col}{r30}*1000/60")
    c_l_30.number_format = FMT_LMIN

    # Color y borde del cuadro (como tabla de referencia visual)
    for c in (col_pct, col_m3, col_lmin):
        hc = ws.cell(fila_enc_pct, c)
        hc.fill = FILL_TABLA_HDR
        hc.font = FONT_TABLA_HDR
        hc.border = BORDE_TABLA
        hc.alignment = Alignment(horizontal="center", vertical="center")
    for r, fill in (
        (r100, FILL_TABLA_ROW_100),
        (r90, FILL_TABLA_ROW_90),
        (r60, FILL_TABLA_ROW_60),
        (r30, FILL_TABLA_ROW_30),
    ):
        for ci, c in enumerate((col_pct, col_m3, col_lmin)):
            cell = ws.cell(r, c)
            cell.fill = fill
            cell.border = BORDE_TABLA
            if ci == 0:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")


def _aplicar_semaforo_matriz(
    ws,
    prom_mat: list[list[object]],
    max_diurno_por_wd: list[float],
    ratio_verde: float,
    ratio_rojo: float,
) -> None:
    for h in range(HORA_INI_REG, HORA_FIN_REG + 1):
        r = h + 3
        for wd in range(7):
            cell = ws.cell(r, 3 + wd)
            v = float(prom_mat[h + 1][1 + wd] or 0)
            if _es_periodo_cero_m3h(v):
                cell.fill = FILL_SIN_SEMAFORO
                continue
            mx = max_diurno_por_wd[wd]
            ratio = (v / mx) if mx > 1e-12 else 0.0
            _lab, fill = _umbral_regulacion(ratio, ratio_verde, ratio_rojo)
            cell.fill = fill


def _crear_hoja_propuesta(
    wb,
    num: int,
    etiqueta: str,
    ratio_verde: float,
    ratio_rojo: float,
    nota: str,
    *,
    nombre: str,
    node_id: str,
    dmin: date | None,
    dmax: date | None,
    prom_mat: list[list[object]],
    max_diurno_por_wd: list[float],
    ref_100: float,
    nota_mascara_lastarria: str = "",
) -> None:
    titulo = f"Consumo promedio — {nombre} ({node_id})"
    if dmin and dmax:
        titulo += f"  ({dmin:%d-%m-%Y} al {dmax:%d-%m-%Y})"
    titulo += f" — Propuesta N°{num}: {etiqueta}"

    ws = wb.create_sheet(f"Propuesta N°{num}")
    ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=13)
    c = ws.cell(1, 3, titulo)
    c.font = Font(bold=True, size=12)
    c.alignment = Alignment(horizontal="center")

    for i, lab in enumerate(WD_HEADER, start=3):
        ws.cell(2, i, lab)
        ws.cell(2, i).font = Font(bold=True)

    for h in range(24):
        r = h + 3
        ws.cell(r, 2, f"{h:02d}:00")
        for wd in range(7):
            v = prom_mat[h + 1][1 + wd]
            if v is None:
                v = 0.0
            fv = _r_m3(v)
            cdata = ws.cell(r, 3 + wd, fv)
            cdata.number_format = FMT_M3

    _aplicar_semaforo_matriz(ws, prom_mat, max_diurno_por_wd, ratio_verde, ratio_rojo)

    # Encabezados tabla % alineados con fila hora 01:00 (fila 4)
    _escribir_tabla_pct_m3_litros(ws, ref_100, fila_enc_pct=4)

    row_leg = 28
    ws.merge_cells(start_row=row_leg, start_column=3, end_row=row_leg, end_column=13)
    texto = (
        f"Semáforo (06:00–23:00) Propuesta N°{num}: verde ≥{ratio_verde:.0%} del máximo diurno de ese día; "
        f"amarillo entre {ratio_rojo:.0%} y {ratio_verde:.0%}; rojo <{ratio_rojo:.0%}. {nota} "
        "Sin relleno si el m³/h promedio de esa celda es ~0 (periodo en cero)."
        f" {nota_mascara_lastarria}".strip()
    )
    tleg = ws.cell(row_leg, 3, texto)
    tleg.font = Font(size=10)
    tleg.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row_leg].height = 36
    ws.cell(row_leg + 1, 3, "Verde ~90%").fill = FILL_VERDE
    ws.cell(row_leg + 1, 4, "Cerca del pico relativo: mantener al menos ~90 % del caudal de referencia.")
    ws.cell(row_leg + 2, 3, "Amarillo ~60%").fill = FILL_AMARILLO
    ws.cell(row_leg + 2, 4, "Intermedio (~60 % del caudal de referencia).")
    ws.cell(row_leg + 3, 3, "Rojo ~30%").fill = FILL_ROJO
    ws.cell(row_leg + 3, 4, "Mayor margen para reducir (~30 % del caudal de referencia).")

    for col_idx in range(1, 14):
        ws.column_dimensions[get_column_letter(col_idx)].width = 11
    ws.column_dimensions["B"].width = 8


def generar(
    path_detalle: Path | None,
    path_promedio: Path | None,
    path_salida: Path | None = None,
    *,
    node_id_filter: str | None = None,
    fecha_min: date | None = None,
    fecha_max: date | None = None,
    usar_api: bool = False,
) -> Path:
    if usar_api:
        if not node_id_filter or fecha_min is None or fecha_max is None:
            raise ValueError("usar_api requiere --node-id, --desde y --hasta")
        node_id, nombre, por_hora_fecha, dmin, dmax = _fetch_detalle_api(
            node_id_filter.strip(), fecha_min, fecha_max
        )
    else:
        if path_detalle is None:
            raise ValueError("Se requiere --detalle si no se usa --api")
        node_id, nombre, por_hora_fecha, dmin, dmax = _cargar_detalle(
            path_detalle,
            node_id_filter=node_id_filter,
            fecha_min=fecha_min,
            fecha_max=fecha_max,
        )
    if not por_hora_fecha:
        raise ValueError("Sin datos: revise nodo, fechas o archivo DETALLE.")

    node_id = _normalizar_id_nodo(node_id)
    mascara_lastarria = _regulacion_mascara_lastarria_aplica(node_id)
    mascara_carmela = _regulacion_mascara_carmela_aplica(node_id)
    if mascara_lastarria:
        por_hora_fecha = _mascara_consumo_regulacion_lastarria(por_hora_fecha)
    elif mascara_carmela:
        por_hora_fecha = _mascara_consumo_regulacion_carmela(por_hora_fecha)

    mascara_regulacion = mascara_lastarria or mascara_carmela
    # Con máscara de regulación recalculamos promedios desde DETALLE/API (no archivo externo)
    if path_promedio is not None and not mascara_regulacion:
        prom_mat = _cargar_promedios(path_promedio)
    else:
        prom_mat = _compute_prom_mat(por_hora_fecha)
    prom_mat = _normalizar_matriz_promedios(prom_mat)

    # Fechas por día de semana (orden cronológico)
    fechas_por_wd: list[list[date]] = [[] for _ in range(7)]
    for (d, _h), _ in por_hora_fecha.items():
        fechas_por_wd[d.weekday()].append(d)
    for wd in range(7):
        fechas_por_wd[wd] = sorted(set(fechas_por_wd[wd]))

    wb = Workbook()
    # --- Consolidado ---
    ws_c = wb.active
    ws_c.title = "Consolidado"

    titulo = f"Consumo promedio — {nombre} ({node_id})"
    if dmin and dmax:
        titulo += f"  ({dmin:%d-%m-%Y} al {dmax:%d-%m-%Y})"
    ws_c.merge_cells(start_row=1, start_column=3, end_row=1, end_column=13)
    c = ws_c.cell(1, 3, titulo)
    c.font = Font(bold=True, size=12)
    c.alignment = Alignment(horizontal="center")

    for i, lab in enumerate(WD_HEADER, start=3):
        ws_c.cell(2, i, lab)
        ws_c.cell(2, i).font = Font(bold=True)

    valores_7: list[float] = []
    for h in range(24):
        r = h + 3
        ws_c.cell(r, 2, f"{h:02d}:00")
        for wd in range(7):
            v = prom_mat[h + 1][1 + wd]
            if v is None:
                v = 0.0
            fv = _r_m3(v)
            c = ws_c.cell(r, 3 + wd, fv)
            c.number_format = FMT_M3
            valores_7.append(fv)

    max_diurno_por_wd: list[float] = []
    for wd in range(7):
        vals_h = [
            float(prom_mat[h + 1][1 + wd] or 0)
            for h in range(HORA_INI_REG, HORA_FIN_REG + 1)
        ]
        max_diurno_por_wd.append(max(vals_h) if vals_h else 0.0)

    _aplicar_semaforo_matriz(ws_c, prom_mat, max_diurno_por_wd, RATIO_VERDE, RATIO_ROJO)

    ref_100 = _r_m3(max(valores_7) if valores_7 else 0.0)
    if ref_100 <= 0:
        ref_100 = 1.0

    # Tabla 100 / 90 / 60 / 30 % con fórmulas (misma fila de encabezado que hora 01:00)
    _escribir_tabla_pct_m3_litros(ws_c, ref_100, fila_enc_pct=4)

    # Leyenda semáforo (06:00-23:00): cada celda coloreada según v / max(06-23 h) de ese día de semana
    row_leg = 28
    ws_c.merge_cells(start_row=row_leg, start_column=3, end_row=row_leg, end_column=12)
    tleg = ws_c.cell(
        row_leg,
        3,
        "Semáforo regulación (06:00-23:00): en cada celda, el color indica el piso de caudal sugerido "
        f"según qué tan cerca está el m³/h al máximo diurno de ese día (≥{RATIO_VERDE:.0%} del pico → verde ~90%; "
        f"{RATIO_ROJO:.0%}–{RATIO_VERDE:.0%} → amarillo ~60%, p. ej. 1,4 m³/h si el pico ese día es ~1,9; "
        f"<{RATIO_ROJO:.0%} del pico → rojo ~30%). "
        "Celdas sin relleno: consumo ~0 m³/h en ese promedio (periodo en cero)."
        + (
            " Ajuste regulación (000006-01): lun–vie 00:00–07:00 y sáb–dom 24 h mostrados como 0 m³/h."
            if mascara_lastarria
            else (
                " Ajuste regulación (000006-02): lun–vie 00:00–04:00 mostrados como 0 m³/h."
                if mascara_carmela
                else ""
            )
        ),
    )
    tleg.font = Font(size=10)
    tleg.alignment = Alignment(wrap_text=True, vertical="top")
    ws_c.row_dimensions[row_leg].height = 48
    ws_c.cell(row_leg + 1, 3, "Verde").fill = FILL_VERDE
    ws_c.cell(row_leg + 1, 4, "Regulación conservadora: mantener al menos ~90 % del caudal de referencia.")
    ws_c.cell(row_leg + 2, 3, "Amarillo").fill = FILL_AMARILLO
    ws_c.cell(row_leg + 2, 4, "Umbral intermedio (~60 %); p. ej. 1,4 m³/h si el máximo 06-23 h de ese día es ~1,9.")
    ws_c.cell(row_leg + 3, 3, "Rojo").fill = FILL_ROJO
    ws_c.cell(row_leg + 3, 4, "Mayor margen para bajar (~30 % del caudal de referencia) en esas horas.")

    # --- Detalles x Dia (estilo ancho) ---
    ws_d = wb.create_sheet("Detalles x Dia")
    col = 2
    ws_d.cell(1, 1, node_id)
    ws_d.cell(1, 1).font = Font(bold=True)

    for wd in range(7):
        fechas = fechas_por_wd[wd]
        if not fechas:
            continue
        ws_d.cell(1, col, f"{WD_HEADER[wd]} ")
        ws_d.cell(1, col).font = Font(bold=True)
        col += 1
        for fd in fechas:
            ws_d.cell(1, col, datetime.combine(fd, time()))
            ws_d.cell(1, col).number_format = "DD/MM/YYYY"
            col += 1
        col += 1
        ws_d.cell(1, col, "Promedio")
        ws_d.cell(1, col).font = Font(bold=True)
        col += 1
        ws_d.cell(1, col, None)
        col += 1

    for h in range(24):
        r = h + 2
        col = 2
        for wd in range(7):
            fechas = fechas_por_wd[wd]
            if not fechas:
                continue
            tcell = ws_d.cell(r, col, time(h, 0))
            tcell.number_format = "HH:MM"
            col += 1
            vals = []
            for fd in fechas:
                v = por_hora_fecha.get((fd, h), 0.0)
                vals.append(v)
                cv = ws_d.cell(r, col, _r_m3(v))
                cv.number_format = FMT_M3
                col += 1
            col += 1
            prom = _r_m3(sum(vals) / len(vals) if vals else 0.0)
            c_prom = ws_d.cell(r, col, prom)
            c_prom.number_format = FMT_M3
            if HORA_INI_REG <= h <= HORA_FIN_REG:
                if _es_periodo_cero_m3h(prom):
                    c_prom.fill = FILL_SIN_SEMAFORO
                else:
                    mx = max_diurno_por_wd[wd]
                    ratio = (float(prom) / mx) if mx > 1e-12 else 0.0
                    _lab, fill = _umbral_regulacion(ratio, RATIO_VERDE, RATIO_ROJO)
                    c_prom.fill = fill
            col += 1
            col += 1

    if mascara_lastarria:
        nota_masc = (
            "Ajuste regulación (000006-01): lun–vie 00:00–07:00 y sáb–dom 24 h como 0 m³/h."
        )
    elif mascara_carmela:
        nota_masc = "Ajuste regulación (000006-02): lun–vie 00:00–04:00 como 0 m³/h."
    else:
        nota_masc = ""
    for num, etiqueta, rv, rr, nota in PROPUESTAS_REGULACION:
        _crear_hoja_propuesta(
            wb,
            num,
            etiqueta,
            rv,
            rr,
            nota,
            nombre=nombre,
            node_id=node_id,
            dmin=dmin,
            dmax=dmax,
            prom_mat=prom_mat,
            max_diurno_por_wd=max_diurno_por_wd,
            ref_100=ref_100,
            nota_mascara_lastarria=nota_masc,
        )

    for col_idx in range(1, ws_c.max_column + 1):
        ws_c.column_dimensions[get_column_letter(col_idx)].width = 11
    ws_c.column_dimensions["B"].width = 8
    for col_idx in range(1, ws_d.max_column + 1):
        ws_d.column_dimensions[get_column_letter(col_idx)].width = 10

    out = path_salida
    if out is None:
        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        safe_node = node_id.replace("-", "_")
        out = CALC_DIR / f"Detalle_consumo_{safe_node}_{stamp}.xlsx"
    out = Path(out).resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


def _parse_fecha(s: str) -> date:
    return datetime.strptime(s.strip(), "%Y-%m-%d").date()


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Excel estilo Lo Valledor desde promedios + DETALLE (o descarga API)"
    )
    ap.add_argument(
        "--promedio",
        type=Path,
        default=None,
        help="Opcional: matriz Promedio_dia_hora; si se omite, se calcula desde DETALLE/API",
    )
    ap.add_argument(
        "--detalle",
        type=Path,
        default=Path(r"c:\Users\aniba\Downloads\e832e010-4508-4e0b-b17d-c3e05da2e0f2.xlsx"),
        help="Excel con hoja DETALLE (ignorado con --api)",
    )
    ap.add_argument(
        "--api",
        action="store_true",
        help="Descargar m³/h desde el API WES (requiere --node-id, --desde, --hasta)",
    )
    ap.add_argument(
        "--node-id",
        type=str,
        default=None,
        help="Filtra DETALLE por nodo (ej. 000006-02) u obligatorio con --api",
    )
    ap.add_argument(
        "--desde",
        type=str,
        default=None,
        help="Fecha mínima inclusive YYYY-MM-DD (con --api o con --detalle)",
    )
    ap.add_argument(
        "--hasta",
        type=str,
        default=None,
        help="Fecha máxima inclusive YYYY-MM-DD",
    )
    ap.add_argument("-o", "--output", type=Path, default=None)
    args = ap.parse_args()

    fecha_min = _parse_fecha(args.desde) if args.desde else None
    fecha_max = _parse_fecha(args.hasta) if args.hasta else None

    if args.api:
        if not args.node_id or fecha_min is None or fecha_max is None:
            print("[ERROR] --api requiere --node-id, --desde y --hasta (YYYY-MM-DD)")
            return 1
        p = generar(
            None,
            args.promedio.resolve() if args.promedio else None,
            args.output,
            node_id_filter=args.node_id,
            fecha_min=fecha_min,
            fecha_max=fecha_max,
            usar_api=True,
        )
        print(p.resolve())
        return 0

    if not args.detalle.is_file():
        print(f"[ERROR] No existe detalle: {args.detalle}")
        return 1
    if args.promedio is not None and not args.promedio.is_file():
        print(f"[ERROR] No existe promedio: {args.promedio}")
        return 1

    prom = args.promedio
    if prom is None:
        prom_path: Path | None = None
    else:
        prom_path = prom.resolve()

    p = generar(
        args.detalle.resolve(),
        prom_path,
        args.output,
        node_id_filter=args.node_id,
        fecha_min=fecha_min,
        fecha_max=fecha_max,
        usar_api=False,
    )
    print(p.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

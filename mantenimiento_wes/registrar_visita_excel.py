# -*- coding: utf-8 -*-
"""
Registra una visita en:
  1) FORMULARIO_MANTENCION_WES_DIGITAL.xlsx (hoja Datos + espejo Ingreso)
  2) Copia local maestro/analisis_falla_google.xlsx (hoja Datos) si existe
  3) Google Sheet «Registro de fallas WES» (append) si hay credenciales Drive
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Border, PatternFill, Side

ROOT = Path(__file__).resolve().parent
XLSX = ROOT / "FORMULARIO_MANTENCION_WES_DIGITAL.xlsx"
MAESTRO_GOOGLE = ROOT / "maestro" / "analisis_falla_google.xlsx"
MAESTRO_ANIBAL = ROOT / "maestro" / "analisis_de_falla.xlsx"

SHEET_REGISTRO_ID = "1GlRn7QXWEre7ziau29ojR5lTl-bZ8T3mCT3cD93HZgM"

FILL_INPUT = PatternFill("solid", fgColor="FFF2CC")
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

# Columnas hoja Datos del formulario digital (B=2 ...)
HEADERS_DIGITAL = [
    "Cliente",
    "Maquina",
    "Tecnico",
    "Fecha",
    "Tipo de Mantenimiento",
    "Tipo de Falla",
    "Falla Expecifica",
    "Solucion y/o Diagnostico",
    "Observaciones",
    "Firma requerida",
    "N OT",
    "Estado visita",
    "Origen",
    "Email cliente",
    "Recibido por",
    "Cargo",
    "Comuna",
    "PDF / Drive",
    "Año",
    "Mes",
]


def _parse_fecha(value: Any):
    if value is None or value == "":
        return datetime.now().date()
    if isinstance(value, datetime):
        return value.date()
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return s


def _row_values(data: Dict[str, Any], *, pdf_link: str = "") -> Dict[str, Any]:
    fecha = _parse_fecha(data.get("fecha"))
    anio = fecha.year if hasattr(fecha, "year") else None
    mes = fecha.month if hasattr(fecha, "month") else None
    firma = (
        f"Sí - {data.get('recibido_por') or 'firmada'}"
        if data.get("firma_png") or data.get("recibido_por")
        else "No - solo digital"
    )
    return {
        "Cliente": data.get("cliente"),
        "Maquina": data.get("maquina"),
        "Tecnico": data.get("tecnico"),
        "Fecha": fecha,
        "Tipo de Mantenimiento": data.get("tipo_mtto"),
        "Tipo de Falla": data.get("tipo_falla"),
        "Falla Expecifica": data.get("falla_especifica"),
        "Solucion y/o Diagnostico": data.get("solucion"),
        "Observaciones": data.get("observaciones"),
        "Firma requerida": firma,
        "N OT": data.get("ot"),
        "Estado visita": data.get("estado_visita") or "cerrada",
        "Origen": f"Formulario web {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "Email cliente": (
            str(data.get("email_cliente") or "")
            + (
                f" | CC: {data.get('email_cc')}"
                if data.get("email_cc")
                else ""
            )
        ),
        "Recibido por": data.get("recibido_por"),
        "Cargo": data.get("cargo"),
        "Comuna": data.get("comuna"),
        "PDF / Drive": pdf_link or data.get("pdf_link") or "",
        "Año": anio,
        "Mes": mes,
    }


def _next_empty_row(ws, col: int = 2, start: int = 2) -> int:
    row = start
    while ws.cell(row, col).value not in (None, ""):
        row += 1
        if row > 20000:
            raise RuntimeError("Hoja Datos llena")
    return row


def _write_digital(
    data: Dict[str, Any],
    *,
    xlsx_path: Path,
    pdf_link: str = "",
) -> Tuple[Path, int]:
    path = Path(xlsx_path)
    if not path.is_file():
        raise FileNotFoundError(f"No existe Excel: {path}")
    wb = openpyxl.load_workbook(path)
    if "Datos" not in wb.sheetnames:
        raise RuntimeError("Falta hoja Datos")

    valores = _row_values(data, pdf_link=pdf_link)

    if "Ingreso" in wb.sheetnames:
        ing = wb["Ingreso"]
        ing["C4"] = valores["Cliente"]
        ing["C6"] = valores["Maquina"]
        ing["C8"] = valores["Tecnico"]
        ing["C10"] = valores["Fecha"]
        ing["C10"].number_format = "DD/MM/YYYY"
        ing["C12"] = valores["Tipo de Mantenimiento"]
        ing["C14"] = valores["Tipo de Falla"]
        ing["C16"] = valores["Falla Expecifica"]
        ing["C18"] = valores["Solucion y/o Diagnostico"]
        ing["C20"] = valores["Observaciones"]
        ing["C22"] = valores["Firma requerida"]
        ing["C24"] = valores["N OT"]
        ing["C26"] = valores["Estado visita"]

    datos = wb["Datos"]
    for i, h in enumerate(HEADERS_DIGITAL, start=2):
        if datos.cell(1, i).value != h:
            datos.cell(1, i, h)
    row = _next_empty_row(datos, 2)
    for i, h in enumerate(HEADERS_DIGITAL, start=2):
        cell = datos.cell(row, i, valores.get(h))
        cell.fill = FILL_INPUT
        cell.border = THIN
        if h == "Fecha":
            cell.number_format = "DD/MM/YYYY"
    wb.save(path)
    return path, row


def _append_local_maestro(path: Path, data: Dict[str, Any]) -> Optional[int]:
    if not path.is_file():
        return None
    wb = openpyxl.load_workbook(path)
    if "Datos" in wb.sheetnames:
        ws = wb["Datos"]
    elif "Data" in wb.sheetnames:
        ws = wb["Data"]
    else:
        return None

    row = _next_empty_row(ws, 2)
    fecha = _parse_fecha(data.get("fecha"))
    vals = [
        data.get("cliente"),
        data.get("maquina"),
        data.get("tecnico"),
        fecha,
        data.get("tipo_mtto"),
        data.get("tipo_falla"),
        data.get("falla_especifica"),
        data.get("solucion"),
        data.get("observaciones"),
    ]
    for i, v in enumerate(vals, start=2):
        ws.cell(row, i, v)
    ws.cell(row, 5).number_format = "DD/MM/YYYY"
    headers = [ws.cell(1, c).value for c in range(1, 15)]
    if "Año" in headers:
        col = headers.index("Año") + 1
        ws.cell(row, col, fecha.year if hasattr(fecha, "year") else None)
    if "Mes" in headers:
        col = headers.index("Mes") + 1
        ws.cell(row, col, fecha.month if hasattr(fecha, "month") else None)
    wb.save(path)
    return row


def _append_google_sheet(data: Dict[str, Any]) -> Dict[str, Any]:
    """Append a row to Registro de fallas WES!Datos (cols B-J)."""
    try:
        import os

        sys.path.insert(0, str(ROOT.parent))
        from google.auth.transport.requests import Request
        from google.oauth2.credentials import Credentials
        from googleapiclient.discovery import build

        creds = Credentials(
            token=None,
            refresh_token=os.environ.get("GOOGLE_DRIVE_REFRESH_TOKEN", "").strip(),
            token_uri="https://oauth2.googleapis.com/token",
            client_id=os.environ.get("GOOGLE_DRIVE_CLIENT_ID", "").strip(),
            client_secret=os.environ.get("GOOGLE_DRIVE_CLIENT_SECRET", "").strip(),
            scopes=["https://www.googleapis.com/auth/drive"],
        )
        creds.refresh(Request())
        sheets = build("sheets", "v4", credentials=creds, cache_discovery=False)
        fecha = _parse_fecha(data.get("fecha"))
        fecha_str = fecha.strftime("%Y-%m-%d") if hasattr(fecha, "strftime") else str(fecha)
        values = [[
            data.get("cliente") or "",
            data.get("maquina") or "",
            data.get("tecnico") or "",
            fecha_str,
            data.get("tipo_mtto") or "",
            data.get("tipo_falla") or "",
            data.get("falla_especifica") or "",
            data.get("solucion") or "",
            data.get("observaciones") or "",
        ]]
        result = (
            sheets.spreadsheets()
            .values()
            .append(
                spreadsheetId=SHEET_REGISTRO_ID,
                range="Datos!B:J",
                valueInputOption="USER_ENTERED",
                insertDataOption="INSERT_ROWS",
                body={"values": values},
            )
            .execute()
        )
        return {
            "ok": True,
            "updatedRange": result.get("updates", {}).get("updatedRange"),
            "updatedRows": result.get("updates", {}).get("updatedRows"),
        }
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": str(exc)}


def _upload_maestro_drive(path: Path) -> Dict[str, Any]:
    """Sube/actualiza la copia xlsx del maestro en Drive (sin Sheets API)."""
    try:
        sys.path.insert(0, str(ROOT.parent))
        from wes_google_drive import subir_a_drive

        info = subir_a_drive(path, subcarpeta="mantenimiento wes/maestro")
        return {"ok": True, "web_view_link": info.get("web_view_link"), "id": info.get("id")}
    except Exception as exc:  # noqa: BLE001
        return {"ok": False, "error": str(exc)}


def registrar_visita_en_excel(
    data: Dict[str, Any],
    xlsx_path: Optional[Path] = None,
    *,
    pdf_link: str = "",
    sync_google: bool = True,
) -> Tuple[Path, int]:
    path = Path(xlsx_path) if xlsx_path else XLSX
    out, row = _write_digital(data, xlsx_path=path, pdf_link=pdf_link)
    _append_local_maestro(MAESTRO_GOOGLE, data)
    _append_local_maestro(MAESTRO_ANIBAL, data)

    google_info: Dict[str, Any] = {"ok": False}
    if sync_google:
        google_info = _append_google_sheet(data)
        # Fallback sin Sheets API: subir Excel maestro actualizado a Drive
        if not google_info.get("ok"):
            up = _upload_maestro_drive(MAESTRO_ANIBAL if MAESTRO_ANIBAL.is_file() else MAESTRO_GOOGLE)
            google_info = {
                "ok": bool(up.get("ok")),
                "mode": "xlsx_upload",
                "web_view_link": up.get("web_view_link"),
                "error": up.get("error") or google_info.get("error"),
            }
            # también sube el digital mejorado
            _upload_maestro_drive(path)
        data["_google_sheet"] = google_info

    return out, row


if __name__ == "__main__":
    demo = {
        "cliente": "CORMUP",
        "maquina": "TOBALABA",
        "tecnico": "Anibal Aranda",
        "fecha": "2026-08-12",
        "tipo_mtto": "Mtto Preventivo",
        "tipo_falla": "Auditoría",
        "falla_especifica": "Validación Data",
        "solucion": "Demo registro maestro",
        "observaciones": "",
        "ot": "OT-DEMO",
        "estado_visita": "cerrada",
        "recibido_por": "Demo",
        "email_cliente": "demo@wes.cl",
        "firma_png": "x",
    }
    p, r = registrar_visita_en_excel(demo, sync_google=False)
    print(f"OK {p} fila {r}")

# -*- coding: utf-8 -*-
"""
Fuente ÚNICA de correos para actas:
  Drive: CONTACTOS_ENVIOS_ACTAS
  Path local Windows: G:\\Mi unidad\\Agente WES\\wes-scripts\\mantenimiento wes\\CONTACTOS_ENVIOS_ACTAS
  Sheet ID: 1Tpjm1eXRXKuKvxachtbYVr9503wICJdsYDTjkbm__o8

Columnas (hoja Contactos):
  Cliente | Máquina | Rol | Nombre | Cargo | Email | Actualizado

Roles:
  general → TO (encargado general del cliente; puede haber varios)
  CC / cc / punto → CC (del punto/máquina; Máquina vacía = CC a nivel cliente)

Uso:
  python contactos_cliente.py --desde-drive
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

ROOT = Path(__file__).resolve().parent
MAESTRO = ROOT / "maestro" / "CONTACTOS_ENVIOS_ACTAS.xlsx"
JSON_OUT = ROOT / "catalogos" / "contactos_cliente.json"
FUENTE_TXT = ROOT / "FUENTE_CONTACTOS_ENVIOS.txt"

SHEET_ID = "1Tpjm1eXRXKuKvxachtbYVr9503wICJdsYDTjkbm__o8"
SHEET_URL = (
    "https://docs.google.com/spreadsheets/d/"
    f"{SHEET_ID}/edit"
)


def _norm(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _email_ok(v: Any) -> bool:
    s = _norm(v)
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", s))


def _rol_tipo(rol: str) -> str:
    r = _norm(rol).lower()
    if r in {"general", "to", "firmante"}:
        return "general"
    if r in {"cc", "punto", "copia"}:
        return "cc"
    # roles descriptivos → CC a nivel cliente si no dicen general
    if r:
        return "cc"
    return ""


def _download_drive(dest: Path = MAESTRO) -> Path:
    sys.path.insert(0, str(ROOT.parent))
    from wes_google_drive import obtener_servicio_drive
    from googleapiclient.http import MediaIoBaseDownload
    import io

    svc = obtener_servicio_drive()
    req = svc.files().export_media(
        fileId=SHEET_ID,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    fh = io.BytesIO()
    dl = MediaIoBaseDownload(fh, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(fh.getvalue())
    return dest


def _read_rows(path: Path) -> List[Dict[str, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    if "Contactos" not in wb.sheetnames:
        raise RuntimeError(f"Falta hoja Contactos en {path}")
    ws = wb["Contactos"]
    # header flexible
    headers = [_norm(ws.cell(1, c).value).lower() for c in range(1, 12)]
    def idx(*names):
        for n in names:
            if n in headers:
                return headers.index(n) + 1
        return None

    c_cli = idx("cliente") or 1
    c_maq = idx("máquina", "maquina", "sitio") or 2
    c_rol = idx("rol") or 3
    c_nom = idx("nombre") or 4
    c_car = idx("cargo") or 5
    c_em = idx("email", "correo") or 6
    c_act = idx("actualizado") or 7

    rows = []
    for r in range(2, (ws.max_row or 1) + 1):
        cli = _norm(ws.cell(r, c_cli).value)
        email = _norm(ws.cell(r, c_em).value)
        if not cli and not email:
            continue
        rows.append({
            "cliente": cli,
            "maquina": _norm(ws.cell(r, c_maq).value),
            "rol": _norm(ws.cell(r, c_rol).value),
            "nombre": _norm(ws.cell(r, c_nom).value),
            "cargo": _norm(ws.cell(r, c_car).value),
            "email": email,
            "actualizado": _norm(ws.cell(r, c_act).value),
        })
    return rows


def export_json(rows: List[Dict[str, str]]) -> Path:
    payload = []
    for r in rows:
        if not _norm(r.get("cliente")):
            continue
        rol_t = _rol_tipo(r.get("rol", ""))
        cargo = r.get("cargo", "")
        nombre = r.get("nombre", "")
        # heurística firmante: linkes / mantencion / supervisión mtto
        blob = f"{cargo} {nombre} {r.get('rol','')}".lower()
        firmante = any(
            k in blob
            for k in ("linkes", "mantencion", "mantención", "mtto", "firmante", "supervis")
        )
        # múltiples general: todos van a TO list (formulario reparte)
        payload.append({
            "cliente": r.get("cliente", ""),
            "sitio": r.get("maquina", ""),  # alias para el form
            "maquina": r.get("maquina", ""),
            "rol": r.get("rol", "") or ("general" if rol_t == "general" else "CC"),
            "rol_tipo": rol_t or "cc",
            "nombre": nombre,
            "cargo": cargo,
            "email": r.get("email", ""),
            "firmante": firmante,
            "enviar_to": rol_t == "general",
            "enviar_cc": rol_t == "cc",
            "activo": True,
            "email_ok": _email_ok(r.get("email")),
            "notas": "",
            "fuente": "CONTACTOS_ENVIOS_ACTAS",
        })
    meta = {
        "_meta": {
            "fuente": "CONTACTOS_ENVIOS_ACTAS",
            "sheet_id": SHEET_ID,
            "web_link": SHEET_URL,
            "actualizado": datetime.now().isoformat(timespec="seconds"),
            "total": len(payload),
        }
    }
    # El form espera un array; guardamos meta en archivo hermano y array limpio.
    JSON_OUT.parent.mkdir(parents=True, exist_ok=True)
    JSON_OUT.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    (JSON_OUT.parent / "contactos_cliente_meta.json").write_text(
        json.dumps(meta["_meta"], ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return JSON_OUT


def contactos_para(cliente: str, sitio: str = "") -> List[Dict[str, Any]]:
    if not JSON_OUT.is_file():
        return []
    data = json.loads(JSON_OUT.read_text(encoding="utf-8"))
    cli = (cliente or "").strip().upper()
    sit = (sitio or "").strip().upper()
    out = []
    for c in data:
        if (c.get("cliente") or "").upper() != cli:
            continue
        c_sit = (c.get("sitio") or c.get("maquina") or "").upper()
        rol_t = c.get("rol_tipo") or _rol_tipo(c.get("rol", ""))
        if rol_t == "general":
            # generales del cliente (máquina vacía) o del mismo sitio
            if c_sit and sit and c_sit != sit:
                continue
            if c_sit and not sit:
                # general ligado a una máquina: solo si se eligió esa
                continue
            out.append(c)
            continue
        # CC / punto
        if c_sit and sit:
            if c_sit == sit:
                out.append(c)
            continue
        if not c_sit:
            # CC a nivel cliente (JO, medio ambiente, etc.)
            out.append(c)
    return out


def emails_destino(cliente: str, sitio: str = "") -> Dict[str, Any]:
    contacts = contactos_para(cliente, sitio)
    to_list, cc_list = [], []
    firmantes = []
    for c in contacts:
        em = (c.get("email") or "").strip()
        if not _email_ok(em):
            continue
        if c.get("enviar_to") or c.get("rol_tipo") == "general":
            to_list.append(em)
        if c.get("enviar_cc") or c.get("rol_tipo") == "cc":
            cc_list.append(em)
        if c.get("firmante"):
            firmantes.append(c)

    def uniq(xs: List[str]) -> List[str]:
        seen, out = set(), []
        for x in xs:
            k = x.lower()
            if k in seen:
                continue
            seen.add(k)
            out.append(x)
        return out

    to_u = uniq(to_list)
    cc_u = [e for e in uniq(cc_list) if e.lower() not in {t.lower() for t in to_u}]
    return {"to": to_u, "cc": cc_u, "firmantes": firmantes, "contactos": contacts}


def escribir_fuente_txt() -> None:
    FUENTE_TXT.write_text(
        f"""Fuente ÚNICA de correos para envío de actas
===========================================
Archivo: CONTACTOS_ENVIOS_ACTAS
Windows: G:\\Mi unidad\\Agente WES\\wes-scripts\\mantenimiento wes\\CONTACTOS_ENVIOS_ACTAS
Drive:   {SHEET_URL}

NO editar correos en FORMULARIO_MANTENCION_WES_DIGITAL ni en otras planillas.
Editá solo CONTACTOS_ENVIOS_ACTAS y pedí sincronizar:
  python contactos_cliente.py --desde-drive

Reglas:
  Rol=general → TO (puede haber varios: JO, Linkes, etc.)
  Rol=CC/punto → CC del punto (Máquina) o del cliente si Máquina vacía
  Aníbal queda siempre en CC adicional desde el formulario.

Sync: {datetime.now().strftime('%Y-%m-%d %H:%M')}
""",
        encoding="utf-8",
    )


def sincronizar(*, desde_drive: bool = False, fuente: Optional[Path] = None) -> dict:
    path = fuente
    if desde_drive:
        path = _download_drive(MAESTRO)
    elif path is None:
        path = MAESTRO if MAESTRO.is_file() else None
        if path is None:
            path = _download_drive(MAESTRO)

    rows = _read_rows(path)
    export_json(rows)
    escribir_fuente_txt()

    con_email = sum(1 for r in rows if _email_ok(r.get("email")))
    clientes = sorted({r["cliente"] for r in rows if r.get("cliente")})
    return {
        "fuente": str(path),
        "sheet_url": SHEET_URL,
        "filas": len(rows),
        "con_email": con_email,
        "clientes": len(clientes),
        "clientes_lista": clientes,
        "json": str(JSON_OUT),
        "sync_at": datetime.now().isoformat(timespec="seconds"),
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--desde-drive", action="store_true")
    ap.add_argument("--fuente", type=Path, default=None)
    args = ap.parse_args()
    info = sincronizar(desde_drive=args.desde_drive, fuente=args.fuente)
    print("OK sync CONTACTOS_ENVIOS_ACTAS")
    print(f"  sheet: {info['sheet_url']}")
    print(f"  filas: {info['filas']} (con email: {info['con_email']})")
    print(f"  clientes: {info['clientes']}")
    print(f"  json: {info['json']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

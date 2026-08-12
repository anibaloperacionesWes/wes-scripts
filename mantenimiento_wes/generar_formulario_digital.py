# -*- coding: utf-8 -*-
"""
Genera FORMULARIO_MANTENCION_WES_DIGITAL.xlsx alineado al Google Sheet
«Registro de fallas WES» (hoja Ingreso + Bases con listas en cascada).

Fuente Google:
  https://docs.google.com/spreadsheets/d/1GlRn7QXWEre7ziau29ojR5lTl-bZ8T3mCT3cD93HZgM
Export local preferido:
  mantenimiento wes/maestro/analisis_falla_google.xlsx
Fallback:
  G:\\Mi unidad\\Anibal\\analisis de falla.xlsx
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "FORMULARIO_MANTENCION_WES_DIGITAL.xlsx"
SRC_GOOGLE = ROOT / "maestro" / "analisis_falla_google.xlsx"
SRC_EXISTENTE = ROOT / "FORMULARIO_MANTENCION_WES_DIGITAL.xlsx"
SRC_LOCAL = Path(r"G:\Mi unidad\Anibal\analisis de falla.xlsx")
SHEET_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "1GlRn7QXWEre7ziau29ojR5lTl-bZ8T3mCT3cD93HZgM"
)

THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_SECTION = PatternFill("solid", fgColor="2E75B6")
FILL_INPUT = PatternFill("solid", fgColor="FFF2CC")
FILL_LIGHT = PatternFill("solid", fgColor="DDEBF7")
FILL_OK = PatternFill("solid", fgColor="E2EFDA")
FONT_WHITE = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
FONT_TITLE = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
FONT_LABEL = Font(name="Calibri", bold=True, size=11, color="1F4E79")
FONT_NORMAL = Font(name="Calibri", size=11)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

TECNICOS = [
    "Maurico Orellana",
    "Anibal Aranda",
    "Jose Otarola",
    "Gabriel Prieto",
    "Mirko Lorca",
    "Cristian Sepulveda",
    "Salvador Cantillana",
    "Jeans Araya",
]
TIPOS_MTTO = [
    "Mtto Correctivo",
    "Mtto Preventivo",
    "Mtto Predictivo",
    "Reubicacion CIR",
    "Instalación Valvula On/Off",
    "Instalación sistema CIR",
    "Instalación sistema CPA",
]


def _pick_source() -> Path:
    if SRC_GOOGLE.is_file():
        return SRC_GOOGLE
    if SRC_LOCAL.is_file():
        return SRC_LOCAL
    if SRC_EXISTENTE.is_file():
        # Regenerar catálogos desde el propio digital si no hay export fresco
        return SRC_EXISTENTE
    raise FileNotFoundError(
        f"No está {SRC_GOOGLE} ni {SRC_LOCAL}. Exportá el Sheet primero."
    )


def _dv(ws, formula: str, cells: str) -> None:
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = "Elegí un valor de la lista"
    validation.errorTitle = "Valor inválido"
    ws.add_data_validation(validation)
    validation.add(cells)


def _copy_base1_base3(src_wb: openpyxl.Workbook, dst: Workbook) -> tuple[int, int]:
    """Copia catálogos Cliente/Máquina y Tipo falla/Falla específica."""
    # Base1
    ws_src = src_wb["Base1"] if "Base1" in src_wb.sheetnames else None
    ws1 = dst.create_sheet("Base1")
    ws1["B1"] = "CLIENTE"
    ws1["C1"] = "MAQUINA"
    ws1["B1"].font = FONT_WHITE
    ws1["C1"].font = FONT_WHITE
    ws1["B1"].fill = FILL_HEADER
    ws1["C1"].fill = FILL_HEADER
    n1 = 0
    if ws_src is not None:
        for r in range(2, ws_src.max_row + 1):
            cli = ws_src.cell(r, 2).value
            maq = ws_src.cell(r, 3).value
            if not cli and not maq:
                continue
            n1 += 1
            ws1.cell(n1 + 1, 2, cli)
            ws1.cell(n1 + 1, 3, maq)
    else:
        # fallback desde Data/Datos
        data_name = "Datos" if "Datos" in src_wb.sheetnames else "Data"
        wsd = src_wb[data_name]
        seen: set[tuple[str, str]] = set()
        for r in range(2, wsd.max_row + 1):
            cli = wsd.cell(r, 2).value
            maq = wsd.cell(r, 3).value
            if not cli or not maq:
                continue
            key = (str(cli).strip(), str(maq).strip())
            if key in seen:
                continue
            seen.add(key)
            n1 += 1
            ws1.cell(n1 + 1, 2, key[0])
            ws1.cell(n1 + 1, 3, key[1])
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 40

    # Base3
    ws3 = dst.create_sheet("Base3")
    ws3["B1"] = "Tipo de falla"
    ws3["C1"] = "Falla Específica"
    ws3["B1"].font = FONT_WHITE
    ws3["C1"].font = FONT_WHITE
    ws3["B1"].fill = FILL_HEADER
    ws3["C1"].fill = FILL_HEADER
    n3 = 0
    if "Base3" in src_wb.sheetnames:
        ws_src3 = src_wb["Base3"]
        for r in range(2, ws_src3.max_row + 1):
            t = ws_src3.cell(r, 2).value
            f = ws_src3.cell(r, 3).value
            if not t and not f:
                continue
            n3 += 1
            ws3.cell(n3 + 1, 2, t)
            ws3.cell(n3 + 1, 3, f)
    ws3.column_dimensions["B"].width = 22
    ws3.column_dimensions["C"].width = 36
    return n1, n3


def _add_filter_helpers(dst: Workbook, n1: int, n3: int) -> None:
    """Base2 / Base 4: FILTER dependientes de Ingreso (Excel 365 / Sheets)."""
    ws2 = dst.create_sheet("Base2")
    # Filtra máquinas del cliente elegido en Ingreso!C4
    ws2["A1"] = (
        f'=IFERROR(FILTER(Base1!B2:C{n1 + 1},Base1!B2:B{n1 + 1}=Ingreso!C4),"")'
    )
    ws2["A1"].font = Font(name="Calibri", size=10, italic=True, color="666666")
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 40

    ws4 = dst.create_sheet("Base 4")
    ws4["A1"] = (
        f'=IFERROR(FILTER(Base3!B2:C{n3 + 1},Base3!B2:B{n3 + 1}=Ingreso!C14),"")'
    )
    ws4["A1"].font = Font(name="Calibri", size=10, italic=True, color="666666")
    ws4.column_dimensions["A"].width = 22
    ws4.column_dimensions["B"].width = 36


def _build_ingreso(dst: Workbook, n1: int, n3: int) -> None:
    ws = dst.create_sheet("Ingreso", 0)
    ws.merge_cells("B2:F2")
    ws["B2"] = "INGRESO AUTOMATIZADO — Registro de fallas WES"
    ws["B2"].font = FONT_TITLE
    ws["B2"].fill = FILL_HEADER
    ws["B2"].alignment = CENTER
    ws.row_dimensions[2].height = 26

    ws.merge_cells("B3:F3")
    ws["B3"] = (
        "Igual que la hoja Ingreso del Google Sheet. "
        "Listas en cascada: Cliente→Máquina y Tipo falla→Falla específica. "
        "Celdas amarillas = completar."
    )
    ws["B3"].font = Font(name="Calibri", italic=True, size=10, color="FFFFFF")
    ws["B3"].fill = FILL_SECTION

    fields = [
        (4, "Cliente", "C4"),
        (6, "Maquina", "C6"),
        (8, "Tecnico", "C8"),
        (10, "Fecha", "C10"),
        (12, "Tipo de Mantenimiento", "C12"),
        (14, "Tipo de Falla", "C14"),
        (16, "Falla Expecifica", "C16"),
        (18, "solucion", "C18"),
        (20, "Observaciones", "C20"),
        (22, "Firma requerida", "C22"),
        (24, "N OT / formulario", "C24"),
        (26, "Estado visita", "C26"),
    ]
    for row, label, _ in fields:
        ws.cell(row, 2, label).font = FONT_LABEL
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        cell = ws.cell(row, 3, "")
        cell.fill = FILL_INPUT
        cell.border = THIN
        cell.alignment = LEFT

    ws["C10"] = date.today()
    ws["C10"].number_format = "DD/MM/YYYY"
    ws["C22"] = "No - solo digital"
    ws["C26"] = "cerrada"

    # Filas altas para texto largo
    ws.row_dimensions[18].height = 45
    ws.row_dimensions[20].height = 45
    ws.merge_cells("C18:F19")
    ws.merge_cells("C20:F21")
    ws["C18"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["C20"].alignment = Alignment(wrap_text=True, vertical="top")

    ws.merge_cells("B28:F28")
    ws["B28"] = (
        "Al terminar: 1) revisar listas  2) ejecutar "
        "`python registrar_desde_ingreso.py` para volcar a Datos  "
        "3) pedir al agente el correo de cierre  "
        f"4) Sheet vivo: {SHEET_URL}"
    )
    ws["B28"].font = Font(name="Calibri", size=10, italic=True)
    ws["B28"].fill = FILL_OK

    for letter, width in {"B": 24, "C": 18, "D": 14, "E": 14, "F": 18}.items():
        ws.column_dimensions[letter].width = width

    # Validaciones (mismas ideas que el Sheet)
    _dv(ws, f"Base1!$B$2:$B${n1 + 1}", "C4")
    _dv(ws, "Base2!$B$1:$B$151", "C6")
    _dv(ws, '"' + ",".join(TECNICOS) + '"', "C8")
    _dv(ws, '"' + ",".join(TIPOS_MTTO) + '"', "C12")
    _dv(ws, f"Base3!$B$2:$B${n3 + 1}", "C14")
    _dv(ws, "'Base 4'!$B$1:$B$31", "C16")
    _dv(ws, '"Sí - acta firmada,No - solo digital"', "C22")
    _dv(ws, '"abierta,en_curso,cerrada"', "C26")


def _build_datos(dst: Workbook) -> None:
    ws = dst.create_sheet("Datos", 1)
    headers = [
        "Cliente",
        "Maquina",
        "Tecnico",
        "Fecha",
        "Tipo de Mantenimiento",
        "Tipo de Falla",
        "Falla Expecifica",
        "Solucion y/o Diagnostico",
        "Observaciones",
        "Firma requerida",
        "N OT",
        "Estado visita",
        "Origen",
    ]
    for i, h in enumerate(headers, start=2):
        cell = ws.cell(1, i, h)
        cell.fill = FILL_HEADER
        cell.font = FONT_WHITE
        cell.alignment = CENTER
        cell.border = THIN
        ws.column_dimensions[get_column_letter(i)].width = 18 if i != 9 else 40
    ws.column_dimensions["I"].width = 40
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = "B1:N2000"
    for row in range(2, 52):
        for col in range(2, 15):
            cell = ws.cell(row, col)
            cell.fill = FILL_INPUT
            cell.border = THIN
        ws.cell(row, 5).number_format = "DD/MM/YYYY"


def _build_formulario_visita(dst: Workbook) -> None:
    ws = dst.create_sheet("Formulario Visita", 2)
    ws.merge_cells("A1:H1")
    ws["A1"] = "FORMULARIO DE MANTENCIÓN WES — ACTA DIGITAL (checklist)"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_HEADER
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:H2")
    ws["A2"] = (
        "Opcional / cuando necesiten checklist CIR-CPA-SAB o firma. "
        "El ingreso operativo diario es la hoja Ingreso."
    )
    ws["A2"].font = Font(name="Calibri", italic=True, color="FFFFFF", size=10)
    ws["A2"].fill = FILL_SECTION

    def section(row: int, title: str) -> None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row, 1, title)
        cell.font = FONT_WHITE
        cell.fill = FILL_SECTION

    section(4, "Checklist CIR / CPA — Estado: OK / Observación / Falla / N/A")
    cir = [
        "Canalización 24V AC",
        "Canalización 220V AC",
        "Poste solar",
        "Panel solar",
        "Batería solar",
        "Cargador solar",
        "Router / Amplimax",
        "Conectividad App",
        "Fuente switching 12V",
        "Batería respaldo CIR",
        "Estado exterior caja CIR",
        "Estado interior caja CIR",
        "Enlace CIR",
        "Voltaje 12V DC CIR",
        "Voltaje 9V DC CIR",
        "Voltaje 5V DC CIR",
        "Equipo ultrasónico",
        "Transductores",
        "Sensor de pulso",
    ]
    cpa = [
        "Presurización cañerías",
        "Cámara albañilería",
        "Tapa metálica",
        "Llaves de paso",
        "Válvula presión alta",
        "Válvula presión baja",
        "Válvula presión nocturna",
        "Solenoide válvula alta",
        "Solenoide válvula baja",
        "Solenoide válvula nocturna",
        "Func. presión alta",
        "Func. presión baja",
        "Func. presión nocturna",
        "Func. medidor",
    ]
    ws["A5"] = "CIR — Eléctrico/Electrónico"
    ws["A5"].fill = FILL_HEADER
    ws["A5"].font = FONT_WHITE
    ws.merge_cells("A5:D5")
    ws["E5"] = "CPA — Hídrico/Cámara"
    ws["E5"].fill = FILL_HEADER
    ws["E5"].font = FONT_WHITE
    ws.merge_cells("E5:H5")
    for col, text in ((1, "Elemento"), (2, "Estado"), (3, "Obs."),
                      (5, "Elemento"), (6, "Estado"), (7, "Obs.")):
        ws.cell(6, col, text).fill = FILL_LIGHT
        ws.cell(6, col).font = FONT_LABEL
    start = 7
    n = max(len(cir), len(cpa))
    for i in range(n):
        row = start + i
        if i < len(cir):
            ws.cell(row, 1, cir[i])
            ws.cell(row, 2, "OK").fill = FILL_INPUT
            ws.cell(row, 2).border = THIN
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
            ws.cell(row, 3).fill = FILL_INPUT
            ws.cell(row, 3).border = THIN
        if i < len(cpa):
            ws.cell(row, 5, cpa[i])
            ws.cell(row, 6, "OK").fill = FILL_INPUT
            ws.cell(row, 6).border = THIN
            ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
            ws.cell(row, 7).fill = FILL_INPUT
            ws.cell(row, 7).border = THIN
    end = start + n - 1
    _dv(ws, '"OK,Observación,Falla,N/A"', f"B{start}:B{end}")
    _dv(ws, '"OK,Observación,Falla,N/A"', f"F{start}:F{end}")

    for letter, width in {
        "A": 28,
        "B": 14,
        "C": 16,
        "D": 12,
        "E": 28,
        "F": 14,
        "G": 16,
        "H": 12,
    }.items():
        ws.column_dimensions[letter].width = width


def _build_instrucciones(dst: Workbook, source: Path) -> None:
    ws = dst.create_sheet("Instrucciones", 0)
    ws["A1"] = "Formulario digital WES = espejo del ingreso automatizado del Sheet"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    lines = [
        "",
        f"Sheet vivo (fuente oficial): {SHEET_URL}",
        f"Catálogos tomados de: {source}",
        "",
        "Hojas:",
        "  • Ingreso — igual al Sheet: Cliente, Máquina, Técnico, Fecha, Tipo mtto,",
        "    Tipo falla, Falla específica, solución, observaciones (+ firma/OT).",
        "    Máquina y Falla específica se filtran según Cliente y Tipo de falla",
        "    (requiere Excel 365 o Google Sheets para las fórmulas FILTER).",
        "  • Datos — historial local (se llena con registrar_desde_ingreso.py).",
        "  • Formulario Visita — checklist CIR/CPA del acta (opcional).",
        "  • Base1 / Base2 / Base3 / Base 4 — catálogos y filtros (como el Sheet).",
        "",
        "Flujo recomendado:",
        "  1) Completar Ingreso (celdas amarillas).",
        "  2) python registrar_desde_ingreso.py",
        "  3) Pedir al agente el correo de cierre.",
        "  4) Resumen semanal → Drive carpeta mantenimiento wes.",
        "",
        "Sin Google Form. Firma: acta PDF si el cliente la pide; si no, solo digital.",
        f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    ]
    for i, line in enumerate(lines, start=2):
        ws.cell(i, 1, line)
    ws.column_dimensions["A"].width = 110


def build() -> Path:
    source = _pick_source()
    src_wb = openpyxl.load_workbook(source, data_only=True)

    wb = Workbook()
    # quitar hoja default luego de crear Instrucciones
    default = wb.active
    wb.remove(default)

    n1, n3 = _copy_base1_base3(src_wb, wb)
    _add_filter_helpers(wb, n1, n3)
    _build_ingreso(wb, n1, n3)
    _build_datos(wb)
    _build_formulario_visita(wb)
    _build_instrucciones(wb, source)

    # Orden de hojas similar al Sheet + extras
    order = [
        "Instrucciones",
        "Ingreso",
        "Datos",
        "Formulario Visita",
        "Base1",
        "Base2",
        "Base3",
        "Base 4",
    ]
    for idx, name in enumerate(order):
        wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))

    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    path = build()
    print(f"OK {path}")
    print(f"size={path.stat().st_size}")

# -*- coding: utf-8 -*-
"""
Crea la planilla permanente de ingreso para técnicos WES y la comparte.

Uso:
  python3 mantenimiento_wes/compartir_acceso_tecnicos.py
"""

from __future__ import annotations

import json
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent
REPO = ROOT.parent
if str(REPO) not in sys.path:
    sys.path.insert(0, str(REPO))

from wes_google_drive import asegurar_carpeta, obtener_servicio_drive  # noqa: E402

FOLDER_TECH = "Tecnicos_WES_Formulario"
SHEET_NAME = "INGRESO_visitas_tecnicos_WES"
XLSX_LOCAL = ROOT / "INGRESO_visitas_tecnicos_WES.xlsx"
STATE = ROOT / "apps_script" / "ESTADO_PUBLICACION.json"

TECNICOS_EMAILS = [
    "anibal.aoperaciones@wes.cl",
    "mauricioorellana@wes.cl",
    "joseotarola@wes.cl",
    "juanlopez@wes.cl",
    "diegocarrasco@wes.cl",
    "benjamingumucio@wes.cl",
]

HEADERS = [
    "Folio / OT",
    "Fecha",
    "Hora",
    "Tecnico",
    "Cliente",
    "Maquina / sitio",
    "Comuna",
    "Motivo",
    "Tecnologia",
    "Tipo de Mantenimiento",
    "Tipo de Falla",
    "Falla especifica",
    "Solucion / diagnostico",
    "Observaciones",
    "Estado visita",
    "Lectura medidor",
    "Recibido por",
    "Cargo",
    "Email cliente",
    "Firma (si/no)",
]


def _load_cats():
    cat = ROOT / "catalogos"
    return {
        "clientes_maquinas": json.loads((cat / "clientes_maquinas.json").read_text(encoding="utf-8")),
        "tipos_falla": json.loads((cat / "tipos_falla.json").read_text(encoding="utf-8")),
        "opciones": json.loads((cat / "opciones.json").read_text(encoding="utf-8")),
    }


def armar_xlsx() -> Path:
    cats = _load_cats()
    wb = Workbook()

    ws_i = wb.active
    ws_i.title = "INSTRUCCIONES"
    ws_i["A1"] = "INGRESO DE VISITAS — TÉCNICOS WES"
    ws_i["A1"].font = Font(bold=True, size=18, color="1F4E79")
    ws_i.merge_cells("A1:F1")
    lineas = [
        "",
        "Este es el lugar permanente para cargar visitas desde el celular o el PC.",
        "No usa el link temporal de Cloudflare ni el editor de Apps Script.",
        "",
        "Cómo entrar",
        "1) Abrí la hoja INGRESO (pestaña de abajo).",
        "2) Completá UNA FILA por visita (hoy ya hay una fila de ejemplo; borrala o pisala).",
        "3) Cliente / máquina / técnico / tipo de falla van con lista desplegable.",
        "4) No borren las filas de encabezado.",
        "",
        "Qué pasa con los datos",
        "Quedan en esta planilla (carpeta Tecnicos_WES_Formulario en Drive).",
        "Anibal / operaciones los consolidan al Registro de fallas WES.",
        "",
        "Si Google dice Access / You need access",
        "Entrá con un correo @wes.cl (el de la empresa). En el celular: cuenta Google → wes.cl.",
        "Este archivo también está abierto con el link (cualquiera que lo tenga puede escribir).",
    ]
    for i, t in enumerate(lineas, start=2):
        ws_i[f"A{i}"] = t
        if t.startswith("Cómo") or t.startswith("Qué") or t.startswith("Si Google"):
            ws_i[f"A{i}"].font = Font(bold=True, color="1F4E79")
    ws_i.column_dimensions["A"].width = 110
    ws_i.row_dimensions[1].height = 28

    # Listas (ocultas)
    ws_l = wb.create_sheet("LISTAS")
    clientes = sorted(cats["clientes_maquinas"].keys())
    tecnicos = cats["opciones"]["tecnicos"]
    tipos_mtto = cats["opciones"]["tipos_mtto"]
    tipos_falla = sorted(cats["tipos_falla"].keys())
    motivos = cats["opciones"]["motivos"]
    tecno = cats["opciones"]["tecnologias"]
    estados = ["cerrada", "en_curso", "abierta"]
    firma = ["si", "no"]
    maquinas = sorted({m for ms in cats["clientes_maquinas"].values() for m in ms})
    fallas_esp = sorted({f for fs in cats["tipos_falla"].values() for f in fs})

    def dump(col, header, values):
        ws_l.cell(1, col, header)
        for i, v in enumerate(values, start=2):
            ws_l.cell(i, col, v)
        return len(values) + 1

    n_cli = dump(1, "Cliente", clientes)
    n_tec = dump(2, "Tecnico", tecnicos)
    n_mtto = dump(3, "TipoMtto", tipos_mtto)
    n_tf = dump(4, "TipoFalla", tipos_falla)
    n_mot = dump(5, "Motivo", motivos)
    n_tecno = dump(6, "Tecnologia", tecno)
    n_est = dump(7, "Estado", estados)
    n_firma = dump(8, "Firma", firma)
    n_maq = dump(9, "Maquina", maquinas)
    n_fe = dump(10, "FallaEsp", fallas_esp)
    ws_l.sheet_state = "hidden"

    ws = wb.create_sheet("INGRESO", 0)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    for i, h in enumerate(HEADERS, start=1):
        cell = ws.cell(1, i, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = 22
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}200"

    # Fila ejemplo
    ejemplo = [
        2250,
        date.today().isoformat(),
        "09:00",
        tecnicos[0],
        clientes[0],
        cats["clientes_maquinas"][clientes[0]][0],
        "",
        motivos[0],
        tecno[0],
        tipos_mtto[0],
        tipos_falla[0],
        cats["tipos_falla"][tipos_falla[0]][0],
        "Ejemplo: reemplazar por la visita real",
        "",
        "cerrada",
        "",
        "",
        "",
        "",
        "si",
    ]
    for i, v in enumerate(ejemplo, start=1):
        c = ws.cell(2, i, v)
        c.fill = input_fill

    # Validaciones columnas 2..200
    def dv(formula, col):
        letter = get_column_letter(col)
        v = DataValidation(type="list", formula1=formula, allow_blank=True)
        v.error = "Elegí un valor de la lista"
        v.errorTitle = "Valor no válido"
        v.prompt = "Elegí de la lista"
        v.showErrorMessage = True
        ws.add_data_validation(v)
        v.add(f"{letter}2:{letter}200")

    dv(f"LISTAS!$B$2:$B${n_tec}", 4)  # Tecnico
    dv(f"LISTAS!$A$2:$A${n_cli}", 5)  # Cliente
    dv(f"LISTAS!$I$2:$I${n_maq}", 6)  # Maquina
    dv(f"LISTAS!$E$2:$E${n_mot}", 8)
    dv(f"LISTAS!$F$2:$F${n_tecno}", 9)
    dv(f"LISTAS!$C$2:$C${n_mtto}", 10)
    dv(f"LISTAS!$D$2:$D${n_tf}", 11)
    dv(f"LISTAS!$J$2:$J${n_fe}", 12)
    dv(f"LISTAS!$G$2:$G${n_est}", 15)
    dv(f"LISTAS!$H$2:$H${n_firma}", 20)

    note = wb.create_sheet("CATALOGO_CLIENTE_MAQUINA")
    note["A1"] = "Cliente"
    note["B1"] = "Maquina"
    note["A1"].font = Font(bold=True)
    note["B1"].font = Font(bold=True)
    r = 2
    for cli, maqs in sorted(cats["clientes_maquinas"].items()):
        for m in maqs:
            note.cell(r, 1, cli)
            note.cell(r, 2, m)
            r += 1
    note.column_dimensions["A"].width = 28
    note.column_dimensions["B"].width = 48

    wb.save(XLSX_LOCAL)
    return XLSX_LOCAL


def compartir(service, file_id: str, *, anyone_writer: bool = False) -> None:
    from googleapiclient.errors import HttpError

    def add(body, notify=False):
        try:
            service.permissions().create(
                fileId=file_id,
                body=body,
                sendNotificationEmail=notify,
                fields="id",
            ).execute()
        except HttpError as e:
            print(f"[WARN] permiso {body}: {e.resp.status} {e._get_reason()}")

    add({"type": "domain", "role": "writer", "domain": "wes.cl", "allowFileDiscovery": True})
    if anyone_writer:
        add({"type": "anyone", "role": "writer", "allowFileDiscovery": False})
    for email in TECNICOS_EMAILS:
        add({"type": "user", "role": "writer", "emailAddress": email}, notify=True)


def main() -> int:
    from googleapiclient.http import MediaFileUpload

    path = armar_xlsx()
    service = obtener_servicio_drive()
    folder = asegurar_carpeta(service, FOLDER_TECH)

    # upsert by name
    safe = SHEET_NAME.replace("'", "\\'")
    q = (
        f"name='{safe}' and '{folder['id']}' in parents and trashed=false "
        "and mimeType='application/vnd.google-apps.spreadsheet'"
    )
    found = (
        service.files()
        .list(q=q, fields="files(id,webViewLink)", pageSize=1)
        .execute()
        .get("files", [])
    )
    media = MediaFileUpload(
        str(path),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=True,
    )
    body = {
        "name": SHEET_NAME,
        "mimeType": "application/vnd.google-apps.spreadsheet",
        "parents": [folder["id"]],
    }
    if found:
        f = (
            service.files()
            .update(
                fileId=found[0]["id"],
                media_body=media,
                fields="id,webViewLink",
            )
            .execute()
        )
    else:
        f = (
            service.files()
            .create(body=body, media_body=media, fields="id,webViewLink")
            .execute()
        )

    compartir(service, f["id"], anyone_writer=True)
    # Maestro de fallas: dominio WES + técnicos (no público)
    maestro = "1GlRn7QXWEre7ziau29ojR5lTl-bZ8T3mCT3cD93HZgM"
    compartir(service, maestro, anyone_writer=False)
    compartir(service, folder["id"], anyone_writer=False)

    link = f.get("webViewLink") or f"https://docs.google.com/spreadsheets/d/{f['id']}/edit"
    state = {}
    if STATE.is_file():
        state = json.loads(STATE.read_text(encoding="utf-8"))
    state["ingreso_tecnicos_id"] = f["id"]
    state["ingreso_tecnicos_link"] = link
    state["folder_link"] = folder["web_view_link"]
    STATE.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")

    reports = REPO / "reports" / "Mantenimientos" / "formulario_visita"
    reports.mkdir(parents=True, exist_ok=True)
    (reports / "LINK_PARA_TECNICOS.txt").write_text(
        "LINK PARA TÉCNICOS WES (permanente)\n"
        "===================================\n\n"
        f"{link}\n\n"
        "Completar la pestaña INGRESO: una fila por visita.\n"
        f"Carpeta: {folder['web_view_link']}\n",
        encoding="utf-8",
    )

    print("=" * 60)
    print("Planilla de ingreso para técnicos lista")
    print(link)
    print("=" * 60)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

# -*- coding: utf-8 -*-
"""
Sincroniza catálogos del formulario web y del Excel digital.

FUENTES (evitar mezclar):
  1) Cliente + Máquina (puntos del formulario)
     → CONTACTOS_ENVIOS_ACTAS · hoja Clientes_catalogo
       Sheet: 1Tpjm1eXRXKuKvxachtbYVr9503wICJdsYDTjkbm__o8
       Path:  G:\\Mi unidad\\Agente WES\\wes-scripts\\mantenimiento wes\\CONTACTOS_ENVIOS_ACTAS
     Si un cliente NO está en ese catálogo, se completa solo con Base1/Datos
     del Registro de fallas (clientes nuevos del historial).

  2) Tipo de falla / falla específica
     → Registro de fallas WES · hoja Base3
       Sheet: 1GlRn7QXWEre7ziau29ojR5lTl-bZ8T3mCT3cD93HZgM

NO usar Base1 del Registro como lista “oficial” de puntos cuando el cliente
ya figura en Clientes_catalogo (ej. RENCA: ICCO/ICCP vs nombres viejos).

Uso:
  python sincronizar_catalogos_desde_maestro.py --desde-drive
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Set, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parent
MAESTRO_DIR = ROOT / "maestro"
CAT_DIR = ROOT / "catalogos"
XLSX_FORM = ROOT / "FORMULARIO_MANTENCION_WES_DIGITAL.xlsx"

SHEET_REGISTRO_ID = "1GlRn7QXWEre7ziau29ojR5lTl-bZ8T3mCT3cD93HZgM"
SHEET_CONTACTOS_ID = "1Tpjm1eXRXKuKvxachtbYVr9503wICJdsYDTjkbm__o8"
FILE_ANALISIS_ID = "1mzIsNG9Kr8PLZkUz_JDDklJu5uv1HJgC"

URL_CONTACTOS = f"https://docs.google.com/spreadsheets/d/{SHEET_CONTACTOS_ID}/edit"
URL_REGISTRO = f"https://docs.google.com/spreadsheets/d/{SHEET_REGISTRO_ID}/edit"

FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FONT_WHITE = Font(name="Calibri", bold=True, color="FFFFFF", size=11)

Pair = Tuple[str, str]


def _download_drive(file_id: str, dest: Path, *, export_xlsx: bool = False) -> Path:
    sys.path.insert(0, str(ROOT.parent))
    from wes_google_drive import obtener_servicio_drive
    from googleapiclient.http import MediaIoBaseDownload
    import io

    svc = obtener_servicio_drive()
    meta = svc.files().get(fileId=file_id, fields="mimeType,name").execute()
    mime = meta["mimeType"]
    if export_xlsx or mime == "application/vnd.google-apps.spreadsheet":
        req = svc.files().export_media(
            fileId=file_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        req = svc.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    dl = MediaIoBaseDownload(fh, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(fh.getvalue())
    return dest


def _norm(v) -> str:
    return "" if v is None else str(v).strip()


def _pairs_from_sheet(ws, cli_col: int, maq_col: int, start_row: int = 2) -> Set[Pair]:
    out: Set[Pair] = set()
    for r in range(start_row, (ws.max_row or 1) + 1):
        c = _norm(ws.cell(r, cli_col).value)
        m = _norm(ws.cell(r, maq_col).value)
        if c and m:
            out.add((c, m))
    return out


def _find_catalogo_sheet(wb) -> str:
    """Prioriza Clientes_catalogo; si la borran, usa 1ª hoja con Cliente+Máquina."""
    names = list(wb.sheetnames)
    for preferred in ("Clientes_catalogo", "Catalogo", "Puntos", "Base1"):
        if preferred in names:
            return preferred
    for sn in names:
        if sn.lower() in {"instrucciones", "contactos", "readme"}:
            continue
        ws = wb[sn]
        headers = [_norm(ws.cell(1, c).value).lower() for c in range(1, 6)]
        if any(h.startswith("cliente") for h in headers) and any(
            "mquina" in h.replace("á", "a") or "sitio" in h for h in headers
        ):
            return sn
    # última opción: primera hoja del libro
    return names[0]


def _pairs_from_contactos_catalogo(path: Path) -> Tuple[Set[Pair], Set[str], str]:
    """Lee catálogo de puntos desde CONTACTOS_ENVIOS_ACTAS. Devuelve (pares, clientes, hoja)."""
    wb = load_workbook(path, data_only=True)
    sn = _find_catalogo_sheet(wb)
    ws = wb[sn]
    headers = [_norm(ws.cell(1, c).value).lower() for c in range(1, 8)]

    def idx(*names: str) -> int:
        for n in names:
            if n in headers:
                return headers.index(n) + 1
        return 0

    c_cli = idx("cliente") or 1
    c_maq = idx("máquina / sitio", "maquina / sitio", "máquina", "maquina", "sitio") or 2
    pairs = _pairs_from_sheet(ws, c_cli, c_maq, start_row=2)
    clients = {c for c, _ in pairs}
    return pairs, clients, sn


def _fallas_from_sheet(ws, tipo_col: int = 2, esp_col: int = 3) -> Dict[str, List[str]]:
    d: Dict[str, List[str]] = defaultdict(list)
    for r in range(2, (ws.max_row or 1) + 1):
        t = _norm(ws.cell(r, tipo_col).value)
        e = _norm(ws.cell(r, esp_col).value)
        if t and e and e not in d[t]:
            d[t].append(e)
    return dict(d)


def _write_base1(ws, pairs: Iterable[Pair]) -> int:
    if ws.max_row and ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    ws["B1"] = "CLIENTE"
    ws["C1"] = "MAQUINA"
    ws["B1"].fill = FILL_HEADER
    ws["C1"].fill = FILL_HEADER
    ws["B1"].font = FONT_WHITE
    ws["C1"].font = FONT_WHITE
    n = 0
    for cli, maq in sorted(pairs, key=lambda x: (x[0].lower(), x[1].lower())):
        n += 1
        ws.cell(n + 1, 2, cli)
        ws.cell(n + 1, 3, maq)
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 42
    return n


def _write_base3(ws, fallas: Dict[str, List[str]]) -> int:
    if ws.max_row and ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)
    ws["B1"] = "Tipo de falla"
    ws["C1"] = "Falla Específica"
    ws["B1"].fill = FILL_HEADER
    ws["C1"].fill = FILL_HEADER
    ws["B1"].font = FONT_WHITE
    ws["C1"].font = FONT_WHITE
    n = 0
    for tipo in sorted(fallas.keys()):
        for esp in fallas[tipo]:
            n += 1
            ws.cell(n + 1, 2, tipo)
            ws.cell(n + 1, 3, esp)
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 36
    return n


def _update_filter_helpers(wb, n1: int, n3: int) -> None:
    if "Base2" not in wb.sheetnames:
        wb.create_sheet("Base2")
    ws2 = wb["Base2"]
    ws2["A1"] = f'=IFERROR(FILTER(Base1!B2:C{n1 + 1},Base1!B2:B{n1 + 1}=Ingreso!C4),"")'
    if "Base 4" not in wb.sheetnames:
        wb.create_sheet("Base 4")
    ws4 = wb["Base 4"]
    ws4["A1"] = f'=IFERROR(FILTER(Base3!B2:C{n3 + 1},Base3!B2:B{n3 + 1}=Ingreso!C14),"")'


def _clientes_dict(pairs: Set[Pair]) -> Dict[str, List[str]]:
    d: Dict[str, List[str]] = defaultdict(list)
    for cli, maq in sorted(pairs):
        if maq not in d[cli]:
            d[cli].append(maq)
    return dict(sorted(d.items()))


def sincronizar(
    *,
    desde_drive: bool = False,
    fuente: Optional[Path] = None,
    fuente_contactos: Optional[Path] = None,
) -> dict:
    MAESTRO_DIR.mkdir(parents=True, exist_ok=True)
    CAT_DIR.mkdir(parents=True, exist_ok=True)

    path_contactos = fuente_contactos or (MAESTRO_DIR / "CONTACTOS_ENVIOS_ACTAS.xlsx")
    path_registro = fuente or (MAESTRO_DIR / "analisis_falla_google.xlsx")

    if desde_drive:
        path_contactos = _download_drive(
            SHEET_CONTACTOS_ID,
            MAESTRO_DIR / "CONTACTOS_ENVIOS_ACTAS.xlsx",
            export_xlsx=True,
        )
        path_registro = _download_drive(
            SHEET_REGISTRO_ID,
            MAESTRO_DIR / "analisis_falla_google.xlsx",
            export_xlsx=True,
        )
        _download_drive(FILE_ANALISIS_ID, MAESTRO_DIR / "analisis_de_falla.xlsx")
    else:
        if not path_contactos.is_file():
            path_contactos = _download_drive(
                SHEET_CONTACTOS_ID,
                MAESTRO_DIR / "CONTACTOS_ENVIOS_ACTAS.xlsx",
                export_xlsx=True,
            )
        if not path_registro.is_file():
            path_registro = _download_drive(
                SHEET_REGISTRO_ID,
                MAESTRO_DIR / "analisis_falla_google.xlsx",
                export_xlsx=True,
            )

    # 1) Puntos oficiales desde CONTACTOS_ENVIOS_ACTAS
    pairs_cat, clients_cat, hoja_cat = _pairs_from_contactos_catalogo(path_contactos)
    pairs: Set[Pair] = set(pairs_cat)

    # 2) Completar solo clientes AUSENTES del catálogo (historial/Base1)
    wb_src = load_workbook(path_registro, data_only=True)
    pairs_hist: Set[Pair] = set()
    if "Base1" in wb_src.sheetnames:
        pairs_hist |= _pairs_from_sheet(wb_src["Base1"], 2, 3)
    for sh_name, ccol, mcol in (("Datos", 2, 3), ("Data", 2, 3)):
        if sh_name in wb_src.sheetnames:
            pairs_hist |= _pairs_from_sheet(wb_src[sh_name], ccol, mcol)

    extra_from_hist = 0
    for cli, maq in pairs_hist:
        if cli in clients_cat:
            continue  # no pisar nombres oficiales (ej. RENCA)
        pairs.add((cli, maq))
        extra_from_hist += 1

    # 3) Fallas desde Registro Base3
    if "Base3" in wb_src.sheetnames:
        fallas = _fallas_from_sheet(wb_src["Base3"])
    else:
        fallas = {}

    clientes = _clientes_dict(pairs)
    (CAT_DIR / "clientes_maquinas.json").write_text(
        json.dumps(clientes, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (CAT_DIR / "tipos_falla.json").write_text(
        json.dumps(fallas, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    fuente_txt = ROOT / "FUENTE_PUNTOS_FORMULARIO.txt"
    fuente_txt.write_text(
        f"""Puntos (Cliente + Máquina) del formulario — fuente ÚNICA
========================================================
Archivo: CONTACTOS_ENVIOS_ACTAS
Hoja:    {hoja_cat}
Windows: G:\\Mi unidad\\Agente WES\\wes-scripts\\mantenimiento wes\\CONTACTOS_ENVIOS_ACTAS
Drive:   {URL_CONTACTOS}

Hojas útiles en ese Excel (el resto se puede borrar):
  • Clientes_catalogo = puntos del formulario
  • Contactos = emails TO/CC
No hace falta copiar esto al FORMULARIO_MANTENCION ni al Registro.

NO uses Base1 del Registro de fallas para editar puntos de clientes que
ya están en este catálogo (ahí quedan nombres viejos, ej. RENCA).

Fallas (tipo / específica): Registro de fallas · Base3
  {URL_REGISTRO}

Sync: {datetime.now().strftime('%Y-%m-%d %H:%M')}
Pares catálogo: {len(pairs_cat)} · extras historial (clientes nuevos): {extra_from_hist}
""",
        encoding="utf-8",
    )

    if XLSX_FORM.is_file():
        wb = load_workbook(XLSX_FORM)
        if "Base1" not in wb.sheetnames:
            wb.create_sheet("Base1")
        if "Base3" not in wb.sheetnames:
            wb.create_sheet("Base3")
        n1 = _write_base1(wb["Base1"], pairs)
        n3 = _write_base3(wb["Base3"], fallas)
        _update_filter_helpers(wb, n1, n3)
        if "Instrucciones" in wb.sheetnames:
            wb["Instrucciones"]["A22"] = (
                f"Catálogos sync {datetime.now().strftime('%Y-%m-%d %H:%M')}: "
                f"puntos desde CONTACTOS_ENVIOS_ACTAS!{hoja_cat} · "
                f"{len(clientes)} clientes · {n1} máquinas · fallas desde Registro Base3 ({n3})"
            )
        wb.save(XLSX_FORM)
    else:
        n1 = len(pairs)
        n3 = sum(len(v) for v in fallas.values())

    resumen = {
        "fuente_puntos": str(path_contactos),
        "fuente_puntos_hoja": hoja_cat,
        "fuente_puntos_url": URL_CONTACTOS,
        "fuente_fallas": str(path_registro),
        "fuente_fallas_url": URL_REGISTRO,
        "pares_catalogo": len(pairs_cat),
        "pares_extra_historial": extra_from_hist,
        "clientes": len(clientes),
        "pares_maquina": n1,
        "tipos_falla": len(fallas),
        "fallas_especificas": n3,
        "clientes_lista": sorted(clientes.keys()),
        "ejemplo_renca": clientes.get("RENCA", []),
        "sync_at": datetime.now().isoformat(timespec="seconds"),
    }
    (CAT_DIR / "ultima_sincronizacion.json").write_text(
        json.dumps(resumen, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return resumen


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Sincroniza catálogos: puntos desde CONTACTOS_ENVIOS_ACTAS, fallas desde Registro"
    )
    parser.add_argument("--desde-drive", action="store_true")
    parser.add_argument("--fuente", type=Path, default=None, help="xlsx Registro (fallas)")
    parser.add_argument(
        "--fuente-contactos",
        type=Path,
        default=None,
        help="xlsx CONTACTOS_ENVIOS_ACTAS (puntos)",
    )
    args = parser.parse_args()
    info = sincronizar(
        desde_drive=args.desde_drive,
        fuente=args.fuente,
        fuente_contactos=args.fuente_contactos,
    )
    print("OK sync catálogos")
    print(f"  puntos: {info['fuente_puntos_url']} · hoja {info['fuente_puntos_hoja']}")
    print(f"  fallas: {info['fuente_fallas_url']} · Base3")
    print(
        f"  clientes: {info['clientes']} · máquinas: {info['pares_maquina']} "
        f"(catálogo {info['pares_catalogo']} + extras {info['pares_extra_historial']})"
    )
    print(f"  fallas: {info['fallas_especificas']} en {info['tipos_falla']} tipos")
    print("  RENCA:", ", ".join(info.get("ejemplo_renca") or []))
    print("  clientes:", ", ".join(info["clientes_lista"]))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

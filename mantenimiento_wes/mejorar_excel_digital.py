# -*- coding: utf-8 -*-
"""
Mejora FORMULARIO_MANTENCION_WES_DIGITAL.xlsx:
  - Instrucciones claras
  - Ingreso (celdas amarillas)
  - Datos ampliado (compatible Sheet + extras de acta)
  - Resumen / Dashboard
  - Formulario Visita (checklist)
  - Base1..Base4 sincronizables

No borra filas ya cargadas en Datos.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "FORMULARIO_MANTENCION_WES_DIGITAL.xlsx"

THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_SECTION = PatternFill("solid", fgColor="2E75B6")
FILL_INPUT = PatternFill("solid", fgColor="FFF2CC")
FILL_LIGHT = PatternFill("solid", fgColor="DDEBF7")
FILL_OK = PatternFill("solid", fgColor="E2EFDA")
FONT_WHITE = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
FONT_TITLE = Font(name="Calibri", bold=True, color="FFFFFF", size=16)
FONT_LABEL = Font(name="Calibri", bold=True, size=11, color="1F4E79")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

TECNICOS = [
    "Mauricio Orellana",
    "Maurico Orellana",
    "Anibal Aranda",
    "Jose Otarola",
    "Gabriel Prieto",
    "Mirko Lorca",
    "Cristian Sepulveda",
    "Salvador Cantillana",
    "Jeans Araya",
]
TIPOS_MTTO = [
    "Mtto Correctivo",
    "Mtto Preventivo",
    "Mtto Predictivo",
    "Reubicacion CIR",
    "Instalación Valvula On/Off",
    "Instalación sistema CIR",
    "Instalación sistema CPA",
]

DATOS_HEADERS = [
    "Cliente",
    "Maquina",
    "Tecnico",
    "Fecha",
    "Tipo de Mantenimiento",
    "Tipo de Falla",
    "Falla Expecifica",
    "Solucion y/o Diagnostico",
    "Observaciones",
    "Firma requerida",
    "N OT",
    "Estado visita",
    "Origen",
    "Email cliente",
    "Recibido por",
    "Cargo",
    "Comuna",
    "PDF / Drive",
    "Año",
    "Mes",
]


def _dv(ws, formula: str, cells: str) -> None:
    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    validation.error = "Elegí un valor de la lista"
    validation.errorTitle = "Valor inválido"
    ws.add_data_validation(validation)
    validation.add(cells)


def _ensure_sheet(wb, name: str, index: int | None = None):
    if name in wb.sheetnames:
        return wb[name]
    ws = wb.create_sheet(name)
    if index is not None:
        wb.move_sheet(name, offset=index - wb.sheetnames.index(name))
    return ws


def _build_instrucciones(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Formulario digital WES — mantención / actas"
    ws["A1"].font = Font(name="Calibri", bold=True, size=18, color="1F4E79")
    lines = [
        "",
        "Este Excel es el maestro local de análisis. El técnico completa el formulario web;",
        "cada visita se agrega sola a la hoja Datos (y también al Registro de fallas en Drive).",
        "",
        "Hojas",
        "  • Ingreso — carga manual rápida (celdas amarillas) si no usan el celular.",
        "  • Datos — historial (compatible con el Sheet Registro de fallas WES).",
        "  • Resumen — totales por cliente / tipo de falla / mes.",
        "  • Formulario Visita — checklist CIR/CPA opcional.",
        "  • Base1 / Base2 / Base3 / Base 4 — catálogos (Cliente→Máquina, Tipo→Falla).",
        "",
        "Actualizar clientes / máquinas que faltan",
        "  1) Editá Base1 en el Google Sheet «Registro de fallas WES» (o este Excel).",
        "  2) python sincronizar_catalogos_desde_maestro.py --desde-drive",
        "  3) Reiniciá el servidor del formulario (servir_formulario_visita.py).",
        "",
        "Flujo recomendado en terreno",
        "  Celular → formulario web → PDF al cliente (acusar recibo) → fila en Datos.",
        "",
        f"Mejorado: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
    ]
    for i, line in enumerate(lines, start=2):
        ws.cell(i, 1, line)
    ws.column_dimensions["A"].width = 110


def _build_ingreso(ws, n1: int, n3: int) -> None:
    ws.merge_cells("B2:F2")
    ws["B2"] = "INGRESO — Registro de fallas / visitas WES"
    ws["B2"].font = FONT_TITLE
    ws["B2"].fill = FILL_HEADER
    ws["B2"].alignment = CENTER
    ws.row_dimensions[2].height = 26

    ws.merge_cells("B3:F3")
    ws["B3"] = "Celdas amarillas = completar. Listas en cascada Cliente→Máquina y Tipo falla→Falla."
    ws["B3"].font = Font(name="Calibri", italic=True, size=10, color="FFFFFF")
    ws["B3"].fill = FILL_SECTION

    fields = [
        (4, "Cliente"),
        (6, "Maquina"),
        (8, "Tecnico"),
        (10, "Fecha"),
        (12, "Tipo de Mantenimiento"),
        (14, "Tipo de Falla"),
        (16, "Falla Expecifica"),
        (18, "solucion"),
        (20, "Observaciones"),
        (22, "Firma requerida"),
        (24, "N OT / formulario"),
        (26, "Estado visita"),
    ]
    for row, label in fields:
        ws.cell(row, 2, label).font = FONT_LABEL
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
        cell = ws.cell(row, 3, ws.cell(row, 3).value if ws.cell(row, 3).value else "")
        cell.fill = FILL_INPUT
        cell.border = THIN
        cell.alignment = LEFT

    if not ws["C10"].value:
        ws["C10"] = date.today()
    ws["C10"].number_format = "DD/MM/YYYY"
    if not ws["C22"].value:
        ws["C22"] = "No - solo digital"
    if not ws["C26"].value:
        ws["C26"] = "cerrada"

    ws.row_dimensions[18].height = 45
    ws.row_dimensions[20].height = 45
    ws.merge_cells("C18:F19")
    ws.merge_cells("C20:F21")

    ws.merge_cells("B28:F28")
    ws["B28"] = (
        "Al terminar (modo Excel): python registrar_desde_ingreso.py  ·  "
        "En terreno preferí el formulario web (PDF + correo + esta hoja Datos)."
    )
    ws["B28"].fill = FILL_OK
    ws["B28"].font = Font(name="Calibri", size=10, italic=True)

    for letter, width in {"B": 24, "C": 18, "D": 14, "E": 14, "F": 18}.items():
        ws.column_dimensions[letter].width = width

    # quitar validaciones viejas y recrear
    ws.data_validations.dataValidation = []
    _dv(ws, f"Base1!$B$2:$B${max(n1 + 1, 2)}", "C4")
    _dv(ws, "Base2!$B$1:$B$300", "C6")
    _dv(ws, '"' + ",".join(TECNICOS) + '"', "C8")
    _dv(ws, '"' + ",".join(TIPOS_MTTO) + '"', "C12")
    _dv(ws, f"Base3!$B$2:$B${max(n3 + 1, 2)}", "C14")
    _dv(ws, "'Base 4'!$B$1:$B$80", "C16")
    _dv(ws, '"Sí - acta firmada,No - solo digital"', "C22")
    _dv(ws, '"abierta,en_curso,cerrada"', "C26")


def _build_datos(ws) -> None:
    # Preserve existing data rows; ensure headers cover new columns
    existing = []
    if ws.max_row and ws.max_row >= 2:
        old_headers = [ws.cell(1, c).value for c in range(2, 22)]
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 2).value in (None, ""):
                continue
            row = {DATOS_HEADERS[i]: ws.cell(r, i + 2).value for i in range(min(len(DATOS_HEADERS), 20))}
            # map by old header name if present
            if old_headers and old_headers[0] == "Cliente":
                row = {}
                for i, h in enumerate(DATOS_HEADERS):
                    # find old col
                    val = None
                    if i < len(old_headers) and old_headers[i]:
                        val = ws.cell(r, i + 2).value
                    row[h] = val
            existing.append(row)

    # rewrite header
    for i, h in enumerate(DATOS_HEADERS, start=2):
        cell = ws.cell(1, i, h)
        cell.fill = FILL_HEADER
        cell.font = FONT_WHITE
        cell.alignment = CENTER
        cell.border = THIN
        ws.column_dimensions[get_column_letter(i)].width = 18 if h not in ("Solucion y/o Diagnostico", "Observaciones", "PDF / Drive") else 36

    # clear body lightly and rewrite preserved
    if ws.max_row > 1:
        for r in range(2, ws.max_row + 1):
            for c in range(2, 2 + len(DATOS_HEADERS)):
                ws.cell(r, c).value = None

    for idx, row in enumerate(existing):
        r = idx + 2
        for i, h in enumerate(DATOS_HEADERS):
            cell = ws.cell(r, i + 2, row.get(h))
            cell.fill = FILL_INPUT
            cell.border = THIN
            if h == "Fecha":
                cell.number_format = "DD/MM/YYYY"
            if h == "Año" and row.get("Fecha") and not row.get("Año"):
                try:
                    cell.value = row["Fecha"].year if hasattr(row["Fecha"], "year") else None
                except Exception:
                    pass
            if h == "Mes" and row.get("Fecha") and not row.get("Mes"):
                try:
                    cell.value = row["Fecha"].month if hasattr(row["Fecha"], "month") else None
                except Exception:
                    pass

    # template empty rows
    start = 2 + len(existing)
    for r in range(start, start + 30):
        for c in range(2, 2 + len(DATOS_HEADERS)):
            cell = ws.cell(r, c)
            if cell.value in (None, ""):
                cell.fill = FILL_INPUT
                cell.border = THIN
        ws.cell(r, 5).number_format = "DD/MM/YYYY"

    ws.freeze_panes = "B2"
    last_col = get_column_letter(1 + len(DATOS_HEADERS))
    ws.auto_filter.ref = f"B1:{last_col}2000"


def _build_resumen(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Resumen operativo (desde hoja Datos)"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    ws["A3"] = "Total visitas registradas"
    ws["B3"] = '=COUNTA(Datos!B:B)-1'
    ws["A4"] = "Clientes distintos"
    ws["B4"] = '=IFERROR(SUMPRODUCT((Datos!B2:B2000<>"")/COUNTIF(Datos!B2:B2000,Datos!B2:B2000&"")),0)'
    ws["A6"] = "Tip: filtrá la hoja Datos por Cliente / Año / Mes para el análisis semanal."
    ws["A6"].font = Font(name="Calibri", italic=True, color="666666")
    for col, w in (("A", 36), ("B", 18)):
        ws.column_dimensions[col].width = w
    ws["B3"].fill = FILL_OK
    ws["B4"].fill = FILL_LIGHT


def _build_formulario_visita(ws) -> None:
    if ws["A1"].value:
        return  # already built; keep checklist
    ws.merge_cells("A1:H1")
    ws["A1"] = "FORMULARIO DE MANTENCIÓN WES — ACTA DIGITAL (checklist)"
    ws["A1"].font = FONT_TITLE
    ws["A1"].fill = FILL_HEADER
    ws["A1"].alignment = CENTER


def mejorar() -> Path:
    if OUT.is_file():
        wb = load_workbook(OUT)
    else:
        wb = Workbook()
        wb.remove(wb.active)

    # sync catalogs first if script available
    try:
        from sincronizar_catalogos_desde_maestro import sincronizar

        info = sincronizar(desde_drive=False)
        n1 = info["pares_maquina"]
        n3 = info["fallas_especificas"]
        # reload after sync
        wb = load_workbook(OUT)
    except Exception:
        n1 = wb["Base1"].max_row - 1 if "Base1" in wb.sheetnames else 2
        n3 = wb["Base3"].max_row - 1 if "Base3" in wb.sheetnames else 2

    ws_inst = _ensure_sheet(wb, "Instrucciones")
    _build_instrucciones(ws_inst)

    ws_ing = _ensure_sheet(wb, "Ingreso")
    _build_ingreso(ws_ing, max(n1, 2), max(n3, 2))

    ws_dat = _ensure_sheet(wb, "Datos")
    _build_datos(ws_dat)

    ws_res = _ensure_sheet(wb, "Resumen")
    _build_resumen(ws_res)

    ws_fv = _ensure_sheet(wb, "Formulario Visita")
    _build_formulario_visita(ws_fv)

    for name in ("Base1", "Base2", "Base3", "Base 4"):
        _ensure_sheet(wb, name)

    order = [
        "Instrucciones",
        "Ingreso",
        "Datos",
        "Resumen",
        "Formulario Visita",
        "Base1",
        "Base2",
        "Base3",
        "Base 4",
    ]
    for idx, name in enumerate(order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=idx - wb.sheetnames.index(name))

    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    path = mejorar()
    print(f"OK mejorado {path}")

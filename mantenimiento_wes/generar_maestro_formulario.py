# -*- coding: utf-8 -*-
"""
Genera el Excel MAESTRO del formulario (configuración).

Arquitectura WES (2 archivos):
  1) MAESTRO_FORMULARIO_WES.xlsx  ← lo que el formulario NECESITA
     Hojas: Puntos | Contactos | Tecnicos | Fallas | Opciones | Instrucciones
  2) Registro de fallas WES        ← historial de visitas + evaluaciones
     Hojas: Datos (+ pivotes/pareto/resúmenes)

Uso:
  python generar_maestro_formulario.py
  python generar_maestro_formulario.py --desde-drive
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
CAT = ROOT / "catalogos"
MAESTRO_DIR = ROOT / "maestro"
OUT = ROOT / "MAESTRO_FORMULARIO_WES.xlsx"

SHEET_CONTACTOS_ID = "1Tpjm1eXRXKuKvxachtbYVr9503wICJdsYDTjkbm__o8"
SHEET_REGISTRO_ID = "1GlRn7QXWEre7ziau29ojR5lTl-bZ8T3mCT3cD93HZgM"

FILL_H = PatternFill("solid", fgColor="1F4E79")
FONT_H = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_T = Font(name="Calibri", bold=True, size=14, color="1F4E79")


def _download(file_id: str, dest: Path) -> Path:
    sys.path.insert(0, str(ROOT.parent))
    from wes_google_drive import obtener_servicio_drive
    from googleapiclient.http import MediaIoBaseDownload
    import io

    svc = obtener_servicio_drive()
    meta = svc.files().get(fileId=file_id, fields="mimeType").execute()
    if meta["mimeType"] == "application/vnd.google-apps.spreadsheet":
        req = svc.files().export_media(
            fileId=file_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        req = svc.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    dl = MediaIoBaseDownload(fh, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(fh.getvalue())
    return dest


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.fill = FILL_H
        cell.font = FONT_H
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _autosize(ws, widths: dict) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def _load_json_catalogs():
    cm = json.loads((CAT / "clientes_maquinas.json").read_text(encoding="utf-8"))
    ct = json.loads((CAT / "contactos_cliente.json").read_text(encoding="utf-8"))
    tf = json.loads((CAT / "tipos_falla.json").read_text(encoding="utf-8"))
    op = json.loads((CAT / "opciones.json").read_text(encoding="utf-8"))
    return cm, ct, tf, op


def _sheet_instrucciones(wb: Workbook) -> None:
    ws = wb.create_sheet("Instrucciones", 0)
    lines = [
        "MAESTRO DEL FORMULARIO WES — configuración",
        "",
        "Este Excel alimenta el formulario de visita (puntos, emails, técnicos, fallas).",
        "NO guardés aquí el historial de visitas: eso va al Registro de fallas.",
        "",
        "Hojas",
        "  • Puntos     → Cliente + Máquina/sitio (lo que elige el técnico)",
        "  • Contactos  → Quién recibe el PDF (TO=general, CC=punto/CC)",
        "  • Tecnicos   → Lista de técnicos WES",
        "  • Fallas     → Árbol Tipo de falla → Falla específica",
        "  • Opciones   → Tipos de mtto, motivos, tecnologías, checklist",
        "",
        "Archivo 2 (historial + evaluaciones)",
        "  Registro de fallas WES",
        f"  https://docs.google.com/spreadsheets/d/{SHEET_REGISTRO_ID}/edit",
        "  Hoja Datos = cada visita. Ahí se hacen tablas dinámicas / pareto / KPIs.",
        "",
        f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "Sync formulario: python sincronizar_desde_maestro_formulario.py --desde-drive",
    ]
    for i, line in enumerate(lines, 1):
        ws.cell(i, 1, line)
        if i == 1:
            ws.cell(i, 1).font = FONT_T
    ws.column_dimensions["A"].width = 100


def _sheet_puntos(wb: Workbook, cm: dict) -> None:
    ws = wb.create_sheet("Puntos")
    ws.append(["Cliente", "Máquina / sitio"])
    _style_header(ws, 2)
    for cli in sorted(cm.keys()):
        for maq in cm[cli]:
            ws.append([cli, maq])
    _autosize(ws, {"A": 22, "B": 48})


def _sheet_contactos(wb: Workbook, ct: list) -> None:
    ws = wb.create_sheet("Contactos")
    ws.append(["Cliente", "Máquina", "Rol", "Nombre", "Cargo", "Email", "Actualizado"])
    _style_header(ws, 7)
    for c in ct:
        rol = c.get("rol") or ("general" if c.get("enviar_to") else "CC")
        ws.append([
            c.get("cliente", ""),
            c.get("maquina") or c.get("sitio") or "",
            rol,
            c.get("nombre", ""),
            c.get("cargo", ""),
            c.get("email", ""),
            c.get("actualizado", ""),
        ])
    _autosize(ws, {"A": 18, "B": 36, "C": 12, "D": 28, "E": 36, "F": 40, "G": 14})


def _sheet_tecnicos(wb: Workbook, op: dict) -> None:
    ws = wb.create_sheet("Tecnicos")
    ws.append(["Nombre", "Email", "Activo", "Notas"])
    _style_header(ws, 4)
    for name in op.get("tecnicos", []):
        ws.append([name, "", "Sí", ""])
    _autosize(ws, {"A": 28, "B": 36, "C": 10, "D": 40})


def _sheet_fallas(wb: Workbook, tf: dict) -> None:
    ws = wb.create_sheet("Fallas")
    ws.append(["Tipo de falla", "Falla específica"])
    _style_header(ws, 2)
    for tipo in sorted(tf.keys()):
        for esp in tf[tipo]:
            ws.append([tipo, esp])
    _autosize(ws, {"A": 22, "B": 40})


def _sheet_opciones(wb: Workbook, op: dict) -> None:
    ws = wb.create_sheet("Opciones")
    ws.append(["Grupo", "Valor", "Orden"])
    _style_header(ws, 3)
    order = 0
    for grupo in (
        "tipos_mtto",
        "motivos",
        "tecnologias",
        "estados_checklist",
        "cir",
        "cpa",
        "sab",
    ):
        for val in op.get(grupo, []):
            order += 1
            ws.append([grupo, val, order])
    _autosize(ws, {"A": 22, "B": 40, "C": 10})


def generar(*, desde_drive: bool = False) -> Path:
    MAESTRO_DIR.mkdir(parents=True, exist_ok=True)
    if desde_drive:
        # refrescar catálogos previos si existen scripts
        try:
            from sincronizar_catalogos_desde_maestro import sincronizar as sync_cat
            from contactos_cliente import sincronizar as sync_cto

            sync_cat(desde_drive=True)
            sync_cto(desde_drive=True)
        except Exception as exc:  # noqa: BLE001
            print(f"Aviso sync previo: {exc}")
            _download(SHEET_CONTACTOS_ID, MAESTRO_DIR / "CONTACTOS_ENVIOS_ACTAS.xlsx")

    cm, ct, tf, op = _load_json_catalogs()
    wb = Workbook()
    # remove default
    wb.remove(wb.active)
    _sheet_instrucciones(wb)
    _sheet_puntos(wb, cm)
    _sheet_contactos(wb, ct)
    _sheet_tecnicos(wb, op)
    _sheet_fallas(wb, tf)
    _sheet_opciones(wb, op)
    OUT.write_bytes(b"")  # touch
    wb.save(OUT)
    # mirror in maestro/
    mirror = MAESTRO_DIR / OUT.name
    wb.save(mirror)
    return OUT


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--desde-drive", action="store_true")
    args = ap.parse_args()
    path = generar(desde_drive=args.desde_drive)
    print(f"OK maestro: {path}")
    print("  Hojas: Instrucciones | Puntos | Contactos | Tecnicos | Fallas | Opciones")
    print("  Historial/evaluaciones → Registro de fallas (archivo separado)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

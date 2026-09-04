# -*- coding: utf-8 -*-
"""
Sincroniza catálogos del formulario DESDE MAESTRO_FORMULARIO_WES.xlsx

Fuente única de configuración del form:
  Puntos | Contactos | Tecnicos | Fallas | Opciones

El historial de visitas NO se lee de aquí (va al Registro de fallas).

Uso:
  python sincronizar_desde_maestro_formulario.py
  python sincronizar_desde_maestro_formulario.py --desde-drive
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
CAT = ROOT / "catalogos"
MAESTRO_DIR = ROOT / "maestro"
MAESTRO_XLSX = ROOT / "MAESTRO_FORMULARIO_WES.xlsx"
MAESTRO_MIRROR = MAESTRO_DIR / "MAESTRO_FORMULARIO_WES.xlsx"

# Si aún no hay maestro propio, se puede apuntar al Sheet de contactos legacy
LEGACY_CONTACTOS_ID = "1Tpjm1eXRXKuKvxachtbYVr9503wICJdsYDTjkbm__o8"


def _norm(v) -> str:
    return "" if v is None else str(v).strip()


def _download_maestro(dest: Path) -> Path:
    """Busca MAESTRO_FORMULARIO_WES en Drive (mantenimiento wes); si no, usa legacy."""
    sys.path.insert(0, str(ROOT.parent))
    from wes_google_drive import obtener_servicio_drive
    from googleapiclient.http import MediaIoBaseDownload
    import io

    svc = obtener_servicio_drive()
    folder = "150GFVtGFlPXb_7bQfe7AS4SClKEXLEuX"
    q = (
        f"'{folder}' in parents and trashed=false and "
        "name contains 'MAESTRO_FORMULARIO_WES'"
    )
    res = (
        svc.files()
        .list(
            q=q,
            fields="files(id,name,mimeType,modifiedTime)",
            orderBy="modifiedTime desc",
            pageSize=5,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        )
        .execute()
    )
    files = res.get("files") or []
    if not files:
        # fallback: CONTACTOS_ENVIOS_ACTAS (solo puntos+contactos)
        file_id = LEGACY_CONTACTOS_ID
        name = "CONTACTOS_ENVIOS_ACTAS (legacy)"
    else:
        file_id = files[0]["id"]
        name = files[0]["name"]

    meta = svc.files().get(fileId=file_id, fields="mimeType").execute()
    if meta["mimeType"] == "application/vnd.google-apps.spreadsheet":
        req = svc.files().export_media(
            fileId=file_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        req = svc.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    dl = MediaIoBaseDownload(fh, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(fh.getvalue())
    print(f"  descargado: {name} → {dest}")
    return dest


def _headers(ws) -> List[str]:
    return [_norm(ws.cell(1, c).value).lower() for c in range(1, (ws.max_column or 1) + 1)]


def _col(headers: List[str], *names: str, default: int = 1) -> int:
    for n in names:
        if n in headers:
            return headers.index(n) + 1
    return default


def _read_puntos(wb) -> Dict[str, List[str]]:
    name = "Puntos" if "Puntos" in wb.sheetnames else (
        "Clientes_catalogo" if "Clientes_catalogo" in wb.sheetnames else None
    )
    if not name:
        return {}
    ws = wb[name]
    h = _headers(ws)
    c_cli = _col(h, "cliente")
    c_maq = _col(h, "máquina / sitio", "maquina / sitio", "máquina", "maquina", "sitio", default=2)
    d: Dict[str, List[str]] = defaultdict(list)
    for r in range(2, (ws.max_row or 1) + 1):
        cli = _norm(ws.cell(r, c_cli).value)
        maq = _norm(ws.cell(r, c_maq).value)
        if cli and maq and maq not in d[cli]:
            d[cli].append(maq)
    return dict(sorted(d.items()))


def _read_contactos(wb) -> List[Dict[str, Any]]:
    if "Contactos" not in wb.sheetnames:
        return []
    ws = wb["Contactos"]
    h = _headers(ws)
    c_cli = _col(h, "cliente")
    c_maq = _col(h, "máquina", "maquina", "sitio", default=2)
    c_rol = _col(h, "rol", default=3)
    c_nom = _col(h, "nombre", default=4)
    c_car = _col(h, "cargo", default=5)
    c_em = _col(h, "email", "correo", default=6)
    out = []
    for r in range(2, (ws.max_row or 1) + 1):
        cli = _norm(ws.cell(r, c_cli).value)
        email = _norm(ws.cell(r, c_em).value)
        if not cli:
            continue
        rol = _norm(ws.cell(r, c_rol).value)
        rol_l = rol.lower()
        rol_tipo = "general" if rol_l in {"general", "to", "firmante"} else ("cc" if rol else "")
        if rol and rol_tipo == "":
            rol_tipo = "cc"
        maq = _norm(ws.cell(r, c_maq).value)
        nombre = _norm(ws.cell(r, c_nom).value)
        cargo = _norm(ws.cell(r, c_car).value)
        blob = f"{cargo} {nombre} {rol}".lower()
        firmante = any(k in blob for k in ("linkes", "mantencion", "mantención", "mtto", "firmante", "supervis"))
        out.append({
            "cliente": cli,
            "sitio": maq,
            "maquina": maq,
            "rol": rol or ("general" if rol_tipo == "general" else "CC"),
            "rol_tipo": rol_tipo or "cc",
            "nombre": nombre,
            "cargo": cargo,
            "email": email,
            "firmante": firmante,
            "enviar_to": rol_tipo == "general",
            "enviar_cc": rol_tipo == "cc",
            "activo": True,
            "email_ok": bool(email and "@" in email and "." in email.split("@")[-1]),
            "notas": "",
            "fuente": "MAESTRO_FORMULARIO_WES",
        })
    return out


def _read_tecnicos(wb) -> List[str]:
    if "Tecnicos" not in wb.sheetnames:
        return []
    ws = wb["Tecnicos"]
    h = _headers(ws)
    c_nom = _col(h, "nombre")
    c_act = _col(h, "activo", default=3)
    out = []
    for r in range(2, (ws.max_row or 1) + 1):
        nom = _norm(ws.cell(r, c_nom).value)
        act = _norm(ws.cell(r, c_act).value).lower()
        if not nom:
            continue
        if act in {"no", "0", "false", "inactivo"}:
            continue
        out.append(nom)
    return out


def _read_fallas(wb) -> Dict[str, List[str]]:
    if "Fallas" not in wb.sheetnames:
        return {}
    ws = wb["Fallas"]
    h = _headers(ws)
    c_t = _col(h, "tipo de falla", "tipo")
    c_e = _col(h, "falla específica", "falla especifica", "falla", default=2)
    d: Dict[str, List[str]] = defaultdict(list)
    for r in range(2, (ws.max_row or 1) + 1):
        t = _norm(ws.cell(r, c_t).value)
        e = _norm(ws.cell(r, c_e).value)
        if t and e and e not in d[t]:
            d[t].append(e)
    return dict(d)


def _read_opciones(wb, tecnicos: List[str]) -> dict:
    base = {}
    if (CAT / "opciones.json").is_file():
        base = json.loads((CAT / "opciones.json").read_text(encoding="utf-8"))
    if "Opciones" in wb.sheetnames:
        ws = wb["Opciones"]
        h = _headers(ws)
        c_g = _col(h, "grupo")
        c_v = _col(h, "valor", default=2)
        grouped: Dict[str, List[str]] = defaultdict(list)
        for r in range(2, (ws.max_row or 1) + 1):
            g = _norm(ws.cell(r, c_g).value)
            v = _norm(ws.cell(r, c_v).value)
            if g and v and v not in grouped[g]:
                grouped[g].append(v)
        base.update(dict(grouped))
    if tecnicos:
        base["tecnicos"] = tecnicos
    return base


def sincronizar(*, desde_drive: bool = False, fuente: Optional[Path] = None) -> dict:
    CAT.mkdir(parents=True, exist_ok=True)
    path = fuente
    if desde_drive:
        path = _download_maestro(MAESTRO_MIRROR)
        # también copia local raíz
        MAESTRO_XLSX.write_bytes(path.read_bytes())
    elif path is None:
        path = MAESTRO_XLSX if MAESTRO_XLSX.is_file() else MAESTRO_MIRROR
        if not path.is_file():
            path = _download_maestro(MAESTRO_MIRROR)

    wb = load_workbook(path, data_only=True)
    puntos = _read_puntos(wb)
    contactos = _read_contactos(wb)
    tecnicos = _read_tecnicos(wb)
    fallas = _read_fallas(wb)
    opciones = _read_opciones(wb, tecnicos)

    (CAT / "clientes_maquinas.json").write_text(
        json.dumps(puntos, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (CAT / "contactos_cliente.json").write_text(
        json.dumps(contactos, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    if fallas:
        (CAT / "tipos_falla.json").write_text(
            json.dumps(fallas, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    (CAT / "opciones.json").write_text(
        json.dumps(opciones, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    # apps script embed refresh
    apps = ROOT / "apps_script_export" / "catalogos.html"
    if apps.parent.is_dir():
        embed = {
            "clientes_maquinas": puntos,
            "tipos_falla": fallas or json.loads((CAT / "tipos_falla.json").read_text(encoding="utf-8")),
            "opciones": opciones,
            "contactos": {},
            "fuente_puntos": "MAESTRO_FORMULARIO_WES!Puntos",
        }
        apps.write_text(json.dumps(embed, ensure_ascii=False, indent=2), encoding="utf-8")

    resumen = {
        "fuente": str(path),
        "hojas": wb.sheetnames,
        "clientes": len(puntos),
        "pares_maquina": sum(len(v) for v in puntos.values()),
        "contactos": len(contactos),
        "contactos_con_email": sum(1 for c in contactos if c.get("email_ok")),
        "tecnicos": len(opciones.get("tecnicos") or []),
        "tipos_falla": len(fallas),
        "ejemplo_renca": puntos.get("RENCA", []),
        "sync_at": datetime.now().isoformat(timespec="seconds"),
        "historial_evaluaciones": (
            "https://docs.google.com/spreadsheets/d/"
            "1GlRn7QXWEre7ziau29ojR5lTl-bZ8T3mCT3cD93HZgM/edit"
        ),
    }
    (CAT / "ultima_sincronizacion.json").write_text(
        json.dumps(resumen, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (ROOT / "FUENTE_MAESTRO_FORMULARIO.txt").write_text(
        f"""MAESTRO DEL FORMULARIO (configuración)
======================================
Archivo: MAESTRO_FORMULARIO_WES.xlsx
Path:    G:\\Mi unidad\\Agente WES\\wes-scripts\\mantenimiento wes\\MAESTRO_FORMULARIO_WES.xlsx
Hojas:   Puntos | Contactos | Tecnicos | Fallas | Opciones

HISTORIAL + EVALUACIONES (archivo separado)
===========================================
Registro de fallas WES → hoja Datos (+ resúmenes/pareto)
https://docs.google.com/spreadsheets/d/1GlRn7QXWEre7ziau29ojR5lTl-bZ8T3mCT3cD93HZgM/edit

Sync: {resumen['sync_at']}
RENCA: {', '.join(resumen['ejemplo_renca'])}
""",
        encoding="utf-8",
    )
    return resumen


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--desde-drive", action="store_true")
    ap.add_argument("--fuente", type=Path, default=None)
    args = ap.parse_args()
    info = sincronizar(desde_drive=args.desde_drive, fuente=args.fuente)
    print("OK sync desde MAESTRO_FORMULARIO_WES")
    print(f"  fuente: {info['fuente']}")
    print(f"  clientes: {info['clientes']} · máquinas: {info['pares_maquina']}")
    print(f"  contactos: {info['contactos']} (email {info['contactos_con_email']})")
    print(f"  técnicos: {info['tecnicos']} · tipos falla: {info['tipos_falla']}")
    print(f"  RENCA: {', '.join(info['ejemplo_renca'])}")
    print(f"  historial/evals: {info['historial_evaluaciones']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

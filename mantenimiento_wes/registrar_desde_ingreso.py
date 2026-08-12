# -*- coding: utf-8 -*-
"""Vuelca la fila de la hoja Ingreso hacia Datos (historial local)."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl

XLSX = Path(__file__).resolve().parent / "FORMULARIO_MANTENCION_WES_DIGITAL.xlsx"

MAP = [
    ("C4", 2),   # Cliente
    ("C6", 3),   # Maquina
    ("C8", 4),   # Tecnico
    ("C10", 5),  # Fecha
    ("C12", 6),  # Tipo mtto
    ("C14", 7),  # Tipo falla
    ("C16", 8),  # Falla especifica
    ("C18", 9),  # Solucion
    ("C20", 10), # Observaciones
    ("C22", 11), # Firma
    ("C24", 12), # N OT
    ("C26", 13), # Estado
]


def main() -> int:
    if not XLSX.is_file():
        print(f"[ERROR] No existe {XLSX}")
        return 1

    wb = openpyxl.load_workbook(XLSX)
    if "Ingreso" not in wb.sheetnames or "Datos" not in wb.sheetnames:
        print("[ERROR] Faltan hojas Ingreso o Datos")
        return 1

    ingreso = wb["Ingreso"]
    datos = wb["Datos"]

    cliente = ingreso["C4"].value
    if not cliente:
        print("[ERROR] Ingreso!C4 (Cliente) vacío — nada que registrar")
        return 1

    # primera fila vacía en columna B (Cliente)
    row = 2
    while datos.cell(row, 2).value not in (None, ""):
        row += 1
        if row > 5000:
            print("[ERROR] Datos lleno")
            return 1

    for src, col in MAP:
        datos.cell(row, col, ingreso[src].value)
    datos.cell(row, 14, f"Ingreso {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    wb.save(XLSX)
    print(f"OK fila {row} en Datos ← Cliente={cliente!r} Máquina={ingreso['C6'].value!r}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

# -*- coding: utf-8 -*-
"""
Servidor local del formulario de visita WES (teléfono en la misma WiFi).

Uso:
  cd mantenimiento_wes
  python servir_formulario_visita.py

Abrí en el celular la URL que imprime (ej. http://192.168.x.x:8787).

Al completar:
  1) Guarda JSON + firma PNG
  2) Agrega fila al Excel FORMULARIO_MANTENCION_WES_DIGITAL.xlsx
  3) Genera PDF del acta
  4) Envía el PDF al correo del cliente pidiendo acusar recibo
  5) (Opcional) Sube PDF a Google Drive
"""

from __future__ import annotations

import json
import os
import re
import socket
import sys
import traceback
from datetime import datetime
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any, Dict
from urllib.parse import unquote

ROOT = Path(__file__).resolve().parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
if str(ROOT.parent) not in sys.path:
    sys.path.insert(0, str(ROOT.parent))

from enviar_acta_cliente_pdf import enviar_acta_pdf_cliente  # noqa: E402
from generar_pdf_acta_visita import generar_pdf_acta  # noqa: E402
from registrar_visita_excel import registrar_visita_en_excel  # noqa: E402

HTML = ROOT / "formulario_visita.html"
CAT = ROOT / "catalogos"
FIRMAS = ROOT / "firmas"
SALIDAS = ROOT / "salidas"
PORT = int(os.environ.get("WES_FORM_PORT", "8787"))


def _lan_ip() -> str:
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except OSError:
        return "127.0.0.1"


def _load_catalogos() -> Dict[str, Any]:
    return {
        "clientes_maquinas": json.loads((CAT / "clientes_maquinas.json").read_text(encoding="utf-8")),
        "tipos_falla": json.loads((CAT / "tipos_falla.json").read_text(encoding="utf-8")),
        "opciones": json.loads((CAT / "opciones.json").read_text(encoding="utf-8")),
    }


def _safe_name(name: str) -> str:
    name = unquote(name or "archivo")
    name = Path(name).name
    name = re.sub(r"[^\w.\-]+", "_", name, flags=re.UNICODE)
    return name[:120] or f"archivo_{datetime.now().strftime('%Y%m%d_%H%M%S')}"


def _save_firma_png(data_url: str, stem: str) -> Path:
    import base64

    FIRMAS.mkdir(parents=True, exist_ok=True)
    if "," not in (data_url or ""):
        raise ValueError("Firma inválida")
    raw = base64.b64decode(data_url.split(",", 1)[1])
    path = FIRMAS / f"{stem}.png"
    path.write_bytes(raw)
    return path


def _maybe_upload_drive(pdf_path: Path, cliente: str = "", fecha: str = "") -> Dict[str, Any]:
    try:
        from wes_google_drive import credenciales_configuradas, subir_acta_mantencion

        if not credenciales_configuradas():
            return {}
        info = subir_acta_mantencion(pdf_path, cliente=cliente or "SIN_CLIENTE", fecha=fecha or None)
        return {"drive_link": info.get("web_view_link"), "drive_id": info.get("id")}
    except Exception as exc:  # noqa: BLE001
        return {"drive_error": str(exc)}


def procesar_visita(data: Dict[str, Any]) -> Dict[str, Any]:
    SALIDAS.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = _safe_name(
        f"visita_{data.get('fecha')}_{data.get('cliente')}_{data.get('maquina')}_{stamp}"
    )

    # JSON crudo
    json_path = SALIDAS / f"{stem}.json"
    to_store = dict(data)
    # no duplicar firma gigante en JSON si se guarda PNG aparte
    firma = to_store.pop("firma_png", None)
    json_path.write_text(json.dumps(to_store, ensure_ascii=False, indent=2), encoding="utf-8")

    firma_path = None
    if firma:
        firma_path = _save_firma_png(firma, stem)
        data = dict(data)
        data["firma_png"] = firma  # para el PDF

    xlsx_path, excel_row = registrar_visita_en_excel(
        data,
        pdf_link="",  # se completa tras Drive si hay link
        sync_google=True,
    )
    pdf_path = generar_pdf_acta(data)

    # también copia a reports/
    reports_dir = ROOT.parent / "reports" / "Mantenimientos" / "formulario_visita"
    reports_dir.mkdir(parents=True, exist_ok=True)
    report_pdf = reports_dir / pdf_path.name
    report_pdf.write_bytes(pdf_path.read_bytes())

    email_info = enviar_acta_pdf_cliente(pdf_path, data)
    drive_info = _maybe_upload_drive(
        report_pdf,
        str(data.get("cliente") or ""),
        str(data.get("fecha") or ""),
    )

    # Actualiza link PDF en la fila digital si hay Drive
    pdf_link = drive_info.get("drive_link") or ""
    if pdf_link:
        try:
            from openpyxl import load_workbook

            wb = load_workbook(xlsx_path)
            ws = wb["Datos"]
            # col 19 = PDF / Drive (B=2 ... R=19)
            ws.cell(excel_row, 19, pdf_link)
            wb.save(xlsx_path)
        except Exception:
            pass

    google_info = data.get("_google_sheet") or {}

    return {
        "ok": True,
        "excel_path": str(xlsx_path),
        "excel_row": excel_row,
        "pdf_path": str(pdf_path),
        "pdf_name": pdf_path.name,
        "pdf_url": f"/salidas/{pdf_path.name}",
        "json_path": str(json_path),
        "firma_path": str(firma_path) if firma_path else None,
        "email_ok": bool(email_info.get("ok")) and not email_info.get("dry_run"),
        "email_skip": email_info.get("skip"),
        "email_to": email_info.get("to"),
        "drive_link": drive_info.get("drive_link"),
        "drive_error": drive_info.get("drive_error"),
        "google_sheet_ok": bool(google_info.get("ok")),
        "google_sheet_range": google_info.get("updatedRange"),
        "google_sheet_error": google_info.get("error"),
        "message": (
            f"PDF generado y correo enviado a {', '.join(email_info.get('to') or [])}"
            if email_info.get("ok") and not email_info.get("dry_run")
            else f"PDF generado. Correo: {email_info.get('skip') or 'pendiente'}"
        ),
    }


class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt: str, *args) -> None:
        print(f"[form] {self.address_string()} - {fmt % args}")

    def _send(self, code: int, body: bytes, content_type: str) -> None:
        self.send_response(code)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(body)

    def do_OPTIONS(self) -> None:  # noqa: N802
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def do_GET(self) -> None:  # noqa: N802
        path = self.path.split("?", 1)[0]
        if path in ("/", "/formulario", "/formulario_visita.html"):
            if not HTML.is_file():
                self._send(404, b"Falta formulario_visita.html", "text/plain; charset=utf-8")
                return
            self._send(200, HTML.read_bytes(), "text/html; charset=utf-8")
            return
        if path == "/api/catalogos":
            body = json.dumps(_load_catalogos(), ensure_ascii=False).encode("utf-8")
            self._send(200, body, "application/json; charset=utf-8")
            return
        if path == "/health":
            self._send(200, b'{"ok":true}', "application/json")
            return
        if path.startswith("/salidas/"):
            rel = unquote(path[len("/salidas/"):])
            target = (SALIDAS / Path(rel).name).resolve()
            if not str(target).startswith(str(SALIDAS.resolve())) or not target.is_file():
                self._send(404, b"Not found", "text/plain")
                return
            ctype = "application/pdf" if target.suffix.lower() == ".pdf" else "application/octet-stream"
            self._send(200, target.read_bytes(), ctype)
            return
        self._send(404, b"Not found", "text/plain; charset=utf-8")

    def do_POST(self) -> None:  # noqa: N802
        path = self.path.split("?", 1)[0]
        length = int(self.headers.get("Content-Length", "0"))
        raw = self.rfile.read(length)
        if path != "/api/completar_visita":
            self._send(404, b'{"error":"Not found"}', "application/json")
            return
        try:
            data = json.loads(raw.decode("utf-8"))
            if not data.get("cliente") or not data.get("maquina"):
                raise ValueError("Cliente y máquina son obligatorios")
            if not data.get("email_cliente"):
                raise ValueError("Correo del cliente es obligatorio")
            if not data.get("solucion"):
                raise ValueError("Solución / diagnóstico es obligatorio")
            if not data.get("firma_png"):
                raise ValueError("Firma obligatoria")
            result = procesar_visita(data)
            body = json.dumps(result, ensure_ascii=False).encode("utf-8")
            self._send(200, body, "application/json; charset=utf-8")
        except Exception as exc:  # noqa: BLE001
            traceback.print_exc()
            body = json.dumps({"error": str(exc)}, ensure_ascii=False).encode("utf-8")
            self._send(400, body, "application/json; charset=utf-8")


def main() -> int:
    if not HTML.is_file():
        print(f"[ERROR] Falta {HTML}")
        return 1
    ip = _lan_ip()
    server = ThreadingHTTPServer(("0.0.0.0", PORT), Handler)
    print("=" * 60)
    print("Formulario visita WES listo")
    print(f"  PC:       http://127.0.0.1:{PORT}")
    print(f"  Teléfono: http://{ip}:{PORT}  (misma WiFi)")
    print("Al completar: PDF + correo (acusar recibo) + Excel Datos")
    print("Ctrl+C para detener")
    print("=" * 60)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nDetenido.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

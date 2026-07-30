"""
Agrega a ``datos para evaluar 2.0.xlsx`` una columna con m³ de la app WES (API medidas)
acotados al rango **01-10-2025 a 30-04-2026** y al período de cada fila.

Criterio por fila (misma convención que las hojas de comparación Renca):
- Columna **A** y **D**: fechas de lectura (``datetime`` o texto reconocible).
- Se toma el intervalo ``[min(A,D), max(A,D)]`` y se intersecta con
  ``[2025-10-01, 2026-04-30]``. Si no hay intersección, la celda queda vacía.
- La suma de ``totalM3`` diarios WES en ese tramo es el valor de la nueva columna.

Hojas reconocidas (nodo WES, alineado con ``exportar_excel_comparacion_facturacion_vs_wes.py``):
  Gimnasio, Facturas_Gimnasio → 000017-05
  Piscina, Facturas_Piscina → 000017-06
  ICCO → 000017-08
  ICCP → 000017-07
  Facturas_Escuela_Lo_Velzaquez → 000017-04

Archivo buscado (en orden) en
``reports/Renca/Coparacion App con Aguas Andinas/``:
  ``datos para evaluar 2.0.xlsx``, ``datos para evaluar_2.0.xlsx``.

Uso:
  python actualizar_datos_para_evaluar_2_columna_wes_oct_abr.py
  python actualizar_datos_para_evaluar_2_columna_wes_oct_abr.py --dry-run
  python actualizar_datos_para_evaluar_2_columna_wes_oct_abr.py --xlsx "C:/ruta/al/archivo.xlsx"
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from generar_reporte_word import (
    acl_node_base_url,
    fetch_json,
    flatten_measures,
    normalize_measures_payload,
    summarize_consumption,
)

ROOT = Path(__file__).resolve().parent
BASE = ROOT / "reports" / "Renca" / "Coparacion App con Aguas Andinas"

CANDIDATE_NAMES = (
    "datos para evaluar 2.0.xlsx",
    "datos para evaluar_2.0.xlsx",
)

# Ventana fija solicitada (inclusive)
WES_WIN_LO = date(2025, 10, 1)
WES_WIN_HI = date(2026, 4, 30)

COL_HEADER = "m³ app WES (oct-2025 a abr-2026)"

SHEET_NODE_ID: dict[str, str] = {
    "Gimnasio": "000017-05",
    "Facturas_Gimnasio": "000017-05",
    "Piscina": "000017-06",
    "Facturas_Piscina": "000017-06",
    "ICCO": "000017-08",
    "ICCP": "000017-07",
    "Facturas_Escuela_Lo_Velzaquez": "000017-04",
}


def _resolve_xlsx_path(cli: Path | None) -> Path:
    if cli is not None:
        p = cli.expanduser()
        if not p.is_file():
            print(f"[ERROR] No existe el archivo: {p}", file=sys.stderr)
            sys.exit(1)
        return p
    for name in CANDIDATE_NAMES:
        p = BASE / name
        if p.is_file():
            return p
    print(
        "[ERROR] No se encontró el Excel. Coloque uno de:\n  "
        + "\n  ".join(str(BASE / n) for n in CANDIDATE_NAMES),
        file=sys.stderr,
    )
    sys.exit(1)


def _as_date(val) -> date | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(s[:10], fmt).date()
            except ValueError:
                continue
        m = re.match(r"^(\d{1,2})-([A-Za-z]{3})-(\d{4})$", s)
        if m:
            meses = {
                "ene": 1,
                "feb": 2,
                "mar": 3,
                "abr": 4,
                "may": 5,
                "jun": 6,
                "jul": 7,
                "ago": 8,
                "sep": 9,
                "oct": 10,
                "nov": 11,
                "dic": 12,
            }
            dd = int(m.group(1))
            mon = meses.get(m.group(2).lower()[:3])
            yy = int(m.group(3))
            if mon:
                return date(yy, mon, dd)
    return None


def _clip_row_to_window(da: date | None, dd: date | None) -> tuple[date, date] | None:
    dates = [x for x in (da, dd) if x is not None]
    if not dates:
        return None
    lo = min(dates)
    hi = max(dates)
    s = max(lo, WES_WIN_LO)
    e = min(hi, WES_WIN_HI)
    if s > e:
        return None
    return s, e


def _iter_date_chunks(sdt: datetime, edt: datetime, chunk_days: int = 31):
    cur = sdt
    while cur <= edt:
        chunk_end = min(edt, cur + timedelta(days=chunk_days - 1))
        yield cur, chunk_end
        cur = chunk_end + timedelta(days=1)


def _fetch_wes_m3_range(node_id: str, lo: date, hi: date) -> float:
    """Suma m³ WES en [lo, hi] inclusive (misma lógica que comparación factura vs app)."""
    start_dt = datetime(lo.year, lo.month, lo.day)
    end_dt = datetime(hi.year, hi.month, hi.day)
    total_days = (hi - lo).days + 1
    usar_chunks = total_days > 45
    base = acl_node_base_url()
    acum: list = []

    if usar_chunks:
        for c_start, c_end in _iter_date_chunks(start_dt, end_dt):
            raw = fetch_json(
                f"{base}/nodes/measures/dates",
                params=[
                    ("id", node_id),
                    ("start", c_start.strftime("%d%m%Y")),
                    ("end", c_end.strftime("%d%m%Y")),
                ],
            )
            norm = normalize_measures_payload(raw, node_id)
            acum.extend(flatten_measures(norm))
    else:
        end_api = hi + timedelta(days=1)
        raw = fetch_json(
            f"{base}/nodes/measures/dates",
            params=[
                ("id", node_id),
                ("start", lo.strftime("%d%m%Y")),
                ("end", end_api.strftime("%d%m%Y")),
            ],
        )
        norm = normalize_measures_payload(raw, node_id)
        acum = flatten_measures(norm)

    meas = []
    for m in acum:
        md = m.date.date() if hasattr(m.date, "date") else m.date
        if lo <= md <= hi:
            meas.append(m)
    s = summarize_consumption(meas)
    return float(s.get("total", 0.0))


def _find_or_create_column(ws, header: str) -> int:
    for c in range(1, (ws.max_column or 0) + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None and str(v).strip() == header:
            return c
    last = 0
    for c in range(1, (ws.max_column or 0) + 1):
        if ws.cell(row=1, column=c).value not in (None, ""):
            last = c
    new_c = last + 1 if last else 1
    cell = ws.cell(row=1, column=new_c, value=header)
    hdr_fill = PatternFill("solid", fgColor="1F4788")
    hdr_font = Font(color="FFFFFF", bold=True)
    cell.fill = hdr_fill
    cell.font = hdr_font
    return new_c


def _process_sheet(ws, node_id: str) -> tuple[int, int]:
    """Escribe m³ en la columna del encabezado; retorna (filas con valor, filas sin intersección)."""
    col = _find_or_create_column(ws, COL_HEADER)
    written = 0
    empty = 0
    cache: dict[tuple[date, date], float] = {}

    for r in range(2, ws.max_row + 1):
        c1 = ws.cell(row=r, column=1)
        c4 = ws.cell(row=r, column=4)
        if isinstance(c1, MergedCell) or isinstance(c4, MergedCell):
            continue
        clip = _clip_row_to_window(_as_date(c1.value), _as_date(c4.value))
        if clip is None:
            ws.cell(row=r, column=col, value=None)
            empty += 1
            continue
        lo, hi = clip
        key = (lo, hi)
        if key not in cache:
            cache[key] = _fetch_wes_m3_range(node_id, lo, hi)
        ws.cell(row=r, column=col, value=round(cache[key], 3))
        written += 1

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col, max_col=col):
        for c in row:
            if c.value is not None:
                c.number_format = "#,##0.###"
    ws.column_dimensions[get_column_letter(col)].width = 22
    return written, empty


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "--xlsx",
        type=Path,
        default=None,
        help="Ruta al xlsx (por defecto: busca en Coparacion App con Aguas Andinas).",
    )
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="No llama a la API ni guarda; solo indica hojas y filas.",
    )
    ap.add_argument(
        "--solo-hoja",
        default="",
        help="Solo procesa una hoja (nombre exacto en el libro).",
    )
    args = ap.parse_args()

    xlsx = _resolve_xlsx_path(args.xlsx)
    wb = load_workbook(xlsx)

    processed = 0
    for name in wb.sheetnames:
        if args.solo_hoja and name != args.solo_hoja:
            continue
        node_id = SHEET_NODE_ID.get(name)
        if not node_id:
            continue
        ws = wb[name]
        if args.dry_run:
            nrows = max(0, ws.max_row - 1)
            print(f"[dry-run] {name}: nodo {node_id}, hasta {nrows} fila(s) de datos, columna «{COL_HEADER}»")
            processed += 1
            continue
        w, e = _process_sheet(ws, node_id)
        print(f"[OK] {name}: nodo {node_id} — {w} fila(s) con valor, {e} sin rango en ventana oct-25..abr-26.")
        processed += 1

    if processed == 0:
        print(
            "[ADVERTENCIA] No se procesó ninguna hoja. Hojas con mapeo WES: "
            + ", ".join(sorted(SHEET_NODE_ID.keys())),
        )

    if not args.dry_run:
        wb.save(xlsx)
        print(f"[OK] Guardado: {xlsx}")
    wb.close()


if __name__ == "__main__":
    main()

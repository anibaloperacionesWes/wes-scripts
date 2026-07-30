from datetime import datetime, timedelta, date
from pathlib import Path

from openpyxl import load_workbook

from control_nocturno import (
    default_excel_path,
    parse_horario_a_horas,
    obtener_datos_horarios_dia,
)

excel_path = default_excel_path()
wb = load_workbook(excel_path, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))

# find header row containing CLIENTE
header_idx = None
for i, row in enumerate(rows):
    if any((c is not None and str(c).strip().upper() == "CLIENTE") for c in row):
        header_idx = i
        break
if header_idx is None:
    raise SystemExit("No se encontró encabezado CLIENTE")

hdr = rows[header_idx]

def idx_of(*keywords):
    # exact contains match (normalized)
    for idx, cell in enumerate(hdr):
        if cell is None:
            continue
        s = str(cell).strip().upper()
        if all(k in s for k in keywords):
            return idx
    # fallback: any keyword in cell
    for idx, cell in enumerate(hdr):
        if cell is None:
            continue
        s = str(cell).strip().upper()
        for k in keywords:
            if k in s:
                return idx
    return None

# columns
(i_cliente, i_nom, i_id) = (idx_of("CLIENTE"), idx_of("NOMBRE", "COLEGIO"), None)
# sometimes header is just "NOMBRE DEL COLEGIO O LICEO"
if i_nom is None:
    i_nom = idx_of("NOMBRE")
if i_nom is None:
    i_nom = idx_of("COLEGIO", "LICEO")

for j, cell in enumerate(hdr):
    if cell is not None and str(cell).strip().upper() == "ID":
        i_id = j
        break
if i_id is None:
    i_id = idx_of("ID")

i_hor = idx_of("HORARIO")
if i_hor is None:
    i_hor = idx_of("CORTE")

if None in (i_cliente, i_nom, i_id, i_hor):
    raise SystemExit(f"Columnas incompletas: cliente={i_cliente} nom={i_nom} id={i_id} hor={i_hor}")

# find Carlos Fernandez Peña row
found = None
for row in rows[header_idx + 1 :]:
    if not row or len(row) <= max(i_nom, i_id, i_hor, i_cliente):
        continue
    nombre = "" if row[i_nom] is None else str(row[i_nom])
    nombre_norm = nombre.replace("�", "Ñ")
    up = nombre_norm.upper()
    if "CARLOS" in up and "FERNANDEZ" in up and ("PE" in up):
        found = row
        break

if found is None:
    raise SystemExit("No encontré el punto CARLOS FERNANDEZ PEÑA en el Excel")

cliente = "" if found[i_cliente] is None else str(found[i_cliente])
nombre = "" if found[i_nom] is None else str(found[i_nom])
node_id = str(found[i_id]).strip()
horario_txt = "" if found[i_hor] is None else str(found[i_hor])
horas_corte = parse_horario_a_horas(horario_txt)

print("Punto encontrado:")
print("  node_id:", node_id)
print("  nombre :", nombre)
print("  cliente:", cliente)
print("  horario:", horario_txt)
print("  horas_corte:", horas_corte)

umbral = 0.0
start_d = date(2026, 2, 1)
end_d = date(2026, 3, 20)

suspect = []
cur = start_d
n = 0
while cur <= end_d:
    n += 1
    try:
        hourly = obtener_datos_horarios_dia(node_id, datetime.combine(cur, datetime.min.time()))
    except Exception as e:
        print("Fallo API", cur.isoformat(), str(e))
        cur += timedelta(days=1)
        continue

    flagged_hours = []
    max_v = 0.0
    for h in horas_corte:
        v = float(hourly.get(h, 0.0))
        if v > umbral:
            max_v = max(max_v, v)
            # el reporte formatea a 3 decimales
            if round(v, 3) == 0.0:
                flagged_hours.append((h, v))

    if flagged_hours:
        suspect.append((cur.isoformat(), max_v, flagged_hours))

    if n % 10 == 0:
        print("progreso:", cur.isoformat())

    cur += timedelta(days=1)

print("\nResultados:")
if not suspect:
    print("No hay días donde el reporte marcaria horas con valor formateado 0.000 (umbral=0.0).")
else:
    for d, max_v, hrs in suspect:
        hrs_fmt = ", ".join([f"{h:02d}:00={v:.6f} (fmt {v:.3f})" for h, v in hrs])
        print(f"- {d}: max={max_v:.6f} -> {hrs_fmt}")

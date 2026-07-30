"""
Reporte Word agregado Parque Arauco — Mall Kennedy.

Incluye todos los puntos de Kennedy según el listado Excel (columna Mall, p. ej. Arauco Kennedy)
o, si no hay Excel, según get_mall_name_for_parque_arauco + API empresa 000025.

Excluye LISTADO_PA_IDS_EXCLUIDOS. Bajas por mall vía pa_nodos_inactivos_por_mall + generate_aggregated_report.

Salida:
  reports/Parque_Arauco/Kennedy/ABREGADO/AGREGADO_<ts>/Reporte_Agregado_Parque_Arauco_<YYYYMMDD>_<YYYYMMDD>.docx

Ejemplos:
  python generar_pa_kennedy_agregado_periodo.py
  python generar_pa_kennedy_agregado_periodo.py --desde 11/03/2026
  python generar_pa_kennedy_agregado_periodo.py --desde 02/02/2026 --hasta 31/03/2026
"""

from __future__ import annotations

import argparse
import os
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path
from typing import List

import requests

from generar_reporte_word import generate_aggregated_report, get_mall_name_for_parque_arauco
from listado_pa_que_esta_instalado import LISTADO_PA_IDS_EXCLUIDOS
from pa_nodos_inactivos_por_mall import filtrar_nodos_activos_mall

ENTITY_BASE = "http://104.248.53.141:7001/wes/api/acl-entities/v1"

DEFAULT_LISTADO = (
    Path("reports")
    / "Parque_Arauco"
    / "Reportes_agregados_reunion_abril_2026"
    / "listado_que_esta_instalado.xlsx"
)

MALL_NOMBRE = "Kennedy"

# Operativos Kennedy (10 nodos): incluye 24; reemplazos 35/36; sin 25/26 (retirados).
KENNEDY_NODOS_OPERATIVOS = frozenset(
    {
        "000025-20",
        "000025-21",
        "000025-22",
        "000025-23",
        "000025-24",
        "000025-27",
        "000025-28",
        "000025-29",
        "000025-35",
        "000025-36",
    }
)


def _carpetas_escritorio() -> List[Path]:
    """Una o dos rutas de Escritorio: OneDrive y/o local (evita confusión si son distintas)."""
    home = Path.home()
    out: List[Path] = []
    one = home / "OneDrive" / "Desktop"
    local = home / "Desktop"
    if one.is_dir():
        out.append(one.resolve())
    if local.is_dir() and (not out or local.resolve() != out[0]):
        out.append(local.resolve())
    if not out:
        out.append(local.resolve())
    return out


def _es_kennedy_texto(mall_cell: str) -> bool:
    s = (mall_cell or "").strip().lower()
    return "kennedy" in s


def _nodos_kennedy_desde_xlsx(path: Path) -> List[str]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(min_row=1, values_only=True)
        header = next(rows, None)
        if not header:
            return []
        headers = [str(h).strip() if h is not None else "" for h in header]
        try:
            i_mall = headers.index("Mall")
            i_id = headers.index("ID_nodo")
        except ValueError as e:
            raise ValueError(
                "El Excel debe tener columnas 'Mall' e 'ID_nodo' (como el listado_que_esta_instalado)."
            ) from e

        ids: List[str] = []
        for row in rows:
            if not row or max(len(headers), len(row or [])) < 2:
                continue
            mall = row[i_mall] if i_mall < len(row) else None
            nid = row[i_id] if i_id < len(row) else None
            if mall is None or nid is None:
                continue
            if not _es_kennedy_texto(str(mall)):
                continue
            nid_s = str(nid).strip()
            if nid_s.startswith("000025-"):
                ids.append(nid_s)
    finally:
        wb.close()

    return sorted(set(ids))


def _nodos_kennedy_desde_api() -> List[str]:
    r = requests.get(f"{ENTITY_BASE}/companies/000025", timeout=120)
    r.raise_for_status()
    empresa = r.json()
    nodes = empresa.get("nodes") or []
    ids: List[str] = []
    for n in nodes:
        nid = str(n.get("nodeId") or "").strip()
        name = str(n.get("name") or "").strip()
        if not nid:
            continue
        if get_mall_name_for_parque_arauco(nid, name) == MALL_NOMBRE:
            ids.append(nid)
    return sorted(set(ids))


def _filtrar_listado_operativo(node_ids: List[str]) -> List[str]:
    return sorted({n for n in node_ids if n not in LISTADO_PA_IDS_EXCLUIDOS})


def _aplicar_lista_kennedy_operativa(node_ids: List[str]) -> List[str]:
    """Lista canónica Kennedy (10 nodos): incluye 24/35/36; excluye 25/26."""
    candidatos = sorted(set(node_ids) & KENNEDY_NODOS_OPERATIVOS)
    faltantes = sorted(KENNEDY_NODOS_OPERATIVOS - set(candidatos))
    if faltantes:
        print(f"[INFO] Kennedy — agregando nodos operativos requeridos: {', '.join(faltantes)}")
        candidatos = sorted(set(candidatos) | KENNEDY_NODOS_OPERATIVOS)
    excluidos = sorted(set(node_ids) - KENNEDY_NODOS_OPERATIVOS)
    if excluidos:
        print(f"[INFO] Kennedy — excluidos (no operativos / fuera de lista): {', '.join(excluidos)}")
    return candidatos


def main() -> int:
    parser = argparse.ArgumentParser(description="Agregado Word PA Kennedy (todos los puntos del mall).")
    parser.add_argument(
        "--desde",
        default="11/03/2026",
        help="DD/MM/YYYY (inicio periodo; default 11/03/2026: post corrección medidor DL/Sandia Antigua)",
    )
    parser.add_argument(
        "--hasta",
        default=None,
        help="DD/MM/YYYY fin periodo (por defecto: hoy)",
    )
    parser.add_argument(
        "--listado-xlsx",
        type=Path,
        default=None,
        help=f"Excel listado por mall (default: {DEFAULT_LISTADO} si existe)",
    )
    parser.add_argument("--solo-api", action="store_true", help="Ignorar Excel; solo mapeo/API Kennedy.")
    parser.add_argument(
        "--no-copia-escritorio",
        action="store_true",
        help="No copiar el .docx al Escritorio al terminar.",
    )
    parser.add_argument(
        "--no-nota-medidor",
        action="store_true",
        help="No añadir nota sobre corrección del medidor ultrasónico (DL / Sandia Antigua).",
    )
    args = parser.parse_args()

    if sys.platform == "win32":
        try:
            sys.stdout.reconfigure(encoding="utf-8")
            sys.stderr.reconfigure(encoding="utf-8")
        except Exception:
            pass

    hasta = args.hasta or datetime.now().strftime("%d/%m/%Y")

    path_xlsx = args.listado_xlsx
    if path_xlsx is None:
        path_xlsx = DEFAULT_LISTADO

    if args.solo_api or not path_xlsx.exists():
        if not args.solo_api and not path_xlsx.exists():
            print(f"[INFO] No se encontró Excel: {path_xlsx.resolve()} — usando API + mapeo Kennedy.")
        node_ids = _nodos_kennedy_desde_api()
        fuente = f"API + mapeo {MALL_NOMBRE}"
    else:
        node_ids = _nodos_kennedy_desde_xlsx(Path(path_xlsx))
        fuente = f"Excel {path_xlsx}"

    node_ids = _filtrar_listado_operativo(node_ids)
    _antes = set(node_ids)
    node_ids = filtrar_nodos_activos_mall(MALL_NOMBRE, node_ids)
    _drop = sorted(_antes - set(node_ids))
    if _drop:
        print(f"[INFO] Excluidos del agregado (ya no activos): {', '.join(_drop)}")
    node_ids = _aplicar_lista_kennedy_operativa(node_ids)

    if not node_ids:
        print("[ERROR] No quedaron nodos Kennedy tras filtros. Revisa el Excel o el mapeo.")
        return 1

    print(f"[INFO] Fuente nodos: {fuente}")
    print(f"[INFO] Periodo: {args.desde} a {hasta}")
    print(f"[INFO] Nodos Kennedy ({len(node_ids)}): {', '.join(node_ids)}")

    nota_periodo = None
    if not args.no_nota_medidor:
        nota_periodo = (
            f"Nota: el periodo analizado comienza el {args.desde} (posterior a la corrección de "
            "configuración del medidor ultrasónico en Distrito de Lujo (DL) y Sala de Bomba "
            "Sandia Antigua)."
        )

    out = generate_aggregated_report(
        company_id="000025",
        node_ids=node_ids,
        start_date=args.desde,
        end_date=hasta,
        output_dir="reports",
        mall_name=MALL_NOMBRE,
        apply_exclusions=False,
        generate_ppt=False,
        nota_contexto_periodo=nota_periodo,
        parallel_node_fetch=True,
        max_parallel_workers=4,
    )
    out_path = Path(out).resolve()
    print(f"[OK] Reporte agregado Kennedy (ruta completa): {out_path}")
    if not args.no_copia_escritorio:
        alias = "Kennedy_Agregado_ULTIMO.docx"
        desks = _carpetas_escritorio()
        for desk in desks:
            try:
                desk.mkdir(parents=True, exist_ok=True)
                dest = desk / out_path.name
                shutil.copy2(out_path, dest)
                dest_alias = desk / alias
                shutil.copy2(out_path, dest_alias)
                print(f"[OK] Copia en Escritorio: {dest}")
                print(f"[OK] Copia fija (siempre la última generada): {dest_alias}")
            except Exception as e:
                print(f"[AVISO] No se pudo copiar a {desk}: {e}")
        if sys.platform == "win32" and desks:
            try:
                primary = desks[0] / alias
                if primary.exists():
                    subprocess.run(
                        ["explorer", "/select,", os.path.normpath(str(primary))],
                        check=False,
                    )
                    print("[INFO] Se abrió el Explorador de archivos con el archivo seleccionado.")
            except Exception:
                pass
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

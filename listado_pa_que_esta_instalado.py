"""
1) ¿Qué está instalado?

Listado Parque Arauco (000025): por punto indica mall (ej. Arauco Maipú), ID, nombre
y la primera fecha en que se registró consumo (total diario > 0) en un rango
configurable. Por defecto el rango comienza el 01/10/2025 (octubre 2025).

Por defecto se excluyen del listado los IDs fuera de operación (LISTADO_PA_IDS_EXCLUIDOS).
Usar --incluir-todos para listar todos los nodos de la API.

Salida por defecto (misma carpeta, se sobrescribe cada vez):
  - listado_que_esta_instalado.xlsx  -> columnas reales en Excel (recomendado)
  - listado_que_esta_instalado.csv   -> separador ";" para Excel regional ES/CL
  --solo-xlsx : solo genera el .xlsx (sin CSV).
Ruta base con --salida (archivo .csv o .xlsx; se genera el par salvo --solo-xlsx).
"""

from __future__ import annotations

import argparse
import csv
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import List, Optional, Tuple

import requests

ENTITY_BASE = "http://104.248.53.141:7001/wes/api/acl-entities/v1"

# Puntos que ya no están en operación (no se incluyen en el listado).
LISTADO_PA_IDS_EXCLUIDOS: frozenset[str] = frozenset(
    {
        "000025-02",
        "000025-03",
        "000025-05",
        "000025-06",
    }
)

from generar_reporte_word import (
    BASE_URL,
    fetch_json,
    flatten_measures,
    get_mall_name_for_parque_arauco,
    get_node_name,
    normalize_measures_payload,
    parse_date,
)
from pa_registro_observaciones import observacion_para_nodo, recepcion_obras_para_mall


def _format_ddmmyyyy(dt: datetime) -> str:
    return dt.strftime("%d%m%Y")


def _iter_date_chunks(sdt: datetime, edt: datetime, chunk_days: int = 31):
    cur = sdt
    while cur <= edt:
        chunk_end = min(edt, cur + timedelta(days=chunk_days - 1))
        yield cur, chunk_end
        cur = chunk_end + timedelta(days=1)


def _medidas_rango(node_id: str, start_dt: datetime, end_dt: datetime) -> List:
    """Descarga medidas diarias en el rango (por tramos si el periodo es largo)."""
    from generar_reporte_word import MeasurePoint

    total_days = (end_dt.date() - start_dt.date()).days + 1
    usar_chunks = total_days > 45
    acum: List[MeasurePoint] = []

    if usar_chunks:
        for c_start, c_end in _iter_date_chunks(start_dt, end_dt):
            measures_payload_raw = fetch_json(
                f"{BASE_URL}/nodes/measures/dates",
                params=[
                    ("id", node_id),
                    ("start", _format_ddmmyyyy(c_start)),
                    ("end", _format_ddmmyyyy(c_end)),
                ],
            )
            measures_payload = normalize_measures_payload(measures_payload_raw, node_id)
            acum.extend(flatten_measures(measures_payload))
    else:
        measures_payload_raw = fetch_json(
            f"{BASE_URL}/nodes/measures/dates",
            params=[
                ("id", node_id),
                ("start", _format_ddmmyyyy(start_dt)),
                ("end", _format_ddmmyyyy(end_dt)),
            ],
        )
        measures_payload = normalize_measures_payload(measures_payload_raw, node_id)
        acum = flatten_measures(measures_payload)

    return sorted(acum, key=lambda m: m.date)


def _primera_fecha_consumo(
    node_id: str, start_dt: datetime, end_dt: datetime
) -> Tuple[Optional[datetime], str]:
    """
    Primera fecha (orden cronológico) con total_m3 > 0 en [start_dt, end_dt].
    Retorna (fecha o None, detalle si hubo error).
    """
    try:
        pts = _medidas_rango(node_id, start_dt, end_dt)
    except Exception as e:
        return None, str(e)

    start_d = start_dt.date()
    end_d = end_dt.date()
    for m in pts:
        d = m.date.date() if hasattr(m.date, "date") else m.date
        if d < start_d or d > end_d:
            continue
        if float(m.total_m3 or 0) > 1e-9:
            return m.date if hasattr(m.date, "replace") else None, ""
    return None, ""


def _mall_arauco(node_id: str, node_name: str) -> str:
    mall = get_mall_name_for_parque_arauco(node_id, node_name).strip()
    if mall:
        return f"Arauco {mall}"
    return "Arauco (sin mall asignado)"


def _cargar_nodos_pa() -> dict:
    r = requests.get(f"{ENTITY_BASE}/companies/000025", timeout=60)
    r.raise_for_status()
    return r.json()


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Listado PA: mall Arauco X, ID, nombre, primera fecha con consumo en rango"
    )
    parser.add_argument(
        "--desde",
        default="01/10/2025",
        help="Inicio del periodo para buscar primera medición con consumo (DD/MM/YYYY). Default: 01/10/2025",
    )
    parser.add_argument(
        "--hasta",
        default="",
        help="Fin del periodo (DD/MM/YYYY). Vacío = hoy",
    )
    parser.add_argument(
        "--salida",
        default="",
        help="Archivo de salida .csv o .xlsx (se crea el par con el mismo nombre base). "
        "Default: carpeta reunión abril 2026 / listado_que_esta_instalado.csv y .xlsx",
    )
    parser.add_argument(
        "--solo-xlsx",
        action="store_true",
        help="Solo escribe el .xlsx (no genera CSV).",
    )
    parser.add_argument(
        "--incluir-todos",
        action="store_true",
        help="Incluir también los nodos en LISTADO_PA_IDS_EXCLUIDOS (listado completo API).",
    )
    args = parser.parse_args()

    start_dt = parse_date(args.desde)
    if args.hasta.strip():
        end_dt = parse_date(args.hasta, end_of_day=True)
    else:
        end_dt = datetime.now(timezone.utc).replace(
            hour=23, minute=59, second=59, microsecond=0
        )

    if end_dt < start_dt:
        print("[ERROR] --hasta debe ser >= --desde")
        return 1

    default_dir = (
        Path("reports")
        / "Parque_Arauco"
        / "Reportes_agregados_reunion_abril_2026"
    )
    if args.salida.strip():
        sal = Path(args.salida)
        if sal.is_dir():
            out_csv = sal / "listado_que_esta_instalado.csv"
            out_xlsx = sal / "listado_que_esta_instalado.xlsx"
        elif sal.suffix.lower() == ".xlsx":
            out_xlsx = sal
            out_csv = sal.with_suffix(".csv")
        else:
            out_csv = sal if sal.suffix.lower() == ".csv" else sal.with_suffix(".csv")
            out_xlsx = out_csv.with_suffix(".xlsx")
    else:
        default_dir.mkdir(parents=True, exist_ok=True)
        out_csv = default_dir / "listado_que_esta_instalado.csv"
        out_xlsx = default_dir / "listado_que_esta_instalado.xlsx"

    payload = _cargar_nodos_pa()
    nodes = sorted(
        payload.get("nodes") or [],
        key=lambda x: str(x.get("nodeId", "")),
    )

    periodo_txt = f"{args.desde} a {end_dt.strftime('%d/%m/%Y')}"

    print(f"[INFO] Periodo para «primera vez con consumo»: {periodo_txt}\n")

    filas = []
    for n in nodes:
        nid = str(n.get("nodeId", "")).strip()
        if not nid:
            continue
        if not args.incluir_todos and nid in LISTADO_PA_IDS_EXCLUIDOS:
            continue
        nombre_api = str(n.get("name", "") or "").strip()
        nombre = nombre_api or get_node_name(nid)
        mall = _mall_arauco(nid, nombre)

        primera, err = _primera_fecha_consumo(nid, start_dt, end_dt)
        if primera:
            fecha_str = primera.strftime("%d-%m-%Y") if hasattr(primera, "strftime") else "—"
        elif err:
            fecha_str = f"(error API: {err[:80]})"
        else:
            fecha_str = "Sin consumo > 0 en el periodo"

        filas.append(
            {
                "Mall": mall,
                "ID_nodo": nid,
                "Nombre_punto": nombre.replace("\n", " ").strip(),
                "Primera_fecha_consumo_registrado": fecha_str,
                "Recepción de obras": recepcion_obras_para_mall(mall),
                "Observaciones": observacion_para_nodo(nid),
            }
        )
        print(f"  {nid} {mall} ... {fecha_str}")

    columnas = [
        "Mall",
        "ID_nodo",
        "Nombre_punto",
        "Primera_fecha_consumo_registrado",
        "Recepción de obras",
        "Observaciones",
    ]

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    if not args.solo_xlsx:
        with open(out_csv, "w", newline="", encoding="utf-8-sig") as f:
            # Punto y coma: en Excel (España / Chile / varios paises) la coma NO separa
            # columnas al abrir doble clic; el separador de listas es ";".
            w = csv.DictWriter(
                f,
                fieldnames=columnas,
                delimiter=";",
                quoting=csv.QUOTE_MINIMAL,
            )
            w.writeheader()
            w.writerows(filas)

    try:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Que_esta_instalado"
        for c, titulo in enumerate(columnas, 1):
            ws.cell(row=1, column=c, value=titulo)
        for r, fila in enumerate(filas, 2):
            for c, key in enumerate(columnas, 1):
                ws.cell(row=r, column=c, value=fila[key])
        from openpyxl.utils import get_column_letter

        anchos = (28, 14, 36, 22, 18, 62)
        for i, w in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        wb.save(out_xlsx)
    except ImportError:
        print(
            "[AVISO] openpyxl no instalado; solo se genero CSV. "
            "pip install openpyxl para obtener el .xlsx con columnas en Excel."
        )

    print(f"\n[OK] {len(filas)} filas")
    if not args.solo_xlsx:
        print(f"     CSV:  {out_csv.resolve()}")
    if out_xlsx.exists():
        print(f"     XLSX: {out_xlsx.resolve()} (abre este en Excel: una columna por dato)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

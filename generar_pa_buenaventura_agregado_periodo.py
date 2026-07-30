"""
Reporte Word agregado Parque Arauco — Mall Buenaventura.

Incluye todos los puntos de Buenaventura según el listado Excel (columna Mall, p. ej. Arauco Buenaventura)
o, si no hay Excel, según get_mall_name_for_parque_arauco + API empresa 000025.

Excluye LISTADO_PA_IDS_EXCLUIDOS. Los nodos dados de baja por mall se aplican en generate_aggregated_report (pa_nodos_inactivos_por_mall).

Salida:
  reports/Parque_Arauco/Buenaventura/ABREGADO/AGREGADO_<ts>/Reporte_Agregado_Parque_Arauco_<YYYYMMDD>_<YYYYMMDD>.docx

Ejemplos:
  python generar_pa_buenaventura_agregado_periodo.py
  python generar_pa_buenaventura_agregado_periodo.py --desde 02/02/2026 --hasta 31/03/2026
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path
from typing import List

import requests

from generar_reporte_word import generate_aggregated_report, get_mall_name_for_parque_arauco
from listado_pa_que_esta_instalado import LISTADO_PA_IDS_EXCLUIDOS
from pa_nodos_inactivos_por_mall import filtrar_nodos_activos_mall

ENTITY_BASE = "http://104.248.53.141:7001/wes/api/acl-entities/v1"

DEFAULT_LISTADO = (
    Path("reports")
    / "Parque_Arauco"
    / "Reportes_agregados_reunion_abril_2026"
    / "listado_que_esta_instalado.xlsx"
)

MALL_NOMBRE = "Buenaventura"


def _es_buenaventura_texto(mall_cell: str) -> bool:
    s = (mall_cell or "").strip().lower()
    return "buenaventura" in s


def _nodos_buenaventura_desde_xlsx(path: Path) -> List[str]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(min_row=1, values_only=True)
        header = next(rows, None)
        if not header:
            return []
        headers = [str(h).strip() if h is not None else "" for h in header]
        try:
            i_mall = headers.index("Mall")
            i_id = headers.index("ID_nodo")
        except ValueError as e:
            raise ValueError(
                "El Excel debe tener columnas 'Mall' e 'ID_nodo' (como el listado_que_esta_instalado)."
            ) from e

        ids: List[str] = []
        for row in rows:
            if not row or max(len(headers), len(row or [])) < 2:
                continue
            mall = row[i_mall] if i_mall < len(row) else None
            nid = row[i_id] if i_id < len(row) else None
            if mall is None or nid is None:
                continue
            if not _es_buenaventura_texto(str(mall)):
                continue
            nid_s = str(nid).strip()
            if nid_s.startswith("000025-"):
                ids.append(nid_s)
    finally:
        wb.close()

    return sorted(set(ids))


def _nodos_buenaventura_desde_api() -> List[str]:
    r = requests.get(f"{ENTITY_BASE}/companies/000025", timeout=120)
    r.raise_for_status()
    empresa = r.json()
    nodes = empresa.get("nodes") or []
    ids: List[str] = []
    for n in nodes:
        nid = str(n.get("nodeId") or "").strip()
        name = str(n.get("name") or "").strip()
        if not nid:
            continue
        if get_mall_name_for_parque_arauco(nid, name) == MALL_NOMBRE:
            ids.append(nid)
    return sorted(set(ids))


def _filtrar_listado_operativo(node_ids: List[str]) -> List[str]:
    return sorted({n for n in node_ids if n not in LISTADO_PA_IDS_EXCLUIDOS})


def main() -> int:
    parser = argparse.ArgumentParser(description="Agregado Word PA Buenaventura (todos los puntos del mall).")
    parser.add_argument("--desde", default="02/02/2026", help="DD/MM/YYYY (inicio periodo)")
    parser.add_argument(
        "--hasta",
        default=None,
        help="DD/MM/YYYY fin periodo (por defecto: hoy)",
    )
    parser.add_argument(
        "--listado-xlsx",
        type=Path,
        default=None,
        help=f"Excel listado por mall (default: {DEFAULT_LISTADO} si existe)",
    )
    parser.add_argument("--solo-api", action="store_true", help="Ignorar Excel; solo mapeo/API Buenaventura.")
    args = parser.parse_args()

    if sys.platform == "win32":
        try:
            sys.stdout.reconfigure(encoding="utf-8")
            sys.stderr.reconfigure(encoding="utf-8")
        except Exception:
            pass

    hasta = args.hasta or datetime.now().strftime("%d/%m/%Y")

    path_xlsx = args.listado_xlsx
    if path_xlsx is None:
        path_xlsx = DEFAULT_LISTADO

    if args.solo_api or not path_xlsx.exists():
        if not args.solo_api and not path_xlsx.exists():
            print(f"[INFO] No se encontró Excel: {path_xlsx.resolve()} — usando API + mapeo Buenaventura.")
        node_ids = _nodos_buenaventura_desde_api()
        fuente = f"API + mapeo {MALL_NOMBRE}"
    else:
        node_ids = _nodos_buenaventura_desde_xlsx(Path(path_xlsx))
        fuente = f"Excel {path_xlsx}"

    node_ids = _filtrar_listado_operativo(node_ids)
    _antes = set(node_ids)
    node_ids = filtrar_nodos_activos_mall(MALL_NOMBRE, node_ids)
    _drop = sorted(_antes - set(node_ids))
    if _drop:
        print(f"[INFO] Excluidos del agregado (ya no activos): {', '.join(_drop)}")

    if not node_ids:
        print("[ERROR] No quedaron nodos Buenaventura tras filtros. Revisa el Excel o el mapeo.")
        return 1

    print(f"[INFO] Fuente nodos: {fuente}")
    print(f"[INFO] Periodo: {args.desde} a {hasta}")
    print(f"[INFO] Nodos Buenaventura ({len(node_ids)}): {', '.join(node_ids)}")

    out = generate_aggregated_report(
        company_id="000025",
        node_ids=node_ids,
        start_date=args.desde,
        end_date=hasta,
        output_dir="reports",
        mall_name=MALL_NOMBRE,
        apply_exclusions=False,
        generate_ppt=False,
    )
    print(f"[OK] Reporte agregado Buenaventura: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

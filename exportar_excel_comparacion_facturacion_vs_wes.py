"""
Exporta un Excel comparando facturación Aguas Andinas vs consumo App WES entre lecturas.

Columnas solicitadas:
- Periodo
- Emision
- N° factura
- Lectura inicial
- Lectura final
- Diferencia de lectura (días)
- m3 facturado
- m3 app entre las lecturas
- días WES
- diferencia fact - WES

Ejemplos:
  python exportar_excel_comparacion_facturacion_vs_wes.py --site iccp
  python exportar_excel_comparacion_facturacion_vs_wes.py --site icco
  python exportar_excel_comparacion_facturacion_vs_wes.py --site gimnasio
  python exportar_excel_comparacion_facturacion_vs_wes.py --site piscina
  python exportar_excel_comparacion_facturacion_vs_wes.py --site icco --desde 26-08-2025
"""

from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from facturacion_aguas_andinas_pdf import listar_periodos_desde_pdf
from generar_reporte_word import (
    acl_node_base_url,
    fetch_json,
    flatten_measures,
    format_number_chilean,
    normalize_measures_payload,
    summarize_consumption,
)


ROOT = Path(__file__).resolve().parent

SITES = {
    "icco": {
        "node_id": "000017-08",
        "empresa": "Renca",
        "facturas_dir": ROOT / "reports" / "Renca" / "Coparacion App con Aguas Andinas" / "ICCO facturas",
        "out_dir": ROOT / "reports" / "Renca" / "Coparacion App con Aguas Andinas" / "reporte_comparacion_Icco",
        "tag": "ICCO_Renca",
        "desde_default": "28-09-2025",
    },
    "iccp": {
        "node_id": "000017-07",
        "empresa": "Renca",
        "facturas_dir": ROOT / "reports" / "Renca" / "Coparacion App con Aguas Andinas" / "ICCP facturas",
        "out_dir": ROOT / "reports" / "Renca" / "Coparacion App con Aguas Andinas" / "reporte_comparacion_Iccp",
        "tag": "ICCP_Renca",
        "desde_default": "26-08-2025",
    },
    "lo_velazquez": {
        "node_id": "000017-04",
        "empresa": "Renca",
        "facturas_dir": ROOT
        / "reports"
        / "Renca"
        / "Coparacion App con Aguas Andinas"
        / "Escuela lo Velzaquez Facturas",
        "out_dir": ROOT
        / "reports"
        / "Renca"
        / "Coparacion App con Aguas Andinas"
        / "reporte_comparacion_Lo_Velasquez",
        "tag": "Lo_Velasquez_Renca",
        "desde_default": "26-08-2025",
    },
    "gimnasio": {
        "node_id": "000017-05",
        "empresa": "Renca",
        "facturas_dir": ROOT
        / "reports"
        / "Renca"
        / "Coparacion App con Aguas Andinas"
        / "Gimnasio Facturaciones",
        "out_dir": ROOT
        / "reports"
        / "Renca"
        / "Coparacion App con Aguas Andinas"
        / "reporte_comparacion_Gimnasio",
        "tag": "Gimnasio_Renca",
        "desde_default": "06-06-2025",
    },
    "piscina": {
        "node_id": "000017-06",
        "empresa": "Renca",
        "facturas_dir": ROOT
        / "reports"
        / "Renca"
        / "Coparacion App con Aguas Andinas"
        / "Piscina Facturaciones",
        "out_dir": ROOT
        / "reports"
        / "Renca"
        / "Coparacion App con Aguas Andinas"
        / "reporte_comparacion_Piscina",
        "tag": "Piscina_Renca",
        "desde_default": "06-06-2025",
    },
}

_MESES = {
    "ENE": 1,
    "JAN": 1,
    "FEB": 2,
    "MAR": 3,
    "ABR": 4,
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AGO": 8,
    "AUG": 8,
    "SEP": 9,
    "OCT": 10,
    "NOV": 11,
    "DIC": 12,
    "DEC": 12,
}


def _parse_fecha_es_dd_mmm_yyyy(s: str) -> datetime:
    p = s.strip().upper().split("-")
    if len(p) != 3:
        raise ValueError(f"Fecha inválida: {s!r}")
    dd = int(p[0])
    mm = _MESES[p[1][:3]]
    yy = int(p[2])
    return datetime(yy, mm, dd)


def _parse_fecha_flexible(s: str) -> datetime:
    s = s.strip()
    if not s:
        raise ValueError("Fecha vacía")
    p = s.split("-")
    if len(p) == 3 and p[0].isdigit() and p[1].isdigit() and p[2].isdigit():
        dd, mm, yy = int(p[0]), int(p[1]), int(p[2])
        return datetime(yy, mm, dd)
    return _parse_fecha_es_dd_mmm_yyyy(s)


def _fmt_periodo_corto(dt: datetime) -> str:
    meses = {
        1: "ene",
        2: "feb",
        3: "mar",
        4: "abr",
        5: "may",
        6: "jun",
        7: "jul",
        8: "ago",
        9: "sep",
        10: "oct",
        11: "nov",
        12: "dic",
    }
    return f"{dt.day:02d}-{meses[dt.month]}-{dt.year}"


def _to_ddmmyyyy(dt: datetime) -> str:
    return dt.strftime("%d%m%Y")


def _parse_cli_fecha_dd_mm_yyyy(s: str) -> date:
    """Parse ``DD-MM-AAAA`` (p. ej. 26-08-2025)."""
    p = s.strip().split("-")
    if len(p) != 3:
        raise ValueError(f"Fecha inválida: {s!r}")
    dd, mm, yy = int(p[0]), int(p[1]), int(p[2])
    return date(yy, mm, dd)


@dataclass(frozen=True)
class FacturaAA:
    pdf: Path
    boleta: str
    emision: datetime
    lectura_inicial: datetime
    lectura_inicial_num: float | None
    lectura_final: datetime
    lectura_final_num: float | None
    m3_facturado: int
    cuenta: str | None = None
    medidor: str | None = None

    @property
    def periodo_txt(self) -> str:
        return f"{_fmt_periodo_corto(self.lectura_inicial)} → {_fmt_periodo_corto(self.lectura_final)}"

    @property
    def api_start(self) -> str:
        return _to_ddmmyyyy(self.lectura_inicial)


CSV_FACTURACION_MANUAL = "facturacion_aguas_andinas.csv"


def _norm_csv_key(k: str) -> str:
    return re.sub(r"\s+", "_", (k or "").strip().lower())


def _cargar_facturas_desde_csv(path: Path) -> list[FacturaAA]:
    """Misma convención CSV que los informes Gimnasio/Piscina (``generar_reporte_comparacion_*_000017_05/06.py``)."""
    out: list[FacturaAA] = []
    with path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            return out
        remap = {_norm_csv_key(k): k for k in reader.fieldnames if k}
        req = ["boleta", "fecha_emision", "lectura_anterior", "lectura_actual", "m3_cuenta"]
        for col in req:
            if col not in remap:
                raise ValueError(
                    f"CSV {path.name}: falta columna «{col}». Encabezados: {reader.fieldnames}"
                )
        for row in reader:
            raw_boleta = (row.get(remap["boleta"]) or "").strip()
            if not raw_boleta or raw_boleta.startswith("#"):
                continue
            emision = _parse_fecha_flexible((row.get(remap["fecha_emision"]) or "").strip())
            la = _parse_fecha_flexible((row.get(remap["lectura_anterior"]) or "").strip())
            lb = _parse_fecha_flexible((row.get(remap["lectura_actual"]) or "").strip())
            m3_raw = (row.get(remap["m3_cuenta"]) or "").strip().replace(".", "").replace(",", ".")
            m3_fact = int(float(m3_raw))
            cuenta_val = None
            medidor_val = None
            if "cuenta" in remap:
                v = (row.get(remap["cuenta"]) or "").strip()
                cuenta_val = v or None
            if "medidor" in remap:
                v = (row.get(remap["medidor"]) or "").strip()
                medidor_val = v or None
            out.append(
                FacturaAA(
                    pdf=path,
                    boleta=raw_boleta,
                    emision=emision,
                    lectura_inicial=la,
                    lectura_inicial_num=None,
                    lectura_final=lb,
                    lectura_final_num=None,
                    m3_facturado=m3_fact,
                    cuenta=cuenta_val,
                    medidor=medidor_val,
                )
            )
    out.sort(key=lambda r: r.lectura_final)
    return out


def _cargar_facturas(dir_: Path) -> list[FacturaAA]:
    csv_path = dir_ / CSV_FACTURACION_MANUAL
    if csv_path.is_file():
        rows = _cargar_facturas_desde_csv(csv_path)
        if not rows:
            raise FileNotFoundError(f"CSV sin filas válidas: {csv_path}")
        return rows
    pdfs = sorted(dir_.glob("*.pdf"))
    if not pdfs:
        raise FileNotFoundError(
            f"No hay {CSV_FACTURACION_MANUAL} ni PDFs en {dir_}"
        )
    rows: list[FacturaAA] = []
    errores: list[str] = []
    for p in pdfs:
        try:
            for per in listar_periodos_desde_pdf(p):
                rows.append(
                    FacturaAA(
                        pdf=per.pdf,
                        boleta=per.boleta,
                        emision=per.emision,
                        lectura_inicial=per.lectura_anterior,
                        lectura_inicial_num=None,
                        lectura_final=per.lectura_actual,
                        lectura_final_num=None,
                        m3_facturado=per.m3_cuenta,
                        cuenta=per.cuenta,
                        medidor=per.medidor,
                    )
                )
        except ValueError as e:
            errores.append(f"{p.name}: {e}")
    if not rows:
        detalle = "\n".join(errores) if errores else ""
        raise FileNotFoundError(
            f"No se pudieron extraer períodos desde PDFs en {dir_}.\n{detalle}"
        )
    rows.sort(key=lambda r: r.lectura_final)
    return rows


def _fetch_wes_total(node_id: str, f: FacturaAA) -> tuple[float, int]:
    # Pedir un día extra y filtrar rango inclusive
    start = f.api_start
    end_api = _to_ddmmyyyy(f.lectura_final + timedelta(days=1))
    base = acl_node_base_url()
    raw = fetch_json(
        f"{base}/nodes/measures/dates",
        params=[("id", node_id), ("start", start), ("end", end_api)],
    )
    norm = normalize_measures_payload(raw, node_id)
    meas = flatten_measures(norm)
    d0, d1 = f.lectura_inicial.date(), f.lectura_final.date()
    meas = [m for m in meas if d0 <= m.date.date() <= d1]
    s = summarize_consumption(meas)
    return float(s.get("total", 0.0)), int(s.get("dias", 0))


def _fmt_num(n: float, dec: int = 1) -> str:
    return format_number_chilean(n, dec)


def _write_xlsx(
    out_path: Path,
    facturas: list[FacturaAA],
    node_id: str,
    *,
    desde: date,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparación"

    headers = [
        "Nodo WES",
        "Archivo PDF",
        "Cuenta Aguas Andinas",
        "N° medidor (boleta/PDF)",
        "Periodo (lectura inicial → lectura final)",
        "Fecha emisión factura",
        "N° factura / boleta",
        "Lectura inicial (fecha contador)",
        "N° lectura inicial contador",
        "Lectura final (fecha contador)",
        "N° lectura final contador",
        "Días entre lecturas (calendario)",
        "m³ facturado en boleta",
        "m³ App WES (suma entre las mismas fechas)",
        "Días con registro WES en el período",
        "Diferencia m³ (facturado − App WES)",
        "% diferencia vs m³ facturado",
    ]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4788")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for f in facturas:
        m3_wes, dias_wes = _fetch_wes_total(node_id, f)
        diff = float(f.m3_facturado) - float(m3_wes)
        pct = (100.0 * diff / float(f.m3_facturado)) if f.m3_facturado else 0.0
        delta_dias = (f.lectura_final.date() - f.lectura_inicial.date()).days
        ws.append(
            [
                node_id,
                f.pdf.name,
                f.cuenta or "",
                f.medidor or "",
                f.periodo_txt,
                f.emision.strftime("%d-%m-%Y"),
                str(f.boleta),
                f.lectura_inicial.strftime("%d-%m-%Y"),
                (float(f.lectura_inicial_num) if f.lectura_inicial_num is not None else ""),
                f.lectura_final.strftime("%d-%m-%Y"),
                (float(f.lectura_final_num) if f.lectura_final_num is not None else ""),
                int(delta_dias),
                int(f.m3_facturado),
                float(m3_wes),
                int(dias_wes),
                float(diff),
                float(pct),
            ]
        )

    # Formato numérico + alineación (columnas 1-based en comentarios)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
        row[8].number_format = "#,##0.0"  # I: N° lectura inicial
        row[10].number_format = "#,##0.0"  # K: N° lectura final
        row[11].number_format = "0"  # L: días período
        row[12].number_format = "#,##0"  # M: m³ fact
        row[13].number_format = "#,##0.0"  # N: m³ WES
        row[14].number_format = "0"  # O: días WES
        row[15].number_format = "#,##0.0"  # P: dif m³
        row[16].number_format = "0.0"  # Q: % (valor numérico; se entiende como porcentaje)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    widths = [12, 28, 18, 20, 30, 14, 16, 16, 16, 16, 16, 22, 16, 28, 22, 26, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    sn = wb.create_sheet("Notas")
    sn["A1"] = "Criterio y hallazgo"
    sn["A2"] = (
        f"Hoja «Comparación»: solo filas cuya lectura inicial y lectura final del período en la factura son "
        f"≥ {desde.strftime('%d-%m-%Y')} (inicio monitoreo de referencia)."
    )
    sn["A3"] = (
        "Los desvíos entre facturación y app en ciclos de transición o estimación suelen explicarse por "
        "calendario de lecturas y alcance del medidor; el detalle por sitio figura en el informe Word correspondiente."
    )
    sn["A4"] = (
        "Cada columna indica un dato usado en la comparación facturación vs App WES; la hoja «Descripcion_columnas» "
        "explica el significado de cada encabezado."
    )
    sn.column_dimensions["A"].width = 100

    dc = wb.create_sheet("Descripcion_columnas")
    dc.append(["Columna (encabezado)", "Qué dato es y cómo se usa en la comparación"])
    h2 = dc.row_dimensions[1]
    h2.height = 28
    for c in dc[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    descripciones = [
        ("Nodo WES", "Identificador del punto en la API WES del cual se suman los m³ de la columna «m³ App WES»."),
        ("Archivo PDF", "Nombre del PDF de Aguas Andinas del que se extrajo el período (o CSV si aplica)."),
        ("Cuenta Aguas Andinas", "Número de cuenta en la boleta/PDF, si consta en el documento."),
        ("N° medidor (boleta/PDF)", "Identificador del medidor en la boleta/PDF, si consta."),
        (
            "Periodo (lectura inicial → lectura final)",
            "Ventana de fechas del ciclo facturado; el consumo WES se suma exactamente entre esas fechas inclusive.",
        ),
        ("Fecha emisión factura", "Fecha de emisión informada en la factura (referencia administrativa)."),
        ("N° factura / boleta", "Folio de la factura o referencia de fila en historial de consumo."),
        ("Lectura inicial (fecha contador)", "Fecha de inicio del período de liquidación según Aguas Andinas."),
        (
            "N° lectura inicial contador",
            "Lectura numérica del medidor al inicio, si el PDF/CSV la trae; vacío en historiales solo con fechas.",
        ),
        ("Lectura final (fecha contador)", "Fecha de término del período de liquidación según Aguas Andinas."),
        (
            "N° lectura final contador",
            "Lectura numérica del medidor al término, si el PDF/CSV la trae; vacío en historiales solo con fechas.",
        ),
        (
            "Días entre lecturas (calendario)",
            "Días civiles entre lectura inicial y final (calendario del período facturado).",
        ),
        ("m³ facturado en boleta", "Consumo total en m³ cobrado en la factura para ese período (base de comparación)."),
        (
            "m³ App WES (suma entre las mismas fechas)",
            "Consumo acumulado según medidas WES del nodo entre la misma lectura inicial y final.",
        ),
        (
            "Días con registro WES en el período",
            "Días con datos de medición WES dentro del rango (puede ser menor que los días de calendario).",
        ),
        (
            "Diferencia m³ (facturado − App WES)",
            "Resta: positivo = la factura lleva más m³ que lo registrado por WES en el mismo intervalo.",
        ),
        (
            "% diferencia vs m³ facturado",
            "100 × (diferencia m³) / (m³ facturado); indica qué fracción del facturado no coincide con WES.",
        ),
    ]
    for titulo, desc in descripciones:
        dc.append([titulo, desc])
    for row in dc.iter_rows(min_row=2, max_row=dc.max_row):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    dc.column_dimensions["A"].width = 42
    dc.column_dimensions["B"].width = 88

    wb.save(out_path)


def main() -> None:
    ap = argparse.ArgumentParser(description="Exportar Excel de comparación facturación vs WES.")
    ap.add_argument(
        "--site",
        choices=sorted(SITES.keys()),
        default="iccp",
        help="Sitio: icco, iccp, lo_velazquez (000017-04), gimnasio (000017-05) o piscina (000017-06); desde por defecto 06-06-2025 en gimnasio y piscina.",
    )
    ap.add_argument(
        "--desde",
        default=None,
        metavar="DD-MM-AAAA",
        help="Solo facturas con lectura inicial y final ≥ esta fecha (por defecto: según sitio en SITES, p. ej. ICCP y Lo Velázquez 26-08-2025, ICCO 28-09-2025).",
    )
    args = ap.parse_args()

    cfg = SITES[args.site]
    desde_str = args.desde if args.desde is not None else cfg.get("desde_default", "26-08-2025")
    try:
        desde = _parse_cli_fecha_dd_mm_yyyy(desde_str)
    except ValueError as e:
        raise SystemExit(f"[ERROR] --desde: {e}") from e
    node_id = cfg["node_id"]
    in_dir: Path = cfg["facturas_dir"]
    out_dir: Path = cfg["out_dir"]
    out_dir.mkdir(parents=True, exist_ok=True)

    facturas_all = _cargar_facturas(in_dir)
    facturas = [
        f
        for f in facturas_all
        if f.lectura_inicial.date() >= desde and f.lectura_final.date() >= desde
    ]
    if not facturas:
        raise SystemExit(
            f"[ERROR] Sin facturas con lectura inicial y lectura final ≥ {desde.strftime('%d-%m-%Y')} en {in_dir}. "
            "Revise PDFs o use otro --desde."
        )
    out_path = out_dir / f"Comparacion_Facturacion_vs_WES_{cfg['tag']}.xlsx"
    _write_xlsx(out_path, facturas, node_id, desde=desde)
    print(f"[OK] XLSX: {out_path} ({len(facturas)} períodos desde {desde.strftime('%d-%m-%Y')})")


if __name__ == "__main__":
    main()


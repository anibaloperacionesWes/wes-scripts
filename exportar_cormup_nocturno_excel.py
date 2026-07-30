"""
Excel consumo nocturno CORMUP — una hoja por colegio.
Ventana: CSV dates.measures.csv por día, marcas TIME UTC 00:00–07:00 (mismo criterio del informe).

Uso:
  python exportar_cormup_nocturno_excel.py
  python exportar_cormup_nocturno_excel.py --desde 01/05/2026 --hasta 31/05/2026
"""
from __future__ import annotations

import argparse
import re
import sys
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from generar_reporte_word import (
    _fetch_csv_colegio_dia,
    _total_m3_from_json_for_chile_day,
    calculate_nocturnal_metrics,
    get_node_name,
    parse_date,
)

COLEGIOS = [f"000008-{i:02d}" for i in range(1, 15)]


def _slug_hoja(nombre: str, node_id: str) -> str:
    base = re.sub(r"[^\w\s\-]", "", nombre, flags=re.UNICODE).strip()[:22]
    suf = node_id.split("-")[-1]
    name = f"{base}_{suf}" if base else node_id
    return name[:31]


def _style_header(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _horas_utc_0_7_por_dia(node_id: str, dia: date) -> tuple[dict[int, float], float]:
    """Horas UTC 0–7 y suma del día (criterio colegios)."""
    try:
        by_time = _fetch_csv_colegio_dia(node_id, dia)
    except Exception:
        return {}, 0.0
    horas: dict[int, float] = {h: 0.0 for h in range(8)}
    for time_str, val in by_time.items():
        try:
            ts_norm = time_str.strip().replace("Z", "+00:00")
            dt_utc = datetime.fromisoformat(ts_norm)
            if dt_utc.tzinfo is None:
                dt_utc = dt_utc.replace(tzinfo=timezone.utc)
            h = int(dt_utc.hour)
            if 0 <= h <= 7:
                horas[h] = horas.get(h, 0.0) + float(val)
        except (ValueError, TypeError):
            continue
    return horas, sum(horas.values())


def _llenar_hoja_colegio(ws, node_id: str, nombre: str, desde: date, hasta: date) -> float:
    ws.append(
        [
            "Fecha",
            "Día",
            "H0 UTC",
            "H1 UTC",
            "H2 UTC",
            "H3 UTC",
            "H4 UTC",
            "H5 UTC",
            "H6 UTC",
            "H7 UTC",
            "Suma 00:00–07:00 UTC",
            "totalM3 día",
        ]
    )
    _style_header(ws)
    total = 0.0
    cur = desde
    while cur <= hasta:
        horas, suma = _horas_utc_0_7_por_dia(node_id, cur)
        tj = _total_m3_from_json_for_chile_day(node_id, cur)
        dias = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"][cur.weekday()]
        ws.append(
            [
                cur.isoformat(),
                dias,
                round(horas.get(0, 0), 2),
                round(horas.get(1, 0), 2),
                round(horas.get(2, 0), 2),
                round(horas.get(3, 0), 2),
                round(horas.get(4, 0), 2),
                round(horas.get(5, 0), 2),
                round(horas.get(6, 0), 2),
                round(horas.get(7, 0), 2),
                round(suma, 2),
                round(float(tj), 2) if tj is not None else None,
            ]
        )
        total += suma
        cur += timedelta(days=1)

    ws.append([])
    r = ws.max_row + 1
    ws.append(["TOTAL periodo", "", "", "", "", "", "", "", "", "", round(total, 2), ""])
    for cell in ws[r]:
        cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 12
    for col in range(3, 12):
        ws.column_dimensions[get_column_letter(col)].width = 10
    return total


def main() -> int:
    ap = argparse.ArgumentParser(description="Excel nocturno CORMUP (hoja por colegio)")
    ap.add_argument("--desde", default="01/05/2026", help="DD/MM/YYYY")
    ap.add_argument("--hasta", default="31/05/2026", help="DD/MM/YYYY")
    args = ap.parse_args()

    desde = parse_date(args.desde).date()
    hasta = parse_date(args.hasta, end_of_day=True).date()
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    out = (
        ROOT
        / "reports"
        / "CORMUP"
        / "ABREGADO"
        / f"CORMUP_nocturno_por_colegio_{desde.strftime('%Y%m%d')}_{hasta.strftime('%Y%m%d')}_{ts}.xlsx"
    )

    wb = Workbook()
    ws_res = wb.active
    ws_res.title = "Resumen"
    ws_res.append(["Colegio", "Nodo", "Consumo nocturno periodo (m³)", "Días con consumo"])
    _style_header(ws_res)

    print(f"[INFO] CORMUP nocturno | {desde} a {hasta} | {len(COLEGIOS)} colegios")

    for node_id in COLEGIOS:
        nombre = get_node_name(node_id)
        hoja = _slug_hoja(nombre, node_id)
        ws = wb.create_sheet(hoja)
        total_hoja = _llenar_hoja_colegio(ws, node_id, nombre, desde, hasta)
        metrics = calculate_nocturnal_metrics(
            node_id,
            parse_date(args.desde),
            parse_date(args.hasta, end_of_day=True),
            company_id="000008",
        )
        dias_con = int(metrics.get("dias_con_consumo_nocturno", 0) or 0)
        ws_res.append([nombre, node_id, round(total_hoja, 2), dias_con])
        print(f"  [OK] {nombre}: {total_hoja:.1f} m³ ({dias_con} días con consumo)")

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"[OK] Excel: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

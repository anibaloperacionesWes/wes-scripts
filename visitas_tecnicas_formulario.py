"""
Visitas técnicas del formulario WES para informes de gestión hídrica.

Fuentes (Google Drive):
  1) Planilla INGRESO_visitas_tecnicos_WES  — formulario de terreno.
  2) Registro de fallas WES / Datos         — consolidado operativo
     (las instrucciones del formulario indican que operaciones copia ahí).

Se omiten filas de ejemplo y se deduplican. El cruce con cada informe usa
cliente + máquina (AGUNSA Lampa vs Intermodal; Inchcape Quilicura vs Lo Boza).
"""

from __future__ import annotations

import io
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

# Planilla del formulario (pestaña INGRESO)
SHEET_FORMULARIO_ID = "1B5gGXua055WO5V9Ff4Tm-ur4msN4fA5WG1XEiKs-RSE"
# Consolidado histórico
SHEET_FALLAS_ID = "1GlRn7QXWEre7ziau29ojR5lTl-bZ8T3mCT3cD93HZgM"

CACHE_DIR = Path("/tmp/visitas_tecnicas_wes")


@dataclass
class VisitaTecnica:
    fecha_iso: str
    fecha: str
    tecnico: str
    punto: str
    motivo: str
    diagnostico: str
    folio: str = ""
    cliente_form: str = ""
    fuente: str = ""


@dataclass
class _Match:
    clientes: Sequence[str]
    maquinas: Sequence[str] = ()
    excluir: Sequence[str] = ()


# Claves = cfg["key"] del lote de gestión hídrica.
MATCH_POR_KEY: Dict[str, _Match] = {
    "zapallar": _Match(clientes=["FUNDO ZAPALLAR", "ZAPALLAR"]),
    "inchcape": _Match(
        clientes=["DERCO", "INCHCAPE"],
        excluir=["LO BOZA", "LOBOZA", "OPEN PLAZA"],
    ),
    "nido": _Match(clientes=["NIDO"]),
    "valledor": _Match(clientes=["LO VALLEDOR"]),
    "udd": _Match(clientes=["UDD"]),
    "club": _Match(clientes=["CLUB PROVIDENCIA"]),
    "lampa": _Match(
        clientes=["AGUNSA"],
        maquinas=["DEPOSITO", "MODULO", "LAMPA"],
    ),
    "intermodal": _Match(
        clientes=["AGUNSA"],
        maquinas=["INTERMODAL"],
    ),
    "renca": _Match(
        clientes=["RENCA"],
        excluir=["GIMNASIO", "PISCINA"],
    ),
    "florida": _Match(clientes=["LA FLORIDA"]),
    "reina": _Match(clientes=["LA REINA"]),
    "cormup": _Match(clientes=["CORMUP"]),
    "copec": _Match(clientes=["COPEC"]),
}


def _norm(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = text.encode("ascii", "ignore").decode("ascii")
    return " ".join(text.upper().split())


def _one_line(value: Any, max_len: int = 280) -> str:
    text = " ".join(str(value or "").replace("\r", " ").split())
    if len(text) > max_len:
        return text[: max_len - 1] + "…"
    return text


def _parse_fecha(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%y"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            continue
    return None


def _fmt_fecha(d: date) -> str:
    return d.strftime("%d/%m/%Y")


def _es_ejemplo(blob: str) -> bool:
    n = blob.lower()
    return (
        "ejemplo: reemplazar" in n
        or "reemplazar por la visita real" in n
        or "fila de ejemplo" in n
    )


def _folio(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _titulo_punto(raw: str) -> str:
    n = _norm(raw)
    known = {
        "DEPOSITO": "Depósito",
        "MODULO D": "Módulo D",
        "MODULO ABC": "Módulo ABC",
        "MODULO E": "Módulo E",
        "INTERMODAL": "Intermodal",
        "MATRIZ ESVAL": "Matriz ESVAL",
        "MATRIZ PRINCIPAL": "Matriz Principal",
        "QUILICURA - MATRIZ PRINCIPAL": "Matriz Principal",
        "QUILICURA - CAMARINES": "Camarines",
        "QUILICURA - CASINO": "Casino",
        "QUILICURA - DERCOMAQ": "Dercomaq",
        "QUILICURA - EDIFICIO JCB": "Edificio JCB",
        "QUILICURA - LAVADO DE MAQUINA": "Lav. Máquinas",
        "QUILICURA - PRODERCO": "Proderco",
    }
    if n in known:
        return known[n]
    if n.startswith("QUILICURA - "):
        return raw.split("-", 1)[-1].strip().title()
    return (raw or "—").strip() or "—"


def _motivo(tipo_mtto: Any, tipo_falla: Any, falla: Any, motivo: Any) -> str:
    parts = []
    for item in (tipo_mtto, tipo_falla, falla, motivo):
        t = _one_line(item, 80)
        if t and t.lower() not in {p.lower() for p in parts}:
            parts.append(t)
    return " · ".join(parts[:3]) if parts else "—"


def _visita_desde_valores(
    *,
    folio: Any,
    fecha: Any,
    tecnico: Any,
    cliente: Any,
    maquina: Any,
    motivo: Any,
    tipo_mtto: Any,
    tipo_falla: Any,
    falla: Any,
    solucion: Any,
    observaciones: Any,
    fuente: str,
) -> Optional[Tuple[str, VisitaTecnica]]:
    blob = " ".join(
        str(x or "")
        for x in (
            folio,
            fecha,
            tecnico,
            cliente,
            maquina,
            motivo,
            tipo_mtto,
            tipo_falla,
            falla,
            solucion,
            observaciones,
        )
    )
    if _es_ejemplo(blob):
        return None
    d = _parse_fecha(fecha)
    if d is None:
        return None
    cli = _one_line(cliente, 80)
    if not cli:
        return None
    diag = _one_line(solucion, 280) or _one_line(observaciones, 280) or _one_line(falla, 280)
    visita = VisitaTecnica(
        fecha_iso=d.isoformat(),
        fecha=_fmt_fecha(d),
        tecnico=_one_line(tecnico, 60) or "—",
        punto=_titulo_punto(_one_line(maquina, 80)),
        motivo=_motivo(tipo_mtto, tipo_falla, falla, motivo),
        diagnostico=diag or "—",
        folio=_folio(folio),
        cliente_form=cli,
        fuente=fuente,
    )
    key = "|".join(
        [
            visita.fecha_iso,
            _norm(cli),
            _norm(maquina),
            _norm(tecnico),
            _norm(solucion)[:80],
        ]
    )
    return key, visita


def _header_map(row: Sequence[Any]) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for i, cell in enumerate(row):
        key = _norm(cell)
        if key:
            out[key] = i
    return out


def _get(row: Sequence[Any], headers: Dict[str, int], *names: str) -> Any:
    for name in names:
        idx = headers.get(_norm(name))
        if idx is not None and idx < len(row):
            return row[idx]
    return None


def _leer_ingreso(path: Path) -> List[Tuple[str, VisitaTecnica]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    if "INGRESO" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["INGRESO"]
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        wb.close()
        return []
    headers = _header_map(header)
    found: List[Tuple[str, VisitaTecnica]] = []
    for row in rows:
        if not row or not any(row):
            continue
        parsed = _visita_desde_valores(
            folio=_get(row, headers, "Folio / OT", "Folio"),
            fecha=_get(row, headers, "Fecha"),
            tecnico=_get(row, headers, "Tecnico", "Técnico"),
            cliente=_get(row, headers, "Cliente"),
            maquina=_get(row, headers, "Maquina / sitio", "Maquina", "Máquina / sitio"),
            motivo=_get(row, headers, "Motivo"),
            tipo_mtto=_get(row, headers, "Tipo de Mantenimiento"),
            tipo_falla=_get(row, headers, "Tipo de Falla"),
            falla=_get(row, headers, "Falla especifica", "Falla específica"),
            solucion=_get(row, headers, "Solucion / diagnostico", "Solución / diagnóstico"),
            observaciones=_get(row, headers, "Observaciones"),
            fuente="formulario",
        )
        if parsed:
            found.append(parsed)
    wb.close()
    return found


def _leer_fallas(path: Path) -> List[Tuple[str, VisitaTecnica]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    sheet = "Datos" if "Datos" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        wb.close()
        return []
    headers = _header_map(header)
    found: List[Tuple[str, VisitaTecnica]] = []
    for row in rows:
        if not row or not any(row):
            continue
        parsed = _visita_desde_valores(
            folio=_get(row, headers, "Folio"),
            fecha=_get(row, headers, "Fecha"),
            tecnico=_get(row, headers, "Tecnico", "Técnico"),
            cliente=_get(row, headers, "Cliente"),
            maquina=_get(row, headers, "Maquina", "Máquina"),
            motivo=None,
            tipo_mtto=_get(row, headers, "Tipo de Mantenimiento"),
            tipo_falla=_get(row, headers, "Tipo de Falla"),
            falla=_get(
                row,
                headers,
                "Falla Expecifica",
                "Falla especifica",
                "Falla específica",
            ),
            solucion=_get(
                row,
                headers,
                "Solucion y/o Diagnostico",
                "Solución y/o Diagnóstico",
            ),
            observaciones=_get(row, headers, "Observaciones"),
            fuente="registro",
        )
        if parsed:
            found.append(parsed)
    wb.close()
    return found


def _descargar_sheet(service, file_id: str, dest: Path) -> Path:
    dest.parent.mkdir(parents=True, exist_ok=True)
    meta = (
        service.files()
        .get(fileId=file_id, fields="id,name,mimeType", supportsAllDrives=True)
        .execute()
    )
    mime = meta.get("mimeType") or ""
    if mime == "application/vnd.google-apps.spreadsheet":
        request = service.files().export_media(
            fileId=file_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        request = service.files().get_media(fileId=file_id, supportsAllDrives=True)
    from googleapiclient.http import MediaIoBaseDownload

    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    dest.write_bytes(fh.getvalue())
    return dest


def _buscar_por_nombre(service, nombre: str) -> Optional[str]:
    safe = nombre.replace("'", "\\'")
    items = (
        service.files()
        .list(
            q=f"name='{safe}' and trashed=false",
            fields="files(id, name)",
            pageSize=5,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        )
        .execute()
        .get("files", [])
    )
    return items[0]["id"] if items else None


def _exportar_fuentes() -> Tuple[Optional[Path], Optional[Path]]:
    """Descarga las dos planillas. Si Drive no está, usa copias locales en /tmp."""
    form_path = CACHE_DIR / "visitas_tecnicos.xlsx"
    fallas_path = CACHE_DIR / "fallas_wes.xlsx"
    try:
        from wes_google_drive import credenciales_configuradas, obtener_servicio_drive

        if credenciales_configuradas():
            service = obtener_servicio_drive()
            form_id = SHEET_FORMULARIO_ID
            fallas_id = SHEET_FALLAS_ID
            try:
                _descargar_sheet(service, form_id, form_path)
            except Exception as exc:
                print(f"[ADVERTENCIA] Formulario INGRESO: {exc}", flush=True)
                alt = _buscar_por_nombre(service, "INGRESO_visitas_tecnicos_WES")
                if alt:
                    _descargar_sheet(service, alt, form_path)
            try:
                _descargar_sheet(service, fallas_id, fallas_path)
            except Exception as exc:
                print(f"[ADVERTENCIA] Registro de fallas: {exc}", flush=True)
                alt = _buscar_por_nombre(service, "Registro de fallas WES")
                if alt:
                    _descargar_sheet(service, alt, fallas_path)
    except Exception as exc:
        print(f"[ADVERTENCIA] No se pudo leer Drive para visitas: {exc}", flush=True)

    if not form_path.is_file() and Path("/tmp/visitas_tecnicos.xlsx").is_file():
        form_path = Path("/tmp/visitas_tecnicos.xlsx")
    if not fallas_path.is_file() and Path("/tmp/fallas_wes.xlsx").is_file():
        fallas_path = Path("/tmp/fallas_wes.xlsx")
    return (
        form_path if form_path.is_file() else None,
        fallas_path if fallas_path.is_file() else None,
    )


def cargar_visitas_periodo(start: datetime, end: datetime) -> List[VisitaTecnica]:
    """Todas las visitas reales entre start y end (inclusive), sin filtrar por cliente."""
    start_d = start.date() if isinstance(start, datetime) else start
    end_d = end.date() if isinstance(end, datetime) else end
    form_path, fallas_path = _exportar_fuentes()
    by_key: Dict[str, VisitaTecnica] = {}
    # Registro primero; el formulario pisa si hay la misma visita (más campos).
    if fallas_path:
        for key, visita in _leer_fallas(fallas_path):
            by_key[key] = visita
    if form_path:
        for key, visita in _leer_ingreso(form_path):
            by_key[key] = visita
    out = []
    for visita in by_key.values():
        d = date.fromisoformat(visita.fecha_iso)
        if start_d <= d <= end_d:
            out.append(visita)
    out.sort(key=lambda v: (v.fecha_iso, v.punto, v.tecnico))
    return out


def visitas_de_cliente(
    todas: Iterable[VisitaTecnica],
    cfg: Dict[str, Any],
) -> List[VisitaTecnica]:
    match = MATCH_POR_KEY.get(cfg.get("key") or "")
    if match is None:
        match = _Match(clientes=[cfg.get("cliente") or ""])
    cli_ok = {_norm(c) for c in match.clientes if c}
    maq_ok = [_norm(m) for m in match.maquinas if m]
    excluir = [_norm(m) for m in match.excluir if m]
    found: List[VisitaTecnica] = []
    for visita in todas:
        cli = _norm(visita.cliente_form)
        if cli not in cli_ok and not any(token and token in cli for token in cli_ok):
            continue
        maq = _norm(visita.punto)
        # El título amigable (Depósito) se normaliza a DEPOSITO.
        if excluir and any(token in maq or token in _norm(visita.cliente_form) for token in excluir):
            # también contra el nombre original vía cliente_form + punto
            blob = _norm(visita.punto + " " + visita.cliente_form)
            if any(token in blob for token in excluir):
                continue
        if maq_ok and not any(token in maq for token in maq_ok):
            continue
        found.append(visita)
    return found

# -*- coding: utf-8 -*-
"""
Listado de usuarios WES con acceso a nodos de la compañía 000025 (Parque Arauco).

La API acl-entities no expone GET /users por compañía; solo GET /users?email=...
Por eso se consulta un universo ampliado de correos (archivos locales + contactos WES).

Salida:
  - CSV por usuario (email, nombre, nodos 000025)
  - CSV por nodo (id, nombre, usuarios con acceso)
  - XLSX con ambas hojas

Uso:
  python listar_usuarios_compania_000025.py
  python listar_usuarios_compania_000025.py --emails extra.txt
"""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Set, Tuple

import requests

from generar_reporte_word import get_mall_name_for_parque_arauco

if __import__("sys").platform == "win32":
    try:
        import sys

        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

ENTITY_BASE = "http://104.248.53.141:7001/wes/api/acl-entities/v1"
COMPANY_ID = "000025"
PREFIX_PA = "000025-"

ROOT = Path(__file__).parent
DEFAULT_EMAIL_FILES = [
    ROOT
    / "reports"
    / "Parque_Arauco"
    / "Reportes_agregados_reunion_abril_2026"
    / "emails_linkes_parauco_encontrados.txt",
    ROOT
    / "reports"
    / "Parque_Arauco"
    / "Reportes_agregados_reunion_abril_2026"
    / "emails_pa_linkes_parauco.txt",
]
OUT_DIR = ROOT / "reports" / "Parque_Arauco" / "usuarios_compania_000025"

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.,_-]+\.[a-zA-Z]{2,}")


def _fetch_user(email: str) -> Tuple[dict | None, str]:
    email = email.strip()
    if not email:
        return None, "vacío"

    def _get(e: str) -> requests.Response:
        return requests.get(f"{ENTITY_BASE}/users", params={"email": e}, timeout=30)

    try:
        r = _get(email)
        if r.status_code == 200:
            u = r.json()
            if isinstance(u, dict):
                u.pop("password", None)
            return u, ""
        if r.status_code == 404 and "," in email:
            alt = email.replace(",", ".")
            r2 = _get(alt)
            if r2.status_code == 200:
                u = r2.json()
                if isinstance(u, dict):
                    u.pop("password", None)
                return u, f"email corregido: {alt}"
        if r.status_code == 404:
            return None, "no existe en API"
        return None, f"HTTP {r.status_code}"
    except requests.RequestException as exc:
        return None, str(exc)


def _nodos_compania() -> Dict[str, str]:
    r = requests.get(f"{ENTITY_BASE}/companies/{COMPANY_ID}", timeout=60)
    r.raise_for_status()
    out: Dict[str, str] = {}
    for n in r.json().get("nodes") or []:
        nid = str(n.get("nodeId", "")).strip()
        if nid.startswith(PREFIX_PA):
            out[nid] = str(n.get("name", "") or "").strip()
    return dict(sorted(out.items()))


def _emails_desde_archivo(path: Path) -> Set[str]:
    if not path.is_file():
        return set()
    found: Set[str] = set()
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "@" in line:
            found.add(line.split()[0].strip().lower())
            continue
        for m in EMAIL_RE.finditer(line):
            found.add(m.group(0).lower())
    return found


def _emails_contactos_wes() -> Set[str]:
    emails: Set[str] = set()
    try:
        from lista_contactos_reportes import CONTACTOS_REPORTES, CORREOS_AUTORIZADOS

        for c in CONTACTOS_REPORTES.values():
            e = (c.get("email") or "").strip().lower()
            if e:
                emails.add(e)
        for row in CORREOS_AUTORIZADOS:
            e = (row.get("email") or "").strip().lower()
            if e:
                emails.add(e)
    except ImportError:
        pass
    return emails


def _recolectar_emails(extra: Iterable[Path]) -> List[str]:
    pool: Set[str] = set()
    pool.update(_emails_contactos_wes())
    for fp in DEFAULT_EMAIL_FILES:
        pool.update(_emails_desde_archivo(fp))
    for fp in extra:
        pool.update(_emails_desde_archivo(fp))
    return sorted(pool)


def _nodos_pa_usuario(user: dict, nodos_nombre: Dict[str, str]) -> List[Tuple[str, str, str]]:
    allowed = user.get("allowedNodes") or []
    if not isinstance(allowed, list):
        return []
    filas: List[Tuple[str, str, str]] = []
    for nid in allowed:
        if not isinstance(nid, str) or not nid.startswith(PREFIX_PA):
            continue
        nombre = nodos_nombre.get(nid, "")
        mall = get_mall_name_for_parque_arauco(nid, nombre).strip()
        filas.append((nid, nombre, mall))
    return sorted(filas, key=lambda x: x[0])


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Listado usuarios WES con nodos 000025 (consulta por email)"
    )
    parser.add_argument(
        "--emails",
        type=Path,
        action="append",
        default=[],
        help="Archivo adicional con correos (uno por línea)",
    )
    parser.add_argument(
        "--salida",
        type=Path,
        default=OUT_DIR,
        help="Carpeta de salida",
    )
    args = parser.parse_args()

    nodos_nombre = _nodos_compania()
    emails = _recolectar_emails(args.emails)
    print(f"[INFO] Compañía {COMPANY_ID} — {len(nodos_nombre)} nodos")
    print(f"[INFO] Consultando {len(emails)} correo(s) candidatos ...")

    filas_usuario: List[dict] = []
    por_nodo: Dict[str, Set[str]] = defaultdict(set)
    sin_usuario: List[str] = []
    sin_nodos_pa: List[str] = []

    for email in emails:
        user, obs = _fetch_user(email)
        if not user:
            sin_usuario.append(f"{email} ({obs})" if obs else email)
            continue
        nodos = _nodos_pa_usuario(user, nodos_nombre)
        email_api = str(user.get("username") or email).strip()
        nombre = f"{user.get('name', '')} {user.get('lastName', '')}".strip()
        user_id = str(user.get("userId") or "").strip()
        if not nodos:
            sin_nodos_pa.append(email_api)
            continue
        detalle = "; ".join(f"{nid} ({nom})" for nid, nom, _m in nodos)
        malls = sorted({m for _nid, _nom, m in nodos if m}, key=str.casefold)
        filas_usuario.append(
            {
                "user_id": user_id,
                "email": email_api,
                "nombre": nombre,
                "nodos_pa_count": len(nodos),
                "malls": ", ".join(malls),
                "nodos_detalle": detalle,
                "observacion": obs,
            }
        )
        for nid, _nom, _mall in nodos:
            por_nodo[nid].add(email_api)
        print(f"[OK] {email_api} — {len(nodos)} nodo(s) PA")

    args.salida.mkdir(parents=True, exist_ok=True)
    csv_usuarios = args.salida / "usuarios_compania_000025_por_usuario.csv"
    csv_nodos = args.salida / "usuarios_compania_000025_por_nodo.csv"
    xlsx = args.salida / "usuarios_compania_000025.xlsx"

    cols_u = [
        "user_id",
        "email",
        "nombre",
        "nodos_pa_count",
        "malls",
        "nodos_detalle",
        "observacion",
    ]
    with open(csv_usuarios, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=cols_u, delimiter=";")
        w.writeheader()
        for row in sorted(filas_usuario, key=lambda r: r["email"].casefold()):
            w.writerow(row)

    with open(csv_nodos, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["ID_nodo", "Nombre_punto", "Mall", "Usuarios_con_acceso"])
        for nid in sorted(nodos_nombre.keys()):
            nombre = nodos_nombre[nid]
            mall = get_mall_name_for_parque_arauco(nid, nombre).strip()
            users = sorted(por_nodo.get(nid, set()), key=str.casefold)
            w.writerow([nid, nombre, mall, ", ".join(users) if users else ""])

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Por_usuario"
        for c, h in enumerate(cols_u, 1):
            ws1.cell(row=1, column=c, value=h)
        for r, row in enumerate(
            sorted(filas_usuario, key=lambda x: x["email"].casefold()), 2
        ):
            for c, h in enumerate(cols_u, 1):
                ws1.cell(row=r, column=c, value=row.get(h, ""))
        widths = (26, 32, 28, 14, 24, 80, 24)
        for i, w in enumerate(widths, 1):
            ws1.column_dimensions[get_column_letter(i)].width = w

        ws2 = wb.create_sheet("Por_nodo")
        headers_n = ["ID_nodo", "Nombre_punto", "Mall", "Usuarios_con_acceso"]
        for c, h in enumerate(headers_n, 1):
            ws2.cell(row=1, column=c, value=h)
        row_idx = 2
        for nid in sorted(nodos_nombre.keys()):
            nombre = nodos_nombre[nid]
            mall = get_mall_name_for_parque_arauco(nid, nombre).strip()
            users = sorted(por_nodo.get(nid, set()), key=str.casefold)
            ws2.cell(row=row_idx, column=1, value=nid)
            ws2.cell(row=row_idx, column=2, value=nombre)
            ws2.cell(row=row_idx, column=3, value=mall)
            ws2.cell(row=row_idx, column=4, value=", ".join(users) if users else "")
            row_idx += 1
        for i, w in enumerate((14, 36, 18, 60), 1):
            ws2.column_dimensions[get_column_letter(i)].width = w
        wb.save(xlsx)
    except ImportError:
        xlsx = None

    resumen = args.salida / "RESUMEN.txt"
    resumen.write_text(
        "\n".join(
            [
                f"Compañía: {COMPANY_ID} (Parque Arauco)",
                f"Nodos en compañía: {len(nodos_nombre)}",
                f"Correos consultados: {len(emails)}",
                f"Usuarios con nodos 000025: {len(filas_usuario)}",
                f"Correos sin usuario en API: {len(sin_usuario)}",
                f"Usuarios sin nodos 000025: {len(sin_nodos_pa)}",
                "",
                "LIMITACIÓN: la API no lista todos los usuarios de la compañía.",
                "Este reporte cubre el universo de correos conocidos en el proyecto.",
                "Para listado exhaustivo solicitar export al backoffice WES.",
            ]
        ),
        encoding="utf-8",
    )

    print(f"\n[OK] Usuarios con nodos 000025: {len(filas_usuario)}")
    print(f"[OK] CSV usuarios: {csv_usuarios.resolve()}")
    print(f"[OK] CSV nodos:    {csv_nodos.resolve()}")
    if xlsx:
        print(f"[OK] XLSX:         {xlsx.resolve()}")
    print(f"[INFO] Sin usuario API: {len(sin_usuario)} | Sin nodos PA: {len(sin_nodos_pa)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

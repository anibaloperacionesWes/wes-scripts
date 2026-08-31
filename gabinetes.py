# -*- coding: utf-8 -*-
"""
Gabinetes con más de un nodo (misma placa / mismo gabinete).

Fuente: Drive «Listado_Equipos_Vigentes_Puntos_En_Cero.xlsx»
https://docs.google.com/spreadsheets/d/1g6rT-qF48UxIZrOxvvrnv7B-7WMtMLpM/
"""

from __future__ import annotations

import io
import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Set, Tuple

DRIVE_GABINETES_ID = "1g6rT-qF48UxIZrOxvvrnv7B-7WMtMLpM"
CACHE_XLSX = (
    Path(__file__).resolve().parent
    / "reports"
    / "Listado_Equipos_Vigentes_Puntos_En_Cero.xlsx"
)
_NODE_RE = re.compile(r"\d{6}-\d{2}")

# Puntos sin SIM 4G: internet lo pone el cliente (no WES).
# AGUNSA: la empresa provee internet en los 5 puntos.
# Nido: Estanque B, High School, Elementary, Teatro, Pozo Profundo.
# Lo Valledor P1: el que tiene el problema.
INTERNET_CLIENTE: Set[str] = {
    "000020-01",
    "000020-02",
    "000020-03",
    "000020-04",
    "000020-05",
    "000007-01",  # Estanque B
    "000007-02",  # Teatro
    "000007-03",  # High School
    "000007-04",  # Elementary
    "000007-06",  # Pozo Profundo
    "000002-01",  # Lo Valledor P1
}

NOTA_SIM = {
    "000020": "Internet AGUNSA (la empresa alimenta los 5 puntos)",
    "000007": (
        "Internet cliente: Estanque B, Teatro, High School, "
        "Elementary, Pozo Profundo"
    ),
    "000002": "P1 / Placa 1 (000002-01): problema",
}


def _descargar_drive(destino: Path) -> Path:
    from googleapiclient.http import MediaIoBaseDownload

    from wes_google_drive import obtener_servicio_drive

    svc = obtener_servicio_drive()
    buf = io.BytesIO()
    request = svc.files().get_media(fileId=DRIVE_GABINETES_ID)
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _status, done = downloader.next_chunk()
    destino.parent.mkdir(parents=True, exist_ok=True)
    destino.write_bytes(buf.getvalue())
    return destino


def _nombre_corto(gabinete: str) -> str:
    t = (gabinete or "").strip()
    if "—" in t:
        t = t.split("—", 1)[1].strip()
    elif " - " in t:
        t = t.split(" - ", 1)[1].strip()
    return t or gabinete


def cargar_gabinetes(
    *,
    excel: Optional[Path] = None,
) -> Tuple[Dict[str, str], Dict[str, List[str]]]:
    """
    Returns:
      gabinete_de: nodeId → nombre del gabinete (solo si el gabinete tiene 2+ nodos)
      miembros: nombre gabinete → nodeIds
    """
    from openpyxl import load_workbook

    path = excel or CACHE_XLSX
    if excel is None:
        try:
            path = _descargar_drive(CACHE_XLSX)
            print(f"[INFO] Gabinetes desde Drive → {path}")
        except Exception as exc:
            if path.is_file():
                print(f"[WARN] No se pudo bajar gabinetes Drive ({exc}); uso caché {path}")
            else:
                raise

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_i = None
    headers: List[str] = []
    for i, row in enumerate(ws.iter_rows(max_col=12, values_only=True), 1):
        vals = [str(v or "").strip().lower() for v in row]
        if "id nodo" in vals or "node_id" in vals:
            header_i = i
            headers = [str(v or "").strip().lower() for v in row]
            break
    if header_i is None:
        raise RuntimeError(f"No hay encabezado de gabinetes en {path}")

    def col(*nombres: str) -> Optional[int]:
        for n in nombres:
            if n in headers:
                return headers.index(n)
        return None

    i_nid = col("id nodo", "node_id", "nodo")
    i_gab = col("gabinete")
    i_nodos = col("nodos que contiene el gabinete", "nodos gabinete")
    if i_nid is None:
        raise RuntimeError("Falta columna ID Nodo en listado de gabinetes")

    crudo: Dict[str, List[str]] = defaultdict(list)
    nid_a_gab: Dict[str, str] = {}
    for row in ws.iter_rows(min_row=header_i + 1, max_col=12, values_only=True):
        if not row or i_nid >= len(row) or not row[i_nid]:
            continue
        nid = str(row[i_nid]).strip()
        if not _NODE_RE.fullmatch(nid):
            m = _NODE_RE.search(nid)
            if not m:
                continue
            nid = m.group(0)
        gab = ""
        if i_gab is not None and i_gab < len(row) and row[i_gab]:
            gab = str(row[i_gab]).strip()
        if gab.lower() in ("", "1 placa", "none", "nan"):
            gab = ""
        extra: List[str] = []
        if i_nodos is not None and i_nodos < len(row) and row[i_nodos]:
            extra = _NODE_RE.findall(str(row[i_nodos]))
        if gab and extra:
            nid_a_gab[nid] = gab
            for x in extra:
                crudo[gab].append(x)
                nid_a_gab.setdefault(x, gab)
        elif gab:
            nid_a_gab[nid] = gab
            crudo[gab].append(nid)

    miembros: Dict[str, List[str]] = {}
    gabinete_de: Dict[str, str] = {}
    for gab, nids in crudo.items():
        uniq = sorted(set(nids))
        if len(uniq) < 2:
            continue
        miembros[gab] = uniq
        for nid in uniq:
            gabinete_de[nid] = gab
    return gabinete_de, miembros


def conteo_gabinetes(
    node_ids: Iterable[str],
    gabinete_de: Dict[str, str],
) -> int:
    vistos = set()
    for nid in node_ids:
        vistos.add(gabinete_de.get(nid, nid))
    return len(vistos)


def texto_gabinetes_numerados(
    node_ids: Sequence[str],
    gabinete_de: Dict[str, str],
    miembros: Dict[str, List[str]],
) -> str:
    """Solo gabinetes unidos (2+ puntos): '1: id, id · 2: id, id'."""
    activos = set(node_ids)
    partes: List[str] = []
    vistos = set()
    n = 0
    for nid in node_ids:
        key = gabinete_de.get(nid)
        if not key or key in vistos:
            continue
        vistos.add(key)
        ids = [x for x in miembros.get(key, []) if x in activos]
        if len(ids) < 2:
            continue
        n += 1
        partes.append(f"{n}: {', '.join(ids)}")
    return " · ".join(partes)


def texto_mismo_gabinete(
    node_ids: Sequence[str],
    gabinete_de: Dict[str, str],
    miembros: Dict[str, List[str]],
) -> str:
    return texto_gabinetes_numerados(node_ids, gabinete_de, miembros)


def nids_con_sim(node_ids: Iterable[str]) -> List[str]:
    """Puntos con SIM 4G WES (excluye internet del cliente)."""
    out: List[str] = []
    for nid in node_ids:
        if nid in INTERNET_CLIENTE or str(nid).startswith("000020-"):
            continue
        out.append(nid)
    return out


def celda_sim_4g(
    node_ids: Sequence[str],
    gabinete_de: Dict[str, str],
    miembros: Dict[str, List[str]],
    cid: str = "",
) -> Tuple[int, str]:
    """
    Misma lógica que gabinetes, solo sobre puntos con SIM 4G.
    Returns (conteo, texto de celda).
    """
    sim = nids_con_sim(node_ids)
    n = conteo_gabinetes(sim, gabinete_de) if sim else 0
    unidos = texto_gabinetes_numerados(sim, gabinete_de, miembros)
    nota = NOTA_SIM.get(cid, "")
    if not nota and any(x.startswith("000020-") for x in node_ids):
        nota = NOTA_SIM["000020"]
    partes: List[str] = [str(n)]
    if unidos:
        partes.append(unidos)
    if nota:
        partes.append(nota)
    return n, "\n".join(partes)

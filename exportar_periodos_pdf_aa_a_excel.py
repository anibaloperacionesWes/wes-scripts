"""
Exporta a Excel el contenido de un PDF Aguas Andinas.

- **Historial de consumo** (escaneado u OCR): hoja con columnas **Fecha Lectura**, **Lectura**, **M3 Consumos**,
  **Facturacion del servicio**, **Total de cuenta** (como el cuadro de facturaciones del PDF).
- **Boleta electrónica** u otro: hoja «Periodos» con lecturas y m³ por período.

Ejemplo:
  python exportar_periodos_pdf_aa_a_excel.py "reports/Renca/.../facturaciones piscina.pdf"
  python exportar_periodos_pdf_aa_a_excel.py ruta.pdf -o salida.xlsx
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from facturacion_aguas_andinas_pdf import listar_filas_historial_consumo_facturacion_desde_pdf, listar_periodos_desde_pdf


def _escribir_historial_tabla(
    wb: Workbook,
    *,
    cuenta: str | None,
    medidor: str | None,
    filas,
) -> None:
    ws = wb.active
    ws.title = "Historial"
    headers = [
        "Fecha Lectura",
        "Lectura",
        "M3 Consumos",
        "Facturacion del servicio",
        "Total de cuenta",
    ]
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F4788")
    fnt = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = fill
        c.font = fnt
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for fl in filas:
        ws.append(
            [
                fl.fecha_lectura.strftime("%d-%m-%Y"),
                fl.lectura,
                fl.m3_consumos,
                fl.facturacion_servicio,
                fl.total_cuenta,
            ]
        )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
        row[1].number_format = "#,##0"
        row[2].number_format = "#,##0"
        row[3].number_format = "#,##0"
        row[4].number_format = "#,##0"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    widths = [14, 14, 14, 26, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    sn = wb.create_sheet("Notas")
    sn["A1"] = "Datos cliente (desde PDF)"
    sn["A2"] = f"Cuenta: {cuenta or '(no detectada)'}"
    sn["A3"] = f"Medidor: {medidor or '(no detectado)'}"
    sn["A4"] = (
        "Columnas según historial de facturación Aguas Andinas. Valores numéricos desde OCR; "
        "revise contra el PDF. «Total de cuenta» es el saldo al cierre de cada fila según el orden de números en el documento."
    )
    sn.column_dimensions["A"].width = 95


def _escribir_periodos_simple(wb: Workbook, periodos) -> None:
    ws = wb.active
    ws.title = "Periodos"
    headers = [
        "Archivo PDF",
        "N° factura / ref.",
        "Cuenta",
        "Medidor",
        "Emisión",
        "Lectura anterior",
        "Lectura actual",
        "Días período",
        "m³ cuenta (facturado)",
    ]
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F4788")
    fnt = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.fill = fill
        c.font = fnt
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for p in sorted(periodos, key=lambda x: x.lectura_actual):
        d0, d1 = p.lectura_anterior.date(), p.lectura_actual.date()
        dias = (d1 - d0).days
        ws.append(
            [
                p.pdf.name,
                str(p.boleta),
                p.cuenta or "",
                p.medidor or "",
                p.emision.strftime("%d-%m-%Y"),
                p.lectura_anterior.strftime("%d-%m-%Y"),
                p.lectura_actual.strftime("%d-%m-%Y"),
                dias,
                int(p.m3_cuenta),
            ]
        )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
        row[7].number_format = "0"
        row[8].number_format = "#,##0"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    widths = [28, 18, 16, 14, 12, 14, 14, 12, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    sn = wb.create_sheet("Notas")
    sn["A1"] = "Origen"
    sn["A2"] = (
        "Filas generadas automáticamente desde el PDF indicado (boleta electrónica o historial / OCR). "
        "Revise fechas y m³ antes de usar en informes oficiales."
    )
    sn.column_dimensions["A"].width = 95


def main() -> None:
    ap = argparse.ArgumentParser(description="PDF Aguas Andinas → Excel (tabla historial o períodos).")
    ap.add_argument("pdf", type=Path, help="Ruta al .pdf")
    ap.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Salida .xlsx (por defecto: *_tabla_historial_pdf.xlsx o *_periodos_desde_pdf.xlsx)",
    )
    args = ap.parse_args()
    pdf: Path = args.pdf.resolve()
    if not pdf.is_file():
        raise SystemExit(f"[ERROR] No existe el archivo: {pdf}")

    wb = Workbook()
    tabla = listar_filas_historial_consumo_facturacion_desde_pdf(pdf)
    if tabla is not None:
        cuenta, medidor, filas = tabla
        if not filas:
            raise SystemExit("[ERROR] Historial detectado pero sin filas de tabla extraíbles.")
        out = (
            args.output.resolve()
            if args.output
            else pdf.with_name(f"{pdf.stem}_tabla_historial_pdf.xlsx")
        )
        _escribir_historial_tabla(wb, cuenta=cuenta, medidor=medidor, filas=filas)
        n = len(filas)
        msg = f"{n} fila(s) tabla historial"
    else:
        periodos = listar_periodos_desde_pdf(pdf)
        if not periodos:
            raise SystemExit("[ERROR] No se extrajo ningún período del PDF.")
        out = (
            args.output.resolve()
            if args.output
            else pdf.with_name(f"{pdf.stem}_periodos_desde_pdf.xlsx")
        )
        _escribir_periodos_simple(wb, periodos)
        n = len(periodos)
        msg = f"{n} periodo(s)"

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"[OK] {msg} -> {out}")


if __name__ == "__main__":
    main()

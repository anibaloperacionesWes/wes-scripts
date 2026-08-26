"""
Exporta a Excel el listado de equipos vigentes del reporte matinal de puntos en cero.

Mismo universo que el reporte matinal (`obtener_todos_los_nodos`: API + exclusiones).
Si no se puede importar `reporte_puntos_en_cero` (p. ej. sin matplotlib), se replica
la consulta con las mismas exclusiones de `exclusiones_reportes.py`.
Columnas pedidas: Cliente, Nombre del nodo, Monitoreo y Control (estas dos vacías para completar).

Uso:
  python exportar_listado_equipos_puntos_en_cero.py
  python exportar_listado_equipos_puntos_en_cero.py -o reports/Puntos_En_Cero/listado.xlsx
"""

from __future__ import annotations

import argparse
from collections import Counter
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Dict, List, Optional

import requests

from exclusiones_reportes import (
    EXCLUDED_COMPANY_IDS_PUNTOS_EN_CERO,
    EXCLUDED_COMPANY_NAME_KEYWORDS,
    EXCLUDED_NODE_IDS_PUNTOS_EN_CERO,
    nombre_cliente_puntos_en_cero,
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from registro_gabinetes import filas_resumen_gabinetes, info_gabinete
from wes_paths import wes_scripts_root

ENTITY_BASE_URL = "http://104.248.53.141:7001/wes/api/acl-entities/v1"

try:
    from zoneinfo import ZoneInfo

    _CHILE_TZ = ZoneInfo("America/Santiago")
except Exception:
    _CHILE_TZ = timezone(timedelta(hours=-4))

WES_AZUL = "1F4788"
WES_AZUL_TABLA = "4472C4"
AMARILLO_EDITAR = "FFF2CC"
GRIS_FILA = "F2F2F2"
BLANCO = "FFFFFF"
VERDE_RESUMEN = "548235"
NARANJA_CONFIRMAR = "FCE4D6"

HEADERS = [
    "N°",
    "Cliente",
    "ID Nodo",
    "Nombre del nodo",
    "Gabinete",
    "Tipo gabinete",
    "Placas",
    "Otros nodos del gabinete",
    "Monitoreo",
    "Control",
    "Observación",
]

OPCIONES_MONITOREO = "Sí,No,Pendiente"
OPCIONES_CONTROL = "Sí,No"


def _now_chile() -> datetime:
    return datetime.now(_CHILE_TZ)


def _obtener_empresas_config() -> List[Dict[str, str]]:
    url = f"{ENTITY_BASE_URL}/configuration/companies"
    try:
        response = requests.get(url, timeout=15)
        if response.status_code == 200:
            data = response.json()
            if isinstance(data, list):
                return data
    except Exception:
        pass
    return []


def obtener_nodos_vigentes_puntos_en_cero() -> List[Dict[str, str]]:
    """
    Mismo universo que el reporte matinal de puntos en cero.

    Prefiere `reporte_puntos_en_cero.obtener_todos_los_nodos` cuando ese módulo
    se puede importar; si faltan dependencias (matplotlib/docx), replica la
    consulta a la API con las mismas exclusiones.
    """
    try:
        from reporte_puntos_en_cero import obtener_todos_los_nodos

        return obtener_todos_los_nodos()
    except ImportError:
        print(
            "[INFO] No se pudo importar reporte_puntos_en_cero; "
            "se consulta la API con las mismas exclusiones."
        )
        return _obtener_nodos_desde_api()


def _obtener_nodos_desde_api() -> List[Dict[str, str]]:
    empresas_excluidas = EXCLUDED_COMPANY_IDS_PUNTOS_EN_CERO
    nombres_empresas_excluidas = EXCLUDED_COMPANY_NAME_KEYWORDS
    nodos_excluidos = EXCLUDED_NODE_IDS_PUNTOS_EN_CERO
    all_nodes: List[Dict[str, str]] = []

    empresas_config = _obtener_empresas_config()
    if not empresas_config:
        print("[ADVERTENCIA] Sin empresas desde API; usando rango 000000-000100.")
        empresas_config = [{"companyId": f"{i:06d}", "name": ""} for i in range(101)]

    print("Obteniendo nodos vigentes (mismo filtro que puntos en cero)...")
    for empresa in empresas_config:
        company_id_raw = str(empresa.get("companyId", "")).strip()
        if not company_id_raw:
            continue
        company_id = company_id_raw.zfill(6)
        if company_id in empresas_excluidas:
            continue

        url = f"{ENTITY_BASE_URL}/companies/{company_id}"
        try:
            response = requests.get(url, timeout=10)
            if response.status_code != 200:
                continue
            data = response.json()
            company_name = data.get("name", "").strip() or str(empresa.get("name", "")).strip()
            if company_name:
                company_name_upper = company_name.upper().strip()
                if any(n in company_name_upper for n in nombres_empresas_excluidas):
                    print(f"[EXCLUIDO] {company_id} ({company_name}) - por nombre")
                    continue
            if not company_name:
                continue
            nodes = data.get("nodes", [])
            incluidos = 0
            for node in nodes:
                node_id = node.get("nodeId", "")
                node_name = (node.get("name") or "").strip()
                if node_id in nodos_excluidos:
                    continue
                if node_id and node_name:
                    all_nodes.append(
                        {
                            "nodeId": node_id,
                            "nodeName": node_name,
                            "companyId": company_id,
                            "companyName": nombre_cliente_puntos_en_cero(node_id, company_name),
                        }
                    )
                    incluidos += 1
            if nodes:
                print(f"[OK] {company_id} ({company_name}): {incluidos} nodos vigentes")
        except requests.RequestException:
            pass
        except Exception as e:
            print(f"[ERROR] {company_id}: {e}")

    print(f"Total nodos vigentes: {len(all_nodes)}")
    return all_nodes


def _fill(color: str) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _thin_border() -> Border:
    side = Side(style="thin", color="B0B0B0")
    return Border(left=side, right=side, top=side, bottom=side)


def _estilo_encabezado(cell) -> None:
    cell.font = Font(bold=True, color=BLANCO, name="Calibri", size=11)
    cell.fill = _fill(WES_AZUL_TABLA)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _thin_border()


def _escribir_titulo(ws: Worksheet, titulo: str, subtitulo: str, ncols: int) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c1 = ws.cell(row=1, column=1, value=titulo)
    c1.font = Font(bold=True, color=BLANCO, name="Calibri", size=14)
    c1.fill = _fill(WES_AZUL)
    c1.alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c2 = ws.cell(row=2, column=1, value=subtitulo)
    c2.font = Font(italic=True, name="Calibri", size=10, color="333333")
    c2.fill = _fill("D6E3F0")
    c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 36
    return 4  # fila de encabezados de tabla


def _aplicar_validacion(ws: Worksheet, col: int, first_data: int, last_data: int, formula: str, titulo: str) -> None:
    letra = get_column_letter(col)
    dv = DataValidation(type="list", formula1=f'"{formula}"', allow_blank=True)
    dv.prompt = f"Seleccione un valor para {titulo} o déjelo en blanco"
    dv.promptTitle = titulo
    dv.error = "Elija una opción de la lista (o borre la celda)."
    dv.errorTitle = titulo
    dv.showErrorMessage = True
    dv.showInputMessage = True
    ws.add_data_validation(dv)
    dv.add(f"{letra}{first_data}:{letra}{last_data}")


def _escribir_equipos(
    ws: Worksheet,
    nodos: List[Dict[str, str]],
    generado: datetime,
    marcas: Optional[Dict[str, Dict[str, str]]] = None,
) -> None:
    marcas = marcas or {}
    subtitulo = (
        f"Equipos vigentes al {generado.strftime('%d/%m/%Y %H:%M')} (hora Chile). "
        "Mismo universo que el reporte matinal de puntos en cero. "
        "Gabinete: 1 nodo/1 placa o varios nodos (hasta 4 placas) en el mismo armario. "
        "Complete Monitoreo y Control (Sí/No)."
    )
    header_row = _escribir_titulo(
        ws,
        "Listado de equipos vigentes — Puntos en cero",
        subtitulo,
        len(HEADERS),
    )

    for col, header in enumerate(HEADERS, start=1):
        _estilo_encabezado(ws.cell(row=header_row, column=col, value=header))

    nodos_ord = sorted(
        nodos,
        key=lambda n: (n.get("companyName") or "", n.get("nodeName") or "", n.get("nodeId") or ""),
    )
    first_data = header_row + 1
    border = _thin_border()
    align_izq = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_cen = Alignment(horizontal="center", vertical="center")
    fill_edit = _fill(AMARILLO_EDITAR)
    fill_alt = _fill(GRIS_FILA)
    fill_conf = _fill(NARANJA_CONFIRMAR)
    font_dato = Font(name="Calibri", size=10)
    cols_edit = {9, 10, 11}  # Monitoreo, Control, Observación
    cols_cen = {1, 3, 6, 7, 9, 10}

    for i, nodo in enumerate(nodos_ord, start=1):
        row = first_data + i - 1
        nid = nodo.get("nodeId") or ""
        gab = info_gabinete(
            nid,
            company_name=nodo.get("companyName") or "",
            node_name=nodo.get("nodeName") or "",
        )
        prev = marcas.get(nid) or {}
        valores = [
            i,
            nodo.get("companyName") or "",
            nid,
            nodo.get("nodeName") or "",
            gab["gabinete"],
            gab["tipo"],
            int(gab["placas"]) if str(gab["placas"]).isdigit() else gab["placas"],
            gab["otros_nodos"],
            prev.get("monitoreo") or "",
            prev.get("control") or "",
            prev.get("observacion") or "",
        ]
        confirmar = (gab.get("confianza") or "") == "media"
        for col, valor in enumerate(valores, start=1):
            cell = ws.cell(row=row, column=col, value=valor)
            cell.font = font_dato
            cell.border = border
            cell.alignment = align_cen if col in cols_cen else align_izq
            if col in cols_edit:
                cell.fill = fill_edit
            elif confirmar and col in (5, 6, 7, 8):
                cell.fill = fill_conf
            elif i % 2 == 0:
                cell.fill = fill_alt

    last_data = first_data + len(nodos_ord) - 1 if nodos_ord else first_data
    if nodos_ord:
        _aplicar_validacion(ws, 9, first_data, last_data, OPCIONES_MONITOREO, "Monitoreo")
        _aplicar_validacion(ws, 10, first_data, last_data, OPCIONES_CONTROL, "Control")

    anchos = (6, 28, 14, 36, 48, 22, 10, 36, 14, 12, 28)
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(HEADERS))}{max(last_data, header_row)}"
    ws.freeze_panes = f"A{first_data}"
    ws.sheet_properties.tabColor = WES_AZUL_TABLA
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.sheet_view.showGridLines = False


def _escribir_resumen(ws: Worksheet, nodos: List[Dict[str, str]], generado: datetime) -> None:
    header_row = _escribir_titulo(
        ws,
        "Resumen por cliente",
        f"Cantidad de equipos vigentes incluidos en el reporte matinal de puntos en cero. "
        f"Generado {generado.strftime('%d/%m/%Y %H:%M')} hora Chile.",
        3,
    )
    headers = ["Cliente", "Equipos vigentes", "% del total"]
    for col, header in enumerate(headers, start=1):
        _estilo_encabezado(ws.cell(row=header_row, column=col, value=header))

    conteo = Counter((n.get("companyName") or "(sin nombre)").strip() for n in nodos)
    total = len(nodos) or 1
    first_data = header_row + 1
    border = _thin_border()
    font_dato = Font(name="Calibri", size=10)

    for i, (cliente, cant) in enumerate(sorted(conteo.items(), key=lambda x: (-x[1], x[0])), start=1):
        row = first_data + i - 1
        ws.cell(row=row, column=1, value=cliente).font = font_dato
        c_cant = ws.cell(row=row, column=2, value=cant)
        c_cant.font = font_dato
        c_cant.alignment = Alignment(horizontal="center")
        c_pct = ws.cell(row=row, column=3, value=cant / total)
        c_pct.number_format = "0.0%"
        c_pct.alignment = Alignment(horizontal="center")
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            if i % 2 == 0:
                cell.fill = _fill(GRIS_FILA)

    total_row = first_data + len(conteo)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color=BLANCO, name="Calibri")
    ws.cell(row=total_row, column=2, value=len(nodos)).font = Font(bold=True, color=BLANCO, name="Calibri")
    ws.cell(row=total_row, column=3, value=1).font = Font(bold=True, color=BLANCO, name="Calibri")
    ws.cell(row=total_row, column=3).number_format = "0.0%"
    for col in range(1, 4):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = _fill(VERDE_RESUMEN)
        cell.border = border
        cell.alignment = Alignment(horizontal="center" if col > 1 else "left")

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14
    ws.freeze_panes = f"A{first_data}"
    ws.auto_filter.ref = f"A{header_row}:C{max(total_row - 1, header_row)}"
    ws.sheet_properties.tabColor = VERDE_RESUMEN


def _escribir_gabinetes(ws: Worksheet, nodos: List[Dict[str, str]], generado: datetime) -> None:
    headers = [
        "Gabinete",
        "Cliente",
        "Nodos",
        "Placas",
        "Tipo",
        "IDs",
        "Nombres",
        "Confianza",
        "Notas",
    ]
    header_row = _escribir_titulo(
        ws,
        "Gabinetes — 1 placa o hasta 4 placas",
        "Un gabinete WES puede llevar 1 nodo (1 placa) o varios (hasta 4 placas). "
        f"Naranja = confirmar en terreno. Generado {generado.strftime('%d/%m/%Y %H:%M')} Chile. "
        "Editar agrupaciones en registro_gabinetes.py.",
        len(headers),
    )
    for col, header in enumerate(headers, start=1):
        _estilo_encabezado(ws.cell(row=header_row, column=col, value=header))

    filas = filas_resumen_gabinetes(nodos)
    # Multi-placa primero, luego 1 placa
    filas.sort(key=lambda f: (-int(f["placas"] or 1), f["cliente"], f["gabinete"]))
    first_data = header_row + 1
    border = _thin_border()
    font_dato = Font(name="Calibri", size=10)
    for i, fila in enumerate(filas, start=1):
        row = first_data + i - 1
        vals = [
            fila["gabinete"],
            fila["cliente"],
            int(fila["nodos"]),
            int(fila["placas"]),
            fila["tipo"],
            fila["ids"],
            fila["nombres"],
            fila["confianza"],
            fila["notas"],
        ]
        confirmar = fila["confianza"] == "media"
        for col, valor in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=col, value=valor)
            cell.font = font_dato
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if col in (3, 4, 8) else "left",
                vertical="center",
                wrap_text=True,
            )
            if confirmar:
                cell.fill = _fill(NARANJA_CONFIRMAR)
            elif int(fila["placas"]) >= 2:
                cell.fill = _fill("E2EFDA")
            elif i % 2 == 0:
                cell.fill = _fill(GRIS_FILA)

    last = first_data + len(filas) - 1 if filas else header_row
    anchos = (46, 22, 10, 10, 22, 42, 55, 14, 55)
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{last}"
    ws.freeze_panes = f"A{first_data}"
    ws.sheet_properties.tabColor = "C65911"


def _escribir_notas(ws: Worksheet, generado: datetime, total: int) -> None:
    ws["A1"] = "Cómo usar este listado"
    ws["A1"].font = Font(bold=True, color=BLANCO, name="Calibri", size=14)
    ws["A1"].fill = _fill(WES_AZUL)
    ws.merge_cells("A1:B1")

    lineas = [
        ("Generado", generado.strftime("%d/%m/%Y %H:%M") + " hora Chile"),
        ("Equipos vigentes", str(total)),
        (
            "Qué incluye",
            "Los mismos puntos que revisa cada mañana el reporte de puntos en cero "
            "(reporte_puntos_en_cero.py → obtener_todos_los_nodos).",
        ),
        (
            "Qué no incluye",
            "Empresas y nodos excluidos (WES, Ejército, Gendarmería, BUPA 01-06 pendiente de instalación, "
            "Corporación Puente Alto, MOP, Lo Boza, TML, MADECO, IDs en exclusiones_reportes.py "
            "y registro_puntos_deshabilitados.txt). Sí incluye BUPA Antofagasta (000029-07..10).",
        ),
        (
            "Cliente / Nombre del nodo",
            "Datos fijos desde la API. No los edite; si un nombre cambió, vuelva a generar el Excel.",
        ),
        (
            "Gabinete / placas",
            "Identifica si el nodo va solo (1 placa) o comparte gabinete con otros "
            "(hasta 4 placas). Naranja = confirmar en terreno. Hoja «Gabinetes» agrupa por armario. "
            "Corregir en registro_gabinetes.py.",
        ),
        (
            "Monitoreo (amarillo)",
            "Columna para completar: Sí / No / Pendiente. Use el desplegable o déjela en blanco.",
        ),
        (
            "Control (amarillo)",
            "Columna para completar: Sí o No. Use el desplegable o déjela en blanco.",
        ),
        (
            "Observación (amarillo)",
            "Texto libre (contacto, visita, ticket, etc.).",
        ),
        (
            "Regenerar",
            "python exportar_listado_equipos_puntos_en_cero.py",
        ),
    ]
    ws["A3"] = "Campo"
    ws["B3"] = "Detalle"
    for col in (1, 2):
        _estilo_encabezado(ws.cell(row=3, column=col))

    border = _thin_border()
    for i, (campo, detalle) in enumerate(lineas, start=4):
        ws.cell(row=i, column=1, value=campo).font = Font(bold=True, name="Calibri", size=10)
        ws.cell(row=i, column=2, value=detalle).font = Font(name="Calibri", size=10)
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        for col in (1, 2):
            ws.cell(row=i, column=col).border = border
        ws.row_dimensions[i].height = 36

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 92
    ws.sheet_properties.tabColor = "7F7F7F"


def _cargar_marcas(path: Path) -> Dict[str, Dict[str, str]]:
    """Lee Monitoreo / Control / Observación de un Excel previo (por ID Nodo)."""
    if not path.is_file():
        return {}
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        header_row = None
        headers: Dict[str, int] = {}
        for r, row in enumerate(ws.iter_rows(max_row=8, values_only=True), start=1):
            vals = [str(c).strip() if c is not None else "" for c in row]
            if "ID Nodo" in vals:
                header_row = r
                headers = {name: i for i, name in enumerate(vals) if name}
                break
        if header_row is None or "ID Nodo" not in headers:
            wb.close()
            return {}
        i_id = headers["ID Nodo"]
        i_mon = headers.get("Monitoreo")
        i_ctl = headers.get("Control")
        i_obs = headers.get("Observación")
        out: Dict[str, Dict[str, str]] = {}
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or i_id >= len(row) or not row[i_id]:
                continue
            nid = str(row[i_id]).strip()
            out[nid] = {
                "monitoreo": "" if i_mon is None or i_mon >= len(row) or row[i_mon] is None else str(row[i_mon]).strip(),
                "control": "" if i_ctl is None or i_ctl >= len(row) or row[i_ctl] is None else str(row[i_ctl]).strip(),
                "observacion": "" if i_obs is None or i_obs >= len(row) or row[i_obs] is None else str(row[i_obs]).strip(),
            }
        wb.close()
        return out
    except Exception:
        return {}


def _descargar_marcas_drive(destino: Path) -> Path | None:
    """Baja el listado de Drive (xlsx nativo o export Sheets) para no perder Control/Monitoreo."""
    try:
        from googleapiclient.http import MediaIoBaseDownload
        from wes_google_drive import obtener_servicio_drive
    except Exception:
        return None
    import io

    file_id = "1g6rT-qF48UxIZrOxvvrnv7B-7WMtMLpM"
    try:
        service = obtener_servicio_drive()
        meta = (
            service.files()
            .get(fileId=file_id, fields="id, mimeType")
            .execute()
        )
        mime = str(meta.get("mimeType") or "")
        if mime == "application/vnd.google-apps.spreadsheet":
            request = service.files().export_media(
                fileId=file_id,
                mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            request = service.files().get_media(fileId=file_id)
        buf = io.BytesIO()
        downloader = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        destino.write_bytes(buf.getvalue())
        return destino
    except Exception as e:
        print(f"[INFO] No se pudo leer marcas desde Drive ({e}).")
        return None


def generar_excel(
    nodos: List[Dict[str, str]],
    salida: Path,
    marcas: Optional[Dict[str, Dict[str, str]]] = None,
) -> Path:
    generado = _now_chile()
    wb = Workbook()
    ws_eq = wb.active
    ws_eq.title = "Equipos vigentes"
    _escribir_equipos(ws_eq, nodos, generado, marcas=marcas)

    ws_gab = wb.create_sheet("Gabinetes")
    _escribir_gabinetes(ws_gab, nodos, generado)

    ws_res = wb.create_sheet("Resumen por cliente")
    _escribir_resumen(ws_res, nodos, generado)

    ws_notas = wb.create_sheet("Cómo usar")
    _escribir_notas(ws_notas, generado, len(nodos))

    salida.parent.mkdir(parents=True, exist_ok=True)
    wb.save(salida)
    return salida


def default_salida() -> Path:
    return wes_scripts_root() / "reports" / "Puntos_En_Cero" / "Listado_Equipos_Vigentes_Puntos_En_Cero.xlsx"


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Excel de equipos vigentes del reporte matinal de puntos en cero"
    )
    parser.add_argument(
        "-o",
        "--salida",
        type=Path,
        default=None,
        help="Ruta del .xlsx (por defecto reports/Puntos_En_Cero/Listado_Equipos_Vigentes_Puntos_En_Cero.xlsx)",
    )
    args = parser.parse_args()
    salida = args.salida or default_salida()

    nodos = obtener_nodos_vigentes_puntos_en_cero()
    if not nodos:
        print("[ERROR] No se obtuvieron nodos vigentes. Revise API y exclusiones.")
        return 1

    tmp_drive = salida.parent / "_marcas_drive_tmp.xlsx"
    drive_path = _descargar_marcas_drive(tmp_drive)
    marcas: Dict[str, Dict[str, str]] = {}
    for src in (drive_path, salida):
        if src and Path(src).is_file():
            got = _cargar_marcas(Path(src))
            llenas = 0
            for nid, vals in got.items():
                cur = marcas.setdefault(
                    nid, {"monitoreo": "", "control": "", "observacion": ""}
                )
                for k, v in vals.items():
                    if v:
                        cur[k] = v
                        llenas += 1
            if llenas:
                print(f"[INFO] Conservando marcas desde {Path(src).name} ({llenas} celdas)")
    if tmp_drive.is_file():
        try:
            tmp_drive.unlink()
        except OSError:
            pass

    path = generar_excel(nodos, salida, marcas=marcas)
    print("=" * 60)
    print(f"[OK] {len(nodos)} equipos vigentes")
    print(f"     Excel: {path.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

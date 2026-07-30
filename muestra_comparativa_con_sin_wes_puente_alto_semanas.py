"""
Muestra comparativa con WES vs sin WES — Corporación Puente Alto (000010).

Compara dos ventanas de 7 días (medición real API, sin estimaciones):
  - Con WES: última semana de mayo (cortes + reducción de caudal diurno activos).
  - Sin WES: semana reciente con estados de corte desactivados (jun).

Para el informe ejecutivo breve use: informe_breve_ahorro_wes_puente_alto.py

Salida: Excel + gráficos PNG en reports/proyeccion ahorre puente 2025/muestra_semanal/

Uso:
  python muestra_comparativa_con_sin_wes_puente_alto_semanas.py
  python muestra_comparativa_con_sin_wes_puente_alto_semanas.py --desde-mayo 26/05/2026 --hasta-mayo 31/05/2026
"""

from __future__ import annotations

import argparse
import csv
import sys
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from generar_consolidado_m3_mensual_puente_alto import _consumo_dia_fallback
from generar_reporte_word import (
    UMBRAL_PCT_DIAS_CONSUMO_NOCTURNO_FILTRACION,
    calculate_nocturnal_metrics,
    parse_date,
)
from reporte_puente_alto_lxm import obtener_nodos_puente_alto

if sys.platform == "win32":
    for _s in (sys.stdout, sys.stderr):
        try:
            _s.reconfigure(encoding="utf-8", errors="replace")
        except (AttributeError, OSError, ValueError):
            pass

ROOT = Path(__file__).resolve().parent
OUT_DIR = ROOT / "reports" / "proyeccion ahorre puente 2025" / "muestra_semanal"
PCT_CSV = ROOT / "reports" / "proyeccion ahorre puente 2025" / "pct_auditoria_informe_pa.csv"
EFICIENCIA_GLOBAL = 49.0
COMPANY_ID = "000010"

DEFAULT_MAYO_INI = date(2026, 5, 25)
DEFAULT_MAYO_FIN = date(2026, 5, 31)
DEFAULT_RECENT_END = date(2026, 6, 15)


def _cargar_pct_auditoria() -> Dict[str, float]:
    out: Dict[str, float] = {}
    if not PCT_CSV.is_file():
        return out
    with PCT_CSV.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            nid = str(row.get("node_id", "")).strip()
            try:
                out[nid] = float(str(row.get("pct_eficiencia_auditoria", "")).replace(",", "."))
            except (TypeError, ValueError):
                continue
    return out


def _rango_dias(ini: date, fin: date) -> List[date]:
    d = ini
    out: List[date] = []
    while d <= fin:
        out.append(d)
        d += timedelta(days=1)
    return out


def _sin_wes_est(con_m3: float, pct: float) -> float:
    """Sin WES = medición / (1 - %%/100), coherente con auditoría Puente Alto."""
    denom = 1.0 - pct / 100.0
    if denom <= 1e-6:
        return con_m3
    return con_m3 / denom


def _consumo_periodo(node_id: str, dias: List[date]) -> Tuple[float, int]:
    total = 0.0
    dias_con = 0
    for d in dias:
        s, _ = _consumo_dia_fallback(node_id, d)
        if s > 1e-6:
            dias_con += 1
        total += s
    return round(total, 2), dias_con


def _metricas_periodo(node_id: str, dias: List[date]) -> Dict[str, object]:
    if not dias:
        return {
            "total_m3": 0.0,
            "dias_con_dato": 0,
            "nocturno_m3": 0.0,
            "diurno_m3": 0.0,
            "dias_con_nocturno": 0,
            "pct_dias_nocturno": 0.0,
            "posible_fuga": False,
        }
    start = datetime.combine(dias[0], datetime.min.time()).replace(tzinfo=timezone.utc)
    end = datetime.combine(dias[-1], datetime.max.time()).replace(tzinfo=timezone.utc)
    nm = calculate_nocturnal_metrics(node_id, start, end, company_id=COMPANY_ID)
    total, dias_con = _consumo_periodo(node_id, dias)
    noct = round(float(nm.get("consumo_nocturno_total", 0) or 0), 2)
    diurno = round(max(total - noct, 0.0), 2)
    dcn = int(nm.get("dias_con_consumo_nocturno", 0) or 0)
    dsd = int(nm.get("dias_sin_consumo_nocturno", 0) or 0)
    total_d = dcn + dsd
    pct_n = (dcn / total_d * 100.0) if total_d > 0 else 0.0
    posible_fuga = pct_n >= UMBRAL_PCT_DIAS_CONSUMO_NOCTURNO_FILTRACION and noct > 0.5
    return {
        "total_m3": total,
        "dias_con_dato": dias_con,
        "nocturno_m3": noct,
        "diurno_m3": diurno,
        "dias_con_nocturno": dcn,
        "pct_dias_nocturno": round(pct_n, 1),
        "posible_fuga": posible_fuga,
    }


def _ultimos_7_dias_con_datos(hoy: date, min_nodos: int = 8) -> Tuple[date, date]:
    nodos = obtener_nodos_puente_alto()
    for offset in range(45):
        fin = hoy - timedelta(days=offset)
        ini = fin - timedelta(days=6)
        ok = True
        for d in _rango_dias(ini, fin):
            cnt = sum(1 for n in nodos if _consumo_dia_fallback(n["nodeId"], d)[0] > 0.01)
            if cnt < min_nodos:
                ok = False
                break
        if ok:
            return ini, fin
    return hoy - timedelta(days=6), hoy


def _fmt(v: float) -> str:
    return f"{v:.1f}".replace(".", ",")


def _grafico_barras_agrupadas(
    labels: List[str],
    serie_a: List[float],
    serie_b: List[float],
    label_a: str,
    label_b: str,
    titulo: str,
    out_png: Path,
) -> None:
    x = np.arange(len(labels))
    w = 0.36
    fig, ax = plt.subplots(figsize=(14, 7), dpi=120)
    b1 = ax.bar(x - w / 2, serie_a, width=w, label=label_a, color="#2563eb")
    b2 = ax.bar(x + w / 2, serie_b, width=w, label=label_b, color="#dc2626")
    ax.set_ylabel("Consumo semanal (m³)")
    ax.set_title(titulo)
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=35, ha="right", fontsize=9)
    ax.legend(loc="upper right")
    ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: _fmt(v)))
    fs = max(6, min(8, int(880 / max(len(labels), 6))))
    ax.bar_label(b1, labels=[_fmt(v) for v in serie_a], fontsize=fs, padding=2, rotation=75)
    ax.bar_label(b2, labels=[_fmt(v) for v in serie_b], fontsize=fs, padding=2, rotation=75)
    ax.grid(axis="y", linestyle="--", alpha=0.35)
    ax.set_axisbelow(True)
    fig.tight_layout()
    out_png.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_png, bbox_inches="tight")
    plt.close(fig)


def _grafico_resumen_totales(
    con_mayo: float,
    sin_mayo: float,
    con_reciente: float,
    noct_mayo: float,
    noct_reciente: float,
    out_png: Path,
    etiqueta_mayo: str,
    etiqueta_reciente: str,
) -> None:
    fig, axes = plt.subplots(1, 2, figsize=(12, 5), dpi=120)
    cats = ["Con WES\n(medido)", "Sin WES\n(estimado)"]
    vals = [con_mayo, sin_mayo]
    colors = ["#16a34a", "#ea580c"]
    bars = axes[0].bar(cats, vals, color=colors)
    axes[0].set_title(f"Agregado municipal — {etiqueta_mayo}")
    axes[0].set_ylabel("m³ semana")
    axes[0].bar_label(bars, labels=[_fmt(v) for v in vals], padding=3)
    ahorro = sin_mayo - con_mayo
    pct = (ahorro / sin_mayo * 100) if sin_mayo > 0 else 0
    axes[0].text(
        0.5,
        0.02,
        f"Ahorro estimado: {_fmt(ahorro)} m³ ({pct:.1f}%)",
        transform=axes[0].transAxes,
        ha="center",
        fontsize=10,
        bbox=dict(boxstyle="round", facecolor="#f0fdf4", alpha=0.9),
    )

    cats2 = [f"Nocturno\n{etiqueta_mayo}", f"Nocturno\n{etiqueta_reciente}", f"Total\n{etiqueta_reciente}"]
    vals2 = [noct_mayo, noct_reciente, con_reciente]
    colors2 = ["#1d4ed8", "#b91c1c", "#64748b"]
    bars2 = axes[1].bar(cats2, vals2, color=colors2)
    axes[1].set_title("Madrugada vs semana reciente (medido)")
    axes[1].set_ylabel("m³")
    axes[1].bar_label(bars2, labels=[_fmt(v) for v in vals2], padding=3)
    fig.suptitle("Puente Alto — muestra comparativa con WES vs sin WES", fontsize=13, fontweight="bold")
    fig.tight_layout()
    out_png.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_png, bbox_inches="tight")
    plt.close(fig)


def _escribir_excel(
    filas: List[Dict[str, object]],
    out_xlsx: Path,
    mayo_ini: date,
    mayo_fin: date,
    rec_ini: date,
    rec_fin: date,
    totales: Dict[str, float],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparativo"

    ws.cell(row=1, column=1, value="Muestra comparativa — Corporación Puente Alto (con WES vs sin WES)")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.cell(
        row=2,
        column=1,
        value=(
            f"Periodo con WES (última semana mayo): {mayo_ini:%d/%m/%Y} – {mayo_fin:%d/%m/%Y} | "
            f"Referencia reciente medida: {rec_ini:%d/%m/%Y} – {rec_fin:%d/%m/%Y}"
        ),
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=16)
    ws.cell(
        row=3,
        column=1,
        value=(
            "Sin WES estimado = consumo medido / (1 − %% auditoría/100). "
            "Nocturno = horas Chile 00:00–06:59. Posible fuga si ≥75% días con consumo nocturno."
        ),
    )
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=16)

    headers = [
        "N°",
        "Node ID",
        "Establecimiento",
        "% auditoría",
        "Con WES mayo (m³)",
        "Sin WES est. mayo (m³)",
        "Ahorro mayo (m³)",
        "Ahorro mayo (%)",
        "Nocturno mayo (m³)",
        "Diurno mayo (m³)",
        "Con WES reciente (m³)",
        "Nocturno reciente (m³)",
        "Diurno reciente (m³)",
        "Δ nocturno (rec − mayo)",
        "Posible fuga reciente",
        "Nota",
    ]
    hr = 5
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, row in enumerate(filas, start=1):
        r = hr + i
        ws.cell(row=r, column=1, value=i)
        for c, key in enumerate(
            [
                "node_id",
                "colegio",
                "pct",
                "con_mayo",
                "sin_mayo",
                "ahorro_m3",
                "ahorro_pct",
                "noct_mayo",
                "diurno_mayo",
                "con_reciente",
                "noct_reciente",
                "diurno_reciente",
                "delta_noct",
                "fuga_reciente",
                "nota",
            ],
            start=2,
        ):
            ws.cell(row=r, column=c, value=row.get(key, ""))

    tr = hr + len(filas) + 2
    ws.cell(row=tr, column=1, value="TOTAL / PROMEDIO").font = Font(bold=True)
    ws.cell(row=tr, column=5, value=round(totales["con_mayo"], 2))
    ws.cell(row=tr, column=6, value=round(totales["sin_mayo"], 2))
    ws.cell(row=tr, column=7, value=round(totales["ahorro_m3"], 2))
    ws.cell(row=tr, column=8, value=round(totales["ahorro_pct"], 1))
    ws.cell(row=tr, column=9, value=round(totales["noct_mayo"], 2))
    ws.cell(row=tr, column=10, value=round(totales["diurno_mayo"], 2))
    ws.cell(row=tr, column=11, value=round(totales["con_reciente"], 2))
    ws.cell(row=tr, column=12, value=round(totales["noct_reciente"], 2))
    ws.cell(row=tr, column=13, value=round(totales["diurno_reciente"], 2))

    ws2 = wb.create_sheet("Metodologia")
    notas = [
        "Objetivo: demostrar el ahorro que genera WES al interrumpir el consumo según horarios de corte programados.",
        "Con WES: suma de mediciones horarias API WES en el periodo (consumo real con cortes activos).",
        "Sin WES estimado: extrapolación contrafactual con % de eficiencia del informe de auditoría por colegio.",
        "La semana reciente muestra el consumo medido actual; un aumento del nocturno puede indicar recintos sin corte efectivo o fugas.",
        f"Agregado municipal ahorro mayo: {totales['ahorro_m3']:.1f} m³ ({totales['ahorro_pct']:.1f}% vs sin WES).",
    ]
    for i, n in enumerate(notas, start=1):
        ws2.cell(row=i, column=1, value=n)

    for col in range(1, 17):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["C"].width = 34
    ws.freeze_panes = f"A{hr + 1}"
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_xlsx))


def main() -> int:
    ap = argparse.ArgumentParser(description="Muestra comparativa con/sin WES — Puente Alto")
    ap.add_argument("--desde-mayo", default="25/05/2026", help="Inicio última semana mayo (dd/mm/aaaa)")
    ap.add_argument("--hasta-mayo", default="31/05/2026", help="Fin última semana mayo")
    ap.add_argument(
        "--hasta-reciente",
        default=None,
        help="Fin ventana reciente (dd/mm/aaaa). Default: auto (últimos 7 días con datos)",
    )
    args = ap.parse_args()

    mayo_ini = parse_date(args.desde_mayo).date()
    mayo_fin = parse_date(args.hasta_mayo, end_of_day=True).date()
    if args.hasta_reciente:
        rec_fin = parse_date(args.hasta_reciente, end_of_day=True).date()
        rec_ini = rec_fin - timedelta(days=6)
    else:
        rec_ini, rec_fin = _ultimos_7_dias_con_datos(DEFAULT_RECENT_END)

    dias_mayo = _rango_dias(mayo_ini, mayo_fin)
    dias_rec = _rango_dias(rec_ini, rec_fin)
    pct_map = _cargar_pct_auditoria()

    nodos = obtener_nodos_puente_alto()
    nodos.sort(key=lambda x: x["nodeName"])

    filas: List[Dict[str, object]] = []
    tot = {
        "con_mayo": 0.0,
        "sin_mayo": 0.0,
        "noct_mayo": 0.0,
        "diurno_mayo": 0.0,
        "con_reciente": 0.0,
        "noct_reciente": 0.0,
        "diurno_reciente": 0.0,
    }
    labels: List[str] = []
    con_mayo_list: List[float] = []
    sin_mayo_list: List[float] = []

    print("=" * 70)
    print("MUESTRA COMPARATIVA PUENTE ALTO — CON WES vs SIN WES")
    print("=" * 70)
    print(f"Mayo (con WES):     {mayo_ini} a {mayo_fin} ({len(dias_mayo)} dias)")
    print(f"Reciente (medido):  {rec_ini} a {rec_fin} ({len(dias_rec)} dias)")
    print(f"Colegios: {len(nodos)}")
    print()

    for n in nodos:
        nid = n["nodeId"]
        nombre = n["nodeName"]
        pct = pct_map.get(nid, EFICIENCIA_GLOBAL)

        m_mayo = _metricas_periodo(nid, dias_mayo)
        m_rec = _metricas_periodo(nid, dias_rec)

        con_m = float(m_mayo["total_m3"])
        sin_m = round(_sin_wes_est(con_m, pct), 2)
        ahorro = round(sin_m - con_m, 2)
        ahorro_pct = round((ahorro / sin_m * 100) if sin_m > 0 else 0.0, 1)

        noct_m = float(m_mayo["nocturno_m3"])
        noct_r = float(m_rec["nocturno_m3"])
        con_r = float(m_rec["total_m3"])
        delta_n = round(noct_r - noct_m, 2)

        nota_parts: List[str] = []
        if m_rec["posible_fuga"]:
            nota_parts.append("Posible fuga/consumo nocturno persistente en periodo reciente")
        if con_m <= 0.01:
            nota_parts.append("Sin datos mayo")
        if con_r <= 0.01:
            nota_parts.append("Sin datos reciente")
        if delta_n > 5:
            nota_parts.append("Nocturno reciente superior a mayo")

        filas.append(
            {
                "node_id": nid,
                "colegio": nombre,
                "pct": round(pct, 2),
                "con_mayo": con_m,
                "sin_mayo": sin_m,
                "ahorro_m3": ahorro,
                "ahorro_pct": ahorro_pct,
                "noct_mayo": noct_m,
                "diurno_mayo": float(m_mayo["diurno_m3"]),
                "con_reciente": con_r,
                "noct_reciente": noct_r,
                "diurno_reciente": float(m_rec["diurno_m3"]),
                "delta_noct": delta_n,
                "fuga_reciente": "Sí" if m_rec["posible_fuga"] else "No",
                "nota": "; ".join(nota_parts),
            }
        )

        if con_m > 0.01:
            labels.append(nombre[:28])
            con_mayo_list.append(con_m)
            sin_mayo_list.append(sin_m)

        tot["con_mayo"] += con_m
        tot["sin_mayo"] += sin_m
        tot["noct_mayo"] += noct_m
        tot["diurno_mayo"] += float(m_mayo["diurno_m3"])
        tot["con_reciente"] += con_r
        tot["noct_reciente"] += noct_r
        tot["diurno_reciente"] += float(m_rec["diurno_m3"])

        print(
            f"  {nombre[:40]:40} | mayo {con_m:7.1f} -> sin est {sin_m:7.1f} m3 | "
            f"reciente {con_r:7.1f} m3 | noct d {delta_n:+6.1f}"
        )

    tot["ahorro_m3"] = tot["sin_mayo"] - tot["con_mayo"]
    tot["ahorro_pct"] = (tot["ahorro_m3"] / tot["sin_mayo"] * 100) if tot["sin_mayo"] > 0 else 0.0

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    xlsx = OUT_DIR / f"muestra_con_sin_wes_pa_{mayo_ini:%Y%m%d}_{rec_fin:%Y%m%d}_{ts}.xlsx"
    png1 = OUT_DIR / f"grafico_con_vs_sin_wes_mayo_{mayo_ini:%Y%m%d}_{ts}.png"
    png2 = OUT_DIR / f"grafico_resumen_agregado_{ts}.png"

    _escribir_excel(filas, xlsx, mayo_ini, mayo_fin, rec_ini, rec_fin, tot)

    if labels:
        _grafico_barras_agrupadas(
            labels,
            con_mayo_list,
            sin_mayo_list,
            "Con WES (medido)",
            "Sin WES (estimado)",
            f"Puente Alto — semana {mayo_ini:%d/%m}–{mayo_fin:%d/%m/%Y} — con vs sin WES por colegio",
            png1,
        )

    _grafico_resumen_totales(
        tot["con_mayo"],
        tot["sin_mayo"],
        tot["con_reciente"],
        tot["noct_mayo"],
        tot["noct_reciente"],
        png2,
        f"{mayo_ini:%d/%m}–{mayo_fin:%d/%m}",
        f"{rec_ini:%d/%m}–{rec_fin:%d/%m}",
    )

    print()
    print("=" * 70)
    print("RESUMEN AGREGADO")
    print("=" * 70)
    print(f"Con WES (mayo):        {tot['con_mayo']:.1f} m³")
    print(f"Sin WES est. (mayo):   {tot['sin_mayo']:.1f} m³")
    print(f"Ahorro estimado mayo:  {tot['ahorro_m3']:.1f} m³ ({tot['ahorro_pct']:.1f}%)")
    print(f"Nocturno mayo:         {tot['noct_mayo']:.1f} m³")
    print(f"Total reciente:        {tot['con_reciente']:.1f} m³ (nocturno {tot['noct_reciente']:.1f} m³)")
    print()
    print(f"Excel:   {xlsx}")
    print(f"Gráfico: {png1}")
    print(f"Gráfico: {png2}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

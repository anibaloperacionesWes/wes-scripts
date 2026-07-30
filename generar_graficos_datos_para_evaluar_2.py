"""
Lee ``datos para evaluar 2.0.xlsx`` (comparación m³ por mes) y genera gráficos PNG.

- Por cada hoja de datos: barras agrupadas (m³ columna C vs columna F) por mes,
  y leyendas tomadas de la fila 1 (bloques «CON/SIN SISTEMA»).
- Consolidado: totales de la fila resumen (sumas m³ C y F, y diferencia G) por establecimiento.

Salida por defecto:
  reports/Renca/Coparacion App con Aguas Andinas/Graficos Consolidados/

Uso:
  python generar_graficos_datos_para_evaluar_2.py
  python generar_graficos_datos_para_evaluar_2.py --xlsx "C:/Users/aniba/Downloads/datos para evaluar 2.0.xlsx"
"""

from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
DEFAULT_OUT = (
    ROOT
    / "reports"
    / "Renca"
    / "Coparacion App con Aguas Andinas"
    / "Graficos Consolidados"
)
DEFAULT_XLSX = Path(r"C:\Users\aniba\Downloads\datos para evaluar 2.0.xlsx")

MESES_ES = (
    "",
    "ene",
    "feb",
    "mar",
    "abr",
    "may",
    "jun",
    "jul",
    "ago",
    "sep",
    "oct",
    "nov",
    "dic",
)


def _fmt_mes(d: date) -> str:
    return f"{d.day:02d}-{MESES_ES[d.month]}-{d.year}"


def _as_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _is_number(x) -> bool:
    return isinstance(x, (int, float)) and not isinstance(x, bool)


def _norm_label(s: str | None) -> str:
    if not s or not isinstance(s, str):
        return ""
    return re.sub(r"\s+", " ", s.strip())


@dataclass
class FilaMensual:
    etiqueta: str
    m3_c: float
    m3_f: float
    diff_g: float | None


@dataclass
class ResumenHoja:
    nombre: str
    etiqueta_c: str
    etiqueta_f: str
    mensual: list[FilaMensual]
    total_c: float | None
    total_f: float | None
    total_g: float | None


def _leer_hoja(ws, nombre: str) -> ResumenHoja | None:
    """Espera fila 1 títulos de bloque, fila 2 encabezados de tabla, datos desde fila 3."""
    a1 = _norm_label(ws.cell(1, 1).value)
    d1 = _norm_label(ws.cell(1, 4).value)
    etiqueta_c = a1 or "Bloque A–C (m³)"
    etiqueta_f = d1 or "Bloque D–F (m³)"

    mensual: list[FilaMensual] = []
    for r in range(3, ws.max_row + 1):
        da = _as_date(ws.cell(r, 1).value)
        dd = _as_date(ws.cell(r, 4).value)
        vc = ws.cell(r, 3).value
        vf = ws.cell(r, 6).value
        vg = ws.cell(r, 7).value

        if da is None and dd is None:
            continue
        if not (_is_number(vc) and _is_number(vf)):
            continue

        ref = max(x for x in (da, dd) if x is not None)
        etiqueta = _fmt_mes(ref)
        diff = float(vg) if _is_number(vg) else None
        mensual.append(FilaMensual(etiqueta, float(vc), float(vf), diff))

    if not mensual:
        return None

    # Fila de totales: sin fecha en A, números en C y F (y a vece «Total» en B o E)
    total_c: float | None = None
    total_f: float | None = None
    total_g: float | None = None
    for r in range(3, ws.max_row + 1):
        da = _as_date(ws.cell(r, 1).value)
        dd = _as_date(ws.cell(r, 4).value)
        if da is not None or dd is not None:
            continue
        vc = ws.cell(r, 3).value
        vf = ws.cell(r, 6).value
        vg = ws.cell(r, 7).value
        if _is_number(vc) and _is_number(vf):
            total_c = float(vc)
            total_f = float(vf)
            total_g = float(vg) if _is_number(vg) else None
            break

    return ResumenHoja(
        nombre=nombre,
        etiqueta_c=etiqueta_c,
        etiqueta_f=etiqueta_f,
        mensual=mensual,
        total_c=total_c,
        total_f=total_f,
        total_g=total_g,
    )


def _safe_name(s: str) -> str:
    return re.sub(r"[^\w\-]+", "_", s, flags=re.UNICODE).strip("_") or "hoja"


def _grafico_mensual(res: ResumenHoja, out_path: Path) -> None:
    labels = [m.etiqueta for m in res.mensual]
    x = np.arange(len(labels))
    w = 0.36
    fig, ax = plt.subplots(figsize=(11, 5.5))
    ax.bar(x - w / 2, [m.m3_c for m in res.mensual], width=w, label=res.etiqueta_c, color="#1F4788")
    ax.bar(x + w / 2, [m.m3_f for m in res.mensual], width=w, label=res.etiqueta_f, color="#C45C26")
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=35, ha="right")
    ax.set_ylabel("m³ (periodo de lectura)")
    ax.set_title(f"{res.nombre} — consumos mensuales comparados")
    ax.legend(loc="upper right")
    ax.grid(axis="y", alpha=0.3)
    fig.tight_layout()
    fig.savefig(out_path, dpi=150)
    plt.close(fig)


def _grafico_diferencia_mensual(res: ResumenHoja, out_path: Path) -> None:
    diffs = [m.diff_g for m in res.mensual if m.diff_g is not None]
    labels = [m.etiqueta for m in res.mensual if m.diff_g is not None]
    if not diffs:
        return
    fig, ax = plt.subplots(figsize=(11, 4.5))
    colors = ["#2E7D32" if d >= 0 else "#C62828" for d in diffs]
    ax.bar(labels, diffs, color=colors)
    ax.axhline(0, color="#333", linewidth=0.8)
    ax.set_ylabel("m³ (diferencia por mes)")
    ax.set_title(f"{res.nombre} — diferencia mensual (columna G)")
    ax.tick_params(axis="x", rotation=35)
    ax.grid(axis="y", alpha=0.3)
    fig.tight_layout()
    fig.savefig(out_path, dpi=150)
    plt.close(fig)


def _grafico_consolidado_totales(resumenes: list[ResumenHoja], out_path: Path) -> None:
    nombres = [r.nombre for r in resumenes]
    tc = [r.total_c if r.total_c is not None else 0.0 for r in resumenes]
    tf = [r.total_f if r.total_f is not None else 0.0 for r in resumenes]
    x = np.arange(len(nombres))
    w = 0.35
    fig, ax = plt.subplots(figsize=(11, 6))
    ax.bar(x - w / 2, tc, width=w, label="Total m³ (columna C)", color="#1F4788")
    ax.bar(x + w / 2, tf, width=w, label="Total m³ (columna F)", color="#C45C26")
    ax.set_xticks(x)
    ax.set_xticklabels(nombres, rotation=20, ha="right")
    ax.set_ylabel("m³ acumulado (fila totales del Excel)")
    ax.set_title("Consolidado — totales por establecimiento (datos para evaluar 2.0)")
    ax.legend()
    ax.grid(axis="y", alpha=0.3)
    fig.tight_layout()
    fig.savefig(out_path, dpi=150)
    plt.close(fig)


def _grafico_consolidado_diferencia(resumenes: list[ResumenHoja], out_path: Path) -> None:
    datos = [(r.nombre, r.total_g) for r in resumenes if r.total_g is not None]
    if not datos:
        return
    nombres, vals = zip(*datos)
    fig, ax = plt.subplots(figsize=(10, 5))
    colors = ["#2E7D32" if v >= 0 else "#C62828" for v in vals]
    ax.barh(nombres, vals, color=colors)
    ax.axvline(0, color="#333", linewidth=0.8)
    ax.set_xlabel("m³ (total diferencia — columna G)")
    ax.set_title("Consolidado — diferencia total por establecimiento")
    ax.grid(axis="x", alpha=0.3)
    fig.tight_layout()
    fig.savefig(out_path, dpi=150)
    plt.close(fig)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="Ruta al Excel")
    ap.add_argument("--out-dir", type=Path, default=DEFAULT_OUT, help="Carpeta de salida PNG")
    args = ap.parse_args()

    xlsx = args.xlsx.expanduser()
    if not xlsx.is_file():
        print(f"[ERROR] No existe: {xlsx}", file=sys.stderr)
        return 1

    out_dir = args.out_dir.resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(xlsx, data_only=True)
    resumenes: list[ResumenHoja] = []

    for name in wb.sheetnames:
        ws = wb[name]
        res = _leer_hoja(ws, name)
        if res is None:
            print(f"[SKIP] {name}: sin filas mensuales reconocibles")
            continue
        resumenes.append(res)
        base = _safe_name(name)
        p1 = out_dir / f"{base}_mensual_m3.png"
        p2 = out_dir / f"{base}_diferencia_mensual_m3.png"
        _grafico_mensual(res, p1)
        _grafico_diferencia_mensual(res, p2)
        print(f"[OK] {name} -> {p1.name}, {p2.name}")

    wb.close()

    if resumenes:
        p_tot = out_dir / "CONSOLIDADO_totales_m3_por_establecimiento.png"
        _grafico_consolidado_totales(resumenes, p_tot)
        print(f"[OK] Consolidado totales -> {p_tot.name}")
        p_dif = out_dir / "CONSOLIDADO_diferencia_total_por_establecimiento.png"
        _grafico_consolidado_diferencia(resumenes, p_dif)
        print(f"[OK] Consolidado diferencias -> {p_dif.name}")

    print(f"[OK] Carpeta: {out_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

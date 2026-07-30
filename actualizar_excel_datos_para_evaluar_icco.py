"""
Actualiza ``datos para evaluar.xlsx`` en Coparacion App con Aguas Andinas.

- Hoja ``Revision``: resumen de Gimnasio/Piscina (hojas manuales) + conteo de PDF por carpeta
  y listado de archivos que no se pudieron leer.
- Hoja ``ICCO``: todas las boletas electrónicas en ``ICCO facturas`` (mismo encabezado A1:G1 que Gimnasio).
- Hoja ``ICCP``: mismos datos leídos desde ``ICCP facturas``.
- Hoja ``Facturas_Escuela_Lo_Velzaquez``: PDF en ``Escuela lo Velzaquez Facturas``.
- Hojas ``Facturas_Gimnasio`` y ``Facturas_Piscina``: datos leídos de **todos** los PDF
  en ``Gimnasio Facturaciones`` y ``Piscina Facturaciones`` (incluye los que agregue),
  mismo formato que ICCO. Si un PDF es historial u otro formato, se usa ``listar_periodos_desde_pdf``
  (lecturas numéricas del medidor pueden quedar vacías si el OCR no las trae).

Vuelva a ejecutar después de agregar PDFs en cualquiera de esas carpetas.

Uso:
  python actualizar_excel_datos_para_evaluar_icco.py
"""

from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parent
BASE = ROOT / "reports" / "Renca" / "Coparacion App con Aguas Andinas"
XLSX = BASE / "datos para evaluar.xlsx"
ICCO_DIR = BASE / "ICCO facturas"
ICCP_DIR = BASE / "ICCP facturas"
GIMNASIO_DIR = BASE / "Gimnasio Facturaciones"
PISCINA_DIR = BASE / "Piscina Facturaciones"
ESCUELA_LO_VELZQUEZ_DIR = BASE / "Escuela lo Velzaquez Facturas"

_MESES = {
    "ENE": 1,
    "JAN": 1,
    "FEB": 2,
    "MAR": 3,
    "ABR": 4,
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AGO": 8,
    "AUG": 8,
    "SEP": 9,
    "OCT": 10,
    "NOV": 11,
    "DIC": 12,
    "DEC": 12,
}


def _parse_fecha_es_dd_mmm_yyyy(s: str) -> datetime:
    p = s.strip().upper().split("-")
    if len(p) != 3:
        raise ValueError(f"Fecha inválida: {s!r}")
    dd = int(p[0])
    mm = _MESES[p[1][:3]]
    yy = int(p[2])
    return datetime(yy, mm, dd)


def _extraer_texto_pdf(path: Path) -> str:
    reader = PdfReader(str(path))
    return "\n".join((pg.extract_text() or "") for pg in reader.pages)


def _parse_lectura_medidor_a_entero(s: str) -> int:
    """Ej. ``157.383`` o ``1.234.567`` (miles con punto) -> entero m³ acumulados."""
    s = s.strip().replace(" ", "")
    if not s:
        raise ValueError("Lectura vacía")
    if "," in s:
        return int(float(s.replace(".", "").replace(",", ".")))
    if "." in s:
        partes = s.split(".")
        if all(p.isdigit() for p in partes):
            return int("".join(partes))
    return int(float(s))


@dataclass(frozen=True)
class FilaFacturaEval:
    """Una fila tipo hoja ICCO (A–F con fechas/lecturas; G libre)."""

    boleta: str
    archivo_pdf: str
    fecha_lectura_actual: datetime
    lectura_actual: int | None
    m3_periodo: int
    fecha_lectura_anterior: datetime
    lectura_anterior: int | None


def _parse_boleta_electronica_aa(path: Path) -> FilaFacturaEval:
    """Factura/boleta estándar Aguas Andinas (texto en PDF)."""
    txt = _extraer_texto_pdf(path)
    boleta_m = re.search(
        r"(?:FACTURA|BOLETA)\s+ELECTR[ÓO]NICA\s*\n?\s*N[º°]\s*([0-9]{6,})",
        txt,
        flags=re.IGNORECASE,
    )
    if not boleta_m:
        boleta_m = re.search(r"\nN[º°]\s*([0-9]{6,})\n", txt, flags=re.IGNORECASE)
    consumo_m = re.search(r"CONSUMO\s+TOTAL\s+([0-9\.\,]+)\s*m3", txt, flags=re.IGNORECASE)
    la_m = re.search(
        r"LECTURA\s+ACTUAL\s+([0-9]{2}-[A-Z]{3}-[0-9]{4})\s+([0-9\.\,]+)\s*m3",
        txt,
        flags=re.IGNORECASE,
    )
    lan_m = re.search(
        r"LECTURA\s+ANTERIOR\s+([0-9]{2}-[A-Z]{3}-[0-9]{4})\s+([0-9\.\,]+)\s*m3",
        txt,
        flags=re.IGNORECASE,
    )
    if not (consumo_m and la_m and lan_m):
        raise ValueError("No es boleta estándar (faltan LECTURA ACTUAL/ANTERIOR o CONSUMO TOTAL)")
    dt_act = _parse_fecha_es_dd_mmm_yyyy(la_m.group(1))
    dt_ant = _parse_fecha_es_dd_mmm_yyyy(lan_m.group(1))
    m3 = int(float(consumo_m.group(1).replace(".", "").replace(",", ".")))
    lect_act = _parse_lectura_medidor_a_entero(la_m.group(2))
    lect_ant = _parse_lectura_medidor_a_entero(lan_m.group(2))
    return FilaFacturaEval(
        boleta=(boleta_m.group(1) if boleta_m else path.stem),
        archivo_pdf=path.name,
        fecha_lectura_actual=dt_act,
        lectura_actual=lect_act,
        m3_periodo=m3,
        fecha_lectura_anterior=dt_ant,
        lectura_anterior=lect_ant,
    )


def _filas_desde_pdf_generico(path: Path) -> list[FilaFacturaEval]:
    """Intenta boleta estándar; si no, períodos desde ``facturacion_aguas_andinas_pdf``."""
    try:
        return [_parse_boleta_electronica_aa(path)]
    except Exception:
        from facturacion_aguas_andinas_pdf import listar_periodos_desde_pdf

        pers = listar_periodos_desde_pdf(path)
        out: list[FilaFacturaEval] = []
        for per in pers:
            out.append(
                FilaFacturaEval(
                    boleta=str(per.boleta),
                    archivo_pdf=path.name,
                    fecha_lectura_actual=per.lectura_actual,
                    lectura_actual=None,
                    m3_periodo=int(per.m3_cuenta),
                    fecha_lectura_anterior=per.lectura_anterior,
                    lectura_anterior=None,
                )
            )
        return out


def _cargar_directorio_facturas(dir_: Path, etiqueta: str) -> tuple[list[FilaFacturaEval], list[str]]:
    if not dir_.is_dir():
        return [], []
    pdfs = sorted(dir_.glob("*.pdf"))
    if not pdfs:
        return [], []
    out: list[FilaFacturaEval] = []
    err: list[str] = []
    for p in pdfs:
        try:
            out.extend(_filas_desde_pdf_generico(p))
        except Exception as e:
            err.append(f"{etiqueta} / {p.name}: {e}")
    out.sort(key=lambda r: (r.fecha_lectura_actual, r.boleta))
    return out, err


def _resumen_hoja(wb, nombre: str) -> str:
    if nombre not in wb.sheetnames:
        return f"{nombre}: (no existe)"
    ws = wb[nombre]
    n = ws.max_row or 0
    if n < 2:
        return f"{nombre}: sin filas de datos (max_row={n})"
    a2 = ws.cell(row=2, column=1).value
    return f"{nombre}: {n - 1} fila(s) bajo encabezado; primera fecha col.A = {a2!r}"


def _escribir_hoja_facturas(
    wb,
    nombre_hoja: str,
    ws_plantilla,
    filas: list[FilaFacturaEval],
) -> None:
    if nombre_hoja in wb.sheetnames:
        del wb[nombre_hoja]
    ws = wb.create_sheet(nombre_hoja)
    for col in range(1, 15):
        ws.cell(row=1, column=col, value=ws_plantilla.cell(row=1, column=col).value)
    hdr_fill = PatternFill("solid", fgColor="1F4788")
    hdr_font = Font(color="FFFFFF", bold=True)
    for c in range(1, 8):
        cell = ws.cell(row=1, column=c)
        cell.fill = hdr_fill
        cell.font = hdr_font

    for i, f in enumerate(filas, start=2):
        ws.cell(row=i, column=1, value=f.fecha_lectura_actual)
        ws.cell(row=i, column=2, value=f.lectura_actual)
        ws.cell(row=i, column=3, value=f.m3_periodo)
        ws.cell(row=i, column=4, value=f.fecha_lectura_anterior)
        ws.cell(row=i, column=5, value=f.lectura_anterior)
        ws.cell(row=i, column=6, value=None)
        ws.cell(row=i, column=7, value=None)
        ws.cell(row=i, column=8, value=f"N° {f.boleta} · {f.archivo_pdf}")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
        for c in row:
            if c.column in (1, 4):
                c.number_format = "DD-MM-YYYY"
            elif c.column in (2, 5) and c.value is not None:
                c.number_format = "#,##0"
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
        for c in row:
            c.number_format = "#,##0"


def main() -> None:
    if not XLSX.is_file():
        print(f"[ERROR] No está el archivo: {XLSX}", file=sys.stderr)
        sys.exit(1)

    filas_icco, err_icco = _cargar_directorio_facturas(ICCO_DIR, "ICCO")
    filas_iccp, err_iccp = _cargar_directorio_facturas(ICCP_DIR, "ICCP")
    filas_gym, err_gym = _cargar_directorio_facturas(GIMNASIO_DIR, "Gimnasio")
    filas_pis, err_pis = _cargar_directorio_facturas(PISCINA_DIR, "Piscina")
    filas_escuela, err_escuela = _cargar_directorio_facturas(
        ESCUELA_LO_VELZQUEZ_DIR, "Escuela lo Velzaquez"
    )

    if not (filas_icco or filas_iccp or filas_gym or filas_pis or filas_escuela):
        print(
            "[ERROR] No se obtuvo ninguna fila desde PDFs. Revise carpetas y errores:\n"
            + "\n".join(err_icco + err_iccp + err_gym + err_pis + err_escuela),
            file=sys.stderr,
        )
        sys.exit(1)

    wb = load_workbook(XLSX)
    if "Gimnasio" not in wb.sheetnames:
        print("[ERROR] El Excel debe tener la hoja «Gimnasio» (encabezado plantilla).", file=sys.stderr)
        sys.exit(1)
    ws_g = wb["Gimnasio"]

    # --- Revision ---
    if "Revision" in wb.sheetnames:
        del wb["Revision"]
    rev = wb.create_sheet("Revision")
    rev["A1"] = "Resumen automático (revisión datos para evaluar)"
    rev["A1"].font = Font(bold=True)
    row = 2
    rev.cell(row=row, column=1, value=_resumen_hoja(wb, "Gimnasio"))
    row += 1
    rev.cell(row=row, column=1, value=_resumen_hoja(wb, "Piscina"))
    row += 1
    rev.cell(row=row, column=1, value=f"PDF en ICCO facturas: {len(list(ICCO_DIR.glob('*.pdf'))) if ICCO_DIR.is_dir() else 0} archivo(s) → {len(filas_icco)} fila(s) generadas")
    row += 1
    rev.cell(
        row=row,
        column=1,
        value=f"PDF en ICCP facturas: {len(list(ICCP_DIR.glob('*.pdf'))) if ICCP_DIR.is_dir() else 0} archivo(s) → {len(filas_iccp)} fila(s) en «ICCP»",
    )
    row += 1
    rev.cell(
        row=row,
        column=1,
        value=f"PDF en Gimnasio Facturaciones: {len(list(GIMNASIO_DIR.glob('*.pdf'))) if GIMNASIO_DIR.is_dir() else 0} archivo(s) → {len(filas_gym)} fila(s) en «Facturas_Gimnasio»",
    )
    row += 1
    rev.cell(
        row=row,
        column=1,
        value=f"PDF en Piscina Facturaciones: {len(list(PISCINA_DIR.glob('*.pdf'))) if PISCINA_DIR.is_dir() else 0} archivo(s) → {len(filas_pis)} fila(s) en «Facturas_Piscina»",
    )
    row += 1
    rev.cell(
        row=row,
        column=1,
        value=(
            f"PDF en Escuela lo Velzaquez Facturas: "
            f"{len(list(ESCUELA_LO_VELZQUEZ_DIR.glob('*.pdf'))) if ESCUELA_LO_VELZQUEZ_DIR.is_dir() else 0} "
            f"archivo(s) → {len(filas_escuela)} fila(s) en «Facturas_Escuela_Lo_Velzaquez»"
        ),
    )
    row += 1
    rev.cell(
        row=row,
        column=1,
        value="Hojas ICCO / ICCP / Facturas_Gimnasio / Facturas_Piscina / Facturas_Escuela_Lo_Velzaquez: encabezado A1:G1 igual que «Gimnasio». Columna H: boleta y nombre de PDF.",
    )
    row += 1
    todos_err = err_icco + err_iccp + err_gym + err_pis + err_escuela
    if todos_err:
        rev.cell(row=row, column=1, value="Errores u omisiones al leer PDF (revisar archivo):")
        rev.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        for e in todos_err:
            rev.cell(row=row, column=1, value=f"  • {e}")
            row += 1
    rev.column_dimensions["A"].width = 110

    if filas_icco:
        _escribir_hoja_facturas(wb, "ICCO", ws_g, filas_icco)
    elif "ICCO" in wb.sheetnames:
        del wb["ICCO"]
    if filas_iccp:
        _escribir_hoja_facturas(wb, "ICCP", ws_g, filas_iccp)
    elif "ICCP" in wb.sheetnames:
        del wb["ICCP"]
    if filas_gym:
        _escribir_hoja_facturas(wb, "Facturas_Gimnasio", ws_g, filas_gym)
    elif "Facturas_Gimnasio" in wb.sheetnames:
        del wb["Facturas_Gimnasio"]
    if filas_pis:
        _escribir_hoja_facturas(wb, "Facturas_Piscina", ws_g, filas_pis)
    elif "Facturas_Piscina" in wb.sheetnames:
        del wb["Facturas_Piscina"]
    if filas_escuela:
        _escribir_hoja_facturas(wb, "Facturas_Escuela_Lo_Velzaquez", ws_g, filas_escuela)
    elif "Facturas_Escuela_Lo_Velzaquez" in wb.sheetnames:
        del wb["Facturas_Escuela_Lo_Velzaquez"]

    wb.save(XLSX)
    print(f"[OK] {XLSX}")
    print(
        f"     ICCO: {len(filas_icco)} filas | ICCP: {len(filas_iccp)} | "
        f"Facturas_Gimnasio: {len(filas_gym)} | Facturas_Piscina: {len(filas_pis)} | "
        f"Facturas_Escuela_Lo_Velzaquez: {len(filas_escuela)}"
    )
    if todos_err:
        print(f"     [ADVERTENCIA] {len(todos_err)} PDF(s) con error (detalle en hoja Revision).")


if __name__ == "__main__":
    main()

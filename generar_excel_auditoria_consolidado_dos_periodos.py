"""
Descarga CSV horario (API dates.measures.csv fusionado por dia Chile) y genera
un Excel aparte con:
  - Hoja "Filas": TIME,VALUE del CSV; **hora y día civil Chile** (TIME en UTC → America/Santiago),
    misma regla que ``get_hourly_measures_for_day`` al leer el CSV fusionado.
  - Hoja "Consolidado": matriz 24 × días con **m³/h por hora Chile 0–23** (etiqueta como la app).

  Si la suma horaria del CSV contradice ``totalM3`` del JSON (CSV casi vacío), se alinea a
  ``totalM3/24`` como en ``get_hourly_measures_for_day`` (p. ej. algunos nodos con ``measures`` vacío).

El CSV guarda marcas en UTC; la columna "01:00" del consolidado es **01:00 Chile**, no el ``T01:00`` UTC.
Ignorar la hora literal del TIME al etiquetar filas desalinea respecto a la app (p. ej. 01:00 Chile
corresponde a ~04:00 UTC en abril, no a la fila ``...T01:00:00Z``).

Por defecto:
  - Periodo 1 (con WES): 23-29 marzo 2026
  - Periodo 2 (sin WES): 06-12 abril 2026

Uso:
  python generar_excel_auditoria_consolidado_dos_periodos.py
  python generar_excel_auditoria_consolidado_dos_periodos.py --solo-desde-csv
  python generar_excel_auditoria_consolidado_dos_periodos.py --toda-la-carpeta

La función ``validar_consolidado_csv_contra_app`` (usada por ``generar_auditoria_wes_cliente``) imprime
una tabla día a día: suma CSV bruta, ``totalM3`` JSON (app) y suma tras reconciliación.
"""
from __future__ import annotations

import argparse
import re
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from descargar_csv_medidas_wes import (
    _fusionar_csv_bloques,
    _get_csv_bloques_utc_dia_chile,
)
from generar_reporte_word import (
    _chile_hours_from_dates_measures_csv_text,
    _dt_to_chile,
    _reconcile_chile_hours_with_total_m3,
    _requests_session,
    _total_m3_from_json_for_chile_day,
    _value_by_time_sum_duplicate_rows,
)

AUDIT_DIR_DEFAULT = (
    Path(__file__).resolve().parent
    / "reports"
    / "reporte de auditoria"
    / "auditoria_puntos_renca_abril_2026"
    / "Auditoria ICCO abril"
)


_re_csv_dia = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def _iter_days(desde: date, hasta: date):
    d = desde
    while d <= hasta:
        yield d
        d += timedelta(days=1)


def _listar_csv_por_dia_ordenados(csv_dir: Path) -> List[Tuple[date, Path]]:
    """Archivos ``YYYY-MM-DD.csv`` en la carpeta, orden cronologico."""
    out: List[Tuple[date, Path]] = []
    for p in csv_dir.glob("*.csv"):
        if not _re_csv_dia.match(p.stem):
            continue
        try:
            d = date.fromisoformat(p.stem)
        except ValueError:
            continue
        out.append((d, p.resolve()))
    out.sort(key=lambda x: x[0])
    return out


def _vector_y_filas_por_dia_en_texto(
    texto_csv: str,
    dia_archivo: date,
    nombre_archivo: str,
    node_id: str,
) -> Tuple[List[float], List[Tuple[str, str, str, str, int, float]]]:
    """
    Vector 0..23 = **hora civil Chile** para ``dia_archivo``; filas con la misma asignación.
    Igual que ``get_hourly_measures_for_day``: CSV + suma por ``TIME`` duplicado + reconciliación
    con ``totalM3`` JSON si la suma horaria del CSV no cuadra con el total del día.
    """
    raw = _chile_hours_from_dates_measures_csv_text(texto_csv, dia_archivo)
    tj = _total_m3_from_json_for_chile_day(node_id, dia_archivo)
    acc_chile, used_fallback = _reconcile_chile_hours_with_total_m3(raw, tj)
    vec = [float(acc_chile.get(h, 0.0)) for h in range(24)]

    filas_out: List[Tuple[str, str, str, str, int, float]] = []
    if used_fallback and tj is not None:
        for h in range(24):
            v = float(acc_chile.get(h, 0.0))
            filas_out.append(
                (
                    nombre_archivo,
                    f"(totalM3 JSON {tj} m³ → reparto uniforme; CSV horario no cuadraba)",
                    dia_archivo.isoformat(),
                    f"{h:02d}:00",
                    h,
                    v,
                )
            )
    else:
        by_t = _value_by_time_sum_duplicate_rows(texto_csv)
        for time_raw in sorted(by_t.keys()):
            val = float(by_t[time_raw])
            if "T" not in time_raw:
                continue
            try:
                ts_norm = time_raw.strip().replace("Z", "+00:00")
                dt = datetime.fromisoformat(ts_norm)
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                ch = _dt_to_chile(dt)
                if ch.date() != dia_archivo:
                    continue
                hi = int(ch.hour)
                filas_out.append(
                    (
                        nombre_archivo,
                        time_raw,
                        ch.date().isoformat(),
                        f"{hi:02d}:00",
                        hi,
                        val,
                    )
                )
            except (ValueError, TypeError):
                continue

    return vec, filas_out


def _descargar_csv_a_disco(
    node_id: str, dia: date, dest_dir: Path, session
) -> Path:
    bloques = _get_csv_bloques_utc_dia_chile(node_id, dia, session=session)
    texto = _fusionar_csv_bloques(bloques)
    dest = dest_dir / f"{dia.isoformat()}.csv"
    dest.write_text(texto, encoding="utf-8")
    return dest


def generar_excel_consolidado(
    out_xlsx: Path,
    node_id: str,
    dias_periodo1: Sequence[date],
    dias_periodo2: Sequence[date],
    *,
    titulo_p1: str,
    titulo_p2: str,
    solo_desde_csv: bool = False,
    csv_dir: Optional[Path] = None,
) -> None:
    dest_csv = (csv_dir or (out_xlsx.parent / "csv_descarga_api")).resolve()
    dias_all = list(dias_periodo1) + list(dias_periodo2)

    mats: List[List[float]] = []
    todas_las_filas: List[Tuple[str, str, str, str, int, float]] = []

    if solo_desde_csv:
        for d in dias_all:
            p = dest_csv / f"{d.isoformat()}.csv"
            if not p.is_file():
                raise FileNotFoundError(
                    f"Falta {p.name} en {dest_csv}. Descarga antes (sin --solo-desde-csv)."
                )
            texto = p.read_text(encoding="utf-8")
            vec, filas = _vector_y_filas_por_dia_en_texto(texto, d, p.name, node_id)
            mats.append(vec)
            todas_las_filas.extend(filas)
    else:
        dest_csv.mkdir(parents=True, exist_ok=True)
        sess = _requests_session()
        for d in dias_all:
            path = _descargar_csv_a_disco(node_id, d, dest_csv, sess)
            texto = path.read_text(encoding="utf-8")
            vec, filas = _vector_y_filas_por_dia_en_texto(texto, d, path.name, node_id)
            mats.append(vec)
            todas_las_filas.extend(filas)

    n1 = len(dias_periodo1)
    n2 = len(dias_periodo2)
    if len(mats) != n1 + n2:
        raise ValueError("conteo dias")

    wb = Workbook()
    ws_f = wb.active
    ws_f.title = "Filas"
    _escribir_hoja_filas(ws_f, todas_las_filas)

    ws = wb.create_sheet("Consolidado", 1)
    ws["A1"] = (
        "Consumo horario (m3/h) — hora civil Chile (TIME UTC → America/Santiago; misma regla que la app). "
        f"Nodo {node_id}."
    )
    ws["A1"].font = Font(italic=True, size=9)

    c0 = 2
    sep_col = c0 + n1
    ws["A2"] = "Hora"
    ws["A2"].font = Font(bold=True)
    ws.merge_cells(start_row=2, start_column=c0, end_row=2, end_column=c0 + n1 - 1)
    ws.cell(row=2, column=c0, value=titulo_p1)
    ws.cell(row=2, column=c0).font = Font(bold=True)
    ws.cell(row=2, column=c0).alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=sep_col, end_row=2, end_column=sep_col + n2 - 1)
    ws.cell(row=2, column=sep_col, value=titulo_p2)
    ws.cell(row=2, column=sep_col).font = Font(bold=True)
    ws.cell(row=2, column=sep_col).alignment = Alignment(horizontal="center")

    r_head = 3
    for i, d in enumerate(dias_periodo1):
        col = c0 + i
        ws.cell(row=r_head, column=col, value=d.strftime("%d-%m-%Y"))
        ws.cell(row=r_head, column=col).font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col)].width = 12
    for j, d in enumerate(dias_periodo2):
        col = sep_col + j
        ws.cell(row=r_head, column=col, value=d.strftime("%d-%m-%Y"))
        ws.cell(row=r_head, column=col).font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col)].width = 12

    r0 = 4
    for h in range(24):
        row = r0 + h
        ws.cell(row=row, column=1, value=f"{h:02d}:00")
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        for i in range(n1):
            val = mats[i][h]
            cell = ws.cell(row=row, column=c0 + i, value=round(val, 4))
            cell.number_format = "0.0000"
        for j in range(n2):
            val = mats[n1 + j][h]
            cell = ws.cell(row=row, column=sep_col + j, value=round(val, 4))
            cell.number_format = "0.0000"

    ws.column_dimensions["A"].width = 10
    ws.freeze_panes = "B4"
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)


def validar_consolidado_csv_contra_app(
    node_id: str,
    csv_dir: Path,
    dias: Sequence[date],
) -> Tuple[int, int]:
    """
    Tras generar el Excel, comprueba cada día que la misma lógica del consolidado
    (CSV fusionado → hora Chile → suma por ``TIME`` duplicado → reconciliación con ``totalM3``)
    deja la **suma diaria** alineada con ``totalM3`` del JSON (referencia habitual de la app).

    Imprime una tabla en consola. Devuelve ``(advertencias, errores_graves)`` donde
    ``errores_graves`` es cantidad de días con ``totalM3`` y suma reconciliada que aún difieren
    > 0,1 m³ (no debería ocurrir si la reconciliación es correcta).
    """
    csv_dir = csv_dir.resolve()
    lineas: List[str] = []
    adv = 0
    err = 0
    lineas.append("")
    lineas.append(
        "  Día          CSV bruto  totalM3 app  Tras reconc.    Δ        Estado"
    )
    lineas.append("  " + "-" * 74)
    for d in dias:
        p = csv_dir / f"{d.isoformat()}.csv"
        ds = d.strftime("%d-%m-%Y")
        if not p.is_file():
            lineas.append(f"  {ds:12}  (no existe {p.name})")
            adv += 1
            continue
        texto = p.read_text(encoding="utf-8")
        raw = _chile_hours_from_dates_measures_csv_text(texto, d)
        s_raw = sum(raw.values())
        tj = _total_m3_from_json_for_chile_day(node_id, d)
        out, _ = _reconcile_chile_hours_with_total_m3(raw, tj)
        s_out = sum(out.values())
        if tj is None:
            st = "sin total JSON"
            adv += 1 if s_raw > 0.01 else 0
            lineas.append(
                f"  {ds:12} {s_raw:10.2f} {'—':>10} {s_out:10.2f}    —       {st}"
            )
            continue
        tjf = float(tj)
        delta = s_out - tjf
        if abs(delta) > 0.1:
            st = "ERROR reconc."
            err += 1
        elif abs(delta) > 0.02:
            st = "revisar"
            adv += 1
        else:
            st = "OK"
        lineas.append(
            f"  {ds:12} {s_raw:10.2f} {tjf:10.2f} {s_out:10.2f} {delta:+8.2f}  {st}"
        )
    for ln in lineas:
        print(ln)
    if err:
        print(
            f"  [validación] {err} día(s) con suma reconciliada ≠ totalM3 (revisar código o API)."
        )
    elif adv and not err:
        print(
            f"  [validación] {adv} aviso(s): sin total JSON o Δ pequeño; revisar tabla arriba."
        )
    else:
        print("  [validación] Todas las sumas post-reconciliación coinciden con totalM3 (±0,1 m³).")
    return adv, err


def _escribir_hoja_filas(
    ws_f,
    todas_las_filas: Sequence[Tuple[str, str, str, str, int, float]],
) -> None:
    hdr = (
        "Archivo CSV",
        "TIME (UTC en CSV)",
        "Día civil Chile",
        "Hora Chile",
        "Hora 0-23 Chile",
        "Consumo m3/h (VALUE)",
    )
    for c, h in enumerate(hdr, start=1):
        cell = ws_f.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
    for r, row in enumerate(todas_las_filas, start=2):
        arch, tutc, fch, hlab, hint, val = row
        ws_f.cell(row=r, column=1, value=arch)
        ws_f.cell(row=r, column=2, value=tutc)
        ws_f.cell(row=r, column=3, value=fch)
        ws_f.cell(row=r, column=4, value=hlab)
        ws_f.cell(row=r, column=5, value=hint)
        ws_f.cell(row=r, column=6, value=round(val, 6))
        ws_f.cell(row=r, column=6).number_format = "0.000000"
    for col in range(1, 7):
        ws_f.column_dimensions[get_column_letter(col)].width = 14
    ws_f.freeze_panes = "A2"


def generar_excel_toda_la_carpeta(
    out_xlsx: Path,
    node_id: str,
    csv_dir: Path,
) -> int:
    """
    Parsea todos los ``YYYY-MM-DD.csv`` en ``csv_dir`` (hora Chile, misma regla que la app) y escribe
    Filas + Consolidado en una sola matriz (una columna por dia, orden cronologico).
    Devuelve cantidad de dias procesados.
    """
    csv_dir = csv_dir.resolve()
    if not csv_dir.is_dir():
        raise NotADirectoryError(str(csv_dir))
    items = _listar_csv_por_dia_ordenados(csv_dir)
    if not items:
        raise FileNotFoundError(
            f"No hay archivos con nombre YYYY-MM-DD.csv en {csv_dir}"
        )

    mats: List[List[float]] = []
    todas_las_filas: List[Tuple[str, str, str, str, int, float]] = []
    dias_ordenados: List[date] = []

    for d, p in items:
        texto = p.read_text(encoding="utf-8")
        vec, filas = _vector_y_filas_por_dia_en_texto(texto, d, p.name, node_id)
        mats.append(vec)
        todas_las_filas.extend(filas)
        dias_ordenados.append(d)

    wb = Workbook()
    ws_f = wb.active
    ws_f.title = "Filas"
    _escribir_hoja_filas(ws_f, todas_las_filas)

    ws = wb.create_sheet("Consolidado", 1)
    n = len(dias_ordenados)
    ws["A1"] = (
        f"Revision: todos los CSV en carpeta ({n} dias). Hora Chile desde TIME UTC. "
        f"Nodo {node_id}. Origen: {csv_dir}"
    )
    ws["A1"].font = Font(italic=True, size=9)

    c0 = 2
    ws["A2"] = "Hora"
    ws["A2"].font = Font(bold=True)
    if n > 0:
        ws.merge_cells(start_row=2, start_column=c0, end_row=2, end_column=c0 + n - 1)
        ws.cell(row=2, column=c0, value=f"Consumo m3/h por dia ({n} columnas)")
        ws.cell(row=2, column=c0).font = Font(bold=True)
        ws.cell(row=2, column=c0).alignment = Alignment(horizontal="center")

    r_head = 3
    for i, d in enumerate(dias_ordenados):
        col = c0 + i
        ws.cell(row=r_head, column=col, value=d.strftime("%d-%m-%Y"))
        ws.cell(row=r_head, column=col).font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col)].width = 11

    r0 = 4
    for h in range(24):
        row = r0 + h
        ws.cell(row=row, column=1, value=f"{h:02d}:00")
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        for i in range(n):
            val = mats[i][h]
            cell = ws.cell(row=row, column=c0 + i, value=round(val, 4))
            cell.number_format = "0.0000"

    ws.column_dimensions["A"].width = 10
    ws.freeze_panes = "B4"
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
    return n


def main() -> int:
    p = argparse.ArgumentParser(
        description="Excel: hoja Filas (parseo TIME,VALUE) + Consolidado (matriz)."
    )
    p.add_argument(
        "--toda-la-carpeta",
        action="store_true",
        help="Procesar todos los YYYY-MM-DD.csv en la carpeta (ignora --p1/--p2). Solo lectura.",
    )
    p.add_argument(
        "--solo-desde-csv",
        action="store_true",
        help="No descargar: leer YYYY-MM-DD.csv desde csv_descarga_api (o --csv-dir).",
    )
    p.add_argument("--csv-dir", type=Path, default=None, help="Carpeta de CSV por dia.")
    p.add_argument("--node-id", default="000017-08", help="Nodo WES (Auditoria ICCO abril)")
    p.add_argument("-o", "--output", type=Path, default=None, help="Ruta del .xlsx de salida.")
    p.add_argument("--p1-desde", type=lambda s: date.fromisoformat(s), default=date(2026, 3, 23))
    p.add_argument("--p1-hasta", type=lambda s: date.fromisoformat(s), default=date(2026, 3, 29))
    p.add_argument("--p2-desde", type=lambda s: date.fromisoformat(s), default=date(2026, 4, 6))
    p.add_argument("--p2-hasta", type=lambda s: date.fromisoformat(s), default=date(2026, 4, 12))
    args = p.parse_args()

    if args.toda_la_carpeta:
        csv_dir = (
            args.csv_dir.resolve()
            if args.csv_dir
            else (AUDIT_DIR_DEFAULT / "csv_descarga_api").resolve()
        )
        out = args.output or (
            AUDIT_DIR_DEFAULT / "consolidado_revision_todos_los_csv_descarga_api.xlsx"
        )
        out = out.resolve()
        n = generar_excel_toda_la_carpeta(out, args.node_id, csv_dir)
        print(out)
        print(f"Carpeta CSV: {csv_dir}")
        print(f"Dias incluidos: {n}")
        return 0

    d1a, d1b = args.p1_desde, args.p1_hasta
    d2a, d2b = args.p2_desde, args.p2_hasta
    dias1 = list(_iter_days(d1a, d1b))
    dias2 = list(_iter_days(d2a, d2b))
    if len(dias1) != 7 or len(dias2) != 7:
        raise SystemExit("Cada periodo debe tener exactamente 7 dias.")

    out = args.output or (
        AUDIT_DIR_DEFAULT
        / "consumo_consolidado_parseo_filas_abr06-12_abr13-19_2026.xlsx"
    )
    out = out.resolve()

    csv_dir = args.csv_dir.resolve() if args.csv_dir else None
    generar_excel_consolidado(
        out,
        args.node_id,
        dias1,
        dias2,
        titulo_p1=f"Con WES: {d1a:%d-%m-%Y} al {d1b:%d-%m-%Y}",
        titulo_p2=f"Sin WES: {d2a:%d-%m-%Y} al {d2b:%d-%m-%Y}",
        solo_desde_csv=args.solo_desde_csv,
        csv_dir=csv_dir,
    )
    print(out)
    base_csv = csv_dir or (out.parent / "csv_descarga_api")
    if args.solo_desde_csv:
        print(f"CSV leidos desde: {base_csv.resolve()}")
    else:
        print(f"CSV guardados en: {base_csv.resolve()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

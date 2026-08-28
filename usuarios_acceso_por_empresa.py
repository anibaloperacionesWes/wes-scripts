# -*- coding: utf-8 -*-
"""
Excel de accesos WES: una hoja por empresa con quienes tienen allowedNodes
en puntos activos de esa empresa.

Reutiliza la misma muestra de correos que contar_usuarios_por_nodo.py
(repo + contactos + receivers FILTRATION).

Uso:
  python usuarios_acceso_por_empresa.py
"""

from __future__ import annotations

import argparse
import re
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Set, Tuple
from zoneinfo import ZoneInfo

from contar_usuarios_por_nodo import (
    _allowed_nodes,
    _emails_desde_alertas,
    _fetch_user,
    _nodos_y_empresas,
    _recolectar_emails,
    _session,
    es_nodo_activo,
    nombre_usuario,
)

ROOT = Path(__file__).resolve().parent
OUT_DIR = ROOT / "reports" / "Usuarios" / "acceso_por_empresa"
TZ_CL = ZoneInfo("America/Santiago")
SHEET_BAD = re.compile(r"[\\/*?:\[\]]+")
NOMBRE_EMPRESA_FALLBACK = {
    "000026": "UDD",
}

# Fuera de este reporte (pedido operativo).
EMPRESAS_EXCLUIDAS = {
    "000010",  # Corporación Puente Alto
    "000028",  # La Florida
}

# Personal WES (compañía 000000 o @wes.cl) se omite, salvo esta excepción.
PERSONAL_WES_EXCEPCIONES = {"go.salass@gmail.com"}
WES_COMPANY_ID = "000000"
DOMINIOS_PERSONALES = {
    "gmail.com",
    "hotmail.com",
    "hotmail.es",
    "outlook.com",
    "outlook.es",
    "yahoo.com",
    "yahoo.es",
    "live.com",
    "icloud.com",
    "wes.cl",
}
DOMINIOS_PA = ("parauco.com", "linkes.cl", "externos.parauco.com")

# Cuentas cliente que no aparecen en el repo (p. ej. Tamara tiene también @parauco.com).
EMAILS_CLIENTES_EXTRA = {
    "tamara.martinez@parauco.com",
    "tamara.martinez.a@linkes.cl",
    "tmartinez@linkes.cl",
}


def _nombre_hoja(cid: str, empresa: str, usados: Set[str]) -> str:
    base = SHEET_BAD.sub(" ", f"{cid} {empresa}").strip()
    base = re.sub(r"\s+", " ", base)[:31] or cid
    name = base
    n = 2
    while name.casefold() in usados:
        suffix = f"_{n}"
        name = (base[: 31 - len(suffix)] + suffix)[:31]
        n += 1
    usados.add(name.casefold())
    return name


def _slug_nombre(texto: str) -> str:
    import unicodedata

    t = unicodedata.normalize("NFKD", texto or "")
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    t = t.lower().strip()
    t = re.sub(r"[^a-z0-9]+", " ", t).strip()
    return t.split()[0] if t else ""


def es_personal_wes(user: dict) -> bool:
    """Solo @wes.cl. companyId 000000 no basta: hay clientes (p. ej. tsaba@parauco.com) ahí."""
    email = str(user.get("username") or "").strip().lower()
    if email in PERSONAL_WES_EXCEPCIONES:
        return False
    return email.endswith("@wes.cl")


def _dominios_clientes_por_empresa(usuarios: Dict[str, dict]) -> Dict[str, Set[str]]:
    dominios: Dict[str, Set[str]] = defaultdict(set)
    for email, user in usuarios.items():
        cid = str(user.get("companyId") or "").strip()
        dom = str(email).partition("@")[2].lower()
        if cid and cid != WES_COMPANY_ID and dom and dom not in DOMINIOS_PERSONALES:
            dominios[cid].add(dom)
    dominios["000025"].update(DOMINIOS_PA)
    return dominios


def _emails_desde_nombres(usuarios: Dict[str, dict]) -> Set[str]:
    """Arma correos alternativos: mismo local en otros dominios de la empresa + first.last."""
    found: Set[str] = set()
    dominios_cia: Dict[str, Set[str]] = defaultdict(set)
    for email, user in usuarios.items():
        cid = str(user.get("companyId") or "").strip()
        dom = str(email).partition("@")[2].lower()
        if cid and dom and dom not in DOMINIOS_PERSONALES:
            dominios_cia[cid].add(dom)
    dominios_cia["000025"].update(DOMINIOS_PA)

    for email, user in usuarios.items():
        cid = str(user.get("companyId") or "").strip()
        dominios = set(dominios_cia.get(cid, set()))
        dom_actual = str(email).partition("@")[2].lower()
        if dom_actual in DOMINIOS_PA:
            dominios.update(DOMINIOS_PA)
        if not dominios:
            continue
        local = str(email).partition("@")[0].strip().lower()
        locales = {local} if local else set()
        first = _slug_nombre(str(user.get("name") or ""))
        last = _slug_nombre(str(user.get("lastName") or ""))
        if first and last and len(first) >= 2 and len(last) >= 2:
            locales.add(f"{first}.{last}")
            locales.add(f"{first[0]}{last}")
        for d in dominios:
            for loc in locales:
                if loc and "." in d:
                    found.add(f"{loc}@{d}")
    return found


def _cargar_usuarios(extra_email_files: List[Path], workers: int):
    session = _session()
    print("[INFO] Obteniendo empresas y nodos...")
    nombres, company_de, companies = _nodos_y_empresas(session)
    for cid, name in list(companies.items()):
        if not name or name.strip() == cid:
            companies[cid] = NOMBRE_EMPRESA_FALLBACK.get(cid, name or cid)
    print(f"[OK] {len(companies)} empresas · {len(nombres)} nodeId")

    emails = _recolectar_emails(extra_email_files)
    emails = sorted(set(emails) | {e.lower() for e in EMAILS_CLIENTES_EXTRA})
    extra = _emails_desde_alertas(session, nombres.keys(), workers=workers)
    if extra:
        n0 = len(emails)
        emails = sorted(set(emails) | extra)
        print(f"[INFO] Correos extra por alertas: {len(emails) - n0} (total {len(emails)})")
    print(f"[INFO] Consultando {len(emails)} correo(s)...")

    usuarios: Dict[str, dict] = {}

    def _ingestar(cands: List[str]) -> None:
        with ThreadPoolExecutor(max_workers=workers) as ex:
            futs = [ex.submit(_fetch_user, session, e) for e in cands]
            for fut in as_completed(futs):
                email, user, _obs = fut.result()
                if not user:
                    continue
                email_api = str(user.get("username") or email).strip().lower()
                if email_api in usuarios:
                    continue
                usuarios[email_api] = user
                print(f"[OK] {email_api}")

    conocidos = {e.lower() for e in emails}
    _ingestar(emails)
    extra_nombres = _emails_desde_nombres(usuarios) - {e.lower() for e in usuarios} - conocidos
    if extra_nombres:
        print(f"[INFO] Probando {len(extra_nombres)} correo(s) alternativos de cliente...")
        antes = set(usuarios)
        _ingestar(sorted(extra_nombres))
        nuevos = sorted(set(usuarios) - antes)
        if nuevos:
            print(f"[INFO] Cuentas cliente extra encontradas: {', '.join(nuevos)}")

    n_wes = sum(1 for u in usuarios.values() if es_personal_wes(u))
    print(f"[INFO] Usuarios API: {len(usuarios)} · personal WES a omitir: {n_wes}")
    return nombres, company_de, companies, usuarios


def _accesos_por_empresa(
    nombres: Dict[str, str],
    company_de: Dict[str, str],
    companies: Dict[str, str],
    usuarios: Dict[str, dict],
    solo_activos: bool,
) -> Tuple[Dict[str, List[str]], Dict[str, List[dict]]]:
    nodos_empresa: Dict[str, List[str]] = defaultdict(list)
    for nid, cid in company_de.items():
        if cid in EMPRESAS_EXCLUIDAS:
            continue
        cname = companies.get(cid, "")
        if solo_activos and not es_nodo_activo(nid, cid, cname):
            continue
        nodos_empresa[cid].append(nid)
    for cid in nodos_empresa:
        nodos_empresa[cid] = sorted(set(nodos_empresa[cid]))

    dominios_cia = _dominios_clientes_por_empresa(usuarios)
    filas: Dict[str, List[dict]] = defaultdict(list)
    for email, user in usuarios.items():
        if es_personal_wes(user):
            continue
        nombre = nombre_usuario(user)
        cid_user = str(user.get("companyId") or "").strip()
        switch = user.get("switchEnabled")
        dom = str(email).partition("@")[2].lower()
        por_cia: Dict[str, List[str]] = defaultdict(list)
        for nid in _allowed_nodes(user):
            cid = company_de.get(nid, "")
            if not cid or cid in EMPRESAS_EXCLUIDAS:
                continue
            cname = companies.get(cid, "")
            if solo_activos and not es_nodo_activo(nid, cid, cname):
                continue
            por_cia[cid].append(nid)
        for cid, nids in por_cia.items():
            nids = sorted(set(nids))
            dominio_de_esta = dom in dominios_cia.get(cid, set())
            es_propia = cid_user == cid or (
                cid_user == WES_COMPANY_ID and cid != WES_COMPANY_ID and dominio_de_esta
            )
            filas[cid].append(
                {
                    "email": email,
                    "nombre": nombre,
                    "company_id_cuenta": cid_user,
                    "empresa_cuenta": companies.get(cid if es_propia else cid_user, cid_user),
                    "switch_on_off": "Sí" if switch else "No",
                    "nodos_count": len(nids),
                    "node_ids": ", ".join(nids),
                    "nombres_puntos": "; ".join(
                        f"{nid} ({nombres.get(nid) or nid})" for nid in nids
                    ),
                    "es_cuenta_de_esta_empresa": "Sí" if es_propia else "No",
                }
            )
    for cid in filas:
        filas[cid].sort(
            key=lambda r: (
                0 if r["es_cuenta_de_esta_empresa"] == "Sí" else 1,
                r["email"].casefold(),
            )
        )
    return dict(nodos_empresa), dict(filas)


def _escribir_xlsx(
    path: Path,
    companies: Dict[str, str],
    nodos_empresa: Dict[str, List[str]],
    filas: Dict[str, List[dict]],
    generado: str,
    nombres: Dict[str, str],
    ids_cpa: Set[str],
    cpa_detalle: List[Tuple[str, str, str]],
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from puntos_cpa import clasificar_nodos
    from gabinetes import cargar_gabinetes, conteo_gabinetes, texto_gabinetes_numerados

    gab_de, gab_miembros = cargar_gabinetes()

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    cpa_fill = PatternFill("solid", fgColor="C6EFCE")
    mon_fill = PatternFill("solid", fgColor="FFF2CC")

    cols_res = [
        "company_id",
        "empresa",
        "puntos_activos",
        "gabinetes",
        "puntos_CPA",
        "puntos_solo_monitoreo",
        "usuarios_con_acceso",
        "ids_CPA",
        "ids_solo_monitoreo",
        "mismo_gabinete",
        "hoja",
    ]
    usados: Set[str] = {"resumen", "puntos", "cpa control"}
    hojas: List[Tuple[str, str, str]] = []  # cid, empresa, sheet name
    for cid, empresa in sorted(companies.items(), key=lambda x: (x[1].casefold(), x[0])):
        if cid in EMPRESAS_EXCLUIDAS or cid not in nodos_empresa:
            continue
        sheet = _nombre_hoja(cid, empresa, usados)
        hojas.append((cid, empresa, sheet))

    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = "Accesos WES por empresa (puntos activos, sin personal WES)"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79")
    ws["A2"] = (
        f"Generado {generado} hora Chile · se omite personal WES (@wes.cl); "
        "se incluye go.salass@gmail.com. "
        "CPA = monitoreo con control (horarios programados Drive). "
        "Monitoreo = punto activo sin CPA. "
        "Gabinetes: 1 estos puntos, 2 estos puntos (mismo gabinete cuenta como 1)."
    )
    ws.merge_cells("A1:K1")
    ws.merge_cells("A2:K2")
    for c, h in enumerate(cols_res, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
    for i, (cid, empresa, sheet) in enumerate(hojas, 5):
        nids = nodos_empresa.get(cid, [])
        cpa, solo = clasificar_nodos(nids, ids_cpa)
        ws.cell(row=i, column=1, value=cid)
        ws.cell(row=i, column=2, value=empresa)
        ws.cell(row=i, column=3, value=len(nids))
        ws.cell(row=i, column=4, value=conteo_gabinetes(nids, gab_de))
        ws.cell(row=i, column=5, value=len(cpa))
        ws.cell(row=i, column=6, value=len(solo))
        ws.cell(row=i, column=7, value=len(filas.get(cid, [])))
        ws.cell(row=i, column=8, value=", ".join(cpa))
        ws.cell(row=i, column=9, value=", ".join(solo))
        ws.cell(row=i, column=10, value=texto_gabinetes_numerados(nids, gab_de, gab_miembros) or "—")
        ws.cell(row=i, column=11, value=sheet)
        ws.cell(row=i, column=8).alignment = wrap
        ws.cell(row=i, column=9).alignment = wrap
        ws.cell(row=i, column=10).alignment = wrap
    for i, w in enumerate((12, 28, 14, 12, 12, 20, 18, 36, 42, 48, 26), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:K{4 + max(1, len(hojas))}"

    tit_pts = [
        "node_id",
        "nombre",
        "company_id",
        "empresa",
        "tipo",
        "nombre_en_excel_horarios",
    ]
    titulo_por_nid: Dict[str, str] = {}
    for titulo, nid, _nom in cpa_detalle:
        if nid:
            titulo_por_nid.setdefault(nid, titulo)
    filas_cpa: List[List[str]] = []
    filas_pts: List[Tuple[List[str], bool]] = []
    for cid, empresa, _sheet in hojas:
        for nid in nodos_empresa.get(cid, []):
            es_cpa = nid in ids_cpa
            vals = [
                nid,
                nombres.get(nid, nid),
                cid,
                empresa,
                "CPA (monitoreo con control)" if es_cpa else "Monitoreo",
                titulo_por_nid.get(nid, ""),
            ]
            filas_pts.append((vals, es_cpa))
            if es_cpa:
                filas_cpa.append(vals)

    header_cpa_fill = PatternFill("solid", fgColor="375623")
    wscpa = wb.create_sheet("CPA control", 1)
    wscpa["A1"] = "Puntos CPA control (monitoreo con control)"
    wscpa["A1"].font = Font(bold=True, size=14, color="375623")
    wscpa.merge_cells("A1:F1")
    wscpa["A2"] = (
        "Hoja 2 · sitios con horario programado en "
        "«Horarios de contron hidrico clientes wes» (Drive)."
    )
    wscpa.merge_cells("A2:F2")
    for c, h in enumerate(tit_pts, 1):
        cell = wscpa.cell(row=4, column=c, value=h)
        cell.fill = header_cpa_fill
        cell.font = header_font
    for r_i, vals in enumerate(filas_cpa, 5):
        for c, v in enumerate(vals, 1):
            cell = wscpa.cell(row=r_i, column=c, value=v)
            cell.fill = cpa_fill
    for i, w in enumerate((14, 40, 12, 28, 28, 40), 1):
        wscpa.column_dimensions[get_column_letter(i)].width = w
    wscpa.freeze_panes = "A5"
    wscpa.auto_filter.ref = f"A4:F{max(5, 4 + len(filas_cpa))}"

    wsp = wb.create_sheet("Puntos", 2)
    wsp["A1"] = "Puntos activos: CPA (monitoreo con control) vs monitoreo"
    wsp["A1"].font = Font(bold=True, size=14, color="1F4E79")
    wsp.merge_cells("A1:F1")
    wsp["A2"] = (
        "CPA toma los sitios con horario programado en "
        "«Horarios de contron hidrico clientes wes» (Drive)."
    )
    wsp.merge_cells("A2:F2")
    for c, h in enumerate(tit_pts, 1):
        cell = wsp.cell(row=4, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
    r_i = 5
    for vals, es_cpa in filas_pts:
        fill = cpa_fill if es_cpa else mon_fill
        for c, v in enumerate(vals, 1):
            cell = wsp.cell(row=r_i, column=c, value=v)
            if c == 5:
                cell.fill = fill
        r_i += 1
    for i, w in enumerate((14, 40, 12, 28, 28, 40), 1):
        wsp.column_dimensions[get_column_letter(i)].width = w
    wsp.freeze_panes = "A5"
    wsp.auto_filter.ref = f"A4:F{max(5, r_i - 1)}"

    cols = [
        "email",
        "nombre",
        "es_cuenta_de_esta_empresa",
        "empresa_cuenta",
        "switch_on_off",
        "nodos_count",
        "node_ids",
        "nombres_puntos",
    ]
    titulos = [
        "Email",
        "Nombre",
        "Cuenta de esta empresa",
        "Empresa de la cuenta",
        "Switch on/off",
        "Puntos con acceso",
        "Node IDs",
        "Puntos (nombre)",
    ]
    widths = (34, 28, 22, 28, 14, 16, 40, 70)

    for cid, empresa, sheet in hojas:
        wsc = wb.create_sheet(sheet)
        wsc["A1"] = f"{empresa} ({cid})"
        wsc["A1"].font = Font(bold=True, size=14, color="1F4E79")
        wsc.merge_cells("A1:H1")
        nids = nodos_empresa.get(cid, [])
        cpa, solo = clasificar_nodos(nids, ids_cpa)
        n_usr = len(filas.get(cid, []))
        wsc["A2"] = (
            f"{n_usr} usuario(s) · {len(nids)} punto(s) activo(s) · "
            f"{len(cpa)} CPA · {len(solo)} monitoreo"
        )
        wsc.merge_cells("A2:H2")
        wsc["A3"] = (
            ("CPA: " + (", ".join(cpa) if cpa else "—"))
            + "  |  Monitoreo: "
            + (", ".join(solo) if solo else "—")
        )
        wsc.merge_cells("A3:H3")
        wsc["A3"].alignment = wrap
        for c, h in enumerate(titulos, 1):
            cell = wsc.cell(row=4, column=c, value=h)
            cell.fill = header_fill
            cell.font = header_font
        rows = filas.get(cid, [])
        if not rows:
            wsc.cell(row=5, column=1, value="Sin usuarios en la muestra consultada")
        for rr, row in enumerate(rows, 5):
            for c, key in enumerate(cols, 1):
                cell = wsc.cell(row=rr, column=c, value=row.get(key, ""))
                cell.alignment = wrap
        for i, w in enumerate(widths, 1):
            wsc.column_dimensions[get_column_letter(i)].width = w
        wsc.freeze_panes = "A5"
        last = 4 + max(1, len(rows))
        wsc.auto_filter.ref = f"A4:H{last}"
        wsc.row_dimensions[3].height = 28
        wsc.row_dimensions[4].height = 18

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def resumen_a_pdf(xlsx_path: Path, pdf_path: Path | None = None) -> Path:
    """PDF de 2 hojas: tabla blanca (resumen) en hoja 1 y tabla verde (CPA) en hoja 2."""
    from openpyxl import load_workbook
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        PageBreak,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    xlsx_path = Path(xlsx_path)
    if pdf_path is None:
        pdf_path = xlsx_path.with_name("accesos_por_empresa_resumen.pdf")
    else:
        pdf_path = Path(pdf_path)

    wb = load_workbook(xlsx_path, data_only=True)
    if "Resumen" not in wb.sheetnames:
        raise SystemExit(f"No hay hoja Resumen en {xlsx_path}")
    ws = wb["Resumen"]
    titulo = str(ws["A1"].value or "Accesos WES por empresa").strip()

    filas_excel = list(ws.iter_rows(min_row=1, max_col=12, values_only=True))
    header_i = None
    for i, row in enumerate(filas_excel):
        first = str(row[0] or "").strip().lower()
        if first == "company_id":
            header_i = i
            break
    if header_i is None:
        raise SystemExit("No se encontró el encabezado de Resumen")

    headers = [str(h or "").strip() for h in filas_excel[header_i]]
    idx = {h: i for i, h in enumerate(headers) if h}

    def _col(row, name, default=0):
        i = idx.get(name)
        if i is None or i >= len(row) or row[i] is None:
            return default
        return row[i]

    from gabinetes import cargar_gabinetes, conteo_gabinetes, texto_gabinetes_numerados

    gab_de, gab_miembros = cargar_gabinetes()
    nodos_por_cia: Dict[str, List[str]] = defaultdict(list)
    if "Puntos" in wb.sheetnames:
        for prow in wb["Puntos"].iter_rows(min_row=5, max_col=4, values_only=True):
            if not prow or not prow[0] or not prow[2]:
                continue
            cid_pt = str(prow[2]).strip()
            if cid_pt in EMPRESAS_EXCLUIDAS:
                continue
            nodos_por_cia[cid_pt].append(str(prow[0]).strip())

    tiene_cpa = "puntos_CPA" in idx
    if tiene_cpa:
        titulos_pdf = [
            "Company ID",
            "Empresa",
            "Activos",
            "Gabinetes",
            "CPA",
            "Monitoreo",
            "Usuarios",
        ]
    else:
        titulos_pdf = ["Company ID", "Empresa", "Puntos activos", "Usuarios con acceso"]
    data_rows = []
    tot_pts = tot_cpa = tot_solo = tot_usr = tot_gab = 0
    for row in filas_excel[header_i + 1 :]:
        if not row or row[0] is None or str(row[0]).strip() == "":
            continue
        cid = str(_col(row, "company_id", "")).strip()
        empresa = str(_col(row, "empresa", "")).strip()
        if cid in EMPRESAS_EXCLUIDAS or "florida" in empresa.casefold():
            continue
        pts = int(_col(row, "puntos_activos", 0) or 0)
        usr = int(_col(row, "usuarios_con_acceso", 0) or 0)
        tot_pts += pts
        tot_usr += usr
        nids = nodos_por_cia.get(cid, [])
        n_gab = conteo_gabinetes(nids, gab_de) if nids else pts
        celda_gab = texto_gabinetes_numerados(nids, gab_de, gab_miembros) or str(n_gab)
        tot_gab += n_gab
        if tiene_cpa:
            n_cpa = int(_col(row, "puntos_CPA", 0) or 0)
            n_solo = int(_col(row, "puntos_solo_monitoreo", 0) or 0)
            tot_cpa += n_cpa
            tot_solo += n_solo
            data_rows.append(
                [cid, empresa, str(pts), celda_gab, str(n_cpa), str(n_solo), str(usr)]
            )
        else:
            data_rows.append([cid, empresa, str(pts), str(usr)])

    cpa_pts = []
    hoja_cpa = "CPA control" if "CPA control" in wb.sheetnames else "Puntos"
    if hoja_cpa in wb.sheetnames:
        for row in wb[hoja_cpa].iter_rows(min_row=5, max_col=6, values_only=True):
            if not row or not row[0]:
                continue
            nid = str(row[0])
            cid_pt = nid.split("-")[0] if "-" in nid else ""
            if cid_pt in EMPRESAS_EXCLUIDAS:
                continue
            tipo = str(row[4] or "")
            if hoja_cpa == "CPA control" or "CPA" in tipo:
                cpa_pts.append(
                    [
                        str(row[0]),
                        str(row[1] or ""),
                        str(row[3] or ""),
                        str(row[5] or ""),
                    ]
                )

    generado = datetime.now(TZ_CL).strftime("%d/%m/%Y %H:%M")
    nota = (
        f"Generado {generado} hora Chile · puntos activos · sin personal WES (@wes.cl); "
        "se incluye go.salass@gmail.com. "
        "CPA = monitoreo con control (horarios programados). "
        "Monitoreo = sin CPA. "
        "Gabinetes: 1 estos puntos, 2 estos puntos (mismo gabinete = 1)."
    )

    navy = colors.HexColor("#1F4E79")
    green_hdr = colors.HexColor("#375623")
    green_body = colors.HexColor("#C6EFCE")
    title_style = ParagraphStyle(
        "titulo",
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=18,
        textColor=navy,
        alignment=TA_LEFT,
    )
    sub_style = ParagraphStyle(
        "sub",
        fontName="Helvetica",
        fontSize=8,
        leading=11,
        textColor=colors.HexColor("#555555"),
    )
    head_style = ParagraphStyle(
        "head",
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        textColor=colors.white,
        alignment=TA_CENTER,
    )
    cell_l = ParagraphStyle(
        "cell_l", fontName="Helvetica", fontSize=8, leading=11, alignment=TA_LEFT
    )
    cell_gab = ParagraphStyle(
        "cell_gab", fontName="Helvetica", fontSize=6.5, leading=8, alignment=TA_LEFT
    )
    cell_c = ParagraphStyle(
        "cell_c", fontName="Helvetica", fontSize=8, leading=11, alignment=TA_CENTER
    )
    tot_style = ParagraphStyle(
        "tot",
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=11,
        alignment=TA_LEFT,
        textColor=navy,
    )
    tot_c = ParagraphStyle(
        "tot_c",
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=11,
        alignment=TA_CENTER,
        textColor=navy,
    )

    def _p(txt, style):
        t = (
            str(txt)
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )
        t = t.replace("\n", "<br/>")
        return Paragraph(t, style)

    table_data = [[_p(h, head_style) for h in titulos_pdf]]
    for vals in data_rows:
        row_cells = []
        for i, v in enumerate(vals):
            if tiene_cpa and i == 3:
                st = cell_gab
            elif i == 1:
                st = cell_l
            else:
                st = cell_c
            row_cells.append(_p(v, st))
        table_data.append(row_cells)
    if tiene_cpa:
        table_data.append(
            [
                _p("", tot_c),
                _p(f"Total ({len(data_rows)} empresas)", tot_style),
                _p(str(tot_pts), tot_c),
                _p(str(tot_gab), tot_c),
                _p(str(tot_cpa), tot_c),
                _p(str(tot_solo), tot_c),
                _p(str(tot_usr), tot_c),
            ]
        )
        weights = (0.11, 0.20, 0.09, 0.24, 0.09, 0.14, 0.13)
    else:
        table_data.append(
            [
                _p("", tot_c),
                _p(f"Total ({len(data_rows)} empresas)", tot_style),
                _p(str(tot_pts), tot_c),
                _p(str(tot_usr), tot_c),
            ]
        )
        weights = (0.18, 0.42, 0.20, 0.20)

    page = landscape(A4) if tiene_cpa else A4
    margin = 12 * mm
    usable = page[0] - 2 * margin
    col_w = [usable * w for w in weights]
    cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), navy),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
        ("VALIGN", (0, 0), (-1, 0), "MIDDLE"),
        ("VALIGN", (0, 1), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#B0B0B0")),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]

    story = [
        _p("Hoja 1 · " + titulo, title_style),
        Spacer(1, 2 * mm),
        _p(nota, sub_style),
        Spacer(1, 5 * mm),
        Table(table_data, colWidths=col_w, repeatRows=1),
    ]
    story[-1].setStyle(TableStyle(cmds))

    title_cpa = ParagraphStyle(
        "titulo_cpa",
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=18,
        textColor=green_hdr,
        alignment=TA_LEFT,
    )
    if cpa_pts:
        story.append(PageBreak())
        story.append(_p("Hoja 2 · Puntos CPA control (monitoreo con control)", title_cpa))
        story.append(Spacer(1, 2 * mm))
        story.append(
            _p(
                "Tabla verde · sitios con horario programado en el Excel de Drive "
                "«Horarios de contron hidrico clientes wes».",
                sub_style,
            )
        )
        story.append(Spacer(1, 4 * mm))
        cpa_cell_l = ParagraphStyle(
            "cpa_l", fontName="Helvetica", fontSize=7, leading=9, alignment=TA_LEFT
        )
        cpa_cell_c = ParagraphStyle(
            "cpa_c", fontName="Helvetica", fontSize=7, leading=9, alignment=TA_CENTER
        )
        cpa_headers = ["Node ID", "Punto", "Empresa", "Nombre en Excel de horarios"]
        cpa_data = [[_p(h, head_style) for h in cpa_headers]]
        for vals in cpa_pts:
            cpa_data.append(
                [
                    _p(vals[0], cpa_cell_c),
                    _p(vals[1], cpa_cell_l),
                    _p(vals[2], cpa_cell_l),
                    _p(vals[3], cpa_cell_l),
                ]
            )
        cpa_w = [usable * w for w in (0.14, 0.34, 0.22, 0.30)]
        cpa_table = Table(cpa_data, colWidths=cpa_w, repeatRows=1)
        cpa_cmds = [
            ("BACKGROUND", (0, 0), (-1, 0), green_hdr),
            ("BACKGROUND", (0, 1), (-1, -1), green_body),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#A9D08E")),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]
        cpa_table.setStyle(TableStyle(cpa_cmds))
        story.append(cpa_table)

    def _folio(canvas, doc):
        canvas.saveState()
        n = canvas.getPageNumber()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#555555"))
        # Página 1 = resumen blanco; el resto = tabla verde CPA.
        label = (
            "Hoja 1 · tabla blanca (resumen)"
            if n == 1
            else "Hoja 2 · tabla verde (CPA control)"
        )
        canvas.drawString(margin, 8 * mm, label)
        canvas.drawRightString(page[0] - margin, 8 * mm, str(n))
        canvas.restoreState()

    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=page,
        leftMargin=margin,
        rightMargin=margin,
        topMargin=margin,
        bottomMargin=14 * mm,
        title=titulo,
        author="WES",
    )
    doc.build(story, onFirstPage=_folio, onLaterPages=_folio)
    return pdf_path


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Excel: una hoja por empresa con usuarios que tienen acceso"
    )
    parser.add_argument("--emails", type=Path, action="append", default=[])
    parser.add_argument("--salida", type=Path, default=OUT_DIR)
    parser.add_argument("--workers", type=int, default=10)
    parser.add_argument(
        "--incluir-inactivos",
        action="store_true",
        help="Incluir puntos dados de baja / fuera de operación",
    )
    parser.add_argument(
        "--pdf-resumen",
        type=Path,
        nargs="?",
        const=OUT_DIR / "accesos_por_empresa.xlsx",
        default=None,
        help="Solo generar PDF de la hoja Resumen (Excel existente, sin consultar API)",
    )
    args = parser.parse_args()

    if args.pdf_resumen is not None:
        src = args.pdf_resumen
        if not src.is_file():
            print(f"[ERROR] No existe el Excel: {src}")
            return 1
        pdf = resumen_a_pdf(src)
        latest_pdf = src.with_name("accesos_por_empresa_resumen.pdf")
        if pdf.resolve() != latest_pdf.resolve():
            latest_pdf.write_bytes(pdf.read_bytes())
        print(f"[OK] PDF: {latest_pdf}")
        return 0

    nombres, company_de, companies, usuarios = _cargar_usuarios(
        args.emails, max(1, min(args.workers, 16))
    )
    nodos_empresa, filas = _accesos_por_empresa(
        nombres, company_de, companies, usuarios, solo_activos=not args.incluir_inactivos
    )

    args.salida.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(TZ_CL).strftime("%Y%m%d_%H%M")
    generado = datetime.now(TZ_CL).strftime("%d/%m/%Y %H:%M")
    xlsx = args.salida / f"accesos_por_empresa_{stamp}.xlsx"
    latest = args.salida / "accesos_por_empresa.xlsx"
    from puntos_cpa import cargar_ids_cpa

    ids_cpa, cpa_detalle = cargar_ids_cpa(nombres)
    _escribir_xlsx(
        xlsx, companies, nodos_empresa, filas, generado, nombres, ids_cpa, cpa_detalle
    )
    latest.write_bytes(xlsx.read_bytes())

    n_emp = sum(1 for cid in companies if cid in nodos_empresa)
    n_usr = len(usuarios)
    print(f"\n[OK] Empresas con puntos activos: {n_emp}")
    print(f"[OK] Usuarios encontrados: {n_usr}")
    for cid, empresa in sorted(companies.items(), key=lambda x: x[1].casefold()):
        if cid not in nodos_empresa:
            continue
        print(
            f"  {cid} {empresa}: {len(filas.get(cid, []))} usuario(s) · "
            f"{len(nodos_empresa[cid])} punto(s) · "
            f"{sum(1 for n in nodos_empresa[cid] if n in ids_cpa)} CPA"
        )
    print(f"[OK] XLSX: {xlsx}")
    print(f"[OK] XLSX: {latest}")
    pdf = resumen_a_pdf(latest, args.salida / f"accesos_por_empresa_resumen_{stamp}.pdf")
    pdf_latest = args.salida / "accesos_por_empresa_resumen.pdf"
    pdf_latest.write_bytes(pdf.read_bytes())
    print(f"[OK] PDF: {pdf}")
    print(f"[OK] PDF: {pdf_latest}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

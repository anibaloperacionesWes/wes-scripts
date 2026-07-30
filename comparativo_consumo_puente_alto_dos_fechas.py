"""
Compara el consumo diario entre dos fechas para cada colegio (nodo) de Puente Alto (API WES).

Interpretación: la API entrega caudal horario en m³/h (promedio de esa hora).
Volumen aproximado de esa hora ≈ m³/h × 1 h → en m³ numéricamente igual al valor horario.
Consumo del día (m³) ≈ suma de todos los valores horarios del día (0–23 h con dato).

Por defecto compara 18-03-2025 con 18-03-2026.

Uso:
  python comparativo_consumo_puente_alto_dos_fechas.py
  python comparativo_consumo_puente_alto_dos_fechas.py --fecha-a 2025-03-18 --fecha-b 2026-03-18 --escritorio
"""

from __future__ import annotations

import argparse
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Reutiliza API Puente Alto
from reporte_puente_alto_lxm import obtener_datos_horarios_dia, obtener_nodos_puente_alto


def consumo_diario_m3_aprox(hourly: Dict[int, float]) -> Tuple[float, int]:
    """
    Suma horaria como volumen aproximado del día (m³).
    Retorna (total_m3, cantidad_horas_con_dato).
    """
    if not hourly:
        return 0.0, 0
    total = sum(float(v) for v in hourly.values())
    return total, len(hourly)


def carpeta_salida(escritorio: bool) -> Path:
    if escritorio:
        for candidate in (
            Path.home() / "OneDrive" / "Desktop",
            Path.home() / "Desktop",
        ):
            if candidate.is_dir():
                sub = candidate / "informes_consumo_puente_alto"
                sub.mkdir(parents=True, exist_ok=True)
                return sub
    base = Path(__file__).resolve().parent / "reports" / "consumo_puente_alto_comparativo"
    base.mkdir(parents=True, exist_ok=True)
    return base


def generar_excel(
    filas: List[Dict[str, object]],
    fecha_a: datetime,
    fecha_b: datetime,
    out_path: Path,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparativo"

    titulo = (
        f"Consumo diario aproximado (m³) — Puente Alto | "
        f"{fecha_a:%d-%m-%Y} vs {fecha_b:%d-%m-%Y}"
    )
    ws.cell(row=1, column=1, value=titulo)
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)

    nota = (
        "Nota: consumo_m3 = suma de valores horarios m³/h del día (aprox. volumen m³ del día). "
        "Si faltan horas, el total puede ser menor al real."
    )
    ws.cell(row=2, column=1, value=nota)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)

    headers = [
        "N°",
        "Node ID",
        "Establecimiento",
        f"Consumo m³ ({fecha_a:%d-%m-%Y})",
        "Horas c/dato A",
        f"Consumo m³ ({fecha_b:%d-%m-%Y})",
        "Horas c/dato B",
        "Diferencia m³ (B − A)",
        "Variación %",
    ]
    hr = 4
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for i, row in enumerate(filas, start=1):
        r = hr + i
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=2, value=row["node_id"])
        ws.cell(row=r, column=3, value=row["nombre"])
        ws.cell(row=r, column=4, value=round(float(row["m3_a"]), 3))
        ws.cell(row=r, column=5, value=row["horas_a"])
        ws.cell(row=r, column=6, value=round(float(row["m3_b"]), 3))
        ws.cell(row=r, column=7, value=row["horas_b"])
        diff = float(row["diff_m3"])
        ws.cell(row=r, column=8, value=round(diff, 3))
        pct = row.get("pct")
        ws.cell(
            row=r,
            column=9,
            value=round(float(pct), 2) if pct is not None else "—",
        )
        for c in range(1, 10):
            ws.cell(row=r, column=c).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="left")

    ws.freeze_panes = f"A{hr+1}"
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["C"].width = 42

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Comparativo consumo diario Puente Alto entre dos fechas"
    )
    parser.add_argument(
        "--fecha-a",
        type=str,
        default="2025-03-18",
        help="Primera fecha YYYY-MM-DD (default: 2025-03-18)",
    )
    parser.add_argument(
        "--fecha-b",
        type=str,
        default="2026-03-18",
        help="Segunda fecha YYYY-MM-DD (default: 2026-03-18)",
    )
    parser.add_argument(
        "--escritorio",
        action="store_true",
        help="Guardar Excel en Escritorio/informes_consumo_puente_alto",
    )
    args = parser.parse_args()

    fecha_a = datetime.strptime(args.fecha_a.strip(), "%Y-%m-%d")
    fecha_b = datetime.strptime(args.fecha_b.strip(), "%Y-%m-%d")

    nodos = obtener_nodos_puente_alto()
    nodos.sort(key=lambda x: x["nodeName"])

    filas: List[Dict[str, object]] = []
    errores: List[str] = []

    for n in nodos:
        nid = n["nodeId"]
        nombre = n["nodeName"]
        try:
            h_a = obtener_datos_horarios_dia(nid, fecha_a)
        except Exception as e:
            errores.append(f"{nid} {nombre} fecha A: {e}")
            h_a = {}
        try:
            h_b = obtener_datos_horarios_dia(nid, fecha_b)
        except Exception as e:
            errores.append(f"{nid} {nombre} fecha B: {e}")
            h_b = {}

        m3_a, horas_a = consumo_diario_m3_aprox(h_a)
        m3_b, horas_b = consumo_diario_m3_aprox(h_b)
        diff = m3_b - m3_a
        if m3_a and m3_a > 1e-9:
            pct = (diff / m3_a) * 100.0
        else:
            pct = None

        filas.append(
            {
                "node_id": nid,
                "nombre": nombre,
                "m3_a": m3_a,
                "horas_a": horas_a,
                "m3_b": m3_b,
                "horas_b": horas_b,
                "diff_m3": diff,
                "pct": pct,
            }
        )

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    base = carpeta_salida(args.escritorio)
    fname = (
        f"comparativo_consumo_puente_alto_"
        f"{fecha_a:%Y%m%d}_vs_{fecha_b:%Y%m%d}_{ts}.xlsx"
    )
    out_path = base / fname

    generar_excel(filas, fecha_a, fecha_b, out_path)

    print("=" * 60)
    print(f"Fecha A: {fecha_a:%Y-%m-%d}  |  Fecha B: {fecha_b:%Y-%m-%d}")
    print(f"Colegios (nodos): {len(filas)}")
    print(f"Excel: {out_path}")
    if errores:
        print("\nAvisos API:")
        for e in errores[:25]:
            print(f"  - {e}")
        if len(errores) > 25:
            print(f"  ... y {len(errores) - 25} mas")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
